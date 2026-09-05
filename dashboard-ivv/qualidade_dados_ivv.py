import pandas as pd
import numpy as np
from datetime import datetime
import os
from tkinter import Tk, filedialog, messagebox
import tkinter as tk
from typing import Dict, List, Tuple, Any
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

class IVVQualityControl:
    def __init__(self):
        # Variação % mínima para ALERTA de VALOR_MEDIO_M2 (mês atual vs anterior)
        self.valor_m2_variation_threshold = 30
        # Variação % mínima para classificar como CRÍTICO (provável erro de digitação)
        self.valor_m2_critical_threshold = 80
        # Definir valores válidos para cada coluna - CORRIGIDO
        self.valid_values = {
            'ORIGEM_RECURSOS': ['Condomínio', 'Cooperativa', 'Finan. Bancário', 'MCMV', 'Próprio'],
            'ESTAGIO_OBRA': ['Planta', 'Fundação', 'Estrutura', 'Acabamento', 'Pronto'],
            'OFERTA_VENDA': ['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS', 'VENDIDOS', 
                           'VENDIDOS - LANCADOS E VENDIDOS', 'DISTRATO'],
            'BAIRRO': ['Águas Claras', 'Asa Norte', 'Asa Sul', 'Ceilândia', 'Gama', 'Guará', 
                      'Jardim Botânico', 'Lago Norte', 'Lago Sul', 'Noroeste', 'Park Sul', 'Planaltina', 
                      'Recanto das Emas', 'Samambaia', 'Santa Maria', 'Sobradinho', 
                      'Sobradinho II', 'Sudoeste', 'Taguatinga', 'SIG'],
            'QTD_QUARTOS': ['1', '2', '3', '4+']
        }

        # Definir colunas obrigatórias para cada aba (QTD_QUARTOS removido)
        self.required_columns = {
            'residencial': [
                'ANO_MES', 'ORIGEM_RECURSOS', 'ESTAGIO_OBRA', 'OFERTA_VENDA', 'BAIRRO',
                'AREA', 'QUANTIDADE', 'QTD_ELEVADORES', 'QTD_GARAGEM',
                'TEMPO_FINANCIAMENTO', 'VALOR_MEDIO_M2', 'AREA_QUANTIDADE', 'AREA_VALOR',
                'AREA_QUANTIDADE_VALOR', 'EMPRESA'
            ],
            'comercial': [
                'ANO_MES', 'ORIGEM_RECURSOS', 'ESTAGIO_OBRA', 'OFERTA_VENDA', 'BAIRRO',
                'AREA', 'QUANTIDADE', 'QTD_GARAGEM', 'QTD_ELEVADORES', 'TEMPO_FINANCIAMENTO',
                'VALOR_MEDIO_M2', 'AREA_QUANTIDADE', 'AREA_VALOR', 'AREA_QUANTIDADE_VALOR',
                'EMPRESA'
            ]
        }

        # Definir colunas numéricas
        self.numeric_columns = [
            'ANO_MES', 'AREA', 'QUANTIDADE', 'QTD_ELEVADORES', 'QTD_GARAGEM',
            'TEMPO_FINANCIAMENTO', 'VALOR_MEDIO_M2', 'AREA_QUANTIDADE', 
            'AREA_VALOR', 'AREA_QUANTIDADE_VALOR'
        ]

        # Mapeamentos para categorização de OFERTA_VENDA
        self.oferta_mapping = {
            'OFERTADOS DISPONIVEIS': 'OFERTA_DISPONIVEL',
            'OFERTADOS LANCAMENTOS': 'OFERTADOS_LANCAMENTOS', 
            'VENDIDOS': 'VENDA',
            'VENDIDOS - LANCADOS E VENDIDOS': 'VENDA',
            'DISTRATO': 'DISTRATO'
        }

        self.oferta_total_mapping = {
            'OFERTADOS DISPONIVEIS': 'OFERTA_TOTAL',
            'OFERTADOS LANCAMENTOS': 'OFERTA_TOTAL', 
            'VENDIDOS': 'VENDA',
            'VENDIDOS - LANCADOS E VENDIDOS': 'VENDA',
            'DISTRATO': 'DISTRATO'
        }

    # === NORMALIZAÇÃO DE QTD_QUARTOS PARA ANÁLISE DE ÁREA ===
    def normalize_quartos(self, x):
        try:
            x_str = str(x).strip()
            if x_str.replace('.0', '').isdigit():
                n = int(float(x_str))
                return "4+" if n >= 4 else str(n)
            if x_str == "4+":
                return "4+"
            return None
        except Exception:
            return None


    def select_file(self) -> str:
        """Abre janela para seleção do arquivo Excel"""
        root = Tk()
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title="Selecione o arquivo Excel da pesquisa IVV",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        
        root.destroy()
        return file_path


    def validate_ano_mes(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida o formato AAAAMM da coluna ANO_MES - VERSÃO MELHORADA"""
        errors = []
        
        if 'ANO_MES' not in df.columns:
            return errors
        
        # Garantir que seja string para validação
        df_copy = df.copy()
        df_copy['ANO_MES_STR'] = df_copy['ANO_MES'].astype(str)
        
        # Verificar formato AAAAMM (6 dígitos)
        invalid_format = ~df_copy['ANO_MES_STR'].str.match(r'^\d{6}$', na=False) & df_copy['ANO_MES'].notna()
        invalid_rows = df_copy[invalid_format].index.tolist()
        
        if invalid_rows:
            invalid_values = df_copy[invalid_format]['ANO_MES_STR'].unique()
            error_msg = f"Aba {sheet_name}: ANO_MES com formato inválido (deve ser AAAAMM) nas linhas: {invalid_rows}"
            error_msg += f"\nValores inválidos encontrados: {list(invalid_values)}"
            errors.append(error_msg)
        
        # Verificar apenas valores com formato válido
        valid_format_mask = df_copy['ANO_MES_STR'].str.match(r'^\d{6}$', na=False)
        valid_data = df_copy[valid_format_mask]
        
        if not valid_data.empty:
            anos = valid_data['ANO_MES_STR'].str[:4].astype(int)
            meses = valid_data['ANO_MES_STR'].str[4:6].astype(int)
            
            # Ampliar faixa de anos válidos
            invalid_years = (anos < 2000) | (anos > 2030)
            invalid_months = (meses < 1) | (meses > 12)
            
            if invalid_years.any():
                invalid_year_rows = valid_data[invalid_years].index.tolist()
                invalid_year_values = valid_data[invalid_years]['ANO_MES_STR'].unique()
                error_msg = f"Aba {sheet_name}: ANO_MES com ano inválido nas linhas: {invalid_year_rows}"
                error_msg += f"\nValores: {list(invalid_year_values)}"
                errors.append(error_msg)
            
            if invalid_months.any():
                invalid_month_rows = valid_data[invalid_months].index.tolist()
                invalid_month_values = valid_data[invalid_months]['ANO_MES_STR'].unique()
                error_msg = f"Aba {sheet_name}: ANO_MES com mês inválido nas linhas: {invalid_month_rows}"
                error_msg += f"\nValores: {list(invalid_month_values)}"
                errors.append(error_msg)
        
        return errors

    def validate_categorical_columns(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida colunas categóricas contra valores esperados"""
        errors = []
        
        for column, valid_vals in self.valid_values.items():
            if column in df.columns:
                # Tratar QTD_QUARTOS especialmente - converter números para strings
                if column == 'QTD_QUARTOS':
                    df[column] = df[column].astype(str)
                    df[column] = df[column].replace({'1.0': '1', '2.0': '2', '3.0': '3', '4.0': '4+', 'nan': None})
                
                # Verificar valores inválidos (excluindo NaN)
                invalid_mask = ~df[column].isin(valid_vals) & df[column].notna()
                
                if invalid_mask.any():
                    invalid_values = df[invalid_mask][column].unique().tolist()
                    invalid_rows = df[invalid_mask].index.tolist()
                    
                    error_msg1 = f"Aba {sheet_name}: Coluna {column} com valores inválidos:"
                    error_msg2 = f"  Valores encontrados: {invalid_values}"
                    error_msg3 = f"  Valores válidos: {valid_vals}"
                    error_msg4 = f"  Linhas com erro: {invalid_rows}"
                    
                    errors.extend([error_msg1, error_msg2, error_msg3, error_msg4])
        
        return errors

    def validate_numeric_columns(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida colunas numéricas"""
        errors = []
        
        for column in self.numeric_columns:
            if column in df.columns:
                # Verificar se valores são numéricos
                non_numeric_mask = pd.to_numeric(df[column], errors='coerce').isna() & df[column].notna()
                
                if non_numeric_mask.any():
                    non_numeric_rows = df[non_numeric_mask].index.tolist()
                    non_numeric_values = df[non_numeric_mask][column].tolist()
                    
                    error_msg1 = f"Aba {sheet_name}: Coluna {column} com valores não numéricos:"
                    error_msg2 = f"  Valores: {non_numeric_values}"
                    error_msg3 = f"  Linhas: {non_numeric_rows}"
                    
                    errors.extend([error_msg1, error_msg2, error_msg3])
                
                # Verificar valores negativos (exceto TEMPO_FINANCIAMENTO)
                if column not in ['TEMPO_FINANCIAMENTO']:
                    numeric_values = pd.to_numeric(df[column], errors='coerce')
                    negative_mask = (numeric_values < 0) & numeric_values.notna()
                    
                    if negative_mask.any():
                        negative_rows = df[negative_mask].index.tolist()
                        error_msg = f"Aba {sheet_name}: Coluna {column} com valores negativos nas linhas: {negative_rows}"
                        errors.append(error_msg)
        
        return errors

    def check_missing_columns(self, df: pd.DataFrame, sheet_name: str, sheet_type: str) -> List[str]:
        """Verifica colunas obrigatórias faltantes"""
        errors = []
        required = set(self.required_columns[sheet_type])
        present = set(df.columns)
        missing = required - present
        
        if missing:
            error_msg = f"Aba {sheet_name}: Colunas obrigatórias faltantes: {list(missing)}"
            errors.append(error_msg)
        
        return errors

    def check_missing_data(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Verifica dados faltantes em colunas críticas"""
        errors = []
        critical_columns = ['ANO_MES', 'BAIRRO', 'AREA', 'QUANTIDADE', 'EMPRESA']
        
        for column in critical_columns:
            if column in df.columns:
                missing_count = df[column].isna().sum()
                if missing_count > 0:
                    missing_rows = df[df[column].isna()].index.tolist()
                    error_msg = f"Aba {sheet_name}: Coluna crítica {column} com {missing_count} valores faltantes nas linhas: {missing_rows}"
                    errors.append(error_msg)
        
        return errors

    def get_last_two_months(self, df: pd.DataFrame) -> List[int]:
        """Retorna os dois últimos meses (maiores valores de ANO_MES) - VERSÃO CORRIGIDA"""
        if 'ANO_MES' not in df.columns:
            return []
        
        # Converter para string primeiro, depois para int
        df_copy = df.copy()
        df_copy['ANO_MES_STR'] = df_copy['ANO_MES'].astype(str)
        
        # Filtrar apenas valores que parecem ser ANO_MES válido (6 dígitos)
        valid_mask = df_copy['ANO_MES_STR'].str.match(r'^\d{6}$', na=False)
        valid_ano_mes = df_copy[valid_mask]['ANO_MES_STR']
        
        if valid_ano_mes.empty:
            print("AVISO: Nenhum valor válido de ANO_MES encontrado")
            return []
        
        # Converter para int e obter únicos
        try:
            ano_mes_numeric = valid_ano_mes.astype(int)
            unique_months = sorted(ano_mes_numeric.unique(), reverse=True)
            
            result = unique_months[:2] if len(unique_months) >= 2 else unique_months
            return result
        except Exception as e:
            print(f"Erro ao converter ANO_MES para numérico: {e}")
            return []

    def validate_data_consistency_filtered(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida consistência dos dados apenas dos dois últimos meses"""
        errors = []
        
        last_months = self.get_last_two_months(df)
        if not last_months:
            return errors
        
        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        
        if df_filtered.empty:
            return errors
        
        # Verificar se AREA_QUANTIDADE = AREA * QUANTIDADE
        if all(col in df_filtered.columns for col in ['AREA', 'QUANTIDADE', 'AREA_QUANTIDADE']):
            area_num = pd.to_numeric(df_filtered['AREA'], errors='coerce')
            quantidade_num = pd.to_numeric(df_filtered['QUANTIDADE'], errors='coerce')
            area_quantidade_num = pd.to_numeric(df_filtered['AREA_QUANTIDADE'], errors='coerce')
            
            calculated = area_num * quantidade_num
            tolerance = 0.01
            
            inconsistent_mask = (
                (abs(calculated - area_quantidade_num) > tolerance) & 
                area_num.notna() & quantidade_num.notna() & area_quantidade_num.notna()
            )
            
            if inconsistent_mask.any():
                inconsistent_rows = df_filtered[inconsistent_mask].index.tolist()
                period_text = ", ".join([self.format_ano_mes(m) for m in last_months])
                error_msg = f"Aba {sheet_name}: Inconsistência entre AREA * QUANTIDADE ≠ AREA_QUANTIDADE nos últimos dois meses ({period_text}), linhas: {inconsistent_rows}"
                errors.append(error_msg)
        
        return errors

    

    def _REMOVED_compute_area_outliers(self, df: pd.DataFrame, sheet_name: str, n_std: int = None) -> List[Dict[str, Any]]:
        """
        Calcula outliers de AREA usando média + N desvios padrão, considerando apenas os dois
        meses mais recentes. Retorna lista de dicionários com detalhes por linha.
        """
        if 'AREA' not in df.columns or 'ANO_MES' not in df.columns:
            return []

        if n_std is None:
            n_std = getattr(self, "area_outlier_std", 2)

        last_months = self.get_last_two_months(df)
        if not last_months:
            return []

        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        if df_filtered.empty:
            return []

        df_filtered['AREA'] = pd.to_numeric(df_filtered['AREA'], errors='coerce')

        # identificar coluna de empreendimento, se existir
        empreendimento_col = None
        for col in ['EMPREENDIMENTO', 'PROJETO', 'NOME_EMPREENDIMENTO', 'NOME_PROJETO']:
            if col in df_filtered.columns:
                empreendimento_col = col
                break

        outliers: List[Dict[str, Any]] = []

        # Caso residencial: usa QTD_QUARTOS
        if 'QTD_QUARTOS' in df_filtered.columns:
            # normalizar número de quartos
            df_filtered['QTD_QUARTOS_NORM'] = df_filtered['QTD_QUARTOS'].apply(self.normalize_quartos)

            for rooms in ['1', '2', '3', '4+']:
                group = df_filtered[df_filtered['QTD_QUARTOS_NORM'] == rooms]
                if group.empty:
                    continue

                mean_val = group['AREA'].mean()
                std_val = group['AREA'].std()

                if pd.isna(std_val) or std_val == 0:
                    continue

                limit = mean_val + n_std * std_val
                mask = group['AREA'] > limit
                out_df = group[mask]

                for idx, row in out_df.iterrows():
                    outliers.append({
                        'sheet': sheet_name,
                        'index': int(idx),
                        'empresa': row.get('EMPRESA'),
                        'empreendimento': row.get(empreendimento_col, None),
                        'bairro': row.get('BAIRRO'),
                        'ano_mes': row.get('ANO_MES'),
                        'oferta_venda': row.get('OFERTA_VENDA'),
                        'quartos': row.get('QTD_QUARTOS'),
                        'area_declarada': row.get('AREA'),
                        'limite_area': float(limit),
                        'media': float(mean_val),
                        'desvio_padrao': float(std_val),
                        'valor_m2': row.get('VALOR_MEDIO_M2'),
                    })
        else:
            # Caso comercial: usa toda a base filtrada
            group = df_filtered
            if group['AREA'].notna().sum() == 0:
                return []

            mean_val = group['AREA'].mean()
            std_val = group['AREA'].std()

            if pd.isna(std_val) or std_val == 0:
                return []

            limit = mean_val + n_std * std_val
            mask = group['AREA'] > limit
            out_df = group[mask]

            for idx, row in out_df.iterrows():
                outliers.append({
                    'sheet': sheet_name,
                    'index': int(idx),
                    'empresa': row.get('EMPRESA'),
                    'empreendimento': row.get(empreendimento_col, None),
                    'bairro': row.get('BAIRRO'),
                    'ano_mes': row.get('ANO_MES'),
                    'oferta_venda': row.get('OFERTA_VENDA'),
                    'area_declarada': row.get('AREA'),
                    'limite_area': float(limit),
                    'media': float(mean_val),
                    'desvio_padrao': float(std_val),
                    'valor_m2': row.get('VALOR_MEDIO_M2'),
                })

        return outliers

    def _REMOVED_validate_area_outliers_std(self, df: pd.DataFrame, sheet_name: str, n_std: int = None) -> List[str]:
        """
        Gera mensagens de erro resumidas sobre outliers de AREA para uso no relatório.
        """
        outliers = self._compute_area_outliers(df, sheet_name, n_std=n_std)
        if not outliers:
            return []

        if n_std is None:
            n_std = getattr(self, "area_outlier_std", 2)

        indices = sorted({o['index'] for o in outliers})
        msg = (
            f"Aba {sheet_name}: {len(outliers)} registro(s) com AREA acima do limite "
            f"média + {n_std} desvio(s) padrão nos dois últimos meses. "
            f"Linhas (índices do arquivo original): {indices}"
        )
        return [msg]


    def _to_numeric_ptbr(self, series: pd.Series) -> pd.Series:
        """Converte série numérica com possíveis formatos PT-BR (1.234,56) para float."""
        if series is None:
            return pd.Series(dtype=float)
        if series.dtype == object:
            s = series.astype(str).str.strip()
            # remover separadores de milhar e converter vírgula decimal
            s = s.str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
            return pd.to_numeric(s, errors='coerce')
        return pd.to_numeric(series, errors='coerce')

    def _compute_valor_m2_variation(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Detecta variações suspeitas de VALOR_MEDIO_M2 comparando os dois meses mais recentes,
        tipologia a tipologia (comparação direta, sem média).

        Chave residencial: EMPREENDIMENTO + QTD_QUARTOS + OFERTA_VENDA + AREA + QTD_GARAGEM
        Chave comercial:   EMPREENDIMENTO + OFERTA_VENDA + AREA + QTD_GARAGEM

        Só compara tipologias presentes nos dois meses com VALOR_MEDIO_M2 > 0.
        Retorna lista de dicionários ordenada por |variação%| decrescente.
        """
        if 'VALOR_MEDIO_M2' not in df.columns or 'ANO_MES' not in df.columns:
            return []

        last_months = self.get_last_two_months(df)
        if len(last_months) < 2:
            return []

        m_atual, m_ant = last_months[0], last_months[1]
        df2 = df[df['ANO_MES'].isin([m_atual, m_ant])].copy()
        df2['VM2'] = self._to_numeric_ptbr(df2['VALOR_MEDIO_M2'])

        if 'QTD_QUARTOS' in df2.columns:
            df2['QTD_QUARTOS'] = df2['QTD_QUARTOS'].astype(str).str.replace('.0', '', regex=False)
            keys = ['EMPREENDIMENTO', 'QTD_QUARTOS', 'OFERTA_VENDA', 'AREA', 'QTD_GARAGEM', 'EMPRESA', 'BAIRRO']
        else:
            keys = ['EMPREENDIMENTO', 'OFERTA_VENDA', 'AREA', 'QTD_GARAGEM', 'EMPRESA', 'BAIRRO']

        keys = [k for k in keys if k in df2.columns]
        df_pos = df2[df2['VM2'] > 0].copy()

        results: List[Dict[str, Any]] = []

        for mes, grupo in [(m_ant, 'ANT'), (m_atual, 'ATUAL')]:
            pass  # só para estruturar — loop real abaixo

        # Indexar cada mês separadamente preservando o índice original
        ant = df_pos[df_pos['ANO_MES'] == m_ant].set_index(keys)[['VM2']].rename(columns={'VM2': 'VM2_ANT'})
        atual = df_pos[df_pos['ANO_MES'] == m_atual].set_index(keys)[['VM2']].rename(columns={'VM2': 'VM2_ATUAL'})

        # Preservar índice original para referência de linha
        ant_idx = df_pos[df_pos['ANO_MES'] == m_ant].copy()
        ant_idx['_KEY'] = list(zip(*[ant_idx[k] for k in keys]))
        atual_idx = df_pos[df_pos['ANO_MES'] == m_atual].copy()
        atual_idx['_KEY'] = list(zip(*[atual_idx[k] for k in keys]))

        ant_line = ant_idx.set_index('_KEY').index
        atual_line = atual_idx.set_index('_KEY').index

        # Linha original no arquivo por chave
        ant_line_map = ant_idx.set_index('_KEY')['VM2']  # não usar — usar index do df original
        ant_orig_idx   = ant_idx.reset_index().groupby('_KEY')['index'].first()
        atual_orig_idx = atual_idx.reset_index().groupby('_KEY')['index'].first()

        # Join: só tipologias presentes nos dois meses
        try:
            merged = ant.join(atual, how='inner')
        except Exception:
            return []

        if merged.empty:
            return []

        var_threshold = getattr(self, 'valor_m2_variation_threshold', 30)
        crit_threshold = getattr(self, 'valor_m2_critical_threshold', 80)

        for idx_tuple, row in merged.iterrows():
            val_ant = row['VM2_ANT']
            val_atual = row['VM2_ATUAL']

            if pd.isna(val_ant) or pd.isna(val_atual) or val_ant == 0:
                continue

            var_pct = (val_atual - val_ant) / val_ant * 100
            var_abs = abs(var_pct)

            if var_abs < var_threshold:
                continue

            nivel = 'CRÍTICO' if var_abs >= crit_threshold else 'ALERTA'

            # Montar dict de campos a partir da chave
            if not isinstance(idx_tuple, tuple):
                idx_tuple = (idx_tuple,)
            campos = dict(zip(keys, idx_tuple))

            # Recuperar índice original (número de linha no arquivo)
            # +2: índice pandas base-0, linha 1 do Excel = cabeçalho
            _i_ant   = int(ant_orig_idx.get(idx_tuple,   -1))
            _i_atual = int(atual_orig_idx.get(idx_tuple, -1))
            linha_ant   = _i_ant   + 2 if _i_ant   >= 0 else -1
            linha_atual = _i_atual + 2 if _i_atual >= 0 else -1

            item: Dict[str, Any] = {
                'nivel': nivel,
                'var_pct': round(var_pct, 1),
                'var_abs': round(var_abs, 1),
                'empresa': campos.get('EMPRESA', ''),
                'empreendimento': campos.get('EMPREENDIMENTO', ''),
                'bairro': campos.get('BAIRRO', ''),
                'qtd_quartos': campos.get('QTD_QUARTOS', None),
                'area': campos.get('AREA', ''),
                'qtd_garagem': campos.get('QTD_GARAGEM', ''),
                'oferta_venda': campos.get('OFERTA_VENDA', ''),
                'mes_ant': m_ant,
                'mes_atual': m_atual,
                'linha_ant': linha_ant,
                'valor_ant': round(val_ant, 2),
                'linha_atual': linha_atual,
                'valor_atual': round(val_atual, 2),
            }
            results.append(item)

        results.sort(key=lambda x: (-({'CRÍTICO': 1, 'ALERTA': 0}[x['nivel']]), -x['var_abs']))
        return results

    def validate_valor_m2_variation(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Gera mensagens de erro resumidas sobre variações de VALOR_MEDIO_M2."""
        casos = self._compute_valor_m2_variation(df, sheet_name)
        if not casos:
            return []

        criticos = [c for c in casos if c['nivel'] == 'CRÍTICO']
        alertas  = [c for c in casos if c['nivel'] == 'ALERTA']
        msgs = []

        if criticos:
            linhas = [c['linha_atual'] for c in criticos]
            msgs.append(
                f"Aba {sheet_name}: {len(criticos)} caso(s) CRÍTICO(S) de VALOR_MEDIO_M2 "
                f"(variação ≥ {getattr(self,'valor_m2_critical_threshold',80)}% em relação ao mês anterior). "
                f"Linhas no arquivo: {linhas}"
            )
        if alertas:
            linhas = [c['linha_atual'] for c in alertas]
            msgs.append(
                f"Aba {sheet_name}: {len(alertas)} caso(s) de ALERTA de VALOR_MEDIO_M2 "
                f"(variação entre {getattr(self,'valor_m2_variation_threshold',30)}% e "
                f"{getattr(self,'valor_m2_critical_threshold',80)-1}% em relação ao mês anterior). "
                f"Linhas no arquivo: {linhas}"
            )
        return msgs

    # ── SITUAÇÃO 1: SALDO NEGATIVO ────────────────────────────────────────────

    def _compute_saldo_negativo(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Detecta tipologias onde OFERTA - VENDA + DISTRATO < 0 nos dois últimos meses.
        Chave residencial: EMPREENDIMENTO + QTD_QUARTOS + AREA + QTD_GARAGEM + EMPRESA + BAIRRO
        Chave comercial:   EMPREENDIMENTO + AREA + QTD_GARAGEM + EMPRESA + BAIRRO
        """
        required = {'ANO_MES', 'OFERTA_VENDA', 'QUANTIDADE', 'EMPRESA', 'BAIRRO', 'AREA'}
        if not required.issubset(df.columns):
            return []

        last_months = self.get_last_two_months(df)
        if not last_months:
            return []

        df2 = df[df['ANO_MES'].isin(last_months)].copy()
        df2['QUANTIDADE'] = pd.to_numeric(df2['QUANTIDADE'], errors='coerce').fillna(0)

        has_quartos = 'QTD_QUARTOS' in df2.columns
        if has_quartos:
            df2['QTD_QUARTOS'] = df2['QTD_QUARTOS'].astype(str).str.replace('.0', '', regex=False)
            grp_keys = ['EMPREENDIMENTO', 'QTD_QUARTOS', 'AREA', 'QTD_GARAGEM', 'EMPRESA', 'BAIRRO']
        else:
            grp_keys = ['EMPREENDIMENTO', 'AREA', 'QTD_GARAGEM', 'EMPRESA', 'BAIRRO']
        grp_keys = [k for k in grp_keys if k in df2.columns]

        results: List[Dict[str, Any]] = []

        for mes in last_months:
            sub = df2[df2['ANO_MES'] == mes]

            oferta   = sub[sub['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])].groupby(grp_keys)['QUANTIDADE'].sum()
            venda    = sub[sub['OFERTA_VENDA'].isin(['VENDIDOS', 'VENDIDOS - LANCADOS E VENDIDOS'])].groupby(grp_keys)['QUANTIDADE'].sum()
            distrato = sub[sub['OFERTA_VENDA'] == 'DISTRATO'].groupby(grp_keys)['QUANTIDADE'].sum()

            chk = pd.DataFrame({'OFERTA': oferta, 'VENDA': venda, 'DISTRATO': distrato}).fillna(0)
            chk['SALDO'] = chk['OFERTA'] - chk['VENDA'] + chk['DISTRATO']
            neg = chk[chk['SALDO'] < 0].reset_index()

            # Índice original da linha de venda (mais provável de conter o erro)
            venda_sub = sub[sub['OFERTA_VENDA'].isin(['VENDIDOS', 'VENDIDOS - LANCADOS E VENDIDOS'])]
            linha_venda_map = venda_sub.reset_index().groupby(grp_keys)['index'].first()

            for _, row in neg.iterrows():
                key_tuple = tuple(row[k] for k in grp_keys)
                linha_ref = int(linha_venda_map.get(key_tuple, -1))
                if linha_ref >= 0:
                    linha_ref += 2  # ajuste: índice pandas base-0 + cabeçalho Excel

                item: Dict[str, Any] = {
                    'mes':          mes,
                    'mes_fmt':      self.format_ano_mes(mes),
                    'empresa':      row.get('EMPRESA', ''),
                    'empreendimento': row.get('EMPREENDIMENTO', ''),
                    'bairro':       row.get('BAIRRO', ''),
                    'qtd_quartos':  row.get('QTD_QUARTOS', None),
                    'area':         row.get('AREA', ''),
                    'qtd_garagem':  row.get('QTD_GARAGEM', ''),
                    'oferta':       int(row['OFERTA']),
                    'venda':        int(row['VENDA']),
                    'distrato':     int(row['DISTRATO']),
                    'saldo':        int(row['SALDO']),
                    'linha_venda':  linha_ref,
                }
                results.append(item)

        results.sort(key=lambda x: (x['mes'], x['saldo']))
        return results

    def create_saldo_worksheet(self, workbook, casos: List[Dict[str, Any]],
                                sheet_name: str, sheet_type: str, tem_quartos: bool):
        """Cria worksheet de saldo negativo com formato didático."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = workbook.create_sheet(sheet_name)

        FILL_HDR   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        FILL_ROW   = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
        FILL_NONE  = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        F_WHITE    = Font(color='FFFFFF', bold=True, size=10)
        F_RED      = Font(color='CC0000', bold=True, size=10)
        F_TITLE    = Font(bold=True, size=13)
        F_SUB      = Font(italic=True, size=9, color='888888')
        F_NORMAL   = Font(size=10)
        CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT       = Alignment(horizontal='left',   vertical='center')

        meses_presentes = sorted({c['mes'] for c in casos}) if casos else []
        periodo = ' e '.join(self.format_ano_mes(m) for m in meses_presentes) if meses_presentes else 'N/A'

        ws.cell(row=1, column=1, value=f'SALDO NEGATIVO — {sheet_type}   |   Período: {periodo}').font = F_TITLE
        ws.cell(row=2, column=1,
                value='Ocorre quando OFERTA - VENDA + DISTRATO < 0 dentro do mesmo mês. '
                      'Indica que a venda foi descontada no mês errado.').font = F_SUB
        ws.cell(row=3, column=1, value=f'Total de ocorrências: {len(casos)}').font = Font(size=10, bold=True)

        headers = ['MÊS', 'EMPRESA', 'EMPREENDIMENTO', 'BAIRRO']
        if tem_quartos:
            headers.append('QUARTOS')
        headers += ['ÁREA (m²)', 'GARAGEM', 'OFERTA', 'VENDA', 'DISTRATO', 'SALDO', 'Nº LINHA\n(VENDA)']

        for col_i, h in enumerate(headers, start=1):
            c = ws.cell(row=5, column=col_i, value=h)
            c.font = F_WHITE; c.fill = FILL_HDR; c.alignment = CENTER

        if not casos:
            ws.cell(row=6, column=1,
                    value='✓ Nenhum saldo negativo encontrado no período.').font = Font(color='006400', bold=True)
        else:
            for r_i, caso in enumerate(casos, start=6):
                fill = FILL_ROW

                col = 1
                for val, aln in [
                    (caso['mes_fmt'],         CENTER),
                    (caso['empresa'],         LEFT),
                    (caso['empreendimento'],  LEFT),
                    (caso['bairro'],          LEFT),
                ]:
                    c = ws.cell(row=r_i, column=col, value=val)
                    c.fill = fill; c.alignment = aln; c.font = F_NORMAL
                    col += 1

                if tem_quartos:
                    c = ws.cell(row=r_i, column=col, value=caso.get('qtd_quartos', ''))
                    c.fill = fill; c.alignment = CENTER; c.font = F_NORMAL
                    col += 1

                for val, aln in [
                    (caso['area'],     CENTER),
                    (caso['qtd_garagem'] if isinstance(caso['qtd_garagem'], int)
                     else (int(caso['qtd_garagem']) if str(caso['qtd_garagem']).replace('.','').isdigit() else caso['qtd_garagem']),
                     CENTER),
                    (caso['oferta'],   CENTER),
                    (caso['venda'],    CENTER),
                    (caso['distrato'], CENTER),
                ]:
                    c = ws.cell(row=r_i, column=col, value=val)
                    c.fill = fill; c.alignment = aln; c.font = F_NORMAL
                    col += 1

                # Saldo em vermelho negrito
                c = ws.cell(row=r_i, column=col, value=caso['saldo'])
                c.fill = fill; c.alignment = CENTER; c.font = F_RED
                col += 1

                linha_ref = caso['linha_venda']
                c = ws.cell(row=r_i, column=col, value=linha_ref if linha_ref >= 0 else 'N/A')
                c.fill = fill; c.alignment = CENTER; c.font = F_NORMAL

        col_widths = [10, 35, 40, 18]
        if tem_quartos:
            col_widths.append(9)
        col_widths += [10, 9, 9, 9, 10, 9, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[5].height = 32
        ws.freeze_panes = 'A6'

    # ── SITUAÇÃO 2: VALOR VENDA > VALOR OFERTA ────────────────────────────────

    def _compute_valor_ov(self, df: pd.DataFrame, sheet_name: str,
                          threshold_pct: float = 10.0) -> List[Dict[str, Any]]:
        """
        Detecta tipologias onde VM2 de VENDA > VM2 de OFERTA em mais de threshold_pct%,
        dentro do mesmo mês, nos dois últimos meses.
        """
        if 'VALOR_MEDIO_M2' not in df.columns or 'OFERTA_VENDA' not in df.columns:
            return []

        last_months = self.get_last_two_months(df)
        if not last_months:
            return []

        df2 = df[df['ANO_MES'].isin(last_months)].copy()
        df2['VM2'] = self._to_numeric_ptbr(df2['VALOR_MEDIO_M2'])

        has_quartos = 'QTD_QUARTOS' in df2.columns
        if has_quartos:
            df2['QTD_QUARTOS'] = df2['QTD_QUARTOS'].astype(str).str.replace('.0', '', regex=False)
            grp_keys = ['EMPREENDIMENTO', 'QTD_QUARTOS', 'AREA', 'QTD_GARAGEM', 'EMPRESA', 'BAIRRO']
        else:
            grp_keys = ['EMPREENDIMENTO', 'AREA', 'QTD_GARAGEM', 'EMPRESA', 'BAIRRO']
        grp_keys = [k for k in grp_keys if k in df2.columns]

        df_pos = df2[df2['VM2'] > 0]
        results: List[Dict[str, Any]] = []

        for mes in last_months:
            sub = df_pos[df_pos['ANO_MES'] == mes]

            oferta_sub = sub[sub['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])]
            venda_sub  = sub[sub['OFERTA_VENDA'].isin(['VENDIDOS', 'VENDIDOS - LANCADOS E VENDIDOS'])]

            vm2_oferta = oferta_sub.groupby(grp_keys)['VM2'].mean()
            vm2_venda  = venda_sub.groupby(grp_keys)['VM2'].mean()
            linha_venda_map = venda_sub.reset_index().groupby(grp_keys)['index'].first()

            comp = pd.DataFrame({
                'VM2_OFERTA': vm2_oferta,
                'VM2_VENDA':  vm2_venda,
                'LINHA_VENDA': linha_venda_map,
            }).dropna(subset=['VM2_OFERTA', 'VM2_VENDA'])

            comp['DIFF_PCT'] = (comp['VM2_VENDA'] - comp['VM2_OFERTA']) / comp['VM2_OFERTA'] * 100
            anom = comp[comp['DIFF_PCT'] > threshold_pct].reset_index()

            for _, row in anom.iterrows():
                linha_ref = int(row['LINHA_VENDA']) + 2  # ajuste base-0 + cabeçalho

                item: Dict[str, Any] = {
                    'mes':            mes,
                    'mes_fmt':        self.format_ano_mes(mes),
                    'empresa':        row.get('EMPRESA', ''),
                    'empreendimento': row.get('EMPREENDIMENTO', ''),
                    'bairro':         row.get('BAIRRO', ''),
                    'qtd_quartos':    row.get('QTD_QUARTOS', None),
                    'area':           row.get('AREA', ''),
                    'qtd_garagem':    row.get('QTD_GARAGEM', ''),
                    'vm2_oferta':     round(row['VM2_OFERTA'], 2),
                    'vm2_venda':      round(row['VM2_VENDA'], 2),
                    'diff_pct':       round(row['DIFF_PCT'], 1),
                    'linha_venda':    linha_ref,
                }
                results.append(item)

        results.sort(key=lambda x: (-x['diff_pct'], x['mes']))
        return results

    def create_valor_ov_worksheet(self, workbook, casos: List[Dict[str, Any]],
                                   sheet_name: str, sheet_type: str, tem_quartos: bool):
        """Cria worksheet de valor venda > valor oferta com formato didático."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = workbook.create_sheet(sheet_name)

        FILL_HDR  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        FILL_ROW  = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        F_WHITE   = Font(color='FFFFFF', bold=True, size=10)
        F_WARN    = Font(color='7D5800', bold=True, size=10)
        F_TITLE   = Font(bold=True, size=13)
        F_SUB     = Font(italic=True, size=9, color='888888')
        F_NORMAL  = Font(size=10)
        CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT      = Alignment(horizontal='left',   vertical='center')

        meses_presentes = sorted({c['mes'] for c in casos}) if casos else []
        periodo = ' e '.join(self.format_ano_mes(m) for m in meses_presentes) if meses_presentes else 'N/A'

        ws.cell(row=1, column=1,
                value=f'VALOR VENDA > VALOR OFERTA — {sheet_type}   |   Período: {periodo}').font = F_TITLE
        ws.cell(row=2, column=1,
                value=f'Casos onde o VM² médio de VENDIDOS supera o VM² de OFERTA em mais de 10%. '
                      f'Em condições normais o valor de oferta é maior que o de venda.').font = F_SUB
        ws.cell(row=3, column=1, value=f'Total de ocorrências: {len(casos)}').font = Font(size=10, bold=True)

        headers = ['MÊS', 'EMPRESA', 'EMPREENDIMENTO', 'BAIRRO']
        if tem_quartos:
            headers.append('QUARTOS')
        headers += ['ÁREA (m²)', 'GARAGEM', 'VM² OFERTA', 'VM² VENDA', 'DIFERENÇA %', 'Nº LINHA\n(VENDA)']

        for col_i, h in enumerate(headers, start=1):
            c = ws.cell(row=5, column=col_i, value=h)
            c.font = F_WHITE; c.fill = FILL_HDR; c.alignment = CENTER

        if not casos:
            ws.cell(row=6, column=1,
                    value='✓ Nenhuma ocorrência encontrada no período.').font = Font(color='006400', bold=True)
        else:
            for r_i, caso in enumerate(casos, start=6):
                col = 1
                for val, aln in [
                    (caso['mes_fmt'],        CENTER),
                    (caso['empresa'],        LEFT),
                    (caso['empreendimento'], LEFT),
                    (caso['bairro'],         LEFT),
                ]:
                    c = ws.cell(row=r_i, column=col, value=val)
                    c.fill = FILL_ROW; c.alignment = aln; c.font = F_NORMAL
                    col += 1

                if tem_quartos:
                    c = ws.cell(row=r_i, column=col, value=caso.get('qtd_quartos', ''))
                    c.fill = FILL_ROW; c.alignment = CENTER; c.font = F_NORMAL
                    col += 1

                for val, aln in [
                    (caso['area'],    CENTER),
                    (caso['qtd_garagem'] if isinstance(caso['qtd_garagem'], int)
                     else (int(caso['qtd_garagem']) if str(caso['qtd_garagem']).replace('.','').isdigit() else caso['qtd_garagem']),
                     CENTER),
                ]:
                    c = ws.cell(row=r_i, column=col, value=val)
                    c.fill = FILL_ROW; c.alignment = aln; c.font = F_NORMAL
                    col += 1

                # VM² oferta e venda com formato numérico
                for val in [caso['vm2_oferta'], caso['vm2_venda']]:
                    c = ws.cell(row=r_i, column=col, value=val)
                    c.fill = FILL_ROW; c.alignment = CENTER; c.font = F_NORMAL
                    c.number_format = '#,##0.00'
                    col += 1

                # Diferença % em destaque
                c = ws.cell(row=r_i, column=col, value=f"▲ {caso['diff_pct']:.1f}%")
                c.fill = FILL_ROW; c.alignment = CENTER; c.font = F_WARN
                col += 1

                linha_ref = caso['linha_venda']
                c = ws.cell(row=r_i, column=col, value=linha_ref if linha_ref >= 0 else 'N/A')
                c.fill = FILL_ROW; c.alignment = CENTER; c.font = F_NORMAL

        col_widths = [10, 35, 40, 18]
        if tem_quartos:
            col_widths.append(9)
        col_widths += [10, 9, 14, 14, 13, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[5].height = 32
        ws.freeze_panes = 'A6'

    def create_valor_m2_worksheet(self, workbook, casos: List[Dict[str, Any]],
                                   sheet_name: str, sheet_type: str, tem_quartos: bool):
        """
        Cria worksheet de validação de VALOR_MEDIO_M2 com formato didático:
        uma linha por caso suspeito, com referência às linhas do arquivo original.
        """
        from openpyxl.styles import Font, PatternFill, Alignment

        ws = workbook.create_sheet(sheet_name)

        RED_FONT   = Font(bold=True, color='CC0000', size=11)
        RED_FILL   = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
        YEL_FONT   = Font(bold=True, color='7D5800', size=11)
        YEL_FILL   = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
        HEADER_FONT = Font(bold=True, size=10)
        TITLE_FONT  = Font(bold=True, size=13)
        WRAP        = Alignment(vertical='top', wrap_text=True)
        CENTER      = Alignment(horizontal='center', vertical='top')

        if casos:
            m_ant   = casos[0]['mes_ant']
            m_atual = casos[0]['mes_atual']
            periodo = f"{self.format_ano_mes(m_ant)} → {self.format_ano_mes(m_atual)}"
            m_ant_fmt   = self.format_ano_mes(m_ant)
            m_atual_fmt = self.format_ano_mes(m_atual)
        else:
            periodo = 'N/A'
            m_ant_fmt   = 'Mês ant.'
            m_atual_fmt = 'Mês atual'

        # Linha 1: título
        ws.cell(row=1, column=1,
                value=f"VALIDAÇÃO VALOR MÉDIO m² — {sheet_type}   |   Período: {periodo}").font = TITLE_FONT

        # Linha 2: legenda
        ws.cell(row=2, column=1,
                value=f"CRÍTICO = variação ≥ {getattr(self,'valor_m2_critical_threshold',80)}%  |  "
                      f"ALERTA = variação entre {getattr(self,'valor_m2_variation_threshold',30)}% e "
                      f"{getattr(self,'valor_m2_critical_threshold',80)-1}%").font = Font(italic=True, size=9)

        # Linha 3: contadores
        criticos = sum(1 for c in casos if c['nivel'] == 'CRÍTICO')
        alertas  = sum(1 for c in casos if c['nivel'] == 'ALERTA')
        ws.cell(row=3, column=1,
                value=f"Total de ocorrências: {len(casos)}   |   Críticos: {criticos}   |   Alertas: {alertas}")

        # Linha 5: cabeçalhos
        headers = ['NÍVEL', 'EMPRESA', 'EMPREENDIMENTO', 'BAIRRO']
        if tem_quartos:
            headers.append('QUARTOS')
        headers += ['ÁREA (m²)', 'GARAGEM', 'OFERTA_VENDA',
                    f'Nº LINHA\n({m_ant_fmt})',
                    f'VALOR m²\n({m_ant_fmt})',
                    f'Nº LINHA\n({m_atual_fmt})',
                    f'VALOR m²\n({m_atual_fmt})',
                    'VARIAÇÃO %']

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Linhas de dados
        for row_idx, caso in enumerate(casos, start=6):
            is_critico = caso['nivel'] == 'CRÍTICO'
            fill = RED_FILL if is_critico else YEL_FILL
            font_nivel = RED_FONT if is_critico else YEL_FONT
            sinal = '▲' if caso['var_pct'] > 0 else '▼'

            col = 1
            # NÍVEL
            c = ws.cell(row=row_idx, column=col, value=caso['nivel'])
            c.font = font_nivel; c.fill = fill; c.alignment = CENTER
            col += 1
            # EMPRESA
            c = ws.cell(row=row_idx, column=col, value=caso['empresa'])
            c.fill = fill; c.alignment = WRAP
            col += 1
            # EMPREENDIMENTO
            c = ws.cell(row=row_idx, column=col, value=caso['empreendimento'])
            c.fill = fill; c.alignment = WRAP
            col += 1
            # BAIRRO
            c = ws.cell(row=row_idx, column=col, value=caso['bairro'])
            c.fill = fill; c.alignment = WRAP
            col += 1
            # QUARTOS (só residencial)
            if tem_quartos:
                c = ws.cell(row=row_idx, column=col, value=caso.get('qtd_quartos', ''))
                c.fill = fill; c.alignment = CENTER
                col += 1
            # ÁREA
            c = ws.cell(row=row_idx, column=col, value=caso['area'])
            c.fill = fill; c.alignment = CENTER
            col += 1
            # GARAGEM
            garagem_val = caso['qtd_garagem']
            try:
                garagem_val = int(garagem_val)
            except (ValueError, TypeError):
                pass
            c = ws.cell(row=row_idx, column=col, value=garagem_val)
            c.fill = fill; c.alignment = CENTER
            col += 1
            # OFERTA_VENDA
            c = ws.cell(row=row_idx, column=col, value=caso['oferta_venda'])
            c.fill = fill; c.alignment = WRAP
            col += 1
            # Nº LINHA mês anterior
            linha_ant = caso['linha_ant']
            c = ws.cell(row=row_idx, column=col, value=linha_ant if linha_ant >= 0 else 'N/A')
            c.fill = fill; c.alignment = CENTER
            col += 1
            # VALOR m² mês anterior
            c = ws.cell(row=row_idx, column=col, value=caso['valor_ant'])
            c.fill = fill; c.alignment = CENTER
            c.number_format = '#,##0.00'
            col += 1
            # Nº LINHA mês atual
            linha_atual = caso['linha_atual']
            c = ws.cell(row=row_idx, column=col, value=linha_atual if linha_atual >= 0 else 'N/A')
            c.fill = fill; c.alignment = CENTER
            col += 1
            # VALOR m² mês atual
            c = ws.cell(row=row_idx, column=col, value=caso['valor_atual'])
            c.fill = fill; c.alignment = CENTER
            c.number_format = '#,##0.00'
            col += 1
            # VARIAÇÃO %
            c = ws.cell(row=row_idx, column=col,
                        value=f"{sinal} {abs(caso['var_pct']):.1f}%")
            c.font = font_nivel; c.fill = fill; c.alignment = CENTER

        # Mensagem quando não há casos
        if not casos:
            ws.cell(row=6, column=1,
                    value='✓ Nenhuma variação suspeita encontrada entre os dois últimos meses.').font = Font(color='006400', bold=True)

        # Larguras de coluna
        col_widths = [12, 35, 40, 18]
        if tem_quartos:
            col_widths.append(9)
        col_widths += [10, 9, 25, 14, 16, 14, 16, 13]

        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Altura do cabeçalho
        ws.row_dimensions[5].height = 32
        ws.freeze_panes = 'A6'

    def _compute_valor_m2_outliers(self, *args, **kwargs):
        """Método removido — use _compute_valor_m2_variation."""
        return []

    def validate_valor_m2_outliers_std(self, *args, **kwargs):
        """Método removido — use validate_valor_m2_variation."""
        return []
        """
        Calcula outliers de VALOR_MEDIO_M2 usando média + N desvios padrão, considerando apenas os dois
        meses mais recentes. Retorna lista de dicionários com detalhes por linha.

        Regras espelhadas das de AREA:
        - Apenas valores ACIMA do limite (média + N*DP)
        - Residencial: segmentação por QTD_QUARTOS (normalizado)
        - Comercial: base toda (últimos 2 meses)
        """
        if 'VALOR_MEDIO_M2' not in df.columns or 'ANO_MES' not in df.columns:
            return []

        if n_std is None:
            n_std = getattr(self, "valor_m2_outlier_std", 3)

        last_months = self.get_last_two_months(df)
        if not last_months:
            return []

        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        if df_filtered.empty:
            return []

        df_filtered['VALOR_MEDIO_M2_NUM'] = self._to_numeric_ptbr(df_filtered['VALOR_MEDIO_M2'])

        # identificar coluna de empreendimento, se existir
        empreendimento_col = None
        for col in ['EMPREENDIMENTO', 'PROJETO', 'NOME_EMPREENDIMENTO', 'NOME_PROJETO']:
            if col in df_filtered.columns:
                empreendimento_col = col
                break

        outliers: List[Dict[str, Any]] = []

        # Caso residencial: usa QTD_QUARTOS
        if 'QTD_QUARTOS' in df_filtered.columns:
            df_filtered['QTD_QUARTOS_NORM'] = df_filtered['QTD_QUARTOS'].apply(self.normalize_quartos)

            for rooms in ['1', '2', '3', '4+']:
                group = df_filtered[df_filtered['QTD_QUARTOS_NORM'] == rooms]
                if group.empty:
                    continue

                valid = group['VALOR_MEDIO_M2_NUM'].dropna()
                if valid.empty:
                    continue

                mean_val = valid.mean()
                std_val = valid.std()

                if pd.isna(std_val) or std_val == 0:
                    continue

                limit = mean_val + n_std * std_val
                out_df = group[group['VALOR_MEDIO_M2_NUM'] > limit]

                for idx, row in out_df.iterrows():
                    outliers.append({
                        'sheet': sheet_name,
                        'index': int(idx),
                        'empresa': row.get('EMPRESA'),
                        'empreendimento': row.get(empreendimento_col, None),
                        'bairro': row.get('BAIRRO'),
                        'ano_mes': row.get('ANO_MES'),
                        'oferta_venda': row.get('OFERTA_VENDA'),
                        'quartos': row.get('QTD_QUARTOS'),
                        'area_declarada': row.get('AREA', None),
                        'valor_m2_declarado': row.get('VALOR_MEDIO_M2_NUM'),
                        'limite_valor_m2': float(limit),
                        'media': float(mean_val),
                        'desvio_padrao': float(std_val),
                    })
        else:
            # Caso comercial: usa toda a base filtrada
            group = df_filtered
            valid = group['VALOR_MEDIO_M2_NUM'].dropna()
            if valid.empty:
                return []

            mean_val = valid.mean()
            std_val = valid.std()

            if pd.isna(std_val) or std_val == 0:
                return []

            limit = mean_val + n_std * std_val
            out_df = group[group['VALOR_MEDIO_M2_NUM'] > limit]

            for idx, row in out_df.iterrows():
                outliers.append({
                    'sheet': sheet_name,
                    'index': int(idx),
                    'empresa': row.get('EMPRESA'),
                    'empreendimento': row.get(empreendimento_col, None),
                    'bairro': row.get('BAIRRO'),
                    'ano_mes': row.get('ANO_MES'),
                    'oferta_venda': row.get('OFERTA_VENDA'),
                    'area_declarada': row.get('AREA', None),
                    'valor_m2_declarado': row.get('VALOR_MEDIO_M2_NUM'),
                    'limite_valor_m2': float(limit),
                    'media': float(mean_val),
                    'desvio_padrao': float(std_val),
                })

        return outliers

    def validate_valor_m2_outliers_std(self, df: pd.DataFrame, sheet_name: str, n_std: int = None) -> List[str]:
        """Gera mensagens de erro resumidas sobre outliers de VALOR_MEDIO_M2 para uso no relatório."""
        outliers = self._compute_valor_m2_outliers(df, sheet_name, n_std=n_std)
        if not outliers:
            return []

        if n_std is None:
            n_std = getattr(self, "valor_m2_outlier_std", 3)

        indices = sorted({o['index'] for o in outliers})
        msg = (
            f"Aba {sheet_name}: {len(outliers)} registro(s) com VALOR_MEDIO_M2 acima do limite "
            f"média + {n_std} desvio(s) padrão nos dois últimos meses. "
            f"Linhas (índices do arquivo original): {indices}"
        )
        return [msg]

    def generate_summary_statistics(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Gera estatísticas resumo dos dados"""
        stats = {
            'total_rows': len(df),
            'empty_rows': df.isnull().all(axis=1).sum(),
            'duplicate_rows': df.duplicated().sum(),
        }
        
        for col in self.numeric_columns:
            if col in df.columns:
                numeric_vals = pd.to_numeric(df[col], errors='coerce')
                stats[f'{col}_mean'] = numeric_vals.mean()
                stats[f'{col}_median'] = numeric_vals.median()
                stats[f'{col}_min'] = numeric_vals.min()
                stats[f'{col}_max'] = numeric_vals.max()
                stats[f'{col}_null_count'] = numeric_vals.isna().sum()
        
        if 'BAIRRO' in df.columns:
            stats['bairro_distribution'] = df['BAIRRO'].value_counts().to_dict()
        
        return stats

    def format_ano_mes(self, ano_mes: int) -> str:
        """Formata ANO_MES para exibição (YYYY-MM)"""
        ano_mes_str = str(ano_mes)
        if len(ano_mes_str) == 6:
            return f"{ano_mes_str[:4]}-{ano_mes_str[4:6]}"
        return str(ano_mes)

    def identify_inactive_companies_safe(self, pivot_table: pd.DataFrame, months_formatted: List[str], df_filtered: pd.DataFrame) -> List[str]:
        """
        Identifica empresas inativas usando a nova fórmula:
        Se a soma da QUANTIDADE para OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS 
        na variável OFERTA_VENDA for 0 no mês atual
        """
        if len(months_formatted) < 1:
            return []
            
        current_month = months_formatted[0]
        inactive_companies = []
                
        try:
            # Filtrar dados do mês atual
            current_month_data = df_filtered[df_filtered['ANO_MES_FORMATTED'] == current_month]
            
            if current_month_data.empty:
                return []
            
            # Para cada empresa, verificar se tem ofertas no mês atual
            for empresa in df_filtered['EMPRESA'].unique():
                if pd.isna(empresa):
                    continue
                    
                empresa_current = current_month_data[current_month_data['EMPRESA'] == empresa]
                
                # Filtrar apenas OFERTADOS DISPONÍVEIS e OFERTADOS LANÇAMENTOS
                ofertas_atuais = empresa_current[
                    empresa_current['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])
                ]
                
                # Somar quantidade de ofertas
                soma_ofertas_atual = ofertas_atuais['QUANTIDADE'].sum()
                
                # Se soma é 0, empresa está inativa
                if soma_ofertas_atual == 0:
                    inactive_companies.append(str(empresa))
                    
        except Exception as e:
            print(f"Erro ao identificar empresas inativas: {e}")
        
        return inactive_companies

    def analyze_business_logic_with_launches_safe(self, df_filtered: pd.DataFrame, last_months: List[int], months_formatted: List[str]) -> Dict[str, Any]:
        """
        Analisa problemas de lógica usando a nova fórmula:
        Se a soma da QUANTIDADE para OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS no mês atual 
        for DIFERENTE da soma da QUANTIDADE no mês anterior para OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS 
        subtraído de VENDIDOS + VENDIDOS - LANÇADOS E VENDIDOS do mês anterior + DISTRATO do mês anterior
        """
        problems = []
        
        if len(months_formatted) < 2:
            print("AVISO: Menos de 2 meses disponíveis para análise de lógica")
            return {'logic_problems': problems}
        
        current_month = months_formatted[0]
        previous_month = months_formatted[1]
        
        
        try:
            # Separar dados por mês
            current_data = df_filtered[df_filtered['ANO_MES_FORMATTED'] == current_month]
            previous_data = df_filtered[df_filtered['ANO_MES_FORMATTED'] == previous_month]
            
            
            # Para cada empresa
            empresas_analisadas = 0
            for empresa in df_filtered['EMPRESA'].unique():
                if pd.isna(empresa) or empresa == 'TOTAL GERAL':
                    continue
                    
                try:
                    empresa_current = current_data[current_data['EMPRESA'] == empresa]
                    empresa_previous = previous_data[previous_data['EMPRESA'] == empresa]
                    
                    # Se não tem dados em pelo menos um dos meses, pular
                    if empresa_current.empty or empresa_previous.empty:
                        continue
                    
                    empresas_analisadas += 1
                    
                    # OFERTAS ATUAIS (OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS)
                    ofertas_atuais = empresa_current[
                        empresa_current['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])
                    ]['QUANTIDADE'].sum()
                    
                    # OFERTAS ANTERIORES (OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS)
                    ofertas_anteriores = empresa_previous[
                        empresa_previous['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])
                    ]['QUANTIDADE'].sum()
                    
                    # VENDAS ANTERIORES (VENDIDOS + VENDIDOS - LANÇADOS E VENDIDOS)
                    vendas_anteriores = empresa_previous[
                        empresa_previous['OFERTA_VENDA'].isin(['VENDIDOS', 'VENDIDOS - LANCADOS E VENDIDOS'])
                    ]['QUANTIDADE'].sum()
                    
                    # DISTRATOS ANTERIORES
                    distratos_anteriores = empresa_previous[
                        empresa_previous['OFERTA_VENDA'] == 'DISTRATO'
                    ]['QUANTIDADE'].sum()
                    
                    # FÓRMULA: Oferta_Esperada = Oferta_Anterior - Vendas_Anterior + Distratos_Anterior
                    oferta_esperada = ofertas_anteriores - vendas_anteriores + distratos_anteriores
                    
                    # Verificar diferença
                    diferenca = ofertas_atuais - oferta_esperada

                                        
                    # Se há diferença significativa, registrar problema
                    if abs(diferenca) > 1:  # tolerância de 1 unidade
                        problem_info = {
                            'empresa': str(empresa),
                            'ofertas_atuais': float(ofertas_atuais),
                            'oferta_esperada': float(oferta_esperada),
                            'diferenca': float(diferenca),
                            'detalhes': {
                                'ofertas_anteriores': float(ofertas_anteriores),
                                'vendas_anteriores': float(vendas_anteriores),
                                'distratos_anteriores': float(distratos_anteriores)
                            }
                        }
                        problems.append(problem_info)
                        
                    
                except Exception as e:
                    print(f"Erro ao analisar empresa {empresa}: {e}")
                    continue
            
                    
        except Exception as e:
            print(f"Erro geral na análise de lógica: {e}")
            import traceback
            traceback.print_exc()
        
        return {'logic_problems': problems}

    def create_advanced_pivot_table(self, df: pd.DataFrame, sheet_type: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Cria tabela dinâmica avançada com validação de lógica de negócio"""
        required_cols = ['ANO_MES', 'EMPRESA', 'OFERTA_VENDA', 'QUANTIDADE']
        
        if not all(col in df.columns for col in required_cols):
            print(f"AVISO: Colunas necessárias faltantes. Requeridas: {required_cols}")
            return pd.DataFrame(), {}
        
        last_months = self.get_last_two_months(df)
        print(f"Últimos dois meses encontrados: {last_months}")
        
        if len(last_months) < 2:
            print("AVISO: Menos de 2 meses de dados disponíveis")
            return pd.DataFrame(), {}
        
        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        print(f"Dados filtrados: {len(df_filtered)} linhas")
        
        if df_filtered.empty:
            print("AVISO: Dados filtrados estão vazios")
            return pd.DataFrame(), {}
        
        # Mapear categorias
        df_filtered['CATEGORIA_DETALHADA'] = df_filtered['OFERTA_VENDA'].map(self.oferta_mapping)
        df_filtered['CATEGORIA_TOTAL'] = df_filtered['OFERTA_VENDA'].map(self.oferta_total_mapping)
        
        df_filtered['QUANTIDADE'] = pd.to_numeric(df_filtered['QUANTIDADE'], errors='coerce').fillna(0)
        df_filtered['ANO_MES_FORMATTED'] = df_filtered['ANO_MES'].apply(self.format_ano_mes)
        
        print("Categorias encontradas:", df_filtered['CATEGORIA_DETALHADA'].unique())
        
        # Criar tabelas pivot
        grouped_detailed = df_filtered.groupby(['EMPRESA', 'ANO_MES', 'ANO_MES_FORMATTED', 'CATEGORIA_DETALHADA'])['QUANTIDADE'].sum().reset_index()
        
        pivot_detailed = grouped_detailed.pivot_table(
            index='EMPRESA',
            columns=['ANO_MES_FORMATTED', 'CATEGORIA_DETALHADA'],
            values='QUANTIDADE',
            aggfunc='sum',
            fill_value=0
        )
        
        grouped_total = df_filtered.groupby(['EMPRESA', 'ANO_MES', 'ANO_MES_FORMATTED', 'CATEGORIA_TOTAL'])['QUANTIDADE'].sum().reset_index()
        
        pivot_total = grouped_total.pivot_table(
            index='EMPRESA',
            columns=['ANO_MES_FORMATTED', 'CATEGORIA_TOTAL'],
            values='QUANTIDADE',
            aggfunc='sum',
            fill_value=0
        )
        
        months_formatted = sorted([self.format_ano_mes(m) for m in last_months], reverse=True)
        
        # Construir tabela final
        all_companies = set()
        if not pivot_detailed.empty:
            all_companies.update(pivot_detailed.index)
        if not pivot_total.empty:
            all_companies.update(pivot_total.index)
        
        if not all_companies:
            print("AVISO: Nenhuma empresa encontrada nos dados")
            return pd.DataFrame(), {}
        
        final_table = pd.DataFrame(index=sorted(all_companies))
        
        for month in months_formatted:
            # Adicionar colunas com verificação de existência
            col_oferta_total = (month, 'OFERTA_TOTAL')
            if col_oferta_total in pivot_total.columns:
                final_table[col_oferta_total] = pivot_total[col_oferta_total].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_oferta_total] = 0
            
            col_lancamentos = (month, 'OFERTADOS_LANCAMENTOS')
            if col_lancamentos in pivot_detailed.columns:
                final_table[col_lancamentos] = pivot_detailed[col_lancamentos].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_lancamentos] = 0
            
            col_venda = (month, 'VENDA')
            if col_venda in pivot_detailed.columns:
                final_table[col_venda] = pivot_detailed[col_venda].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_venda] = 0
            
            col_distrato = (month, 'DISTRATO')
            if col_distrato in pivot_detailed.columns:
                final_table[col_distrato] = pivot_detailed[col_distrato].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_distrato] = 0
        
        print(f"Tabela final criada com {len(final_table)} empresas e {len(final_table.columns)} colunas")
        
        # Realizar análises com as NOVAS FÓRMULAS
        analysis_results = {}
        try:
            analysis_results = self.analyze_business_logic_with_launches_safe(df_filtered, last_months, months_formatted)
        except Exception as e:
            print(f"Erro na análise de lógica de negócio: {e}")
            analysis_results = {'logic_problems': []}
        
        try:
            launches_analysis = self.analyze_launches_by_company_and_neighborhood_with_empreendimentos(df_filtered, last_months)
            analysis_results['launches_analysis'] = launches_analysis
        except Exception as e:
            print(f"Erro na análise de lançamentos: {e}")
            analysis_results['launches_analysis'] = {'by_company': {}, 'by_neighborhood': {}, 'total_launches': 0}
        
        # USAR A NOVA FÓRMULA para empresas inativas
        inactive_companies = self.identify_inactive_companies_safe(final_table, months_formatted, df_filtered)
        analysis_results['inactive_companies'] = inactive_companies
        analysis_results['current_month'] = months_formatted[0] if months_formatted else None
        
        # Adicionar linha de total
        if not final_table.empty:
            total_row = final_table.sum()
            total_row.name = 'TOTAL GERAL'
            final_table = pd.concat([final_table, total_row.to_frame().T])
        
        return final_table, analysis_results

    def analyze_launches_by_company_and_neighborhood(self, df: pd.DataFrame, last_months: List[int]) -> Dict[str, Any]:
        """Analisa empreendimentos lançados por empresa e bairro"""
        launches = df[
            (df['OFERTA_VENDA'] == 'OFERTADOS LANCAMENTOS') & 
            (df['ANO_MES'].isin(last_months)) &
            (df['QUANTIDADE'] > 0)
        ].copy()
        
        if launches.empty:
            return {
                'by_company': {},
                'by_neighborhood': {},
                'total_launches': 0
            }
        
        launches['ANO_MES_FORMATTED'] = launches['ANO_MES'].apply(self.format_ano_mes)
        launches['QUANTIDADE'] = pd.to_numeric(launches['QUANTIDADE'], errors='coerce').fillna(0)
        
        # Análise por empresa
        by_company = {}
        for empresa in launches['EMPRESA'].unique():
            empresa_launches = launches[launches['EMPRESA'] == empresa]
            
            by_company[empresa] = {
                'total_quantidade': empresa_launches['QUANTIDADE'].sum(),
                'por_mes': {},
                'bairros': list(empresa_launches['BAIRRO'].unique())
            }
            
            for mes in empresa_launches['ANO_MES_FORMATTED'].unique():
                mes_data = empresa_launches[empresa_launches['ANO_MES_FORMATTED'] == mes]
                by_company[empresa]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'bairros': list(mes_data['BAIRRO'].unique())
                }
        
        # Análise por bairro
        by_neighborhood = {}
        for bairro in launches['BAIRRO'].unique():
            bairro_launches = launches[launches['BAIRRO'] == bairro]
            
            by_neighborhood[bairro] = {
                'total_quantidade': bairro_launches['QUANTIDADE'].sum(),
                'empresas': list(bairro_launches['EMPRESA'].unique()),
                'por_mes': {}
            }
            
            for mes in bairro_launches['ANO_MES_FORMATTED'].unique():
                mes_data = bairro_launches[bairro_launches['ANO_MES_FORMATTED'] == mes]
                by_neighborhood[bairro]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'empresas': list(mes_data['EMPRESA'].unique())
                }
        
        return {
            'by_company': by_company,
            'by_neighborhood': by_neighborhood,
            'total_launches': launches['QUANTIDADE'].sum()
        }

    def extract_empreendimento_name(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Extrai e normaliza o nome do empreendimento de forma determinística,
        evitando agrupamentos por similaridade que possam misturar empreendimentos
        entre empresas (ex.: FIRENZE x FIKEI).

        Regras:
        - Correções ortográficas básicas
        - Remoção de prefixos genéricos
        - Remoção de sufixos padronizados (bloco/torre/unidade/fase, etc.)
        - NÃO aplica fuzzy matching / difflib para agrupar termos
        """
        df = df.copy()
        
        # Verificar se a coluna existe (pode ter nomes diferentes)
        empreendimento_col = None
        possible_names = ['EMPREENDIMENTO', 'PROJETO', 'NOME_EMPREENDIMENTO', 'NOME_PROJETO']
        for col_name in possible_names:
            if col_name in df.columns:
                empreendimento_col = col_name
                break
        
        if empreendimento_col is None:
            print("AVISO: Coluna de empreendimento não encontrada")
            df['EMPREENDIMENTO_AGRUPADO'] = 'N/A'
            return df
        
        df['EMPREENDIMENTO_AGRUPADO'] = df[empreendimento_col].fillna('N/A').astype(str)
        
        # === CORREÇÕES ORTOGRÁFICAS CRÍTICAS ===
        df['EMPREENDIMENTO_AGRUPADO'] = (
            df['EMPREENDIMENTO_AGRUPADO']
              .str.replace('EMPPREENDIMENTO', 'EMPREENDIMENTO', case=False, regex=False)
              .str.replace('EMPREEENDIMENTO', 'EMPREENDIMENTO', case=False, regex=False)
              .str.replace('EMPRENDIMENTO', 'EMPREENDIMENTO', case=False, regex=False)
        )
        
        # === REMOÇÃO DE PREFIXOS GENÉRICOS ===
        df['EMPREENDIMENTO_AGRUPADO'] = (
            df['EMPREENDIMENTO_AGRUPADO']
              .str.replace(r'^EMPREENDIMENTO\s+', '', regex=True, case=False)
              .str.replace(r'^RESIDENCIAL\s+', '', regex=True, case=False)
              .str.replace(r'^RES\s+', '', regex=True, case=False)
              .str.strip()
        )
        
        # Normalizar: remover sufixos padronizados (BL A, TORRE 1, etc.)
        patterns_to_remove = [
            r'\s+BL\s+[A-Z0-9]+',
            r'\s+BLOCO\s+[A-Z0-9]+',
            r'\s+TORRE\s+[A-Z0-9]+',
            r'\s+TIPO\b',
            r'\s+APTO\s+[A-Z0-9]+',
            r'\s+APT\s+[A-Z0-9]+',
            r'\s+APARTAMENTO\s+[A-Z0-9]+',
            r'\s+SALA\s+[A-Z0-9]+',
            r'\s+LOJA\s+[A-Z0-9]+',
            r'\s+COBERTURA\b',
            r'\s+GARDEN\b',
            r'\s+DUPLEX\b',
            r'\s+TRIPLEX\b',
            r'\s+[0-9]+Q\b',
            r'\s+[0-9]+\s+QUARTOS?',
            r'\s+COM\s+TERRAÇO',
            r'\s+STUDIO\b',
            r'\s+LOFT\b',
            r'\s+[0-9]+\s+SUÍTES?\b',
            r'\s+D[0-9]+\b',
            r'\s+C[0-9]+\b',
            r'\s+(FASE|ETAPA)\s*\d+\b',
            r'\s+(I|II|III|IV|V|VI|VII|VIII|IX|X)\b$',
        ]
        
        df['TERMO_PRINCIPAL'] = df['EMPREENDIMENTO_AGRUPADO'].str.upper().str.strip()
        for pattern in patterns_to_remove:
            df['TERMO_PRINCIPAL'] = df['TERMO_PRINCIPAL'].str.replace(pattern, '', regex=True).str.strip()
        
        # Criar máscara de códigos sequenciais (EMP_123, etc.)
        df['IS_SEQUENTIAL_CODE'] = df['TERMO_PRINCIPAL'].str.match(r'^EMP_\d+$', na=False)
        real_names_mask = ~df['IS_SEQUENTIAL_CODE'] & (df['TERMO_PRINCIPAL'] != 'N/A')
        if real_names_mask.any():
            df.loc[real_names_mask, 'EMPREENDIMENTO_AGRUPADO'] = df.loc[real_names_mask, 'TERMO_PRINCIPAL']
        else:
            df['EMPREENDIMENTO_AGRUPADO'] = df['TERMO_PRINCIPAL']
        
        df = df.drop(['TERMO_PRINCIPAL', 'IS_SEQUENTIAL_CODE'], axis=1)
        return df
    def terms_are_similar(self, term1, term2, threshold=0.7):
        """
        Verifica se dois termos são similares o suficiente para serem agrupados
        """
        import difflib
        
        # Converter para minúsculas para comparação
        t1 = term1.lower().strip()
        t2 = term2.lower().strip()
        
        # Se um termo está contido no outro
        if t1 in t2 or t2 in t1:
            return True
        
        # Usar ratio de similaridade
        similarity = difflib.SequenceMatcher(None, t1, t2).ratio()
        return similarity >= threshold

    def analyze_launches_by_company_and_neighborhood_with_empreendimentos(self, df: pd.DataFrame, last_months: List[int]) -> Dict[str, Any]:
        """
        Analisa lançamentos residenciais por empresa, bairro e empreendimento.

        ALINHADO AO GERADOR DO DASHBOARD:
        - 'Lançamento' é contado no PRIMEIRO mês (ANO_MES) em que a tríade
          (EMPRESA, BAIRRO, EMPREENDIMENTO_AGRUPADO) aparece como 'OFERTADOS LANCAMENTOS'.
        - Fases/edições posteriores são ignoradas na contagem de lançamentos.

        last_months:
        - A saída permanece focada nos 'últimos meses' (ex.: últimos 2 meses),
          MAS a identificação do primeiro mês é feita olhando TODO o histórico do df.
        """
        # Extrair/normalizar empreendimentos (determinístico)
        df_with_empreendimentos = self.extract_empreendimento_name(df)
        
        # Filtrar linhas de lançamentos no histórico completo (para achar o PRIMEIRO mês)
        launches_hist = df_with_empreendimentos[
            (df_with_empreendimentos['OFERTA_VENDA'] == 'OFERTADOS LANCAMENTOS') &
            (df_with_empreendimentos['QUANTIDADE'] > 0)
        ].copy()
        
        if launches_hist.empty:
            return {
                'by_company': {},
                'by_neighborhood': {},
                'by_empreendimento': {},
                'total_launches': 0
            }
        
        # Garantir tipos
        launches_hist['ANO_MES'] = pd.to_numeric(launches_hist['ANO_MES'], errors='coerce')
        launches_hist = launches_hist.dropna(subset=['ANO_MES'])
        launches_hist['ANO_MES'] = launches_hist['ANO_MES'].astype(int)
        launches_hist['QUANTIDADE'] = pd.to_numeric(launches_hist['QUANTIDADE'], errors='coerce').fillna(0)
        
        key_cols = ['EMPRESA', 'BAIRRO', 'EMPREENDIMENTO_AGRUPADO']
        
        # Primeiro mês por tríade
        first_month = (
            launches_hist.groupby(key_cols)['ANO_MES']
            .min()
            .reset_index()
            .rename(columns={'ANO_MES': 'ANO_MES_PRIMEIRO'})
        )
        
        # Manter apenas empreendimentos cujo PRIMEIRO mês está em last_months
        first_month_filtered = first_month[first_month['ANO_MES_PRIMEIRO'].isin(last_months)].copy()
        if first_month_filtered.empty:
            return {
                'by_company': {},
                'by_neighborhood': {},
                'by_empreendimento': {},
                'total_launches': 0
            }
        
        # Trazer quantidade do primeiro mês (somar linhas daquele mês)
        launches_first = launches_hist.merge(first_month_filtered, on=key_cols, how='inner')
        launches_first = launches_first[launches_first['ANO_MES'] == launches_first['ANO_MES_PRIMEIRO']].copy()
        
        launches_first['ANO_MES_FORMATTED'] = launches_first['ANO_MES_PRIMEIRO'].apply(self.format_ano_mes)
        
        # Agregar para garantir 1 linha por tríade/mês
        agg = (
            launches_first.groupby(key_cols + ['ANO_MES_PRIMEIRO', 'ANO_MES_FORMATTED'])['QUANTIDADE']
            .sum()
            .reset_index()
        )
        
        total_launches = agg['QUANTIDADE'].sum()
        
        # =========================
        # Análise por empresa
        # =========================
        by_company: Dict[str, Any] = {}
        for empresa in agg['EMPRESA'].unique():
            if pd.isna(empresa):
                continue
            emp_launches = agg[agg['EMPRESA'] == empresa]
            by_company[empresa] = {
                'total_quantidade': emp_launches['QUANTIDADE'].sum(),
                'bairros': list(emp_launches['BAIRRO'].unique()),
                'empreendimentos': list(emp_launches['EMPREENDIMENTO_AGRUPADO'].unique()),
                'por_mes': {}
            }
            for mes in emp_launches['ANO_MES_FORMATTED'].unique():
                mes_data = emp_launches[emp_launches['ANO_MES_FORMATTED'] == mes]
                by_company[empresa]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'bairros': list(mes_data['BAIRRO'].unique()),
                    'empreendimentos': list(mes_data['EMPREENDIMENTO_AGRUPADO'].unique()),
                }
        
        # =========================
        # Análise por bairro
        # =========================
        by_neighborhood: Dict[str, Any] = {}
        for bairro in agg['BAIRRO'].unique():
            if pd.isna(bairro):
                continue
            bairro_launches = agg[agg['BAIRRO'] == bairro]
            by_neighborhood[bairro] = {
                'total_quantidade': bairro_launches['QUANTIDADE'].sum(),
                'empresas': list(bairro_launches['EMPRESA'].unique()),
                'empreendimentos': list(bairro_launches['EMPREENDIMENTO_AGRUPADO'].unique()),
                'por_mes': {}
            }
            for mes in bairro_launches['ANO_MES_FORMATTED'].unique():
                mes_data = bairro_launches[bairro_launches['ANO_MES_FORMATTED'] == mes]
                by_neighborhood[bairro]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'empresas': list(mes_data['EMPRESA'].unique()),
                    'empreendimentos': list(mes_data['EMPREENDIMENTO_AGRUPADO'].unique()),
                }
        
        # =========================
        # Análise por empreendimento
        # =========================
        by_empreendimento: Dict[str, Any] = {}
        for emp in agg['EMPREENDIMENTO_AGRUPADO'].unique():
            if pd.isna(emp) or emp == 'N/A':
                continue
            emp_launches = agg[agg['EMPREENDIMENTO_AGRUPADO'] == emp]
            by_empreendimento[emp] = {
                'total_quantidade': emp_launches['QUANTIDADE'].sum(),
                'empresas': list(emp_launches['EMPRESA'].unique()),
                'bairros': list(emp_launches['BAIRRO'].unique()),
                'por_mes': {}
            }
            for mes in emp_launches['ANO_MES_FORMATTED'].unique():
                mes_data = emp_launches[emp_launches['ANO_MES_FORMATTED'] == mes]
                by_empreendimento[emp]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'empresas': list(mes_data['EMPRESA'].unique()),
                    'bairros': list(mes_data['BAIRRO'].unique()),
                }
        
        return {
            'by_company': by_company,
            'by_neighborhood': by_neighborhood,
            'by_empreendimento': by_empreendimento,
            'total_launches': total_launches,
        }


    def create_launches_worksheet_with_empreendimentos(self, workbook, launches_data, sheet_name, sheet_type):
        """
        Cria worksheet de lançamentos: uma linha por lançamento, com cor por empresa.
        Colunas: MÊS | EMPREENDIMENTO | EMPRESA | BAIRRO | UNIDADES
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            ws = workbook.create_sheet(sheet_name)

            # Paleta de cores pastel por empresa (10 tons)
            PALETA = [
                'DDEEFF', 'D5F5E3', 'FFE8CC', 'FCE4EC', 'EDE7F6',
                'FFF9C4', 'E0F2F1', 'F3E5F5', 'E8EAF6', 'FBE9E7',
            ]
            FILL_HDR   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            FILL_TOTAL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            F_WHITE    = Font(color='FFFFFF', bold=True, size=10)
            F_TITLE    = Font(bold=True, size=13)
            F_SUB      = Font(bold=True, size=11)
            F_NOTE     = Font(italic=True, size=9, color='888888')
            F_NORMAL   = Font(size=10)
            CENTER     = Alignment(horizontal='center', vertical='center')
            LEFT       = Alignment(horizontal='left',   vertical='center')

            total_launches    = launches_data.get('total_launches', 0)
            by_empreendimento = launches_data.get('by_empreendimento', {}) or {}

            # ── Título ────────────────────────────────────────────────────────
            ws.cell(row=1, column=1, value=f'LANÇAMENTOS — {sheet_type}').font = F_TITLE
            ws.cell(row=2, column=1, value=f'Total: {total_launches} unidades lançadas').font = F_SUB
            ws.cell(row=3, column=1,
                    value='Contabiliza apenas o primeiro mês de lançamento de cada empreendimento').font = F_NOTE

            # ── Cabeçalhos ────────────────────────────────────────────────────
            headers = ['MÊS', 'EMPREENDIMENTO', 'EMPRESA', 'BAIRRO', 'UNIDADES']
            for col_i, h in enumerate(headers, start=1):
                c = ws.cell(row=5, column=col_i, value=h)
                c.font = F_WHITE; c.fill = FILL_HDR; c.alignment = CENTER

            if not by_empreendimento:
                ws.cell(row=6, column=1,
                        value='Nenhum lançamento encontrado no período.').font = Font(italic=True)
                return

            # ── Achatar dados: 1 linha por (empreendimento, mês) ─────────────
            rows = []
            for emp_name, emp_data in by_empreendimento.items():
                for mes_key, mes_data in emp_data.get('por_mes', {}).items():
                    empresa_str = ', '.join(mes_data.get('empresas', emp_data.get('empresas', [])))
                    bairro_str  = ', '.join(mes_data.get('bairros',  emp_data.get('bairros',  [])))
                    rows.append({
                        'mes':            mes_key,
                        'empreendimento': emp_name,
                        'empresa':        empresa_str,
                        'bairro':         bairro_str,
                        'unidades':       mes_data.get('quantidade', 0),
                    })

            def _mes_int(mes_key):
                s = ''.join(ch for ch in str(mes_key) if ch.isdigit())[:6]
                try: return int(s)
                except: return 0

            rows.sort(key=lambda r: (_mes_int(r['mes']), r['empresa'], r['empreendimento']))

            # ── Cor por empresa ───────────────────────────────────────────────
            empresas_ord = []
            for r in rows:
                if r['empresa'] not in empresas_ord:
                    empresas_ord.append(r['empresa'])
            cor_emp = {
                emp: PatternFill(start_color=PALETA[i % len(PALETA)],
                                 end_color=PALETA[i % len(PALETA)],
                                 fill_type='solid')
                for i, emp in enumerate(empresas_ord)
            }

            # ── Escrita das linhas ────────────────────────────────────────────
            for r_i, r in enumerate(rows, start=6):
                fill = cor_emp.get(r['empresa'],
                                   PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'))

                mes_fmt = r['mes']
                s = ''.join(ch for ch in str(mes_fmt) if ch.isdigit())
                if len(s) == 6:
                    mes_fmt = f"{s[:4]}-{s[4:6]}"

                valores = [mes_fmt, r['empreendimento'], r['empresa'], r['bairro'], r['unidades']]
                aligns  = [CENTER, LEFT, LEFT, LEFT, CENTER]

                for col_i, (val, aln) in enumerate(zip(valores, aligns), start=1):
                    c = ws.cell(row=r_i, column=col_i, value=val)
                    c.fill = fill; c.alignment = aln; c.font = F_NORMAL

            # ── Linha de total ────────────────────────────────────────────────
            tot_row = 6 + len(rows)
            ws.cell(row=tot_row, column=1, value='TOTAL').font = F_WHITE
            ws.cell(row=tot_row, column=1).fill = FILL_TOTAL
            ws.cell(row=tot_row, column=1).alignment = CENTER
            ws.cell(row=tot_row, column=5, value=total_launches).font = F_WHITE
            ws.cell(row=tot_row, column=5).fill = FILL_TOTAL
            ws.cell(row=tot_row, column=5).alignment = CENTER
            for col_i in [2, 3, 4]:
                ws.cell(row=tot_row, column=col_i).fill = FILL_TOTAL

            # ── Larguras e ajustes ────────────────────────────────────────────
            for col_i, w in enumerate([10, 45, 45, 18, 10], start=1):
                ws.column_dimensions[get_column_letter(col_i)].width = w
            ws.row_dimensions[5].height = 22
            ws.freeze_panes = 'A6'

        except Exception as e:
            print(f"Erro ao criar worksheet de lançamentos: {e}")
            import traceback
            traceback.print_exc()

    def create_summary_tables_excel(self, file_path: str, df_residencial: pd.DataFrame, 
                                  df_comercial: pd.DataFrame, residencial_name: str, 
                                  comercial_name: str) -> str:
        """Cria arquivo Excel com tabelas resumo e análise de lançamentos"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)
        output_path = os.path.join(directory, f"{base_name}_resumo_{timestamp}.xlsx")
        
        print(f"Iniciando criação do arquivo: {output_path}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Criar tabela para residencial
            print("Processando dados residenciais...")
            pivot_residencial, analysis_res = self.create_advanced_pivot_table(df_residencial, 'residencial')
            
            if not pivot_residencial.empty:
                print(f"Tabela residencial criada com {len(pivot_residencial)} linhas")
                
                ws_res = wb.create_sheet('Resumo_Residencial')
                
                last_months = self.get_last_two_months(df_residencial)
                period_text = ", ".join([self.format_ano_mes(m) for m in last_months]) if last_months else "N/A"
                
                header_font = Font(bold=True, size=12)
                subheader_font = Font(bold=True, size=10)
                
                ws_res['A1'] = f"RESUMO RESIDENCIAL - Período: {period_text}"
                ws_res['A1'].font = header_font
                
                ws_res['A2'] = "Empresa x Mês x Tipo de Oferta/Venda (Soma de QUANTIDADE)"
                ws_res['A2'].font = subheader_font
                
                ws_res['A3'] = "OFERTA_TOTAL = OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS"
                ws_res['A3'].font = subheader_font
                
                self.insert_pivot_table_with_formatting(ws_res, pivot_residencial, analysis_res, start_row=5)
                
                print("Dados residenciais escritos no Excel com formatação")
            else:
                print("AVISO: Tabela residencial está vazia")
            
            # Criar tabela para comercial
            print("Processando dados comerciais...")
            pivot_comercial, analysis_com = self.create_advanced_pivot_table(df_comercial, 'comercial')
            
            if not pivot_comercial.empty:
                print(f"Tabela comercial criada com {len(pivot_comercial)} linhas")
                
                ws_com = wb.create_sheet('Resumo_Comercial')
                
                last_months = self.get_last_two_months(df_comercial)
                period_text = ", ".join([self.format_ano_mes(m) for m in last_months]) if last_months else "N/A"
                
                ws_com['A1'] = f"RESUMO COMERCIAL - Período: {period_text}"
                ws_com['A1'].font = header_font
                
                ws_com['A2'] = "Empresa x Mês x Tipo de Oferta/Venda (Soma de QUANTIDADE)"
                ws_com['A2'].font = subheader_font
                
                ws_com['A3'] = "OFERTA_TOTAL = OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS"
                ws_com['A3'].font = subheader_font
                
                self.insert_pivot_table_with_formatting(ws_com, pivot_comercial, analysis_com, start_row=5)
                
                print("Dados comerciais escritos no Excel com formatação")
            else:
                print("AVISO: Tabela comercial está vazia")
            

            # Criar abas de validação de VALOR_MEDIO_M2 (variação mês a mês)
            casos_valor_res = self._compute_valor_m2_variation(df_residencial, residencial_name)
            print(f"Criando aba ValorM2_Residencial ({len(casos_valor_res)} caso(s))...")
            self.create_valor_m2_worksheet(wb, casos_valor_res, 'ValorM2_Residencial',
                                           'RESIDENCIAL', tem_quartos=True)

            casos_valor_com = self._compute_valor_m2_variation(df_comercial, comercial_name)
            print(f"Criando aba ValorM2_Comercial ({len(casos_valor_com)} caso(s))...")
            self.create_valor_m2_worksheet(wb, casos_valor_com, 'ValorM2_Comercial',
                                           'COMERCIAL', tem_quartos=False)

            # Criar abas de saldo negativo
            casos_saldo_res = self._compute_saldo_negativo(df_residencial, residencial_name)
            print(f"Criando aba Saldo_Residencial ({len(casos_saldo_res)} caso(s))...")
            self.create_saldo_worksheet(wb, casos_saldo_res, 'Saldo_Residencial',
                                        'RESIDENCIAL', tem_quartos=True)

            casos_saldo_com = self._compute_saldo_negativo(df_comercial, comercial_name)
            print(f"Criando aba Saldo_Comercial ({len(casos_saldo_com)} caso(s))...")
            self.create_saldo_worksheet(wb, casos_saldo_com, 'Saldo_Comercial',
                                        'COMERCIAL', tem_quartos=False)

            # Criar abas de valor venda > valor oferta
            casos_ov_res = self._compute_valor_ov(df_residencial, residencial_name)
            print(f"Criando aba ValorOV_Residencial ({len(casos_ov_res)} caso(s))...")
            self.create_valor_ov_worksheet(wb, casos_ov_res, 'ValorOV_Residencial',
                                           'RESIDENCIAL', tem_quartos=True)

            casos_ov_com = self._compute_valor_ov(df_comercial, comercial_name)
            print(f"Criando aba ValorOV_Comercial ({len(casos_ov_com)} caso(s))...")
            self.create_valor_ov_worksheet(wb, casos_ov_com, 'ValorOV_Comercial',
                                           'COMERCIAL', tem_quartos=False)


            # Criar abas de lançamentos sempre (exibe mensagem quando não há dados)
            if 'launches_analysis' in analysis_res:
                print("Criando aba de lançamentos residencial...")
                self.create_launches_worksheet_with_empreendimentos(wb, analysis_res['launches_analysis'], 'Lancamentos_Residencial', 'RESIDENCIAL')
                print("Lançamentos residenciais adicionados")
            
            if 'launches_analysis' in analysis_com:
                print("Criando aba de lançamentos comercial...")
                self.create_launches_worksheet_with_empreendimentos(wb, analysis_com['launches_analysis'], 'Lancamentos_Comercial', 'COMERCIAL')
                print("Lançamentos comerciais adicionados")
            
            wb.save(output_path)
            wb.close()
            
            print(f"Arquivo salvo: {output_path}")
            
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"✅ Arquivo criado com sucesso! Tamanho: {file_size} bytes")
                return output_path
            else:
                print("❌ ERRO: Arquivo não foi encontrado após criação")
                return ""
                
        except Exception as e:
            print(f"❌ ERRO ao criar arquivo Excel: {str(e)}")
            import traceback
            print("Stack trace completo:")
            traceback.print_exc()
            return ""

    def insert_pivot_table_with_formatting(self, worksheet, pivot_table, analysis_data, start_row=5):
        """Insere tabela pivot com layout visual melhorado: STATUS, cabeçalho em 2 níveis e cores por linha."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            import pandas as pd

            # ── Paleta ──────────────────────────────────────────────────────────
            F_WHITE  = Font(color='FFFFFF', bold=True, size=10)
            F_OK     = Font(color='1A6B3A', bold=True, size=10)
            F_DIV    = Font(color='7D5800', bold=True, size=10)
            F_INAT   = Font(color='CC0000', bold=True, size=10)
            F_NEW    = Font(color='666666', bold=True, size=10)
            F_NORMAL = Font(size=10)
            F_LEGEND = Font(size=9, italic=True)

            FILL_HDR_DARK  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            FILL_HDR_MED   = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            FILL_HDR_LIGHT = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            FILL_OK        = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            FILL_DIV       = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
            FILL_INAT      = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
            FILL_NEW       = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            FILL_TOTAL     = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

            CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
            LEFT   = Alignment(horizontal='left',   vertical='center')

            # ── Dados de análise ─────────────────────────────────────────────────
            inactive_set   = {str(c).strip() for c in analysis_data.get('inactive_companies', [])}
            problem_set    = {str(p['empresa']).strip() for p in analysis_data.get('logic_problems', [])}

            # ── Estrutura de colunas ─────────────────────────────────────────────
            months = sorted(
                {col[0] for col in pivot_table.columns if isinstance(col, tuple)},
                reverse=True
            )
            metrics_order  = ['OFERTA_TOTAL', 'OFERTADOS_LANCAMENTOS', 'VENDA', 'DISTRATO']
            metric_labels  = {'OFERTA_TOTAL': 'OFERTA', 'OFERTADOS_LANCAMENTOS': 'LANÇAM.',
                              'VENDA': 'VENDAS', 'DISTRATO': 'DISTRATO'}
            n_met    = len(metrics_order)
            COL_EMP  = 1
            COL_STA  = 2
            COL_DATA = 3
            total_cols = COL_DATA + len(months) * n_met - 1

            # ── Legenda (start_row) ──────────────────────────────────────────────
            leg_row = start_row
            legenda = [
                ('✓ OK',            F_OK),
                ('⚠ DIVERGÊNCIA',   F_DIV),
                ('✗ AUSENTE',       F_INAT),
                ('○ SEM HISTÓRICO', F_NEW),
            ]
            for i, (txt, fnt) in enumerate(legenda):
                c = worksheet.cell(row=leg_row, column=1 + i * 2, value=txt)
                c.font = fnt

            # ── Cabeçalho de meses (start_row+1) ────────────────────────────────
            mhdr = start_row + 1

            for col, lbl in [(COL_EMP, 'EMPRESA'), (COL_STA, 'STATUS')]:
                c = worksheet.cell(row=mhdr, column=col, value=lbl)
                c.font = F_WHITE; c.fill = FILL_HDR_MED; c.alignment = CENTER
                worksheet.merge_cells(
                    start_row=mhdr, start_column=col,
                    end_row=mhdr + 1, end_column=col
                )

            for m_i, month in enumerate(months):
                cs = COL_DATA + m_i * n_met
                ce = cs + n_met - 1
                worksheet.merge_cells(start_row=mhdr, start_column=cs,
                                      end_row=mhdr,   end_column=ce)
                c = worksheet.cell(row=mhdr, column=cs, value=month)
                c.font = F_WHITE; c.fill = FILL_HDR_MED; c.alignment = CENTER

            # ── Sub-cabeçalhos de métricas (start_row+2) ────────────────────────
            shdr = start_row + 2
            for col in [COL_EMP, COL_STA]:
                worksheet.cell(row=shdr, column=col).fill = FILL_HDR_LIGHT

            for m_i, month in enumerate(months):
                for mt_i, metric in enumerate(metrics_order):
                    col = COL_DATA + m_i * n_met + mt_i
                    c = worksheet.cell(row=shdr, column=col,
                                       value=metric_labels.get(metric, metric))
                    c.font = Font(bold=True, size=9, color='1F4E79')
                    c.fill = FILL_HDR_LIGHT; c.alignment = CENTER

            # ── Linhas de dados (start_row+3 em diante) ──────────────────────────
            data_row = start_row + 3

            for r_i, (empresa, row_data) in enumerate(pivot_table.iterrows()):
                cur = data_row + r_i
                emp = str(empresa).strip()
                is_total = (emp == 'TOTAL GERAL')

                if is_total:
                    fill = FILL_TOTAL
                    f_emp = f_sta = f_dat = F_WHITE
                    sta_txt = ''
                else:
                    is_inat = emp in inactive_set
                    is_div  = emp in problem_set
                    is_new  = False
                    if len(months) > 1:
                        prev = months[1]
                        prev_vals = [row_data.get((prev, m), 0) for m in metrics_order
                                     if (prev, m) in pivot_table.columns]
                        is_new = all(v == 0 for v in prev_vals)

                    if is_inat:
                        fill = FILL_INAT; f_emp = f_sta = F_INAT
                        sta_txt = '✗ AUSENTE'
                    elif is_div:
                        fill = FILL_DIV; f_emp = f_sta = F_DIV
                        sta_txt = '⚠ DIVERGÊNCIA'
                    elif is_new:
                        fill = FILL_NEW; f_emp = f_sta = F_NEW
                        sta_txt = '○ SEM HISTÓRICO'
                    else:
                        fill = FILL_OK; f_emp = f_sta = F_OK
                        sta_txt = '✓ OK'
                    f_dat = F_NORMAL

                c = worksheet.cell(row=cur, column=COL_EMP, value=emp)
                c.font = f_emp; c.fill = fill; c.alignment = LEFT

                c = worksheet.cell(row=cur, column=COL_STA, value=sta_txt)
                c.font = f_sta; c.fill = fill; c.alignment = CENTER

                for m_i, month in enumerate(months):
                    for mt_i, metric in enumerate(metrics_order):
                        col = COL_DATA + m_i * n_met + mt_i
                        val = row_data.get((month, metric), 0)
                        try:
                            val = int(float(val)) if pd.notna(val) else 0
                        except (ValueError, TypeError):
                            val = 0
                        c = worksheet.cell(row=cur, column=col, value=val)
                        c.fill = fill; c.font = f_dat if not is_total else F_WHITE
                        c.alignment = CENTER

            # ── Larguras ─────────────────────────────────────────────────────────
            worksheet.column_dimensions[get_column_letter(COL_EMP)].width = 40
            worksheet.column_dimensions[get_column_letter(COL_STA)].width = 17
            for col in range(COL_DATA, total_cols + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 11

            worksheet.row_dimensions[mhdr].height = 22
            worksheet.row_dimensions[shdr].height = 22
            worksheet.freeze_panes = worksheet.cell(row=data_row, column=COL_DATA)

        except Exception as e:
            print(f"Erro ao inserir tabela com formatação: {e}")
            import traceback
            traceback.print_exc()

    def create_launches_worksheet(self, workbook, launches_data, sheet_name, sheet_type):
        """Cria worksheet de lançamentos com formatação adequada"""
        try:
            from openpyxl.styles import Font, Alignment
            
            ws = workbook.create_sheet(sheet_name)
            
            title_font = Font(bold=True, size=14)
            section_font = Font(bold=True, size=12)
            header_font = Font(bold=True, size=10)
            
            row = 1
            
            ws.cell(row=row, column=1, value=f"ANÁLISE DE LANÇAMENTOS - {sheet_type}").font = title_font
            row += 2
            
            total_launches = launches_data.get('total_launches', 0)
            cell = ws.cell(row=row, column=1, value=f"TOTAL DE LANÇAMENTOS: {total_launches} unidades")
            cell.font = section_font
            row += 2
            
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR EMPRESA:").font = section_font
            row += 1
            
            by_company = launches_data.get('by_company', {})
            if by_company:
                ws.cell(row=row, column=1, value="Empresa").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Bairros").font = header_font
                ws.cell(row=row, column=4, value="Detalhes por Mês").font = header_font
                row += 1
                
                for empresa, data in by_company.items():
                    ws.cell(row=row, column=1, value=str(empresa))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['bairros']))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        bairros_mes = ', '.join(mes_data['bairros'])
                        detalhe = f"{mes}: {mes_data['quantidade']} un. ({bairros_mes})"
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=4, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum lançamento encontrado")
                row += 1
            
            row += 1
            
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR BAIRRO:").font = section_font
            row += 1
            
            by_neighborhood = launches_data.get('by_neighborhood', {})
            if by_neighborhood:
                ws.cell(row=row, column=1, value="Bairro").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Empresas").font = header_font
                ws.cell(row=row, column=4, value="Detalhes por Mês").font = header_font
                row += 1
                
                for bairro, data in by_neighborhood.items():
                    ws.cell(row=row, column=1, value=str(bairro))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['empresas']))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        empresas_mes = ', '.join(mes_data['empresas'])
                        detalhe = f"{mes}: {mes_data['quantidade']} un. ({empresas_mes})"
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=4, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum lançamento encontrado por bairro")
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 60
            
        except Exception as e:
            print(f"Erro ao criar worksheet de lançamentos: {e}")

    def process_file(self, file_path: str) -> Tuple[List[str], Dict[str, Any], str]:
        """Processa o arquivo Excel e retorna erros, estatísticas e caminho do resumo"""
        errors = []
        statistics = {}
        summary_path = ""
        
        try:
            print("Iniciando processamento do arquivo...")
            
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            print(f"Abas encontradas: {sheet_names}")
            
            if len(sheet_names) < 2:
                errors.append("Arquivo deve conter pelo menos 2 abas (residencial e comercial)")
                return errors, statistics, summary_path
            
            # Processar aba residencial
            residencial_name = sheet_names[0]
            print(f"Processando aba residencial: {residencial_name}")
            df_residencial = pd.read_excel(file_path, sheet_name=residencial_name)
            print(f"Dados residenciais: {len(df_residencial)} linhas")
            
            errors.extend(self.check_missing_columns(df_residencial, residencial_name, 'residencial'))
            errors.extend(self.validate_ano_mes(df_residencial, residencial_name))
            errors.extend(self.validate_categorical_columns(df_residencial, residencial_name))
            errors.extend(self.validate_numeric_columns(df_residencial, residencial_name))
            errors.extend(self.check_missing_data(df_residencial, residencial_name))
            errors.extend(self.validate_data_consistency_filtered(df_residencial, residencial_name))
            errors.extend(self.validate_valor_m2_variation(df_residencial, residencial_name))
            statistics[f'residencial_{residencial_name}'] = self.generate_summary_statistics(df_residencial, residencial_name)
            
            # Processar aba comercial
            comercial_name = sheet_names[1]
            print(f"Processando aba comercial: {comercial_name}")
            df_comercial = pd.read_excel(file_path, sheet_name=comercial_name)
            print(f"Dados comerciais: {len(df_comercial)} linhas")
            
            errors.extend(self.check_missing_columns(df_comercial, comercial_name, 'comercial'))
            errors.extend(self.validate_ano_mes(df_comercial, comercial_name))
            errors.extend(self.validate_categorical_columns(df_comercial, comercial_name))
            errors.extend(self.validate_numeric_columns(df_comercial, comercial_name))
            errors.extend(self.check_missing_data(df_comercial, comercial_name))
            errors.extend(self.validate_data_consistency_filtered(df_comercial, comercial_name))
            errors.extend(self.validate_valor_m2_variation(df_comercial, comercial_name))
            statistics[f'comercial_{comercial_name}'] = self.generate_summary_statistics(df_comercial, comercial_name)
            
            print("Validações concluídas. Iniciando criação do arquivo de resumo...")
            
            summary_path = self.create_summary_tables_excel(
                file_path, df_residencial, df_comercial, 
                residencial_name, comercial_name
            )
            
            if summary_path:
                print(f"✅ Arquivo de resumo criado: {summary_path}")
            else:
                print("❌ Falha na criação do arquivo de resumo")
            
        except Exception as e:
            error_msg = f"Erro ao processar arquivo: {str(e)}"
            errors.append(error_msg)
            print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
        
        return errors, statistics, summary_path

    def run(self):
        """Executa o controle de qualidade"""
        print("=== CONTROLE DE QUALIDADE - PESQUISA IVV ===")
        print("Selecione o arquivo Excel para análise...\n")
        
        file_path = self.select_file()
        
        if not file_path:
            print("Nenhum arquivo selecionado. Operação cancelada.")
            return
        
        print(f"Processando arquivo: {os.path.basename(file_path)}")
        
        errors, statistics, summary_path = self.process_file(file_path)
        
        print("\n" + "=" * 50)
        print("RESULTADOS DO CONTROLE DE QUALIDADE")
        print("=" * 50)
        
        if errors:
            print(f"❌ {len(errors)} erro(s) encontrado(s):")
            for i, error in enumerate(errors[:10], 1):
                print(f"  {i}. {error}")
            if len(errors) > 10:
                print(f"  ... e mais {len(errors) - 10} erro(s)")
        else:
            print("✅ Nenhum erro encontrado! Arquivo válido.")
        
        if summary_path:
            print(f"\n📊 Tabelas resumo geradas em:")
            print(f"   {summary_path}")
            print(f"\n🎨 Formatação aplicada:")
            print(f"   • Empresas inativas (mês atual): texto em VERMELHO")
            print(f"   • Problemas de lógica de oferta: fundo AMARELO")
            print(f"   • Lógica: Oferta(atual) = Oferta(anterior) - Venda(anterior) + Distrato(anterior)")
            print(f"\n📋 Abas geradas:")
            print(f"   • Resumo_Residencial / Resumo_Comercial: pivot por empresa e mês")
            print(f"   • ValorM2_Residencial / ValorM2_Comercial: variações suspeitas de VALOR_MEDIO_M2")
            print(f"     (CRÍTICO ≥{self.valor_m2_critical_threshold}%  |  ALERTA ≥{self.valor_m2_variation_threshold}%)")
            print(f"   • Lancamentos_Residencial / Lancamentos_Comercial: detalhes por empreendimento")
        
        root = tk.Tk()
        root.withdraw()
        
        message = "Processamento concluído!\n\n"
        
        if summary_path:
            message += f"📊 Tabelas resumo: {os.path.basename(summary_path)}\n"
            message += "\n🎨 Formatação condicional aplicada:\n"
            message += "• Empresas inativas: texto VERMELHO\n"
            message += "• Problemas de lógica: fundo AMARELO\n"
            message += f"\n📋 Validação de VALOR_MEDIO_M2:\n"
            message += f"• CRÍTICO ≥{self.valor_m2_critical_threshold}%  |  ALERTA ≥{self.valor_m2_variation_threshold}%\n"
            message += "• Abas ValorM2_Residencial e ValorM2_Comercial\n"
            message += "\n📋 Inclui análise de lançamentos:\n"
            message += "• Por empresa e por bairro\n"
            message += "• Detalhamento por mês"
        
        if errors:
            message += f"\n\n⚠️ {len(errors)} erro(s) encontrado(s)"
            messagebox.showwarning("Controle de Qualidade", message)
        else:
            message += "\n\n✅ Nenhum erro encontrado!"
            messagebox.showinfo("Controle de Qualidade", message)
        
        root.destroy()

if __name__ == "__main__":
    controller = IVVQualityControl()
    controller.run()
