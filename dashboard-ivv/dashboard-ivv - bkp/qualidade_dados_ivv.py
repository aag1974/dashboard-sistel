import pandas as pd
import numpy as np
from datetime import datetime
import os
from tkinter import Tk, filedialog, messagebox
import tkinter as tk
from typing import Dict, List, Tuple, Any
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

class IVVQualityControl:
    def __init__(self):
        # Número de desvios padrão configurável para outliers de AREA
        self.area_outlier_std = 3

        # Definir valores válidos para cada coluna - CORRIGIDO
        self.valid_values = {
            'ORIGEM_RECURSOS': ['Condomínio', 'Cooperativa', 'Finan. Bancário', 'MCMV', 'Próprio'],
            'ESTAGIO_OBRA': ['Planta', 'Fundação', 'Estrutura', 'Acabamento', 'Pronto'],
            'OFERTA_VENDA': ['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS', 'VENDIDOS', 
                           'VENDIDOS - LANCADOS E VENDIDOS', 'DISTRATO'],
            'BAIRRO': ['Águas Claras', 'Asa Norte', 'Asa Sul', 'Ceilândia', 'Gama', 'Guará', 
                      'Jardim Botânico', 'Lago Norte', 'Noroeste', 'Park Sul', 'Planaltina', 
                      'Recanto das Emas', 'Samambaia', 'Santa Maria', 'Sobradinho', 
                      'Sobradinho II', 'Sudoeste', 'Taguatinga', 'SIG'],
            'QTD_QUARTOS': ['1', '2', '3', '4+']
        }

        # Definir colunas obrigatórias para cada aba (QTD_QUARTOS removido)
        self.required_columns = {
            'residencial': [
                'ANO_MES', 'ORIGEM_RECURSOS', 'ESTAGIO_OBRA', 'OFERTA_VENDA', 'BAIRRO',
                'AREA', 'QUANTIDADE', 'QTD_ELEVADORES', 'QTD_GARAGEM',
                'TEMPO_FINANCIAMENTO', 'VALOR_MEDIO_M2', 'AREA_QUANTIDADE', 'AREA_VALOR',
                'AREA_QUANTIDADE_VALOR', 'EMPRESA'
            ],
            'comercial': [
                'ANO_MES', 'ORIGEM_RECURSOS', 'ESTAGIO_OBRA', 'OFERTA_VENDA', 'BAIRRO',
                'AREA', 'QUANTIDADE', 'QTD_GARAGEM', 'QTD_ELEVADORES', 'TEMPO_FINANCIAMENTO',
                'VALOR_MEDIO_M2', 'AREA_QUANTIDADE', 'AREA_VALOR', 'AREA_QUANTIDADE_VALOR',
                'EMPRESA'
            ]
        }

        # Definir colunas numéricas
        self.numeric_columns = [
            'ANO_MES', 'AREA', 'QUANTIDADE', 'QTD_ELEVADORES', 'QTD_GARAGEM',
            'TEMPO_FINANCIAMENTO', 'VALOR_MEDIO_M2', 'AREA_QUANTIDADE', 
            'AREA_VALOR', 'AREA_QUANTIDADE_VALOR'
        ]

        # Mapeamentos para categorização de OFERTA_VENDA
        self.oferta_mapping = {
            'OFERTADOS DISPONIVEIS': 'OFERTA_DISPONIVEL',
            'OFERTADOS LANCAMENTOS': 'OFERTADOS_LANCAMENTOS', 
            'VENDIDOS': 'VENDA',
            'VENDIDOS - LANCADOS E VENDIDOS': 'VENDA',
            'DISTRATO': 'DISTRATO'
        }

        self.oferta_total_mapping = {
            'OFERTADOS DISPONIVEIS': 'OFERTA_TOTAL',
            'OFERTADOS LANCAMENTOS': 'OFERTA_TOTAL', 
            'VENDIDOS': 'VENDA',
            'VENDIDOS - LANCADOS E VENDIDOS': 'VENDA',
            'DISTRATO': 'DISTRATO'
        }

    # === NORMALIZAÇÃO DE QTD_QUARTOS PARA ANÁLISE DE ÁREA ===
    def normalize_quartos(self, x):
        try:
            x_str = str(x).strip()
            if x_str.replace('.0', '').isdigit():
                n = int(float(x_str))
                return "4+" if n >= 4 else str(n)
            if x_str == "4+":
                return "4+"
            return None
        except Exception:
            return None


    def select_file(self) -> str:
        """Abre janela para seleção do arquivo Excel"""
        root = Tk()
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title="Selecione o arquivo Excel da pesquisa IVV",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        
        root.destroy()
        return file_path


    def validate_ano_mes(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida o formato AAAAMM da coluna ANO_MES - VERSÃO MELHORADA"""
        errors = []
        
        if 'ANO_MES' not in df.columns:
            return errors
        
        # Garantir que seja string para validação
        df_copy = df.copy()
        df_copy['ANO_MES_STR'] = df_copy['ANO_MES'].astype(str)
        
        # Verificar formato AAAAMM (6 dígitos)
        invalid_format = ~df_copy['ANO_MES_STR'].str.match(r'^\d{6}$', na=False) & df_copy['ANO_MES'].notna()
        invalid_rows = df_copy[invalid_format].index.tolist()
        
        if invalid_rows:
            invalid_values = df_copy[invalid_format]['ANO_MES_STR'].unique()
            error_msg = f"Aba {sheet_name}: ANO_MES com formato inválido (deve ser AAAAMM) nas linhas: {invalid_rows}"
            error_msg += f"\nValores inválidos encontrados: {list(invalid_values)}"
            errors.append(error_msg)
        
        # Verificar apenas valores com formato válido
        valid_format_mask = df_copy['ANO_MES_STR'].str.match(r'^\d{6}$', na=False)
        valid_data = df_copy[valid_format_mask]
        
        if not valid_data.empty:
            anos = valid_data['ANO_MES_STR'].str[:4].astype(int)
            meses = valid_data['ANO_MES_STR'].str[4:6].astype(int)
            
            # Ampliar faixa de anos válidos
            invalid_years = (anos < 2000) | (anos > 2030)
            invalid_months = (meses < 1) | (meses > 12)
            
            if invalid_years.any():
                invalid_year_rows = valid_data[invalid_years].index.tolist()
                invalid_year_values = valid_data[invalid_years]['ANO_MES_STR'].unique()
                error_msg = f"Aba {sheet_name}: ANO_MES com ano inválido nas linhas: {invalid_year_rows}"
                error_msg += f"\nValores: {list(invalid_year_values)}"
                errors.append(error_msg)
            
            if invalid_months.any():
                invalid_month_rows = valid_data[invalid_months].index.tolist()
                invalid_month_values = valid_data[invalid_months]['ANO_MES_STR'].unique()
                error_msg = f"Aba {sheet_name}: ANO_MES com mês inválido nas linhas: {invalid_month_rows}"
                error_msg += f"\nValores: {list(invalid_month_values)}"
                errors.append(error_msg)
        
        return errors

    def validate_categorical_columns(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida colunas categóricas contra valores esperados"""
        errors = []
        
        for column, valid_vals in self.valid_values.items():
            if column in df.columns:
                # Tratar QTD_QUARTOS especialmente - converter números para strings
                if column == 'QTD_QUARTOS':
                    df[column] = df[column].astype(str)
                    df[column] = df[column].replace({'1.0': '1', '2.0': '2', '3.0': '3', '4.0': '4+', 'nan': None})
                
                # Verificar valores inválidos (excluindo NaN)
                invalid_mask = ~df[column].isin(valid_vals) & df[column].notna()
                
                if invalid_mask.any():
                    invalid_values = df[invalid_mask][column].unique().tolist()
                    invalid_rows = df[invalid_mask].index.tolist()
                    
                    error_msg1 = f"Aba {sheet_name}: Coluna {column} com valores inválidos:"
                    error_msg2 = f"  Valores encontrados: {invalid_values}"
                    error_msg3 = f"  Valores válidos: {valid_vals}"
                    error_msg4 = f"  Linhas com erro: {invalid_rows}"
                    
                    errors.extend([error_msg1, error_msg2, error_msg3, error_msg4])
        
        return errors

    def validate_numeric_columns(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida colunas numéricas"""
        errors = []
        
        for column in self.numeric_columns:
            if column in df.columns:
                # Verificar se valores são numéricos
                non_numeric_mask = pd.to_numeric(df[column], errors='coerce').isna() & df[column].notna()
                
                if non_numeric_mask.any():
                    non_numeric_rows = df[non_numeric_mask].index.tolist()
                    non_numeric_values = df[non_numeric_mask][column].tolist()
                    
                    error_msg1 = f"Aba {sheet_name}: Coluna {column} com valores não numéricos:"
                    error_msg2 = f"  Valores: {non_numeric_values}"
                    error_msg3 = f"  Linhas: {non_numeric_rows}"
                    
                    errors.extend([error_msg1, error_msg2, error_msg3])
                
                # Verificar valores negativos (exceto TEMPO_FINANCIAMENTO)
                if column not in ['TEMPO_FINANCIAMENTO']:
                    numeric_values = pd.to_numeric(df[column], errors='coerce')
                    negative_mask = (numeric_values < 0) & numeric_values.notna()
                    
                    if negative_mask.any():
                        negative_rows = df[negative_mask].index.tolist()
                        error_msg = f"Aba {sheet_name}: Coluna {column} com valores negativos nas linhas: {negative_rows}"
                        errors.append(error_msg)
        
        return errors

    def check_missing_columns(self, df: pd.DataFrame, sheet_name: str, sheet_type: str) -> List[str]:
        """Verifica colunas obrigatórias faltantes"""
        errors = []
        required = set(self.required_columns[sheet_type])
        present = set(df.columns)
        missing = required - present
        
        if missing:
            error_msg = f"Aba {sheet_name}: Colunas obrigatórias faltantes: {list(missing)}"
            errors.append(error_msg)
        
        return errors

    def check_missing_data(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Verifica dados faltantes em colunas críticas"""
        errors = []
        critical_columns = ['ANO_MES', 'BAIRRO', 'AREA', 'QUANTIDADE', 'EMPRESA']
        
        for column in critical_columns:
            if column in df.columns:
                missing_count = df[column].isna().sum()
                if missing_count > 0:
                    missing_rows = df[df[column].isna()].index.tolist()
                    error_msg = f"Aba {sheet_name}: Coluna crítica {column} com {missing_count} valores faltantes nas linhas: {missing_rows}"
                    errors.append(error_msg)
        
        return errors

    def get_last_two_months(self, df: pd.DataFrame) -> List[int]:
        """Retorna os dois últimos meses (maiores valores de ANO_MES) - VERSÃO CORRIGIDA"""
        if 'ANO_MES' not in df.columns:
            return []
        
        # Converter para string primeiro, depois para int
        df_copy = df.copy()
        df_copy['ANO_MES_STR'] = df_copy['ANO_MES'].astype(str)
        
        # Filtrar apenas valores que parecem ser ANO_MES válido (6 dígitos)
        valid_mask = df_copy['ANO_MES_STR'].str.match(r'^\d{6}$', na=False)
        valid_ano_mes = df_copy[valid_mask]['ANO_MES_STR']
        
        if valid_ano_mes.empty:
            print("AVISO: Nenhum valor válido de ANO_MES encontrado")
            return []
        
        # Converter para int e obter únicos
        try:
            ano_mes_numeric = valid_ano_mes.astype(int)
            unique_months = sorted(ano_mes_numeric.unique(), reverse=True)
            
            result = unique_months[:2] if len(unique_months) >= 2 else unique_months
            return result
        except Exception as e:
            print(f"Erro ao converter ANO_MES para numérico: {e}")
            return []

    def validate_data_consistency_filtered(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Valida consistência dos dados apenas dos dois últimos meses"""
        errors = []
        
        last_months = self.get_last_two_months(df)
        if not last_months:
            return errors
        
        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        
        if df_filtered.empty:
            return errors
        
        # Verificar se AREA_QUANTIDADE = AREA * QUANTIDADE
        if all(col in df_filtered.columns for col in ['AREA', 'QUANTIDADE', 'AREA_QUANTIDADE']):
            area_num = pd.to_numeric(df_filtered['AREA'], errors='coerce')
            quantidade_num = pd.to_numeric(df_filtered['QUANTIDADE'], errors='coerce')
            area_quantidade_num = pd.to_numeric(df_filtered['AREA_QUANTIDADE'], errors='coerce')
            
            calculated = area_num * quantidade_num
            tolerance = 0.01
            
            inconsistent_mask = (
                (abs(calculated - area_quantidade_num) > tolerance) & 
                area_num.notna() & quantidade_num.notna() & area_quantidade_num.notna()
            )
            
            if inconsistent_mask.any():
                inconsistent_rows = df_filtered[inconsistent_mask].index.tolist()
                period_text = ", ".join([self.format_ano_mes(m) for m in last_months])
                error_msg = f"Aba {sheet_name}: Inconsistência entre AREA * QUANTIDADE ≠ AREA_QUANTIDADE nos últimos dois meses ({period_text}), linhas: {inconsistent_rows}"
                errors.append(error_msg)
        
        return errors

    

    def _compute_area_outliers(self, df: pd.DataFrame, sheet_name: str, n_std: int = None) -> List[Dict[str, Any]]:
        """
        Calcula outliers de AREA usando média + N desvios padrão, considerando apenas os dois
        meses mais recentes. Retorna lista de dicionários com detalhes por linha.
        """
        if 'AREA' not in df.columns or 'ANO_MES' not in df.columns:
            return []

        if n_std is None:
            n_std = getattr(self, "area_outlier_std", 2)

        last_months = self.get_last_two_months(df)
        if not last_months:
            return []

        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        if df_filtered.empty:
            return []

        df_filtered['AREA'] = pd.to_numeric(df_filtered['AREA'], errors='coerce')

        # identificar coluna de empreendimento, se existir
        empreendimento_col = None
        for col in ['EMPREENDIMENTO', 'PROJETO', 'NOME_EMPREENDIMENTO', 'NOME_PROJETO']:
            if col in df_filtered.columns:
                empreendimento_col = col
                break

        outliers: List[Dict[str, Any]] = []

        # Caso residencial: usa QTD_QUARTOS
        if 'QTD_QUARTOS' in df_filtered.columns:
            # normalizar número de quartos
            df_filtered['QTD_QUARTOS_NORM'] = df_filtered['QTD_QUARTOS'].apply(self.normalize_quartos)

            for rooms in ['1', '2', '3', '4+']:
                group = df_filtered[df_filtered['QTD_QUARTOS_NORM'] == rooms]
                if group.empty:
                    continue

                mean_val = group['AREA'].mean()
                std_val = group['AREA'].std()

                if pd.isna(std_val) or std_val == 0:
                    continue

                limit = mean_val + n_std * std_val
                mask = group['AREA'] > limit
                out_df = group[mask]

                for idx, row in out_df.iterrows():
                    outliers.append({
                        'sheet': sheet_name,
                        'index': int(idx),
                        'empresa': row.get('EMPRESA'),
                        'empreendimento': row.get(empreendimento_col, None),
                        'bairro': row.get('BAIRRO'),
                        'ano_mes': row.get('ANO_MES'),
                        'oferta_venda': row.get('OFERTA_VENDA'),
                        'quartos': row.get('QTD_QUARTOS'),
                        'area_declarada': row.get('AREA'),
                        'limite_area': float(limit),
                        'media': float(mean_val),
                        'desvio_padrao': float(std_val),
                        'valor_m2': row.get('VALOR_MEDIO_M2'),
                    })
        else:
            # Caso comercial: usa toda a base filtrada
            group = df_filtered
            if group['AREA'].notna().sum() == 0:
                return []

            mean_val = group['AREA'].mean()
            std_val = group['AREA'].std()

            if pd.isna(std_val) or std_val == 0:
                return []

            limit = mean_val + n_std * std_val
            mask = group['AREA'] > limit
            out_df = group[mask]

            for idx, row in out_df.iterrows():
                outliers.append({
                    'sheet': sheet_name,
                    'index': int(idx),
                    'empresa': row.get('EMPRESA'),
                    'empreendimento': row.get(empreendimento_col, None),
                    'bairro': row.get('BAIRRO'),
                    'ano_mes': row.get('ANO_MES'),
                    'oferta_venda': row.get('OFERTA_VENDA'),
                    'area_declarada': row.get('AREA'),
                    'limite_area': float(limit),
                    'media': float(mean_val),
                    'desvio_padrao': float(std_val),
                    'valor_m2': row.get('VALOR_MEDIO_M2'),
                })

        return outliers

    def validate_area_outliers_std(self, df: pd.DataFrame, sheet_name: str, n_std: int = None) -> List[str]:
        """
        Gera mensagens de erro resumidas sobre outliers de AREA para uso no relatório.
        """
        outliers = self._compute_area_outliers(df, sheet_name, n_std=n_std)
        if not outliers:
            return []

        if n_std is None:
            n_std = getattr(self, "area_outlier_std", 2)

        indices = sorted({o['index'] for o in outliers})
        msg = (
            f"Aba {sheet_name}: {len(outliers)} registro(s) com AREA acima do limite "
            f"média + {n_std} desvio(s) padrão nos dois últimos meses. "
            f"Linhas (índices do arquivo original): {indices}"
        )
        return [msg]

    def generate_summary_statistics(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Gera estatísticas resumo dos dados"""
        stats = {
            'total_rows': len(df),
            'empty_rows': df.isnull().all(axis=1).sum(),
            'duplicate_rows': df.duplicated().sum(),
        }
        
        for col in self.numeric_columns:
            if col in df.columns:
                numeric_vals = pd.to_numeric(df[col], errors='coerce')
                stats[f'{col}_mean'] = numeric_vals.mean()
                stats[f'{col}_median'] = numeric_vals.median()
                stats[f'{col}_min'] = numeric_vals.min()
                stats[f'{col}_max'] = numeric_vals.max()
                stats[f'{col}_null_count'] = numeric_vals.isna().sum()
        
        if 'BAIRRO' in df.columns:
            stats['bairro_distribution'] = df['BAIRRO'].value_counts().to_dict()
        
        return stats

    def format_ano_mes(self, ano_mes: int) -> str:
        """Formata ANO_MES para exibição (YYYY-MM)"""
        ano_mes_str = str(ano_mes)
        if len(ano_mes_str) == 6:
            return f"{ano_mes_str[:4]}-{ano_mes_str[4:6]}"
        return str(ano_mes)

    def identify_inactive_companies_safe(self, pivot_table: pd.DataFrame, months_formatted: List[str], df_filtered: pd.DataFrame) -> List[str]:
        """
        Identifica empresas inativas usando a nova fórmula:
        Se a soma da QUANTIDADE para OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS 
        na variável OFERTA_VENDA for 0 no mês atual
        """
        if len(months_formatted) < 1:
            return []
            
        current_month = months_formatted[0]
        inactive_companies = []
                
        try:
            # Filtrar dados do mês atual
            current_month_data = df_filtered[df_filtered['ANO_MES_FORMATTED'] == current_month]
            
            if current_month_data.empty:
                return []
            
            # Para cada empresa, verificar se tem ofertas no mês atual
            for empresa in df_filtered['EMPRESA'].unique():
                if pd.isna(empresa):
                    continue
                    
                empresa_current = current_month_data[current_month_data['EMPRESA'] == empresa]
                
                # Filtrar apenas OFERTADOS DISPONÍVEIS e OFERTADOS LANÇAMENTOS
                ofertas_atuais = empresa_current[
                    empresa_current['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])
                ]
                
                # Somar quantidade de ofertas
                soma_ofertas_atual = ofertas_atuais['QUANTIDADE'].sum()
                
                # Se soma é 0, empresa está inativa
                if soma_ofertas_atual == 0:
                    inactive_companies.append(str(empresa))
                    
        except Exception as e:
            print(f"Erro ao identificar empresas inativas: {e}")
        
        return inactive_companies

    def analyze_business_logic_with_launches_safe(self, df_filtered: pd.DataFrame, last_months: List[int], months_formatted: List[str]) -> Dict[str, Any]:
        """
        Analisa problemas de lógica usando a nova fórmula:
        Se a soma da QUANTIDADE para OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS no mês atual 
        for DIFERENTE da soma da QUANTIDADE no mês anterior para OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS 
        subtraído de VENDIDOS + VENDIDOS - LANÇADOS E VENDIDOS do mês anterior + DISTRATO do mês anterior
        """
        problems = []
        
        if len(months_formatted) < 2:
            print("AVISO: Menos de 2 meses disponíveis para análise de lógica")
            return {'logic_problems': problems}
        
        current_month = months_formatted[0]
        previous_month = months_formatted[1]
        
        
        try:
            # Separar dados por mês
            current_data = df_filtered[df_filtered['ANO_MES_FORMATTED'] == current_month]
            previous_data = df_filtered[df_filtered['ANO_MES_FORMATTED'] == previous_month]
            
            
            # Para cada empresa
            empresas_analisadas = 0
            for empresa in df_filtered['EMPRESA'].unique():
                if pd.isna(empresa) or empresa == 'TOTAL GERAL':
                    continue
                    
                try:
                    empresa_current = current_data[current_data['EMPRESA'] == empresa]
                    empresa_previous = previous_data[previous_data['EMPRESA'] == empresa]
                    
                    # Se não tem dados em pelo menos um dos meses, pular
                    if empresa_current.empty or empresa_previous.empty:
                        continue
                    
                    empresas_analisadas += 1
                    
                    # OFERTAS ATUAIS (OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS)
                    ofertas_atuais = empresa_current[
                        empresa_current['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])
                    ]['QUANTIDADE'].sum()
                    
                    # OFERTAS ANTERIORES (OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS)
                    ofertas_anteriores = empresa_previous[
                        empresa_previous['OFERTA_VENDA'].isin(['OFERTADOS DISPONIVEIS', 'OFERTADOS LANCAMENTOS'])
                    ]['QUANTIDADE'].sum()
                    
                    # VENDAS ANTERIORES (VENDIDOS + VENDIDOS - LANÇADOS E VENDIDOS)
                    vendas_anteriores = empresa_previous[
                        empresa_previous['OFERTA_VENDA'].isin(['VENDIDOS', 'VENDIDOS - LANCADOS E VENDIDOS'])
                    ]['QUANTIDADE'].sum()
                    
                    # DISTRATOS ANTERIORES
                    distratos_anteriores = empresa_previous[
                        empresa_previous['OFERTA_VENDA'] == 'DISTRATO'
                    ]['QUANTIDADE'].sum()
                    
                    # FÓRMULA: Oferta_Esperada = Oferta_Anterior - Vendas_Anterior + Distratos_Anterior
                    oferta_esperada = ofertas_anteriores - vendas_anteriores + distratos_anteriores
                    
                    # Verificar diferença
                    diferenca = ofertas_atuais - oferta_esperada

                                        
                    # Se há diferença significativa, registrar problema
                    if abs(diferenca) > 1:  # tolerância de 1 unidade
                        problem_info = {
                            'empresa': str(empresa),
                            'ofertas_atuais': float(ofertas_atuais),
                            'oferta_esperada': float(oferta_esperada),
                            'diferenca': float(diferenca),
                            'detalhes': {
                                'ofertas_anteriores': float(ofertas_anteriores),
                                'vendas_anteriores': float(vendas_anteriores),
                                'distratos_anteriores': float(distratos_anteriores)
                            }
                        }
                        problems.append(problem_info)
                        
                    
                except Exception as e:
                    print(f"Erro ao analisar empresa {empresa}: {e}")
                    continue
            
                    
        except Exception as e:
            print(f"Erro geral na análise de lógica: {e}")
            import traceback
            traceback.print_exc()
        
        return {'logic_problems': problems}

    def create_advanced_pivot_table(self, df: pd.DataFrame, sheet_type: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Cria tabela dinâmica avançada com validação de lógica de negócio"""
        required_cols = ['ANO_MES', 'EMPRESA', 'OFERTA_VENDA', 'QUANTIDADE']
        
        if not all(col in df.columns for col in required_cols):
            print(f"AVISO: Colunas necessárias faltantes. Requeridas: {required_cols}")
            return pd.DataFrame(), {}
        
        last_months = self.get_last_two_months(df)
        print(f"Últimos dois meses encontrados: {last_months}")
        
        if len(last_months) < 2:
            print("AVISO: Menos de 2 meses de dados disponíveis")
            return pd.DataFrame(), {}
        
        df_filtered = df[df['ANO_MES'].isin(last_months)].copy()
        print(f"Dados filtrados: {len(df_filtered)} linhas")
        
        if df_filtered.empty:
            print("AVISO: Dados filtrados estão vazios")
            return pd.DataFrame(), {}
        
        # Mapear categorias
        df_filtered['CATEGORIA_DETALHADA'] = df_filtered['OFERTA_VENDA'].map(self.oferta_mapping)
        df_filtered['CATEGORIA_TOTAL'] = df_filtered['OFERTA_VENDA'].map(self.oferta_total_mapping)
        
        df_filtered['QUANTIDADE'] = pd.to_numeric(df_filtered['QUANTIDADE'], errors='coerce').fillna(0)
        df_filtered['ANO_MES_FORMATTED'] = df_filtered['ANO_MES'].apply(self.format_ano_mes)
        
        print("Categorias encontradas:", df_filtered['CATEGORIA_DETALHADA'].unique())
        
        # Criar tabelas pivot
        grouped_detailed = df_filtered.groupby(['EMPRESA', 'ANO_MES', 'ANO_MES_FORMATTED', 'CATEGORIA_DETALHADA'])['QUANTIDADE'].sum().reset_index()
        
        pivot_detailed = grouped_detailed.pivot_table(
            index='EMPRESA',
            columns=['ANO_MES_FORMATTED', 'CATEGORIA_DETALHADA'],
            values='QUANTIDADE',
            aggfunc='sum',
            fill_value=0
        )
        
        grouped_total = df_filtered.groupby(['EMPRESA', 'ANO_MES', 'ANO_MES_FORMATTED', 'CATEGORIA_TOTAL'])['QUANTIDADE'].sum().reset_index()
        
        pivot_total = grouped_total.pivot_table(
            index='EMPRESA',
            columns=['ANO_MES_FORMATTED', 'CATEGORIA_TOTAL'],
            values='QUANTIDADE',
            aggfunc='sum',
            fill_value=0
        )
        
        months_formatted = sorted([self.format_ano_mes(m) for m in last_months], reverse=True)
        
        # Construir tabela final
        all_companies = set()
        if not pivot_detailed.empty:
            all_companies.update(pivot_detailed.index)
        if not pivot_total.empty:
            all_companies.update(pivot_total.index)
        
        if not all_companies:
            print("AVISO: Nenhuma empresa encontrada nos dados")
            return pd.DataFrame(), {}
        
        final_table = pd.DataFrame(index=sorted(all_companies))
        
        for month in months_formatted:
            # Adicionar colunas com verificação de existência
            col_oferta_total = (month, 'OFERTA_TOTAL')
            if col_oferta_total in pivot_total.columns:
                final_table[col_oferta_total] = pivot_total[col_oferta_total].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_oferta_total] = 0
            
            col_lancamentos = (month, 'OFERTADOS_LANCAMENTOS')
            if col_lancamentos in pivot_detailed.columns:
                final_table[col_lancamentos] = pivot_detailed[col_lancamentos].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_lancamentos] = 0
            
            col_venda = (month, 'VENDA')
            if col_venda in pivot_detailed.columns:
                final_table[col_venda] = pivot_detailed[col_venda].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_venda] = 0
            
            col_distrato = (month, 'DISTRATO')
            if col_distrato in pivot_detailed.columns:
                final_table[col_distrato] = pivot_detailed[col_distrato].reindex(final_table.index, fill_value=0)
            else:
                final_table[col_distrato] = 0
        
        print(f"Tabela final criada com {len(final_table)} empresas e {len(final_table.columns)} colunas")
        
        # Realizar análises com as NOVAS FÓRMULAS
        analysis_results = {}
        try:
            analysis_results = self.analyze_business_logic_with_launches_safe(df_filtered, last_months, months_formatted)
        except Exception as e:
            print(f"Erro na análise de lógica de negócio: {e}")
            analysis_results = {'logic_problems': []}
        
        try:
            launches_analysis = self.analyze_launches_by_company_and_neighborhood_with_empreendimentos(df_filtered, last_months)
            analysis_results['launches_analysis'] = launches_analysis
        except Exception as e:
            print(f"Erro na análise de lançamentos: {e}")
            analysis_results['launches_analysis'] = {'by_company': {}, 'by_neighborhood': {}, 'total_launches': 0}
        
        # USAR A NOVA FÓRMULA para empresas inativas
        inactive_companies = self.identify_inactive_companies_safe(final_table, months_formatted, df_filtered)
        analysis_results['inactive_companies'] = inactive_companies
        analysis_results['current_month'] = months_formatted[0] if months_formatted else None
        
        # Adicionar linha de total
        if not final_table.empty:
            total_row = final_table.sum()
            total_row.name = 'TOTAL GERAL'
            final_table = pd.concat([final_table, total_row.to_frame().T])
        
        return final_table, analysis_results

    def analyze_launches_by_company_and_neighborhood(self, df: pd.DataFrame, last_months: List[int]) -> Dict[str, Any]:
        """Analisa empreendimentos lançados por empresa e bairro"""
        launches = df[
            (df['OFERTA_VENDA'] == 'OFERTADOS LANCAMENTOS') & 
            (df['ANO_MES'].isin(last_months)) &
            (df['QUANTIDADE'] > 0)
        ].copy()
        
        if launches.empty:
            return {
                'by_company': {},
                'by_neighborhood': {},
                'total_launches': 0
            }
        
        launches['ANO_MES_FORMATTED'] = launches['ANO_MES'].apply(self.format_ano_mes)
        launches['QUANTIDADE'] = pd.to_numeric(launches['QUANTIDADE'], errors='coerce').fillna(0)
        
        # Análise por empresa
        by_company = {}
        for empresa in launches['EMPRESA'].unique():
            empresa_launches = launches[launches['EMPRESA'] == empresa]
            
            by_company[empresa] = {
                'total_quantidade': empresa_launches['QUANTIDADE'].sum(),
                'por_mes': {},
                'bairros': list(empresa_launches['BAIRRO'].unique())
            }
            
            for mes in empresa_launches['ANO_MES_FORMATTED'].unique():
                mes_data = empresa_launches[empresa_launches['ANO_MES_FORMATTED'] == mes]
                by_company[empresa]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'bairros': list(mes_data['BAIRRO'].unique())
                }
        
        # Análise por bairro
        by_neighborhood = {}
        for bairro in launches['BAIRRO'].unique():
            bairro_launches = launches[launches['BAIRRO'] == bairro]
            
            by_neighborhood[bairro] = {
                'total_quantidade': bairro_launches['QUANTIDADE'].sum(),
                'empresas': list(bairro_launches['EMPRESA'].unique()),
                'por_mes': {}
            }
            
            for mes in bairro_launches['ANO_MES_FORMATTED'].unique():
                mes_data = bairro_launches[bairro_launches['ANO_MES_FORMATTED'] == mes]
                by_neighborhood[bairro]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'empresas': list(mes_data['EMPRESA'].unique())
                }
        
        return {
            'by_company': by_company,
            'by_neighborhood': by_neighborhood,
            'total_launches': launches['QUANTIDADE'].sum()
        }

    def extract_empreendimento_name(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Extrai e identifica nomes de empreendimentos baseado em padrões comuns
        """
        import re
        from collections import Counter
        
        # Verificar se a coluna existe (pode ter nomes diferentes)
        empreendimento_col = None
        possible_names = ['EMPREENDIMENTO', 'PROJETO', 'NOME_EMPREENDIMENTO', 'NOME_PROJETO']
        
        for col_name in possible_names:
            if col_name in df.columns:
                empreendimento_col = col_name
                break
        
        if empreendimento_col is None:
            print("AVISO: Coluna de empreendimento não encontrada")
            df['EMPREENDIMENTO_AGRUPADO'] = 'N/A'
            return df
        
        df = df.copy()
        df['EMPREENDIMENTO_AGRUPADO'] = df[empreendimento_col]
        
        # Função para extrair termo principal do empreendimento
        def extract_main_term(name):
            if pd.isna(name) or name == '':
                return 'N/A'
            
            name = str(name).strip()
            
            # Remover palavras comuns no final que indicam variações
            suffixes_to_remove = [
                r'\s+(Torre|Bloco|Torre de|Bloco de)\s+[A-Z]+$',
                r'\s+(Apartamento|Apto|Apt)\s*\d*$',
                r'\s+(Duplex|Cobertura|Penthouse)$',
                r'\s+(Fase|Etapa)\s*\d+$',
                r'\s+(I|II|III|IV|V|\d+)$',
                r'\s+[A-Z]$',  # Letra única no final
                r'\s+\d+$'     # Número no final
            ]
            
            cleaned_name = name
            for suffix_pattern in suffixes_to_remove:
                cleaned_name = re.sub(suffix_pattern, '', cleaned_name, flags=re.IGNORECASE)
            
            # Se depois da limpeza restou algo muito pequeno, usar o nome original
            if len(cleaned_name.strip()) < 3:
                return name
            
            return cleaned_name.strip()
        
        # Aplicar extração do termo principal
        df['TERMO_PRINCIPAL'] = df[empreendimento_col].apply(extract_main_term)
        
        # Agrupar termos similares (considerando pequenas variações)
        termo_groups = {}
        processed_terms = set()
        
        unique_terms = df['TERMO_PRINCIPAL'].unique()
        
        for term1 in unique_terms:
            if term1 in processed_terms or pd.isna(term1):
                continue
            
            # Criar grupo com o termo atual
            group = [term1]
            processed_terms.add(term1)
            
            # Procurar termos similares
            for term2 in unique_terms:
                if term2 in processed_terms or pd.isna(term2) or term1 == term2:
                    continue
                
                # Verificar similaridade
                if self.terms_are_similar(term1, term2):
                    group.append(term2)
                    processed_terms.add(term2)
            
            # Usar o termo mais comum como representativo do grupo
            if len(group) > 1:
                # Contar quantas vezes cada termo aparece
                term_counts = {}
                for term in group:
                    term_counts[term] = len(df[df['TERMO_PRINCIPAL'] == term])
                
                # Usar o mais frequente como representativo
                representative_term = max(term_counts, key=term_counts.get)
            else:
                representative_term = term1
            
            # Mapear todos os termos do grupo para o representativo
            for term in group:
                termo_groups[term] = representative_term
        
        # Aplicar o agrupamento
        df['EMPREENDIMENTO_AGRUPADO'] = df['TERMO_PRINCIPAL'].map(termo_groups).fillna(df['TERMO_PRINCIPAL'])
        
        return df

    def terms_are_similar(self, term1, term2, threshold=0.7):
        """
        Verifica se dois termos são similares o suficiente para serem agrupados
        """
        import difflib
        
        # Converter para minúsculas para comparação
        t1 = term1.lower().strip()
        t2 = term2.lower().strip()
        
        # Se um termo está contido no outro
        if t1 in t2 or t2 in t1:
            return True
        
        # Usar ratio de similaridade
        similarity = difflib.SequenceMatcher(None, t1, t2).ratio()
        return similarity >= threshold

    def analyze_launches_by_company_and_neighborhood_with_empreendimentos(self, df: pd.DataFrame, last_months: List[int]) -> Dict[str, Any]:
        """
        Analisa empreendimentos lançados por empresa e bairro, incluindo nomes dos empreendimentos
        """
        # Primeiro extrair/agrupar empreendimentos
        df_with_empreendimentos = self.extract_empreendimento_name(df)
        
        launches = df_with_empreendimentos[
            (df_with_empreendimentos['OFERTA_VENDA'] == 'OFERTADOS LANCAMENTOS') & 
            (df_with_empreendimentos['ANO_MES'].isin(last_months)) &
            (df_with_empreendimentos['QUANTIDADE'] > 0)
        ].copy()
        
        if launches.empty:
            return {
                'by_company': {},
                'by_neighborhood': {},
                'by_empreendimento': {},
                'total_launches': 0
            }
        
        launches['ANO_MES_FORMATTED'] = launches['ANO_MES'].apply(self.format_ano_mes)
        launches['QUANTIDADE'] = pd.to_numeric(launches['QUANTIDADE'], errors='coerce').fillna(0)
        
        # Análise por empresa (incluindo empreendimentos)
        by_company = {}
        for empresa in launches['EMPRESA'].unique():
            if pd.isna(empresa):
                continue
                
            empresa_launches = launches[launches['EMPRESA'] == empresa]
            
            # Listar empreendimentos únicos da empresa
            empreendimentos = list(empresa_launches['EMPREENDIMENTO_AGRUPADO'].unique())
            empreendimentos = [emp for emp in empreendimentos if emp != 'N/A']
            
            by_company[empresa] = {
                'total_quantidade': empresa_launches['QUANTIDADE'].sum(),
                'por_mes': {},
                'bairros': list(empresa_launches['BAIRRO'].unique()),
                'empreendimentos': empreendimentos
            }
            
            for mes in empresa_launches['ANO_MES_FORMATTED'].unique():
                mes_data = empresa_launches[empresa_launches['ANO_MES_FORMATTED'] == mes]
                empreendimentos_mes = list(mes_data['EMPREENDIMENTO_AGRUPADO'].unique())
                empreendimentos_mes = [emp for emp in empreendimentos_mes if emp != 'N/A']
                
                by_company[empresa]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'bairros': list(mes_data['BAIRRO'].unique()),
                    'empreendimentos': empreendimentos_mes
                }
        
        # Análise por bairro (incluindo empreendimentos)
        by_neighborhood = {}
        for bairro in launches['BAIRRO'].unique():
            if pd.isna(bairro):
                continue
                
            bairro_launches = launches[launches['BAIRRO'] == bairro]
            
            # Listar empreendimentos únicos do bairro
            empreendimentos = list(bairro_launches['EMPREENDIMENTO_AGRUPADO'].unique())
            empreendimentos = [emp for emp in empreendimentos if emp != 'N/A']
            
            by_neighborhood[bairro] = {
                'total_quantidade': bairro_launches['QUANTIDADE'].sum(),
                'empresas': list(bairro_launches['EMPRESA'].unique()),
                'empreendimentos': empreendimentos,
                'por_mes': {}
            }
            
            for mes in bairro_launches['ANO_MES_FORMATTED'].unique():
                mes_data = bairro_launches[bairro_launches['ANO_MES_FORMATTED'] == mes]
                empreendimentos_mes = list(mes_data['EMPREENDIMENTO_AGRUPADO'].unique())
                empreendimentos_mes = [emp for emp in empreendimentos_mes if emp != 'N/A']
                
                by_neighborhood[bairro]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'empresas': list(mes_data['EMPRESA'].unique()),
                    'empreendimentos': empreendimentos_mes
                }
        
        # Análise por empreendimento
        by_empreendimento = {}
        for emp in launches['EMPREENDIMENTO_AGRUPADO'].unique():
            if pd.isna(emp) or emp == 'N/A':
                continue
                
            emp_launches = launches[launches['EMPREENDIMENTO_AGRUPADO'] == emp]
            
            by_empreendimento[emp] = {
                'total_quantidade': emp_launches['QUANTIDADE'].sum(),
                'empresas': list(emp_launches['EMPRESA'].unique()),
                'bairros': list(emp_launches['BAIRRO'].unique()),
                'por_mes': {}
            }
            
            for mes in emp_launches['ANO_MES_FORMATTED'].unique():
                mes_data = emp_launches[emp_launches['ANO_MES_FORMATTED'] == mes]
                by_empreendimento[emp]['por_mes'][mes] = {
                    'quantidade': mes_data['QUANTIDADE'].sum(),
                    'empresas': list(mes_data['EMPRESA'].unique()),
                    'bairros': list(mes_data['BAIRRO'].unique())
                }
        
        return {
            'by_company': by_company,
            'by_neighborhood': by_neighborhood,
            'by_empreendimento': by_empreendimento,
            'total_launches': launches['QUANTIDADE'].sum()
        }

    def create_launches_worksheet_with_empreendimentos(self, workbook, launches_data, sheet_name, sheet_type):
        """
        Cria worksheet de lançamentos com empreendimentos incluídos
        """
        try:
            from openpyxl.styles import Font, Alignment
            
            ws = workbook.create_sheet(sheet_name)
            
            title_font = Font(bold=True, size=14)
            section_font = Font(bold=True, size=12)
            header_font = Font(bold=True, size=10)
            
            row = 1
            
            ws.cell(row=row, column=1, value=f"ANÁLISE DE LANÇAMENTOS - {sheet_type}").font = title_font
            row += 2
            
            total_launches = launches_data.get('total_launches', 0)
            cell = ws.cell(row=row, column=1, value=f"TOTAL DE LANÇAMENTOS: {total_launches} unidades")
            cell.font = section_font
            row += 2
            
            # LANÇAMENTOS POR EMPRESA
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR EMPRESA:").font = section_font
            row += 1
            
            by_company = launches_data.get('by_company', {})
            if by_company:
                ws.cell(row=row, column=1, value="Empresa").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Bairros").font = header_font
                ws.cell(row=row, column=4, value="Empreendimentos").font = header_font
                ws.cell(row=row, column=5, value="Detalhes por Mês").font = header_font
                row += 1
                
                for empresa, data in by_company.items():
                    ws.cell(row=row, column=1, value=str(empresa))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['bairros']))
                    ws.cell(row=row, column=4, value=', '.join(data.get('empreendimentos', [])))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        empreendimentos_mes = ', '.join(mes_data.get('empreendimentos', []))
                        if empreendimentos_mes:
                            detalhe = f"{mes}: {mes_data['quantidade']} un. ({empreendimentos_mes})"
                        else:
                            detalhe = f"{mes}: {mes_data['quantidade']} un."
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=5, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum lançamento encontrado")
                row += 1
            
            row += 1
            
            # LANÇAMENTOS POR EMPREENDIMENTO
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR EMPREENDIMENTO:").font = section_font
            row += 1
            
            by_empreendimento = launches_data.get('by_empreendimento', {})
            if by_empreendimento:
                ws.cell(row=row, column=1, value="Empreendimento").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Empresa").font = header_font
                ws.cell(row=row, column=4, value="Bairros").font = header_font
                ws.cell(row=row, column=5, value="Detalhes por Mês").font = header_font
                row += 1
                
                for empreendimento, data in by_empreendimento.items():
                    ws.cell(row=row, column=1, value=str(empreendimento))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['empresas']))
                    ws.cell(row=row, column=4, value=', '.join(data['bairros']))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        empresas_mes = ', '.join(mes_data['empresas'])
                        detalhe = f"{mes}: {mes_data['quantidade']} un. ({empresas_mes})"
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=5, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum empreendimento encontrado")
                row += 1
            
            row += 1
            
            # LANÇAMENTOS POR BAIRRO
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR BAIRRO:").font = section_font
            row += 1
            
            by_neighborhood = launches_data.get('by_neighborhood', {})
            if by_neighborhood:
                ws.cell(row=row, column=1, value="Bairro").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Empresas").font = header_font
                ws.cell(row=row, column=4, value="Empreendimentos").font = header_font
                ws.cell(row=row, column=5, value="Detalhes por Mês").font = header_font
                row += 1
                
                for bairro, data in by_neighborhood.items():
                    ws.cell(row=row, column=1, value=str(bairro))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['empresas']))
                    ws.cell(row=row, column=4, value=', '.join(data.get('empreendimentos', [])))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        empreendimentos_mes = ', '.join(mes_data.get('empreendimentos', []))
                        if empreendimentos_mes:
                            detalhe = f"{mes}: {mes_data['quantidade']} un. ({empreendimentos_mes})"
                        else:
                            detalhe = f"{mes}: {mes_data['quantidade']} un."
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=5, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum lançamento encontrado por bairro")
            
            # Ajustar larguras das colunas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 80
            
        except Exception as e:
            print(f"Erro ao criar worksheet de lançamentos: {e}")

    def create_summary_tables_excel(self, file_path: str, df_residencial: pd.DataFrame, 
                                  df_comercial: pd.DataFrame, residencial_name: str, 
                                  comercial_name: str) -> str:
        """Cria arquivo Excel com tabelas resumo e análise de lançamentos"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)
        output_path = os.path.join(directory, f"{base_name}_resumo_{timestamp}.xlsx")
        
        print(f"Iniciando criação do arquivo: {output_path}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Criar tabela para residencial
            print("Processando dados residenciais...")
            pivot_residencial, analysis_res = self.create_advanced_pivot_table(df_residencial, 'residencial')
            
            if not pivot_residencial.empty:
                print(f"Tabela residencial criada com {len(pivot_residencial)} linhas")
                
                ws_res = wb.create_sheet('Resumo_Residencial')
                
                last_months = self.get_last_two_months(df_residencial)
                period_text = ", ".join([self.format_ano_mes(m) for m in last_months]) if last_months else "N/A"
                
                header_font = Font(bold=True, size=12)
                subheader_font = Font(bold=True, size=10)
                
                ws_res['A1'] = f"RESUMO RESIDENCIAL - Período: {period_text}"
                ws_res['A1'].font = header_font
                
                ws_res['A2'] = "Empresa x Mês x Tipo de Oferta/Venda (Soma de QUANTIDADE)"
                ws_res['A2'].font = subheader_font
                
                ws_res['A3'] = "OFERTA_TOTAL = OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS"
                ws_res['A3'].font = subheader_font
                
                self.insert_pivot_table_with_formatting(ws_res, pivot_residencial, analysis_res, start_row=5)
                
                print("Dados residenciais escritos no Excel com formatação")
            else:
                print("AVISO: Tabela residencial está vazia")
            
            # Criar tabela para comercial
            print("Processando dados comerciais...")
            pivot_comercial, analysis_com = self.create_advanced_pivot_table(df_comercial, 'comercial')
            
            if not pivot_comercial.empty:
                print(f"Tabela comercial criada com {len(pivot_comercial)} linhas")
                
                ws_com = wb.create_sheet('Resumo_Comercial')
                
                last_months = self.get_last_two_months(df_comercial)
                period_text = ", ".join([self.format_ano_mes(m) for m in last_months]) if last_months else "N/A"
                
                ws_com['A1'] = f"RESUMO COMERCIAL - Período: {period_text}"
                ws_com['A1'].font = header_font
                
                ws_com['A2'] = "Empresa x Mês x Tipo de Oferta/Venda (Soma de QUANTIDADE)"
                ws_com['A2'].font = subheader_font
                
                ws_com['A3'] = "OFERTA_TOTAL = OFERTADOS DISPONÍVEIS + OFERTADOS LANÇAMENTOS"
                ws_com['A3'].font = subheader_font
                
                self.insert_pivot_table_with_formatting(ws_com, pivot_comercial, analysis_com, start_row=5)
                
                print("Dados comerciais escritos no Excel com formatação")
            else:
                print("AVISO: Tabela comercial está vazia")
            

            # Criar abas de outliers de AREA (Residencial e Comercial) para os dois últimos meses
            outliers_res = self._compute_area_outliers(df_residencial, residencial_name)
            if outliers_res:
                print("Criando aba de outliers de AREA residencial...")
                ws_area_res = wb.create_sheet('Area_Residencial')
                ws_area_res['A1'] = "OUTLIERS DE AREA - RESIDENCIAL (dois últimos meses)"
                ws_area_res['A1'].font = Font(bold=True, size=14)

                headers = [
                    "Linha (índice)", "Empresa", "Empreendimento", "Bairro", "ANO_MES",
                    "Oferta/Venda", "Quartos", "Área Declarada", "Limite Aceitável",
                    "Média", "Desvio Padrão", "Valor Médio m²"
                ]
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws_area_res.cell(row=3, column=col_idx, value=h)
                    cell.font = Font(bold=True)

                row_idx = 4
                for item in outliers_res:
                    ws_area_res.cell(row=row_idx, column=1, value=item['index'])
                    ws_area_res.cell(row=row_idx, column=2, value=item['empresa'])
                    ws_area_res.cell(row=row_idx, column=3, value=item['empreendimento'])
                    ws_area_res.cell(row=row_idx, column=4, value=item['bairro'])
                    ws_area_res.cell(row=row_idx, column=5, value=item['ano_mes'])
                    ws_area_res.cell(row=row_idx, column=6, value=item['oferta_venda'])
                    ws_area_res.cell(row=row_idx, column=7, value=item['quartos'])
                    ws_area_res.cell(row=row_idx, column=8, value=item['area_declarada'])
                    ws_area_res.cell(row=row_idx, column=9, value=item['limite_area'])
                    ws_area_res.cell(row=row_idx, column=10, value=item['media'])
                    ws_area_res.cell(row=row_idx, column=11, value=item['desvio_padrao'])
                    ws_area_res.cell(row=row_idx, column=12, value=item['valor_m2'])
                    row_idx += 1

            outliers_com = self._compute_area_outliers(df_comercial, comercial_name)
            if outliers_com:
                print("Criando aba de outliers de AREA comercial...")
                ws_area_com = wb.create_sheet('Area_Comercial')
                ws_area_com['A1'] = "OUTLIERS DE AREA - COMERCIAL (dois últimos meses)"
                ws_area_com['A1'].font = Font(bold=True, size=14)

                headers = [
                    "Linha (índice)", "Empresa", "Empreendimento", "Bairro", "ANO_MES",
                    "Oferta/Venda", "Área Declarada", "Limite Aceitável",
                    "Média", "Desvio Padrão", "Valor Médio m²"
                ]
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws_area_com.cell(row=3, column=col_idx, value=h)
                    cell.font = Font(bold=True)

                row_idx = 4
                for item in outliers_com:
                    ws_area_com.cell(row=row_idx, column=1, value=item['index'])
                    ws_area_com.cell(row=row_idx, column=2, value=item['empresa'])
                    ws_area_com.cell(row=row_idx, column=3, value=item['empreendimento'])
                    ws_area_com.cell(row=row_idx, column=4, value=item['bairro'])
                    ws_area_com.cell(row=row_idx, column=5, value=item['ano_mes'])
                    ws_area_res.cell(row=row_idx, column=6, value=item['oferta_venda'])
                    ws_area_com.cell(row=row_idx, column=7, value=item['area_declarada'])
                    ws_area_com.cell(row=row_idx, column=8, value=item['limite_area'])
                    ws_area_com.cell(row=row_idx, column=9, value=item['media'])
                    ws_area_com.cell(row=row_idx, column=10, value=item['desvio_padrao'])
                    ws_area_com.cell(row=row_idx, column=11, value=item['valor_m2'])
                    row_idx += 1

            # Criar abas de lançamentos se houver dados
            if 'launches_analysis' in analysis_res and analysis_res['launches_analysis'].get('total_launches', 0) > 0:
                print("Criando aba de lançamentos residencial...")
                self.create_launches_worksheet_with_empreendimentos(wb, analysis_res['launches_analysis'], 'Lancamentos_Residencial', 'RESIDENCIAL')
                print("Lançamentos residenciais adicionados")
            
            if 'launches_analysis' in analysis_com and analysis_com['launches_analysis'].get('total_launches', 0) > 0:
                print("Criando aba de lançamentos comercial...")
                self.create_launches_worksheet_with_empreendimentos(wb, analysis_com['launches_analysis'], 'Lancamentos_Comercial', 'COMERCIAL')
                print("Lançamentos comerciais adicionados")
            
            wb.save(output_path)
            wb.close()
            
            print(f"Arquivo salvo: {output_path}")
            
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"✅ Arquivo criado com sucesso! Tamanho: {file_size} bytes")
                return output_path
            else:
                print("❌ ERRO: Arquivo não foi encontrado após criação")
                return ""
                
        except Exception as e:
            print(f"❌ ERRO ao criar arquivo Excel: {str(e)}")
            import traceback
            print("Stack trace completo:")
            traceback.print_exc()
            return ""

    def insert_pivot_table_with_formatting(self, worksheet, pivot_table, analysis_data, start_row=5):
        """Insere tabela pivot no worksheet com formatação condicional"""
        try:
            from openpyxl.styles import Font, PatternFill
            
            red_font = Font(color="FF0000", bold=True, size=11)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
            
            inactive_companies = analysis_data.get('inactive_companies', [])
            logic_problems = analysis_data.get('logic_problems', [])
            problem_companies = [p['empresa'] for p in logic_problems]
            
            
            # Inserir cabeçalhos das colunas
            col = 1
            worksheet.cell(row=start_row, column=col, value="EMPRESA").font = bold_font
            col += 1
            
            for col_name in pivot_table.columns:
                if isinstance(col_name, tuple):
                    header = f"{col_name[0]} {col_name[1]}"
                else:
                    header = str(col_name)
                cell = worksheet.cell(row=start_row, column=col, value=header)
                cell.font = bold_font
                col += 1
            
            # Inserir dados das empresas
            for row_idx, (empresa, row_data) in enumerate(pivot_table.iterrows()):
                current_row = start_row + 1 + row_idx
                empresa_str = str(empresa).strip()
                                
                empresa_cell = worksheet.cell(row=current_row, column=1, value=empresa_str)
                
                # Verificar se empresa deve ser formatada
                is_inactive = any(empresa_str == str(inactive).strip() for inactive in inactive_companies)
                is_problem = any(empresa_str == str(problem).strip() for problem in problem_companies)
                                
                # Aplicar formatação condicional
                try:
                    if is_inactive:
                        empresa_cell.font = red_font
                    
                    if is_problem:
                        empresa_cell.fill = yellow_fill
                        
                except Exception as style_error:
                    print(f"  - ERRO ao aplicar formatação: {style_error}")
                
                # Inserir dados da linha
                for col_idx, value in enumerate(row_data):
                    data_cell = worksheet.cell(row=current_row, column=col_idx + 2)
                    
                    try:
                        numeric_value = float(value) if pd.notna(value) and value != 0 else 0
                        data_cell.value = numeric_value
                    except (ValueError, TypeError):
                        data_cell.value = 0
                    
                    # Aplicar formatação amarela para toda a linha com problemas de lógica
                    try:
                        if is_problem:
                            data_cell.fill = yellow_fill
                    except Exception as cell_error:
                        print(f"  - ERRO ao formatar célula de dados: {cell_error}")
            
            # Ajustar largura das colunas
            worksheet.column_dimensions['A'].width = 25
            for col_idx in range(2, col + 1):
                try:
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[col_letter].width = 15
                except:
                    pass
                                
        except Exception as e:
            print(f"Erro ao inserir tabela com formatação: {e}")
            import traceback
            traceback.print_exc()

    def create_launches_worksheet(self, workbook, launches_data, sheet_name, sheet_type):
        """Cria worksheet de lançamentos com formatação adequada"""
        try:
            from openpyxl.styles import Font, Alignment
            
            ws = workbook.create_sheet(sheet_name)
            
            title_font = Font(bold=True, size=14)
            section_font = Font(bold=True, size=12)
            header_font = Font(bold=True, size=10)
            
            row = 1
            
            ws.cell(row=row, column=1, value=f"ANÁLISE DE LANÇAMENTOS - {sheet_type}").font = title_font
            row += 2
            
            total_launches = launches_data.get('total_launches', 0)
            cell = ws.cell(row=row, column=1, value=f"TOTAL DE LANÇAMENTOS: {total_launches} unidades")
            cell.font = section_font
            row += 2
            
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR EMPRESA:").font = section_font
            row += 1
            
            by_company = launches_data.get('by_company', {})
            if by_company:
                ws.cell(row=row, column=1, value="Empresa").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Bairros").font = header_font
                ws.cell(row=row, column=4, value="Detalhes por Mês").font = header_font
                row += 1
                
                for empresa, data in by_company.items():
                    ws.cell(row=row, column=1, value=str(empresa))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['bairros']))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        bairros_mes = ', '.join(mes_data['bairros'])
                        detalhe = f"{mes}: {mes_data['quantidade']} un. ({bairros_mes})"
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=4, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum lançamento encontrado")
                row += 1
            
            row += 1
            
            ws.cell(row=row, column=1, value="LANÇAMENTOS POR BAIRRO:").font = section_font
            row += 1
            
            by_neighborhood = launches_data.get('by_neighborhood', {})
            if by_neighborhood:
                ws.cell(row=row, column=1, value="Bairro").font = header_font
                ws.cell(row=row, column=2, value="Total Unidades").font = header_font
                ws.cell(row=row, column=3, value="Empresas").font = header_font
                ws.cell(row=row, column=4, value="Detalhes por Mês").font = header_font
                row += 1
                
                for bairro, data in by_neighborhood.items():
                    ws.cell(row=row, column=1, value=str(bairro))
                    ws.cell(row=row, column=2, value=f"{data['total_quantidade']} unidades")
                    ws.cell(row=row, column=3, value=', '.join(data['empresas']))
                    
                    detalhes_mes = []
                    for mes, mes_data in data['por_mes'].items():
                        empresas_mes = ', '.join(mes_data['empresas'])
                        detalhe = f"{mes}: {mes_data['quantidade']} un. ({empresas_mes})"
                        detalhes_mes.append(detalhe)
                    ws.cell(row=row, column=4, value='; '.join(detalhes_mes))
                    row += 1
            else:
                ws.cell(row=row, column=1, value="Nenhum lançamento encontrado por bairro")
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 60
            
        except Exception as e:
            print(f"Erro ao criar worksheet de lançamentos: {e}")

    def process_file(self, file_path: str) -> Tuple[List[str], Dict[str, Any], str]:
        """Processa o arquivo Excel e retorna erros, estatísticas e caminho do resumo"""
        errors = []
        statistics = {}
        summary_path = ""
        
        try:
            print("Iniciando processamento do arquivo...")
            
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            print(f"Abas encontradas: {sheet_names}")
            
            if len(sheet_names) < 2:
                errors.append("Arquivo deve conter pelo menos 2 abas (residencial e comercial)")
                return errors, statistics, summary_path
            
            # Processar aba residencial
            residencial_name = sheet_names[0]
            print(f"Processando aba residencial: {residencial_name}")
            df_residencial = pd.read_excel(file_path, sheet_name=residencial_name)
            print(f"Dados residenciais: {len(df_residencial)} linhas")
            
            errors.extend(self.check_missing_columns(df_residencial, residencial_name, 'residencial'))
            errors.extend(self.validate_ano_mes(df_residencial, residencial_name))
            errors.extend(self.validate_categorical_columns(df_residencial, residencial_name))
            errors.extend(self.validate_numeric_columns(df_residencial, residencial_name))
            errors.extend(self.check_missing_data(df_residencial, residencial_name))
            errors.extend(self.validate_data_consistency_filtered(df_residencial, residencial_name))
            errors.extend(self.validate_area_outliers_std(df_residencial, residencial_name))
            
            statistics[f'residencial_{residencial_name}'] = self.generate_summary_statistics(df_residencial, residencial_name)
            
            # Processar aba comercial
            comercial_name = sheet_names[1]
            print(f"Processando aba comercial: {comercial_name}")
            df_comercial = pd.read_excel(file_path, sheet_name=comercial_name)
            print(f"Dados comerciais: {len(df_comercial)} linhas")
            
            errors.extend(self.check_missing_columns(df_comercial, comercial_name, 'comercial'))
            errors.extend(self.validate_ano_mes(df_comercial, comercial_name))
            errors.extend(self.validate_categorical_columns(df_comercial, comercial_name))
            errors.extend(self.validate_numeric_columns(df_comercial, comercial_name))
            errors.extend(self.check_missing_data(df_comercial, comercial_name))
            errors.extend(self.validate_data_consistency_filtered(df_comercial, comercial_name))
            errors.extend(self.validate_area_outliers_std(df_comercial, comercial_name))
            
            statistics[f'comercial_{comercial_name}'] = self.generate_summary_statistics(df_comercial, comercial_name)
            
            print("Validações concluídas. Iniciando criação do arquivo de resumo...")
            
            summary_path = self.create_summary_tables_excel(
                file_path, df_residencial, df_comercial, 
                residencial_name, comercial_name
            )
            
            if summary_path:
                print(f"✅ Arquivo de resumo criado: {summary_path}")
            else:
                print("❌ Falha na criação do arquivo de resumo")
            
        except Exception as e:
            error_msg = f"Erro ao processar arquivo: {str(e)}"
            errors.append(error_msg)
            print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
        
        return errors, statistics, summary_path

    def save_report(self, file_path: str, errors: List[str], statistics: Dict[str, Any]) -> str:
        """Salva relatório de qualidade"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        directory = os.path.dirname(file_path)
        report_path = os.path.join(directory, f"{base_name}_qualidade_{timestamp}.txt")
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=== RELATÓRIO DE CONTROLE DE QUALIDADE - PESQUISA IVV ===\n")
            f.write(f"Arquivo: {os.path.basename(file_path)}\n")
            f.write(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            
            f.write("ERROS ENCONTRADOS:\n")
            f.write("-" * 20 + "\n")
            if errors:
                for i, error in enumerate(errors, 1):
                    f.write(f"{i}. {error}\n")
            else:
                f.write("✓ Nenhum erro encontrado!\n")
            
            f.write("\n" + "=" * 60 + "\n\n")
            
            f.write("ESTATÍSTICAS RESUMO:\n")
            f.write("-" * 20 + "\n")
            for sheet_name, stats in statistics.items():
                f.write(f"\n{sheet_name.upper()}:\n")
                f.write(f"  Total de linhas: {stats.get('total_rows', 'N/A')}\n")
                f.write(f"  Linhas vazias: {stats.get('empty_rows', 'N/A')}\n")
                f.write(f"  Linhas duplicadas: {stats.get('duplicate_rows', 'N/A')}\n")
                
                if 'bairro_distribution' in stats:
                    f.write("  Distribuição por bairro:\n")
                    for bairro, count in stats['bairro_distribution'].items():
                        f.write(f"    {bairro}: {count}\n")
            
            f.write("\n" + "=" * 60 + "\n")
            f.write("Relatório gerado automaticamente pelo Sistema de Controle de Qualidade IVV\n")
        
        return report_path

    def run(self):
        """Executa o controle de qualidade"""
        print("=== CONTROLE DE QUALIDADE - PESQUISA IVV ===")
        print("Selecione o arquivo Excel para análise...\n")
        
        file_path = self.select_file()
        
        if not file_path:
            print("Nenhum arquivo selecionado. Operação cancelada.")
            return
        
        print(f"Processando arquivo: {os.path.basename(file_path)}")
        
        errors, statistics, summary_path = self.process_file(file_path)
        
        report_path = self.save_report(file_path, errors, statistics)
        
        print("\n" + "=" * 50)
        print("RESULTADOS DO CONTROLE DE QUALIDADE")
        print("=" * 50)
        
        if errors:
            print(f"❌ {len(errors)} erro(s) encontrado(s):")
            for i, error in enumerate(errors[:10], 1):
                print(f"  {i}. {error}")
            if len(errors) > 10:
                print(f"  ... e mais {len(errors) - 10} erro(s)")
        else:
            print("✅ Nenhum erro encontrado! Arquivo válido.")
        
        print(f"\n📄 Relatório de qualidade salvo em:")
        print(f"   {report_path}")
        
        if summary_path:
            print(f"\n📊 Tabelas resumo geradas em:")
            print(f"   {summary_path}")
            print(f"\n🎨 Formatação aplicada:")
            print(f"   • Empresas inativas (mês atual): texto em VERMELHO")
            print(f"   • Problemas de lógica de oferta: fundo AMARELO")
            print(f"   • Lógica: Oferta(atual) = Oferta(anterior) - Venda(anterior) + Distrato(anterior)")
            print(f"\n📋 Abas adicionais:")
            print(f"   • Lançamentos_Residencial: Detalhes por empresa e bairro")
            print(f"   • Lançamentos_Comercial: Detalhes por empresa e bairro")
        
        root = tk.Tk()
        root.withdraw()
        
        message = "Processamento concluído!\n\n"
        message += f"📄 Relatório de qualidade: {os.path.basename(report_path)}\n"
        
        if summary_path:
            message += f"📊 Tabelas resumo: {os.path.basename(summary_path)}\n"
            message += "\n🎨 Formatação condicional aplicada:\n"
            message += "• Empresas inativas: texto VERMELHO\n"
            message += "• Problemas de lógica: fundo AMARELO\n"
            message += "• Nova lógica considera ofertas esperadas\n"
            message += "\n📋 Inclui análise de lançamentos:\n"
            message += "• Por empresa e por bairro\n"
            message += "• Detalhamento por mês"
        
        if errors:
            message += f"\n\n⚠️ {len(errors)} erro(s) encontrado(s)"
            messagebox.showwarning("Controle de Qualidade", message)
        else:
            message += "\n\n✅ Nenhum erro encontrado!"
            messagebox.showinfo("Controle de Qualidade", message)
        
        root.destroy()

if __name__ == "__main__":
    controller = IVVQualityControl()
    controller.run()
