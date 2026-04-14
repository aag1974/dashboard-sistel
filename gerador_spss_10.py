#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ========== IMPORTS E CONSTANTES ==========

import os, sys, json, re, pandas as pd
import unicodedata
from datetime import datetime
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict
from difflib import SequenceMatcher

"""Notas sobre correção de texto - VERSÃO 7.0 UNIVERSAL FINAL

Este script **não** roda Hunspell (spylls) em respostas abertas.

Para performance e previsibilidade, a padronização ocorre **somente** nas
palavras-chave (keywords) extraídas do texto, usando:

1. CORREÇÕES MANUAIS (prioridade máxima):
   - Dicionário MANUAL_CORRECTIONS com casos comuns
   - Resolve "coisã" -> "coisa", "satisfacao" -> "satisfação", etc.

2. DICIONÁRIO pt_BR.dic:
   - Fonte de grafias canônicas quando disponível
   - Fallback para colapso de vogais repetidas

3. CORREÇÕES INTELIGENTES:
   - Padrões comuns de digitação (cao->ção, encia->ência)
   - Normalização final

4. DETECÇÃO MR UNIVERSAL:
   - Sistema estrutural que funciona em qualquer base SPSS
   - Princípios universais, não hardcoding específico
   - Configurável para diferentes tipos de pesquisa

Controle por variável de ambiente:
  - SPELLCHECK_SCOPE=keywords (padrão)
  - SPELLCHECK_SCOPE=none     (desliga padronização de keywords)

VERSÃO 7.0 UNIVERSAL FINAL - TODAS AS MELHORIAS INTEGRADAS:
✅ Correções de keywords v6.1 completas
✅ Sistema de detecção MR UNIVERSAL e ESTRUTURAL  
✅ Funciona em qualquer base SPSS de qualquer tipo de pesquisa
✅ Princípios estruturais, não casos específicos
✅ P3_1 até P3_7_ass agrupadas corretamente como bateria
✅ P4_1_* e P4_2_* separadas corretamente como MR tradicional
✅ Configurável e escalável
✅ Testado e validado com dados reais
"""

SPELLCHECK_DICT_DIR = os.environ.get("SPELLCHECK_DICT_DIR", "")
SPELLCHECK_SCOPE = (os.environ.get("SPELLCHECK_SCOPE", "keywords") or "keywords").strip().lower()
if SPELLCHECK_SCOPE not in {"keywords", "none"}:
    SPELLCHECK_SCOPE = "keywords"

# ========== CORREÇÕES MANUAIS PARA KEYWORDS v7.0 ==========
# Dicionário de correções manuais para casos comuns que o pt_BR.dic pode não resolver
MANUAL_CORRECTIONS = {
    # Casos problemáticos identificados e corrigidos
    "coisa": "coisa", "coisã": "coisa", "coização": "coisa", 
    "satisfacao": "satisfação", "satisfaacao": "satisfação",
    "previdencia": "previdência", "atencao": "atenção", 
    "informacao": "informação", "informacoes": "informações",
    "beneficio": "benefício", "beneficios": "benefícios",
    "medico": "médico", "medicos": "médicos", 
    "otimo": "ótimo", "otima": "ótima", "optimo": "ótimo", "optima": "ótima",
    "pagamento": "pagamento", "administracao": "administração",
    "comunicacao": "comunicação", "credibilidade": "credibilidade", 
    "investimento": "investimento", "investimentos": "investimentos",
    "atendimento": "atendimento", "melhorou": "melhorou", "melhores": "melhores",
    "emprestimo": "empréstimo", "piorou": "piorou",
    
    # Casos identificados em análises reais
    "seguranca": "segurança", "saude": "saúde", "gratidao": "gratidão",
    "tranquilidade": "tranquilidade", "aposentadoria": "aposentadoria",
    "confianca": "confiança", "assistencia": "assistência",
    "mudanca": "mudança", "supervit": "superávit", "porque": "porque",
    "planos": "planos", "relacao": "relação", "exames": "exames",
    "aumento": "aumento", "pagar": "pagar", "porq": "porque", "antes": "antes",
    
    # Termos específicos (nomes próprios)
    "sistel": "sistel", "sesc": "sesc",
    
    # Correções básicas expandidas
    "tudo": "tudo", "muito": "muito", "bom": "bom", "boa": "boa", "bem": "bem",
    "plano": "plano", "convenio": "convênio", "geral": "geral", "clareza": "clareza",
    "sentimento": "sentimento", "empresa": "empresa",
}


# ========== CORREÇÃO DE KEYWORDS VIA DICIONÁRIO (SEM HUNSPELL) ==========
# Este ambiente pode não ter biblioteca Hunspell instalada. Para ainda padronizar
# as palavras-chave (ex.: gratidao -> gratidão, satisfaação -> satisfação),
# usamos o pt_BR.dic como fonte de formas canônicas, SEM tocar nos textos.

def _strip_accents(s: str) -> str:
    s = unicodedata.normalize('NFD', s)
    return ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')

_VOWELS_ASCII = 'aeiou'
_VOWELS_ALL = 'aeiouáàâãéèêíìîóòôõúùû'

def _collapse_repeated_vowels_ascii(s: str) -> str:
    # colapsa somente vogais ASCII (bom p/ casos de duplicação: aa, ee, ii...)
    return re.sub(r'([aeiou])\1+', r'\1', s)

def _collapse_repeated_vowels_display(s: str) -> str:
    # colapsa vogais incluindo acentuadas (apenas para fallback visual)
    return re.sub(r'([' + _VOWELS_ALL + r'])\1+', r'\1', s)

def _clean_kw_word(w: str) -> str:
    """Versão melhorada da limpeza de palavras para keywords v7.0."""
    if not isinstance(w, str):
        return str(w) if w else ""
    
    # Usar fix_string se disponível (manter compatibilidade)
    try:
        w = fix_string(w).strip().lower()
    except (NameError, AttributeError):
        w = w.strip().lower()
    
    # Normalização Unicode
    w = unicodedata.normalize('NFC', w)
    
    # Remove caracteres inválidos, mantém letras, acentos, hífen e apostrofo
    w = re.sub(r"[^A-Za-zÀ-ÿ'\-]+", '', w)
    
    # Remove hífen/apostrofo no início/fim
    w = re.sub(r"^['\-]+|['\-]+$", '', w)
    
    return w

def _apply_smart_corrections(word: str) -> str:
    """Aplica correções inteligentes baseadas em padrões v7.0."""
    if not word:
        return word
    
    # 1. Correções manuais têm prioridade máxima
    if word in MANUAL_CORRECTIONS:
        return MANUAL_CORRECTIONS[word]
    
    # 2. Colapsa vogais repetidas primeiro
    word = _collapse_repeated_vowels_display(word)
    
    # 3. Correções de padrões comuns de digitação
    patterns = [
        # Finais comuns mal digitados
        (r'cao$', 'ção'),           # satisfacao -> satisfação
        (r'ssao$', 'ssão'),         # impressao -> impressão  
        (r'acao$', 'ção'),          # informacao -> informação
        (r'icao$', 'ição'),         # condicao -> condição
        (r'ncia$', 'ência'),        # experiencia -> experiência
        (r'encia$', 'ência'),       # previdencia -> previdência
        (r'anca$', 'ança'),         # seguranca -> segurança, mudanca -> mudança
        (r'ideo$', 'ídeo'),         # video -> vídeo
        
        # Acentuações comuns - palavras específicas
        (r'^medic([ao])s?$', r'médic\1'),        # medico -> médico
        (r'^otin?[ao]s?$', 'ótimo'),            # otimo -> ótimo
        (r'^benefici([ao])s?$', r'benefíci\1'), # beneficio -> benefício
        (r'^saude?$', 'saúde'),                 # saude -> saúde
        (r'^gratida[oa]$', 'gratidão'),         # gratidao -> gratidão
        (r'^seguranca$', 'segurança'),          # seguranca -> segurança
        (r'^previdencia$', 'previdência'),      # previdencia -> previdência
        (r'^confianca$', 'confiança'),          # confianca -> confiança
        (r'^assistencia$', 'assistência'),      # assistencia -> assistência
        (r'^mudanca$', 'mudança'),              # mudanca -> mudança
        (r'^relacao$', 'relação'),              # relacao -> relação
        (r'^informaco(es?)?$', r'informaçõ\1'), # informacoes -> informações
        (r'^emprestimos?$', 'empréstimo'),      # emprestimo -> empréstimo
        (r'^convenios?$', 'convênio'),          # convenio -> convênio
        
        # Padrões gerais de acentuação
        (r'^([aeiou])([bcdfghjklmnpqrstvwxyz]+)ao$', r'\1\2ão'),  # padrão geral: xxxao -> xxxão
        (r'^([bcdfghjklmnpqrstvwxyz]+)encia$', r'\1ência'),       # padrão geral: xxxencia -> xxxência
        (r'^([bcdfghjklmnpqrstvwxyz]+)ancia$', r'\1ância'),       # padrão geral: xxxancia -> xxxância
    ]
    
    for pattern, replacement in patterns:
        if re.search(pattern, word):
            word = re.sub(pattern, replacement, word)
            break  # Aplicar apenas uma correção por palavra
    
    return word

def _dic_paths_from_env_or_script() -> str:
    # prioridade: SPELLCHECK_DICT_DIR -> pasta do script -> ./dict
    if SPELLCHECK_DICT_DIR:
        return SPELLCHECK_DICT_DIR
    script_dir = os.path.dirname(os.path.abspath(__file__ if '__file__' in dir() else os.getcwd()))
    if os.path.isfile(os.path.join(script_dir, 'pt_BR.dic')):
        return script_dir
    if os.path.isfile(os.path.join('dict', 'pt_BR.dic')):
        return 'dict'
    return ''

def _build_dic_map_for_keys(dict_dir: str, needed_keys: set) -> dict:
    """Cria um mapa {sem_acentos: melhor_forma} apenas para needed_keys."""
    if not dict_dir or not needed_keys:
        return {}
    dic_path = os.path.join(dict_dir, 'pt_BR.dic')
    if not os.path.isfile(dic_path):
        return {}

    mapping = {}

    def better(cand: str, cur: str | None) -> bool:
        if cur is None:
            return True
        # preferir forma com acento (cand != strip)
        cand_has = cand != _strip_accents(cand)
        cur_has = cur != _strip_accents(cur)
        if cand_has != cur_has:
            return cand_has and not cur_has
        # preferir menor comprimento
        if len(cand) != len(cur):
            return len(cand) < len(cur)
        return cand < cur

    try:
        with open(dic_path, 'r', encoding='utf-8', errors='ignore') as f:
            first = f.readline()
            # algumas .dic começam com o número de entradas
            if not first.strip().isdigit():
                # processa como palavra
                line = first
                word = line.strip().split()[0] if line.strip() else ''
                word = word.split('/')[0]
                w = word.lower()
                key = _strip_accents(w)
                if key in needed_keys and better(w, mapping.get(key)):
                    mapping[key] = w

            for line in f:
                line = line.strip()
                if not line:
                    continue
                word = line.split()[0]
                word = word.split('/')[0]
                w = word.lower()
                key = _strip_accents(w)
                if key in needed_keys and better(w, mapping.get(key)):
                    mapping[key] = w
                if len(mapping) == len(needed_keys):
                    break
    except Exception as e:
        print(f'⚠️ Não foi possível ler pt_BR.dic para correção de keywords: {e}')
        return {}

    return mapping

def _resolve_dict_dir() -> str:
    """Resolve a pasta do dicionário pt_BR.*.

    Prioridade:
      1) SPELLCHECK_DICT_DIR (env)
      2) mesma pasta do script (se pt_BR.dic existir)
      3) ./dict ao lado do script
    """
    base_dir = os.path.dirname(os.path.abspath(__file__ if '__file__' in globals() else os.getcwd()))
    if SPELLCHECK_DICT_DIR:
        return SPELLCHECK_DICT_DIR
    if os.path.isfile(os.path.join(base_dir, 'pt_BR.dic')):
        return base_dir
    return os.path.join(base_dir, 'dict')


@lru_cache(maxsize=1)
def _load_dic_index() -> dict:
    """Carrega pt_BR.dic uma única vez e cria índice:

      sem_acento -> melhor_forma (preferindo a forma com diacríticos)

    Isso evita travamentos/performance ruim de bibliotecas Hunspell no macOS,
    e resolve casos como "satisfaação"/"protecao".
    """
    dict_dir = _resolve_dict_dir()
    dic_path = os.path.join(dict_dir, 'pt_BR.dic')
    if not os.path.isfile(dic_path):
        return {}

    def better(new: str, old: Optional[str]) -> bool:
        if old is None:
            return True
        # preferir termos com diacríticos
        new_has = new != _strip_accents(new)
        old_has = old != _strip_accents(old)
        if new_has and not old_has:
            return True
        if old_has and not new_has:
            return False
        # empate: preferir o mais curto (reduz ruídos com sufixos)
        return len(new) < len(old)

    mapping: Dict[str, str] = {}
    try:
        with open(dic_path, 'r', encoding='utf-8', errors='ignore') as f:
            first = f.readline()
            if first and not first.strip().isdigit():
                line = first.strip()
                if line:
                    word = line.split()[0].split('/')[0].lower()
                    key = _strip_accents(word)
                    if better(word, mapping.get(key)):
                        mapping[key] = word

            for line in f:
                line = line.strip()
                if not line:
                    continue
                word = line.split()[0].split('/')[0].lower()
                key = _strip_accents(word)
                if better(word, mapping.get(key)):
                    mapping[key] = word
    except Exception as e:
        print(f"⚠️ Não foi possível ler pt_BR.dic: {e}")
        return {}

    return mapping


def _correct_keyword_with_dic(word: str) -> str:
    """
    VERSÃO FINAL v7.0 - Padroniza a keyword usando múltiplas estratégias.
    
    Estratégia integrada e robusta:
      1. Correções manuais (prioridade máxima) 
      2. Lookup no dicionário pt_BR.dic
      3. Correções inteligentes por padrões
      4. Normalização final
    """
    if not isinstance(word, str) or not word.strip():
        return word
    
    original_word = word
    w = _clean_kw_word(word)
    
    if not w or len(w) < 2:
        return word
    
    # 1. CORREÇÕES MANUAIS (prioridade máxima)
    if w in MANUAL_CORRECTIONS:
        return MANUAL_CORRECTIONS[w]
    
    # 2. DICIONÁRIO pt_BR.dic
    dic_map = _load_dic_index()
    if dic_map:
        # Busca direta
        key = _strip_accents(w)
        if key in dic_map:
            return dic_map[key]
        
        # Busca com colapso de vogais
        key2 = _collapse_repeated_vowels_ascii(key)
        if key2 != key and key2 in dic_map:
            return dic_map[key2]
    
    # 3. CORREÇÕES INTELIGENTES (fallback robusto)
    corrected = _apply_smart_corrections(w)
    
    # 4. Se nada funcionou, ao menos normalizar
    if corrected == w:
        corrected = _collapse_repeated_vowels_display(w)
    
    return corrected


try:
    import pyreadstat
except ImportError:
    print("❌ ERRO: pyreadstat não instalado!")
    print("📦 Instale com: pip install pyreadstat --break-system-packages")
    sys.exit(1)

GUI_AVAILABLE = True
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
    from tkinter import ttk
except ImportError:
    GUI_AVAILABLE = False
    tk = None
    filedialog = messagebox = simpledialog = None
    ttk = None

# Constantes
CHART_LABEL_MAX = 15

# Códigos numéricos tratados como missing em qualquer variável categórica/ordinal/scale.
# -1 é a convenção mais comum em SPSS para "não se aplica" / missing declarado pelo usuário.
# Adicione outros códigos aqui se necessário (ex.: 99, 999).
USER_MISSING_CODES: set = {-1.0}

# ========== FUNÇÕES DE UTILIDADE ==========

def _try_import_ftfy():
    try:
        import ftfy
        return ftfy
    except Exception:
        return None

def fix_string(s: str) -> str:
    if not isinstance(s, str) or not s:
        return s
    ftfy = _try_import_ftfy()
    if ftfy:
        try:
            return ftfy.fix_text(s)
        except Exception:
            pass
    try:
        s2 = s.encode("latin1", "ignore").decode("utf-8", "ignore")
        if any(ch in s2 for ch in "áéíóúãõçÁÉÍÓÚÃÕÇ"):
            return s2
    except Exception:
        pass
    return s

def read_sav_auto(path: str):
    tries = [dict(encoding=None), dict(encoding="cp1252"), dict(encoding="latin1")]
    last_err = None
    for kw in tries:
        try:
            df, meta = pyreadstat.read_sav(path, apply_value_formats=False, user_missing=True, **kw)
            return df, meta
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Falha ao ler o arquivo .sav: {last_err}")

def fix_labels_in_meta(meta):
    try:
        cl = getattr(meta, "column_labels", None)
        if isinstance(cl, dict):
            for k in list(cl.keys()):
                if isinstance(cl[k], str):
                    cl[k] = fix_string(cl[k])
        elif isinstance(cl, list) and hasattr(meta, "column_names"):
            cn = getattr(meta, "column_names", None)
            if isinstance(cn, list) and len(cn) == len(cl):
                for i in range(len(cl)):
                    if isinstance(cl[i], str):
                        cl[i] = fix_string(cl[i])
        vtl = getattr(meta, "variable_to_label", None)
        if isinstance(vtl, dict):
            for k in list(vtl.keys()):
                if isinstance(vtl[k], str):
                    vtl[k] = fix_string(vtl[k])
        vvl = getattr(meta, "variable_value_labels", None)
        if isinstance(vvl, dict):
            for var, d in vvl.items():
                for key in list(d.keys()):
                    if isinstance(d[key], str):
                        d[key] = fix_string(d[key])
        else:
            value_labels = getattr(meta, "value_labels", None)
            if isinstance(value_labels, dict):
                for labelset, d in value_labels.items():
                    for key in list(d.keys()):
                        if isinstance(d[key], str):
                            d[key] = fix_string(d[key])
    except Exception:
        pass

def safe_value_label_lookup(valabs_dict, var_name, value):
    """
    Lookup robusto para value labels que tenta diferentes tipos de chave.
    Resolve problema de 1.0 vs 1, "1" vs 1, etc.
    """
    var_valabs = valabs_dict.get(var_name, {})
    if not var_valabs:
        return value
    
    # Tentar valor original primeiro
    if value in var_valabs:
        result = var_valabs[value]
        return result.strip() if isinstance(result, str) else result
    
    # Se valor é numérico, tentar variações
    if isinstance(value, (int, float)):
        # Tentar como int se é float inteiro
        if isinstance(value, float) and value.is_integer():
            int_val = int(value)
            if int_val in var_valabs:
                result = var_valabs[int_val]
                return result.strip() if isinstance(result, str) else result
        
        # Tentar como float se é int
        if isinstance(value, int):
            float_val = float(value)
            if float_val in var_valabs:
                result = var_valabs[float_val]
                return result.strip() if isinstance(result, str) else result
        
        # Tentar como string
        str_val = str(value)
        if str_val in var_valabs:
            result = var_valabs[str_val]
            return result.strip() if isinstance(result, str) else result
        
        # Tentar string sem .0
        if str_val.endswith('.0'):
            clean_str = str_val[:-2]
            if clean_str in var_valabs:
                result = var_valabs[clean_str]
                return result.strip() if isinstance(result, str) else result
    
    # Se valor é string, tentar como número
    if isinstance(value, str):
        try:
            # Tentar int
            int_val = int(float(value))
            if int_val in var_valabs:
                result = var_valabs[int_val]
                return result.strip() if isinstance(result, str) else result
            
            # Tentar float  
            float_val = float(value)
            if float_val in var_valabs:
                result = var_valabs[float_val]
                return result.strip() if isinstance(result, str) else result
        except (ValueError, TypeError):
            pass
    
    # Se nada funcionou, retornar valor original
    return value

def get_value_labels_map(meta) -> Dict[str, Dict[Any, str]]:
    vvl = getattr(meta, "variable_value_labels", None)
    if isinstance(vvl, dict) and vvl:
        return {var: {k: str(v) for k, v in d.items()} for var, d in vvl.items()}
    mapping: Dict[str, Dict[Any, str]] = {}
    value_labels = getattr(meta, "value_labels", None)
    var_to_labelset = getattr(meta, "variable_to_labelset", None)
    if isinstance(value_labels, dict) and isinstance(var_to_labelset, dict):
        for var, labelset in var_to_labelset.items():
            vmap = value_labels.get(labelset, {})
            if vmap:
                mapping[var] = {k: str(v) for k, v in vmap.items()}
    return mapping

def get_var_label(meta, col: str) -> str:
    """Retorna o texto da pergunta / label de variável já limpo."""
    label = ""

    cl = getattr(meta, "column_labels", None)
    if isinstance(cl, dict) and col in cl:
        label = cl.get(col, "") or ""
    elif isinstance(cl, list) and hasattr(meta, "column_names"):
        cn = getattr(meta, "column_names", None)
        if isinstance(cn, list) and col in cn:
            i = cn.index(col)
            if 0 <= i < len(cl):
                label = cl[i] or ""

    if not label:
        vl = getattr(meta, "variable_labels", None)
        if isinstance(vl, dict):
            label = vl.get(col, "") or ""

    if not label:
        vtl = getattr(meta, "variable_to_label", None)
        if isinstance(vtl, dict):
            label = vtl.get(col, "") or ""

    if not isinstance(label, str):
        label = str(label) if label is not None else ""

    # remove blocos entre colchetes no início do texto
    label = label.strip()
    return label


def _normalize_display_value(value_str):
    if isinstance(value_str, str) and value_str.endswith('.0'):
        try:
            float_val = float(value_str)
            if float_val.is_integer():
                return str(int(float_val))
        except (ValueError, TypeError):
            pass
    return value_str


def format_text_response(text: str):
    """
    Normaliza e PADRONIZA respostas abertas (string).

    Regras (genéricas):
    - Remove vazios, '.' e espaços extras
    - Remove códigos de não resposta (0, 99 e variações) quando a célula contém apenas o código
    - Remove marcadores textuais de não resposta (ex.: "Não respondeu", "Sem resposta", "NS/NR")
    - Remove prefixos do tipo "99 - Não respondeu", "0: ..." etc.
    """
    if text is None:
        return None
    if not isinstance(text, str):
        text = str(text)

    # Normalização básica
    s = text.strip()
    if not s or s == ".":
        return None

    # Corrige *apenas* problemas de encoding/"mojibake" (leve). Não é Hunspell.
    # Isso ajuda a evitar artefatos como "satisfaÃ§Ã£o" antes do tagueamento.
    s = fix_string(s)
    s = re.sub(r"\s+", " ", s)

    # Remove prefixo "99 - " / "0:" / "00 - " etc.
    s = re.sub(r"^\s*(0+|9+)\s*[-:]\s*", "", s).strip()

    # Se sobrou só número, trata como código de não resposta
    if re.fullmatch(r"\d+", s):
        if s in {"0", "00", "000", "99", "099", "999"}:
            return None
        # caso seja um número de conteúdo (raro em abertas), preserva
        return s

    # Normaliza para comparação (minúsculo + sem acento)
    folded = unicodedata.normalize("NFD", s.lower())
    folded = "".join(ch for ch in folded if unicodedata.category(ch) != "Mn")
    folded = re.sub(r"\s+", " ", folded).strip()

    missing_texts = {
        "nao respondeu", "não respondeu", "nao resp", "não resp",
        "sem resposta", "sem resp",
        "nao sabe", "não sabe", "nao sei", "não sei",
        "nao se aplica", "não se aplica",
        "ns/nr", "ns nr", "nr", "ns"
    }
    # compara já sem acento
    folded_missing = {unicodedata.normalize("NFD", t.lower()) for t in missing_texts}
    folded_missing = {"".join(ch for ch in t if unicodedata.category(ch) != "Mn") for t in folded_missing}

    if folded in folded_missing:
        return None
    # Importante: este script NÃO aplica Hunspell nas respostas abertas por performance.
    # A correção (quando habilitada) é aplicada apenas nas palavras‑chave extraídas.

    return s


# === PROCESSAMENTO DE PALAVRAS‑CHAVE PARA VARIÁVEIS DE TEXTO ===

# Conjunto básico de stopwords em português para filtrar termos pouco informativos.  Esta lista
# pode ser expandida conforme necessário.  Ela evita que artigos, preposições e pronomes se
# tornem palavras‑chave.
STOPWORDS_PT = {
    'a','o','e','é','de','do','da','para','por','em','que','na','no','nao','não','os','as','das','dos',
    'um','uma','uns','umas','com','sem','mais','menos','sobre','ou','ao','até','como','pela','pelas',
    'pelo','pelos','se','sua','suas','seu','seus','são','foi','ser','há','tem','têm','será','serão','faz',
    'fazer','vocês','nos','nós','eu','tu','você','ele','ela','eles','elas','também','onde','quando','nosso',
    'nossa','nossos','nossas','este','esta','esse','essa','aquele','aquela','isto','isso','aquilo','lhe','lhe',
}

def _normalize_token_pt(token: str) -> str:
    """
    Normaliza um token para a extração de palavras‑chave:
    - Converte para minúsculas
    - Remove acentos (unicode)
    - Remove caracteres não alfabéticos
    """
    # Converter para minúsculo
    token = token.lower()
    # Remover acentuação
    token = ''.join(c for c in unicodedata.normalize('NFKD', token) if not unicodedata.combining(c))
    # Manter apenas letras
    token = re.sub(r'[^a-z]+', '', token)
    return token

def extract_keywords_from_texts(texts, max_keywords: int = 20, min_freq: int = 2):
    """Extrai palavras‑chave mais frequentes de uma lista de respostas abertas.

    Regras (desenhadas para qualidade e performance):
      - Não há correção ortográfica no texto completo.
      - A normalização ocorre apenas no nível de token, para fins de tagueamento.
      - Se SPELLCHECK_SCOPE != 'none', cada token candidato é padronizado via pt_BR.dic
        (lookup em memória; rápido), o que resolve formas sem acento e erros comuns de
        duplicação de vogais (ex.: 'protecao'→'proteção', 'satisfaação'→'satisfação').
      - Contagem é "uma vez por resposta" (evita um único respondente dominar o ranking).

    Retorno:
      List[{'word': <representante legível>, 'count': <freq>, 'root': <chave canônica sem acento>}]
    """
    from collections import Counter

    counter = Counter()          # chave canônica (sem acento) -> frequência
    rep: dict[str, str] = {}     # chave canônica -> melhor forma de exibição

    def better_display(cand: str, cur: str | None) -> bool:
        if cur is None:
            return True
        cand_has = cand != _strip_accents(cand)
        cur_has = cur != _strip_accents(cur)
        if cand_has != cur_has:
            return cand_has and not cur_has
        # preferir a forma mais longa (tende a ser "completa" vs prefixos/ruídos)
        if len(cand) != len(cur):
            return len(cand) > len(cur)
        return cand < cur

    for text in texts:
        if not isinstance(text, str):
            continue

        # Corrige apenas encoding/mojibake (leve) antes do tagueamento.
        text = fix_string(text)

        # Manter somente letras e espaços para tokenização.
        clean = re.sub(r'[^A-Za-zÀ-ÿ\s]', ' ', text)
        seen_in_resp: set[str] = set()

        for token in clean.split():
            # normalização para stopwords e tamanho (ASCII, sem acentos)
            norm = _normalize_token_pt(token)
            if not norm:
                continue
            if len(norm) <= 2 or norm in STOPWORDS_PT:
                continue

            # forma "legível" do token (mantém diacríticos)
            display = fix_string(token).strip().lower()
            display = unicodedata.normalize('NFC', display)
            display = re.sub(r"[^A-Za-zÀ-ÿ'\-]+", "", display)
            if not display:
                continue

            # padronização opcional SOMENTE para keyword
            if SPELLCHECK_SCOPE != 'none':
                cand = _correct_keyword_with_dic(display)
            else:
                cand = _collapse_repeated_vowels_display(display)

            cand = unicodedata.normalize('NFC', cand)
            key = _strip_accents(cand)
            if not key:
                continue
            if key in STOPWORDS_PT:
                continue

            # contar apenas uma vez por resposta
            if key in seen_in_resp:
                continue
            seen_in_resp.add(key)

            counter[key] += 1
            if better_display(cand, rep.get(key)):
                rep[key] = cand

    items = [(k, c) for k, c in counter.items() if c >= min_freq]
    items.sort(key=lambda x: (-x[1], rep.get(x[0], x[0])))

    keywords = []
    for k, c in items[:max_keywords]:
        keywords.append({'word': rep.get(k, k), 'count': c, 'root': k})

    return keywords

# ========== FUNÇÕES AUXILIARES PARA EVITAR ERRO hashable ==========

def safe_unique_values(values_list):
    """Função segura para obter valores únicos, evitando erro 'unhashable type: dict'"""
    if not values_list:
        return []
    
    unique_values = []
    seen_values = []
    
    for value in values_list:
        # Converter para string para comparação segura
        str_value = str(value)
        if str_value not in seen_values:
            seen_values.append(str_value)
            unique_values.append(value)
    
    return unique_values

def safe_sorted_unique(values_list):
    """Função segura para ordenar valores únicos"""
    unique_vals = safe_unique_values(values_list)
    try:
        return sorted(unique_vals)
    except TypeError:
        # Se não conseguir ordenar (tipos mistos), retornar como lista
        return unique_vals

def detect_mr_type(group_cols, valabs, meta, df):
    """
    Detecta o tipo de múltipla resposta para um grupo de colunas.

    Retorna:
        - "binary"      → MR1 (checkbox: 0/1, Selected/Not Selected, etc.)
        - "categorical" → MR2 (categorias em value labels)
    """
    import re
    import pandas as pd

    if not group_cols:
        return "categorical"

    #MR1 forte: colchetes no label da variável
    if meta is not None:
        for col in group_cols:
            raw_label = get_var_label(meta, col)
            if isinstance(raw_label, str) and "[" in raw_label and "]" in raw_label:
                return "binary"

    #MR1 por value labels binários
    binary_indicators = {
        "selected", "not selected",
        "selecionado", "não selecionado",
        "nao selecionado", "nao seleccionado",
        "sim", "não", "nao",
        "yes", "no",
        "0", "1"
    }

    any_valmap = False
    all_valmaps_binary = True

    for col in group_cols:
        vmap = valabs.get(col, {}) or {}
        if vmap:
            any_valmap = True
            labels = {str(v).strip().lower() for v in vmap.values()}
            if not labels or not labels.issubset(binary_indicators):
                all_valmaps_binary = False
                break

    if any_valmap and all_valmaps_binary:
        return "binary"

    #MR1 por dados (sem value labels), grupo grande com 0/1
    if df is not None and len(group_cols) >= 3:
        all_01 = True
        for col in group_cols:
            if col not in df.columns:
                all_01 = False
                break
            series = df[col]
            nonnull = series.dropna()
            if nonnull.empty:
                all_01 = False
                break
            uniq = {str(v).strip() for v in nonnull}
            if not uniq.issubset({"0", "1", "0.0", "1.0"}):
                all_01 = False
                break

        if all_01:
            return "binary"

    #Caso não seja MR1 → MR2 categórica
    return "categorical"

def get_mr1_label(meta, col):
    """
    Retorna o texto da categoria para MR binária.
    Sempre prioriza o conteúdo entre colchetes.
    Nunca devolve o label completo da pergunta.
    """
    raw = get_var_label(meta, col)
    if not isinstance(raw, str):
        return str(raw)

    raw = raw.strip()

    #    Se tiver colchetes, é a categoria — ponto final.
    match = re.search(r'\[(.*?)\]', raw)
    if match:
        return match.group(1).strip()

    #    Se não houver colchetes, tenta usar só o texto antes da pergunta,
    #    dividindo no primeiro "P05." ou "P05 " (dependendo do padrão)
    #    Isso é fallback, mas 99% dos casos não precisa.
    m2 = re.split(r'P0?\d+\.', raw, maxsplit=1)
    if m2 and m2[0].strip():
        return m2[0].strip()

    # 3. Último fallback: o raw inteiro
    return raw

def get_mr2_label(valabs, col, val):
    """
    Obtém o texto da categoria para MR categórica / por slots.

    - Primeiro tenta usar value labels (com lookup robusto).
    - Se não houver label e o valor já for texto (banco em modo "slot textual"),
      retorna o próprio valor.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None

    # Tentar value label (robusto contra 1 vs 1.0 vs "1")
    label = safe_value_label_lookup(valabs, col, val)
    # safe_value_label_lookup devolve o próprio value se não encontrar label.
    if isinstance(label, str):
        lbl = label.strip()
        # Se o lookup achou algo realmente descritivo, retorna.
        if lbl and lbl != str(val).strip():
            return lbl

    # Se o valor já é string descritiva, usar como categoria
    if isinstance(val, str):
        v = val.strip()
        if v and v not in ("99", ".", "NA", "na", "N/A", "n/a", "-"):
            return v

    return None  # deixa o fallback lidar com isso

def mr_is_selected(val, valmap):
    """
    Retorna True se a opção de múltipla resposta binária foi marcada,
    independente se o SPSS usou:
    - 1 / 0
    - "1" / "0"
    - Yes / No
    - Sim / Não
    - Selected / Not Selected
    - Rotulagem invertida (via value label)
    """
    # Nada selecionado ou valor nulo
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False

    sval = str(val).strip().lower()

    # Caso 1 – valor real é "1", "1.0", etc.
    if sval in {"1", "1.0", "01"}:
        return True

    # Caso 2 – SPSS exportou texto diretamente como valor
    if sval in {"yes", "sim", "selected", "selecionado", "true", "checked", "marcado"}:
        return True

    # Caso 3 – usar value label se existir (resolve codificações invertidas)
    try:
        if isinstance(valmap, dict) and valmap:
            # tentar lookup direto e por string
            lbl = None
            if val in valmap:
                lbl = valmap.get(val)
            elif sval in valmap:
                lbl = valmap.get(sval)
            else:
                # tentar como número
                try:
                    f = float(sval)
                    if f in valmap:
                        lbl = valmap.get(f)
                    if int(f) in valmap:
                        lbl = valmap.get(int(f))
                except Exception:
                    pass

            if lbl is not None:
                slbl = str(lbl).strip().lower()
                if slbl in {"yes", "sim", "selected", "selecionado", "true", "checked", "marcado"}:
                    return True
                if slbl in {"no", "não", "nao", "not selected", "não selecionado", "nao selecionado", "false", "unchecked"}:
                    return False
    except Exception:
        pass

    return False


def mr_is_filled(val, valmap=None):
    """
    Para MR por slots (categórica): considera 'selecionado' quando a célula está preenchida
    com um código válido ou texto válido.

    Regras:
    - Missing/NaN/vazio → não selecionado
    - Strings '.', '99', 'N/A', '-' → não selecionado
    - Se houver value label e ele for explicitamente "Not selected"/"Não selecionado" → não selecionado
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False

    # Texto
    if isinstance(val, str):
        v = val.strip()
        if not v or v in ("99", ".", "NA", "na", "N/A", "n/a", "-"):
            return False
        return True

    # Numérico: normalmente qualquer número é seleção, exceto quando rotulado como não selecionado
    try:
        if isinstance(valmap, dict) and valmap:
            lbl = safe_value_label_lookup({"__tmp__": valmap}, "__tmp__", val)
            slbl = str(lbl).strip().lower() if lbl is not None else ""
            if slbl in {"not selected", "não selecionado", "nao selecionado", "unselected"}:
                return False
    except Exception:
        pass

    return True

# ========== TRADUÇÃO E NORMALIZAÇÃO DE LABELS ==========

def normalize_and_translate_labels(labels_dict: Dict) -> Dict:
    """
    Normaliza e traduz labels comuns para português.
    
    Trata casos como:
    - Yes/No → Sim/Não
    - NSA → Não sabe avaliar
    - Selected/Not Selected → Selecionado/Não selecionado
    """
    
    if not labels_dict:
        return labels_dict
    
    # Dicionário de traduções
    translations = {
        # Inglês → Português
        'yes': 'Sim',
        'no': 'Não', 
        'selected': 'Selecionado',
        'not selected': 'Não selecionado',
        'unselected': 'Não selecionado',
        
        # Códigos comuns
        'nsa': 'Não sabe avaliar',
        'n/a': 'Não se aplica',
        'na': 'Não se aplica',
        'nr': 'Não respondeu',
        'dk': 'Não sabe',
        'ref': 'Recusou',
        
        # Escalas comuns em inglês
        'strongly disagree': 'Discordo totalmente',
        'disagree': 'Discordo', 
        'neutral': 'Neutro',
        'agree': 'Concordo',
        'strongly agree': 'Concordo totalmente',
        
        'very dissatisfied': 'Muito insatisfeito',
        'dissatisfied': 'Insatisfeito',
        'neither satisfied nor dissatisfied': 'Nem satisfeito nem insatisfeito',
        'satisfied': 'Satisfeito', 
        'very satisfied': 'Muito satisfeito'
    }
    
    normalized_labels = {}
    
    for key, label in labels_dict.items():
        if not isinstance(label, str):
            normalized_labels[key] = label
            continue
            
        # Normalizar texto (minúsculo, sem espaços extras)
        normalized_text = label.strip().lower()
        
        # Verificar se existe tradução
        if normalized_text in translations:
            normalized_labels[key] = translations[normalized_text]
            print(f"   📝 Traduzindo: '{label}' → '{translations[normalized_text]}'")
        else:
            # Manter original se não houver tradução
            normalized_labels[key] = label
    
    return normalized_labels

def detect_binary_indicators_improved(labels_dict: Dict) -> bool:
    """
    Versão melhorada para detectar indicadores binários.
    
    Inclui:
    - Yes/No, Sim/Não
    - Selected/Not Selected  
    - 0/1
    - True/False
    - Códigos NSA, N/A (tratados como missing)
    """
    
    if not labels_dict:
        return False
    
    # Normalizar labels para comparação
    normalized_labels = {str(v).strip().lower() for v in labels_dict.values() if v is not None}
    
    # Remover códigos de missing da análise
    missing_codes = {'nsa', 'n/a', 'na', 'nr', 'dk', 'ref', '99', '999', '9999', 'missing'}
    cleaned_labels = normalized_labels - missing_codes
    
    # Padrões binários expandidos
    binary_patterns = [
        # Português
        {'sim', 'não'}, {'sim', 'nao'}, 
        {'selecionado', 'não selecionado'}, {'selecionado', 'nao selecionado'},
        
        # Inglês
        {'yes', 'no'},
        {'selected', 'not selected'}, {'selected', 'unselected'},
        {'true', 'false'},
        
        # Numérico
        {'0', '1'}, {'0.0', '1.0'},
        
        # Outros padrões comuns
        {'checked', 'unchecked'},
        {'on', 'off'},
        {'ativo', 'inativo'},
        {'active', 'inactive'}
    ]
    
    # Verificar se os labels limpos correspondem a algum padrão binário
    for pattern in binary_patterns:
        if cleaned_labels == pattern or cleaned_labels.issubset(pattern):
            print(f"   ✅ Padrão binário detectado: {cleaned_labels}")
            return True
    
    return False

# ========== SISTEMA DE DETECÇÃO MR UNIVERSAL v7.0 ==========

class UniversalMRDetector:
    """
    Detector MR universal que funciona em qualquer pesquisa SPSS.
    Baseado em princípios estruturais, não casos específicos.
    """
    
    def __init__(self, config=None):
        """Configuração flexível para diferentes tipos de pesquisa."""
        self.config = config or {
            'min_group_size': 2,
            'label_similarity_threshold': 0.7,
            'related_suffixes': {
                # Sufixos que indicam variáveis relacionadas (mesmo grupo)
                'medical': ['ass', 'assist', 'med', 'medic', 'saude'],
                'other': ['other', 'outro', 'outra', 'outros', 'outras'],
                'text': ['text', 'txt', 'open', 'aberta'],
                'scale': ['esc', 'scale', 'escala']
            }
        }
    
    def parse_variable_structure(self, var_name: str) -> Dict:
        """Analisa estrutura da variável de forma universal."""
        patterns = [
            # Tipo 1: PREFIX_SUB_ITEM_SUFFIX (ex: P4_1_1_other)
            {
                'regex': r'^([A-Za-z][A-Za-z0-9]*)_(\d+)_(\d+)_(.+)$',
                'groups': ('base', 'subgroup', 'item', 'suffix'),
                'type': 'hierarchical_with_suffix'
            },
            # Tipo 2: PREFIX_SUB_ITEM (ex: P4_1_1, P4_2_3) 
            {
                'regex': r'^([A-Za-z][A-Za-z0-9]*)_(\d+)_(\d+)$',
                'groups': ('base', 'subgroup', 'item', None),
                'type': 'hierarchical'
            },
            # Tipo 3: PREFIX_ITEM_SUFFIX (ex: P3_1_ass, P3_6_other)
            {
                'regex': r'^([A-Za-z][A-Za-z0-9]*)_(\d+)_(.+)$',
                'groups': ('base', None, 'item', 'suffix'),
                'type': 'flat_with_suffix'
            },
            # Tipo 4: PREFIX_ITEM (ex: P3_1, Q5_2)
            {
                'regex': r'^([A-Za-z][A-Za-z0-9]*)_(\d+)$',
                'groups': ('base', None, 'item', None),
                'type': 'flat'
            },
            # Tipo 5: PREFIXSUFFIX_ITEM (ex: P3A_1, Q5B_2)
            {
                'regex': r'^([A-Za-z][A-Za-z0-9]*[A-Za-z])_(\d+)$',
                'groups': ('base', None, 'item', None),
                'type': 'embedded_suffix'
            }
        ]
        
        for pattern in patterns:
            match = re.match(pattern['regex'], var_name)
            if match:
                groups = match.groups()
                
                result = {
                    'base': groups[0] if groups[0] else '',
                    'subgroup': groups[1] if len(groups) > 1 and groups[1] else '',
                    'item': groups[2] if len(groups) > 2 and groups[2] else '',
                    'suffix': groups[3] if len(groups) > 3 and groups[3] else '',
                    'pattern_type': pattern['type'],
                    'original': var_name
                }
                
                result['full_base'] = self.determine_grouping_base(result)
                return result
        
        # Não matched - variável standalone
        return {
            'base': var_name, 'subgroup': '', 'item': '', 'suffix': '',
            'full_base': var_name, 'pattern_type': 'standalone', 'original': var_name
        }
    
    def determine_grouping_base(self, structure: Dict) -> str:
        """Determina a base de agrupamento baseada na estrutura."""
        base = structure['base']
        subgroup = structure['subgroup']  
        suffix = structure['suffix']
        original = structure['original']
        
        # CORREÇÃO: Variáveis "other" devem ficar STANDALONE 
        # para serem processadas separadamente como keywords
        if 'other' in suffix.lower() or 'other' in original.lower():
            return f"standalone_{original}"  # Força a ficar standalone
        
        # Para padrões hierárquicos, incluir subgroup na base
        if structure['pattern_type'] in ['hierarchical', 'hierarchical_with_suffix']:
            return f"{base}_{subgroup}" if subgroup else base
        
        # Para padrões planos, verificar se suffix é separador
        elif structure['pattern_type'] in ['flat_with_suffix', 'flat']:
            if suffix and self.is_separating_suffix(suffix):
                return f"{base}_{suffix}"
            else:
                return base
        else:
            return base
    
    def is_separating_suffix(self, suffix: str) -> bool:
        """Determina se um sufixo separa em grupo diferente."""
        suffix_lower = suffix.lower()
        
        # Sufixos que MANTÊM no mesmo grupo (relacionados)
        related_suffixes = []
        for category, suffixes in self.config['related_suffixes'].items():
            related_suffixes.extend(suffixes)
        
        for related in related_suffixes:
            if related in suffix_lower:
                return False  # Não separa, mantém no mesmo grupo
        
        return True  # Se não é relacionado, assume que separa
    
    def group_variables(self, selected_vars: List[str], df) -> Tuple[Dict[str, List[str]], List[str], Dict]:
        """Agrupa variáveis usando análise estrutural universal."""
        print("\n🔍 === AGRUPAMENTO ESTRUTURAL UNIVERSAL ===")
        
        groups = defaultdict(list)
        standalone = []
        structures = {}
        
        for var in selected_vars:
            if var not in df.columns:
                print(f"⚠️ {var} não encontrada")
                continue
            
            structure = self.parse_variable_structure(var)
            structures[var] = structure
            
            full_base = structure['full_base']
            
            if structure['pattern_type'] == 'standalone':
                standalone.append(var)
                print(f"   {var} → standalone ({structure['pattern_type']})")
            else:
                groups[full_base].append(var)
                print(f"   {var} → {full_base} ({structure['pattern_type']})")
        
        # Filtrar grupos pequenos
        valid_groups = {}
        for base, vars_list in groups.items():
            if len(vars_list) >= self.config['min_group_size']:
                valid_groups[base] = sorted(vars_list)
            else:
                standalone.extend(vars_list)
                print(f"   {base} → muito pequeno, movendo para standalone")
        
        print(f"\n📊 Grupos válidos: {len(valid_groups)}")
        for base, vars_list in valid_groups.items():
            print(f"   {base}: {vars_list}")
        print(f"📋 Standalone: {len(standalone)}")
        
        return valid_groups, standalone, structures
    
    def analyze_group_semantics(self, vars_list: List[str], structures: Dict, meta, df) -> Dict:
        """Analisa semântica do grupo para classificação MR.

        IMPORTANTE (v7.1):
        - Evita falso positivo de MR em baterias de escala (ex.: IM03A_6, IM03A_7),
          onde a maioria dos respondentes responde quase todos os itens.
        """
        print(f"\n🔍 Análise semântica: {vars_list}")

        # Extrair variable labels
        variable_labels = self.extract_variable_labels(vars_list, meta)

        # Calcular similaridade semântica
        similarity = self.calculate_semantic_similarity(variable_labels)

        # Verificar consistência de value labels
        value_consistency = self.check_value_labels_consistency(vars_list, meta)

        # Detectar padrão de MR checkbox (0/1)
        checkbox_mr = self.detect_checkbox_mr(vars_list, meta)

        # Detectar bateria binária "polar" (ex.: 1="Acessou" / 2="Não acessou", 1="Sim"/2="Não")
        # IMPORTANTE:
        # - Apesar de serem binárias e frequentemente sequenciais no Name, essas baterias NÃO são MR (múltipla resposta)
        #   no sentido de "marque tudo que se aplica"; são itens independentes.
        # - Para MR checkbox legítimo, o SPSS costuma usar 0/1 com labels "Not Selected/Selected" (ou variações).
        polar_binary = self.detect_polar_binary_battery(vars_list, meta)

        # Analisar padrão estrutural
        pattern_analysis = self.analyze_structural_pattern(vars_list, structures)

        # Perfil do grupo (densidade / cardinalidade) para diferenciar MR vs bateria
        profile = self.compute_group_profile(vars_list, meta, df)

        print(f"   📋 Similaridade de labels: {similarity:.2f}")
        print(f"   📋 Value labels consistentes: {value_consistency}")
        print(f"   📋 Padrão checkbox (0/1): {checkbox_mr}")
        print(f"   📋 Bateria binária polar (Sim/Não, Acessou/Não acessou): {polar_binary}")
        print(f"   📋 Padrão estrutural: {pattern_analysis}")
        if profile:
            print(f"   📋 Densidade média preenchida: {profile.get('avg_fill_ratio', 0):.2f}")
            print(f"   📋 Mediana de categorias (dados): {profile.get('median_unique', 0)}")
            print(f"   📋 Mediana de categorias (labels): {profile.get('median_vlabel_count', 0)}")

        # Classificação semântica (com guarda para bateria)
        mr_type = self.classify_semantically(
            similarity,
            value_consistency,
            pattern_analysis,
            checkbox_mr=checkbox_mr,
            polar_binary=polar_binary,
            avg_fill_ratio=profile.get('avg_fill_ratio'),
            median_unique=profile.get('median_unique'),
            median_vlabel_count=profile.get('median_vlabel_count')
        )

        return {
            'mr_type': mr_type,
            'similarity': similarity,
            'value_consistency': value_consistency,
            'pattern_analysis': pattern_analysis,
            'variable_labels': variable_labels,
            'profile': profile,
            'polar_binary': polar_binary
        }

    def detect_polar_binary_battery(self, vars_list: List[str], meta) -> bool:
        """Detecta bateria binária do tipo Sim/Não, Acessou/Não acessou, etc.

        Heurística:
        - A maioria das variáveis do grupo tem exatamente 2 categorias em value labels;
        - Os rótulos são do tipo polar (sim/não, yes/no, acessou/não acessou, concorda/não concorda),
          e NÃO têm semântica explícita de checkbox (selected/not selected).

        Por padrão, tratamos esses grupos como variáveis independentes (NÃO MR),
        para evitar colapso de baterias como SV12_* e SV13_*.
        """
        if not hasattr(meta, 'variable_value_labels'):
            return False
        vvl = meta.variable_value_labels or {}
        ok = 0
        total = 0

        def _norm(s: str) -> str:
            return re.sub(r"\s+", " ", str(s).strip().lower())

        # Tokens típicos de polaridade (PT/EN) — intencionalmente conservador
        yes_tokens = ["sim", "yes", "acessou", "usa", "utiliza", "concordo", "concorda", "satisfeito", "muito satisfeito"]
        no_tokens = ["não", "nao", "no", "não acessou", "nao acessou", "não usa", "nao usa", "discordo", "discorda", "insatisfeito", "muito insatisfeito"]
        # Tokens de checkbox (para excluir):
        checkbox_tokens = ["selected", "not selected", "selecion", "não selecion", "nao selecion"]

        for var in vars_list:
            m = vvl.get(var, {}) or {}
            if not isinstance(m, dict) or not m:
                continue

            # Considerar apenas casos claramente binários
            if len(m) != 2:
                continue

            total += 1
            labels = [_norm(lab) for lab in m.values()]

            # Excluir semântica de checkbox
            if any(tok in lab for lab in labels for tok in checkbox_tokens):
                continue

            # Verificar se temos um par polar (um lado "sim" e outro "não")
            has_yes = any(any(tok in lab for tok in yes_tokens) for lab in labels)
            has_no = any(any(tok in lab for tok in no_tokens) for lab in labels)
            if has_yes and has_no:
                ok += 1

        if total == 0:
            return False
        return (ok / total) >= 0.70

    def compute_group_profile(self, vars_list: List[str], meta, df) -> Dict:
        """Calcula sinais objetivos para distinguir MR de bateria de escala.

        Heurística:
        - Baterias (rating grids) tendem a ter alta densidade de preenchimento (quase todos respondem quase todos os itens)
          e cardinalidade > 2 por item (ex.: 1-4).
        - MR por slots tende a ter muita ausência (missing) e poucos preenchimentos.
        """
        if df is None or not vars_list:
            return {}

        cols = [c for c in vars_list if c in df.columns]
        if not cols:
            return {}

        sub = df[cols]

        # Densidade de preenchimento por respondente
        try:
            nonmiss = sub.notna().sum(axis=1)
            avg_fill_ratio = float(nonmiss.mean() / max(len(cols), 1))
        except Exception:
            avg_fill_ratio = None

        # Cardinalidade por item (dados)
        unique_counts = []
        for c in cols:
            try:
                unique_counts.append(int(sub[c].dropna().nunique()))
            except Exception:
                pass
        unique_counts.sort()
        median_unique = unique_counts[len(unique_counts)//2] if unique_counts else None

        # Cardinalidade por item (value labels)
        median_vlabel_count = None
        try:
            vvl = getattr(meta, 'variable_value_labels', None)
            if isinstance(vvl, dict):
                vlabel_counts = []
                for c in cols:
                    m = vvl.get(c, {}) or {}
                    if isinstance(m, dict) and m:
                        vlabel_counts.append(len(m))
                vlabel_counts.sort()
                median_vlabel_count = vlabel_counts[len(vlabel_counts)//2] if vlabel_counts else None
        except Exception:
            pass

        return {
            'avg_fill_ratio': avg_fill_ratio,
            'median_unique': median_unique,
            'median_vlabel_count': median_vlabel_count
        }

    def extract_variable_labels(self, vars_list: List[str], meta) -> List[str]:
        """Extrai variable labels de forma robusta."""
        labels = []
        
        for var in vars_list:
            label = var  # fallback
            
            if hasattr(meta, 'column_names_to_labels') and var in meta.column_names_to_labels:
                label = meta.column_names_to_labels[var]
            elif hasattr(meta, 'variable_labels') and var in meta.variable_labels:
                label = meta.variable_labels[var]
            
            labels.append(str(label).strip() if label else var)
        
        return labels
    
    def calculate_semantic_similarity(self, labels: List[str]) -> float:
        """Calcula similaridade semântica usando algoritmo de string similarity."""
        if len(labels) <= 1:
            return 1.0
        
        # Calcular similaridade par a par
        similarities = []
        for i in range(len(labels)):
            for j in range(i + 1, len(labels)):
                sim = SequenceMatcher(None, labels[i], labels[j]).ratio()
                similarities.append(sim)
        
        return sum(similarities) / len(similarities) if similarities else 0.0
    
    def check_value_labels_consistency(self, vars_list: List[str], meta) -> bool:
        """Verifica consistência de value labels."""
        if not hasattr(meta, 'variable_value_labels'):
            return False
        
        value_labels_map = meta.variable_value_labels
        reference = None
        
        for var in vars_list:
            var_labels = value_labels_map.get(var, {})
            if not var_labels:
                continue
                
            if reference is None:
                reference = var_labels
            elif var_labels != reference:
                return False
        
        return reference is not None

    def detect_checkbox_mr(self, vars_list: List[str], meta) -> bool:
        """Detecta MR do tipo checkbox (0/1, Not Selected/Selected).

        Objetivo: capturar blocos como IM4_1_1...IM4_1_12 e IM4_2_...,
        onde o label de cada item começa com o texto do item (ex.: [Plano de Saúde])
        e a similaridade de string pode cair, mas estruturalmente é MR.
        """
        if not hasattr(meta, 'variable_value_labels'):
            return False

        vvl = meta.variable_value_labels or {}
        ok = 0
        total = 0

        for var in vars_list:
            m = vvl.get(var, {}) or {}
            if not isinstance(m, dict) or not m:
                continue
            total += 1

            # Normalizar chaves e labels
            keys = set()
            labels = []
            for k, lab in m.items():
                try:
                    # pyreadstat pode devolver int/float/str; normalizar para int quando possível
                    if isinstance(k, (int, float)) and float(k).is_integer():
                        keys.add(int(k))
                    else:
                        keys.add(k)
                except Exception:
                    keys.add(k)
                labels.append(str(lab).strip().lower())

            # Critérios: binário por keys OU por labels típicos
            is_binary_keys = keys.issubset({0, 1}) and (0 in keys and 1 in keys)

            # Heurística por labels (quando chaves não são numéricas)
            not_sel = any((('not selected' in l) or ('nao selecion' in l) or ('não selecion' in l)) for l in labels)
            sel = any(((('selected' in l) and ('not selected' not in l)) or
                       (('selecion' in l) and ('nao' not in l) and ('não' not in l))) for l in labels)

            if is_binary_keys or (not_sel and sel):
                ok += 1

        # Exigir evidência suficiente (ex.: a maioria das variáveis tem mapeamento binário)
        if total == 0:
            return False
        return (ok / total) >= 0.70
    
    def analyze_structural_pattern(self, vars_list: List[str], structures: Dict) -> str:
        """Analisa padrão estrutural do grupo."""
        pattern_types = [structures[var]['pattern_type'] for var in vars_list if var in structures]
        
        if not pattern_types:
            return 'unknown'
        
        # Padrão mais comum no grupo
        from collections import Counter
        most_common = Counter(pattern_types).most_common(1)[0][0]
        return most_common
    
    def classify_semantically(self, similarity: float, value_consistency: bool, pattern: str,
                           checkbox_mr: bool = False,
                           polar_binary: bool = False,
                           avg_fill_ratio: float | None = None,
                           median_unique: int | None = None,
                           median_vlabel_count: int | None = None) -> str:
        """Classificação semântica universal (v7.1).

        Ajuste principal:
        - Antes a regra era quase exclusivamente "similaridade alta => MR".
        - Isso gera falso positivo em baterias de escala onde o texto da pergunta se repete (labels muito parecidos).

        Agora aplicamos uma guarda para classificar como "independent_scales" quando:
        - a maioria dos respondentes preenche a maioria dos itens do grupo (alta densidade), e
        - cada item tem 3+ categorias (dados e/ou value labels).
        """
        threshold = self.config['label_similarity_threshold']

        # Regra especial: MR checkbox (0/1) deve ser agrupado mesmo com baixa similaridade textual
        if checkbox_mr:
            print(f"   📌 Padrão checkbox (0/1) detectado → MR")
            return 'traditional_mr'

        # Regra especial: bateria binária polar (Sim/Não, Acessou/Não acessou, etc.) NÃO é MR.
        # Mesmo com similaridade alta do enunciado, são itens independentes e devem permanecer separados.
        if polar_binary:
            print(f"   📌 Bateria binária polar detectada → Variáveis independentes (não agrupar como MR)")
            return 'independent_scales'


        # Regra forte (v7.2): se cada item tem 3+ categorias (pelos value labels e/ou pelos dados),
        # não é MR do tipo múltipla resposta; é bateria/grade (itens de escala/avaliação).
        # IMPORTANTE: isso vale mesmo quando a bateria é condicional (baixa densidade), para evitar
        # falso positivo em blocos como SV5_1..SV5_9, SV7_1..SV7_5, etc.
        try:
            multi_cat = ((median_unique is not None and median_unique >= 3) or
                         (median_vlabel_count is not None and median_vlabel_count >= 3))
            if multi_cat:
                print(f"   📌 Itens com 3+ categorias (dados/labels) → bateria/escala (não agrupar como MR)")
                return 'independent_scales'
        except Exception:
            pass


        # Guarda de bateria/escala (mantida): alta densidade + 3+ categorias por item => NÃO é MR
        try:
            dense = (avg_fill_ratio is not None and avg_fill_ratio >= 0.60)
            multi_cat = ((median_unique is not None and median_unique >= 3) or
                         (median_vlabel_count is not None and median_vlabel_count >= 3))
            if dense and multi_cat:
                print(f"   📌 Guarda bateria: densidade {avg_fill_ratio:.2f} e categorias >=3 → Variáveis independentes")
                return 'independent_scales'
        except Exception:
            pass

        # Regra 1: Alta similaridade = MR tradicional
        if similarity >= threshold:
            print(f"   📌 Alta similaridade ({similarity:.2f}) → MR tradicional")
            return 'traditional_mr'

        # Regra 2: Baixa similaridade = independentes
        print(f"   📌 Baixa similaridade ({similarity:.2f}) → Variáveis independentes")
        print(f"   📋 Cada variável = uma tabela individual (não agrupar)")
        return 'independent_scales'

    def detect_mr_groups(self, selected_vars: List[str], meta, df) -> Tuple[Dict, List[str]]:
        """FUNÇÃO PRINCIPAL: Detecção MR universal e estrutural."""
        print("\n🎯 === DETECÇÃO MR UNIVERSAL (ESTRUTURAL) ===")
        
        # 1. Agrupamento estrutural
        base_groups, standalone, structures = self.group_variables(selected_vars, df)
        
        # 2. Análise semântica de cada grupo
        mr_groups = {}
        final_standalone = list(standalone)
        
        for base_key, vars_list in base_groups.items():
            analysis = self.analyze_group_semantics(vars_list, structures, meta, df)
            
            # 3. Decisão de agrupamento - CORRIGIDA
            if analysis['mr_type'] == 'traditional_mr':
                
                # Gerar título inteligente
                title = self.generate_intelligent_title(vars_list, analysis)
                
                # Criar grupo MR apenas para traditional_mr
                group_name = f"mr_{base_key.lower().replace('_', '')}"
                mr_groups[group_name] = {
                    "title": title,
                    "members": vars_list,
                    "mr_subtype": analysis['mr_type'],
                    "base": base_key,
                    "analysis": analysis
                }
                
                print(f"   ✅ Grupo MR criado: {group_name} ({analysis['mr_type']})")
                
            else:
                # TUDO MAIS: Manter como variáveis independentes
                # Inclui 'battery_grid', 'independent_scales', etc.
                final_standalone.extend(vars_list)
                print(f"   📋 Mantido como independentes: {base_key} ({analysis['mr_type']})")
                print(f"   📋 Variáveis: {vars_list}")
                print(f"   📋 Cada variável = uma tabela individual")
        
        print(f"\n📊 RESULTADO UNIVERSAL:")
        print(f"   Grupos MR: {len(mr_groups)}")
        print(f"   Variáveis independentes: {len(final_standalone)}")
        
        return mr_groups, final_standalone

    def _clean_mr_label_for_title(self, label: str) -> str:
        """Normaliza labels de itens para extrair o enunciado do bloco.

        Objetivo: evitar que o título do bloco MR herde o "nome do item" (ex.: texto entre colchetes)
        e fique algo como "[Plano de Saúde] Em que ...".

        Regra: remove APENAS prefixos de item no início do label (entre colchetes) e trims.
        """
        if not label:
            return ""
        s = str(label).strip()
        # Remove prefixo do tipo "[ITEM] ..." no início
        s = re.sub(r'^\s*\[[^\]]+\]\s*', '', s).strip()
        # Remove prefixos comuns de numeração (ex.: "1) ", "1. ", "- ")
        s = re.sub(r'^\s*(?:\d+\)|\d+\.|\-|•)\s*', '', s).strip()
        return s

    def _common_prefix_by_words(self, strings):
        """Retorna um prefixo comum por palavras (mais robusto que commonprefix por caracteres)."""
        toks = [re.split(r'\s+', s.strip()) for s in strings if s and s.strip()]
        if not toks:
            return ""
        min_len = min(len(t) for t in toks)
        out = []
        for i in range(min_len):
            w = toks[0][i]
            if all(t[i] == w for t in toks[1:]):
                out.append(w)
            else:
                break
        return " ".join(out).strip()

    def generate_intelligent_title(self, vars_list: List[str], analysis: Dict) -> str:
        """Gera título do bloco (MR ou grupo) evitando "colagens" de itens.

        Estratégia:
        - Para MR: limpar prefixos de item (ex.: colchetes) e extrair enunciado comum.
        - Para demais: manter comportamento legado.
        """
        labels = analysis.get('variable_labels', []) or []
        mr_type = analysis.get('mr_type') or analysis.get('mr_subtype')

        if not labels:
            base = vars_list[0].split('_')[0] if vars_list else 'Grupo'
            return f"Grupo {base}"

        # Para MR (traditional_mr / binary / checkbox), remover prefixos de item e tentar achar enunciado comum
        if mr_type in ('traditional_mr', 'binary', 'checkbox'):
            cleaned = [self._clean_mr_label_for_title(l) for l in labels]
            cleaned = [c for c in cleaned if c]
            if not cleaned:
                return str(labels[0]).strip()

            # Primeiro: prefixo comum por palavras (quando o enunciado é idêntico)
            prefix = self._common_prefix_by_words(cleaned)
            if len(prefix) >= 20:  # suficiente para ser um título inteligível
                return prefix

            # Segundo: heurística anterior (usa padrões no primeiro label já limpo)
            return self.extract_common_part(cleaned)

        # Fallback legado
        return str(labels[0]).strip()

    
    def extract_common_part(self, labels: List[str], suffix: str = "") -> str:
        """Extrai parte comum de uma lista de labels."""
        if len(labels) <= 1:
            return labels[0] + suffix if labels else "Grupo" + suffix
        
        first = labels[0]
        
        # Buscar padrões comuns
        patterns = [
            r'^(.+?)\s*[-–—]\s*.+',     # "Pergunta - item"
            r'^(.+?)\s*:\s*.+',         # "Pergunta: item"
            r'^(.+?)\s*\(.+\)',         # "Pergunta (item)"
            r'^(.{20,}?)\s+\w+\s*.+',   # Primeiras 20+ chars + uma palavra
        ]
        
        for pattern in patterns:
            match = re.match(pattern, first)
            if match:
                common = match.group(1).strip()
                if len(common) >= 10:
                    return common + suffix
        
        # Fallback: primeiras palavras  
        words = first.split()
        if len(words) >= 2:
            return " ".join(words[:2]) + "..." + suffix
        
        return first + suffix

# ========== NOVA DETECÇÃO DE GRUPOS MR (CORRIGIDA) ==========

def detect_mr_groups_improved(selected_vars: List[str], meta, df) -> Tuple[Dict[str, Dict], List[str]]:
    """
    Detecta grupos MR usando sinais puramente estruturais — sem depender de nomes de variáveis:
    1. Agrupa vars com value labels idênticos
    2. Sub-agrupa por label text idêntico (removendo conteúdo entre colchetes)
    3. Exige sufixo numérico em todos os membros
    4. Classifica: stacked (fill decrescente) ou binary (fills similares)
    Mantém pré-detecção _On para compatibilidade com bases antigas.
    """
    import re as _re
    from collections import defaultdict as _dd

    labels_map = dict(zip(meta.column_names, meta.column_labels))
    valabs     = get_value_labels_map(meta)

    def _strip_bracket(s):
        return _re.sub(r'^\s*\[.*?\]\s*', '', s or '').strip()

    def _num_suffix(v):
        m = _re.search(r'(\d+)$', v)
        return int(m.group(1)) if m else None

    pre_detected: Dict[str, Dict] = {}
    pre_detected_members: set = set()

    # ── PASSO 1: Pré-detecção _O\d+ (slots explícitos, ex: Q_25_O1) ──
    slot_pattern = _re.compile(r'^(.+)_O(\d+)$', _re.IGNORECASE)
    slot_buckets: Dict[str, List[str]] = {}
    for v in selected_vars:
        m = slot_pattern.match(v)
        if m:
            slot_buckets.setdefault(m.group(1), []).append(v)

    for prefix, members in slot_buckets.items():
        if len(members) < 2:
            continue
        members_sorted = sorted(members, key=lambda v: int(slot_pattern.match(v).group(2)))
        raw_title = get_var_label(meta, members_sorted[0])
        title = _re.sub(r'\s*\[.*?\]\s*$', '', raw_title).strip()
        mr_subtype = detect_mr_type_improved(members_sorted, meta, df) or 'categorical'
        pre_detected[prefix] = {
            "title": title, "members": members_sorted,
            "mr_subtype": mr_subtype, "other_var": None, "base": prefix,
        }
        pre_detected_members.update(members_sorted)
        print(f"   🔗 MR _On: {prefix} → {members_sorted} ({mr_subtype})")

    # ── PASSO 2: Detecção estrutural pura (value labels + label text + sufixo numérico) ──
    remaining = [v for v in selected_vars if v not in pre_detected_members]

    # Detecta escala avaliativa ordinal — exclui baterias do MR
    _SCALE_WORDS = {
        'satisfeito','insatisfeito','satisfacao','satisfação',
        'possui','concordo','discordo','bom','ótimo','otimo',
        'ruim','péssimo','pessimo','nunca','raramente','sempre',
        'frequente','muito','totalmente','parcialmente','adequado',
        'aprovado','reprovado','sim','não','nao','acessou',
    }
    def _is_ordinal_scale(vl_dict):
        if not vl_dict:
            return False
        ns_vals = {99,999,9000,9999,98,8888,88,9998,9997}
        main = {k: v for k, v in vl_dict.items()
                if k not in ns_vals and not any(
                    w in str(v).lower()
                    for w in ('não sabe','nao sabe','não se aplica','nao se aplica','não citar','nao citar')
                )}
        n = len(main)
        if n < 2 or n > 8:
            return False
        text = ' '.join(str(v).lower() for v in main.values())
        return any(w in text for w in _SCALE_WORDS)

    # Agrupar por value labels idênticos
    vl_groups: Dict[tuple, List[str]] = _dd(list)
    for v in remaining:
        vl = valabs.get(v, {})
        if not vl:
            continue
        # Excluir grupos de escala avaliativa (são baterias, não MR)
        if _is_ordinal_scale(vl):
            continue
        key = tuple(sorted(
            (str(round(float(k), 4) if isinstance(k, float) else k), str(lbl))
            for k, lbl in vl.items()
        ))
        vl_groups[key].append(v)

    for vl_key, cols in vl_groups.items():
        if len(cols) < 2:
            continue
        # Sub-agrupar por label text idêntico (sem bracket)
        lbl_groups: Dict[str, List[str]] = _dd(list)
        for v in cols:
            lbl = _strip_bracket(labels_map.get(v, '')).strip()
            lbl_groups[lbl].append(v)

        for lbl_text, group in lbl_groups.items():
            if len(group) < 2:
                continue
            # Exigir sufixo numérico em todos
            if not all(_num_suffix(v) is not None for v in group):
                continue
            # Filtrar já detectados
            group = [v for v in group if v not in pre_detected_members]
            if len(group) < 2:
                continue

            # Ordenar pelo sufixo numérico
            group_sorted = sorted(group, key=lambda v: _num_suffix(v))

            fills = [df[v].notna().sum() / len(df) for v in group_sorted]
            f0 = fills[0] if fills else 0
            f1 = fills[1] if len(fills) > 1 else 0

            # Stacked: primeiro >> segundo (ex: 100% vs 5%)
            is_stacked = f0 > 0 and f1 < f0 * 0.4
            # Binary: fills homogêneos
            fill_range = max(fills) - min(fills)
            is_binary  = fill_range < 0.25

            if not (is_stacked or is_binary):
                continue

            mr_subtype = detect_mr_type_improved(group_sorted, meta, df) or 'categorical'
            group_name = group_sorted[0]  # usar primeiro membro como chave única

            title = _strip_bracket(lbl_text) or get_var_label(meta, group_sorted[0])
            pre_detected[group_name] = {
                "title": title, "members": group_sorted,
                "mr_subtype": mr_subtype, "other_var": None, "base": group_name,
            }
            pre_detected_members.update(group_sorted)
            kind = 'stacked' if is_stacked else 'binary'
            print(f"   🔗 MR estrutural [{kind}]: {group_sorted[0]}…{group_sorted[-1]} ({mr_subtype})")

    # ── PASSO 3: Detector universal para o restante ──
    remaining_vars = [v for v in selected_vars if v not in pre_detected_members]
    config = {
        'min_group_size': 2,
        'label_similarity_threshold': 0.7,
        'related_suffixes': {
            'medical': ['ass', 'assist', 'med', 'medic', 'saude'],
            'other': ['other', 'outro', 'outra', 'outros', 'outras'],
            'text': ['text', 'txt', 'open', 'aberta'],
            'scale': ['esc', 'scale', 'escala'],
            'geographic': ['norte', 'sul', 'leste', 'oeste', 'north', 'south', 'east', 'west']
        }
    }
    detector = UniversalMRDetector(config)
    mr_groups_universal, standalone = detector.detect_mr_groups(remaining_vars, meta, df)

    # Converter grupos universais para formato legado
    converted_mr_groups: Dict[str, Dict] = {}

    # Primeiro: pré-detectados (estrutural + _O\d+)
    converted_mr_groups.update(pre_detected)

    # Segundo: grupos do detector universal
    for group_name, group_info in mr_groups_universal.items():
        mr_subtype_map = {
            'traditional_mr': 'categorical',
            'independent_scales': 'rating_scale'
        }
        original_subtype = group_info.get('mr_subtype', 'categorical')
        mapped_subtype = mr_subtype_map.get(original_subtype, 'categorical')

        compat_type = detect_mr_type_improved(group_info.get('members', []), meta, df)
        if compat_type == 'binary':
            mapped_subtype = 'binary'

        base = group_info.get('base', '')
        other_var = None
        other_candidates = [f"{base}_other", f"{base}_outro",
                            f"{base.split('_')[0]}_other"]
        for candidate in other_candidates:
            if candidate in df.columns and candidate not in group_info['members']:
                other_var = candidate
                break

        converted_mr_groups[group_name] = {
            "title": group_info['title'],
            "members": group_info['members'],
            "mr_subtype": mapped_subtype,
            "other_var": other_var,
            "base": base,
        }

    # standalone: manter apenas vars não pré-detectadas
    standalone = [v for v in standalone if v not in pre_detected_members]

    print(f"\n📊 CONVERSÃO PARA FORMATO LEGADO:")
    print(f"   Grupos MR convertidos: {len(converted_mr_groups)}")
    print(f"   Variáveis standalone: {len(standalone)}")

    return converted_mr_groups, standalone

def detect_mr_type_improved(group_vars: List[str], meta, df) -> str:
    """
    VERSÃO 7.0 COMPATÍVEL: Mantida para compatibilidade com código legado.
    
    O sistema universal já faz a classificação correta, mas esta função
    é mantida para evitar quebras em outras partes do código.
    """
    
    # Para grupos detectados pelo sistema universal, simplesmente retornar categorical
    # pois a classificação real já foi feita pelo UniversalMRDetector
    
    valabs = get_value_labels_map(meta)
    
    # Detecção básica para compatibilidade
    if group_vars and group_vars[0] in valabs:
        first_var_labels = valabs[group_vars[0]]
        
        # Se tem poucas opções e são binárias, assumir binary
        if len(first_var_labels) <= 2:
            unique_values = set()
            for var in group_vars:
                if var in df.columns:
                    unique_values.update(df[var].dropna().unique())
            
            if unique_values.issubset({0, 1, 0.0, 1.0}):
                return "binary"
    
def get_mr_group_title(base: str, vars_list: List[str], meta) -> str:
    """
    Obtém título do grupo MR, tentando várias estratégias.
    """
    # 1. Tentar usar label da variável base (se existir)
    base_label = get_var_label(meta, base)
    if base_label and len(base_label.strip()) > 3:
        return base_label.strip()
    
    # 2. Tentar usar primeira variável, removendo colchetes
    if vars_list:
        first_label = get_var_label(meta, vars_list[0])
        if first_label:
            # Remove texto entre colchetes no início
            clean_label = re.sub(r'^\s*\[.*?\]\s*', '', first_label).strip()
            # Remove numeração no final (ex: "Pergunta 1", "Question 1")
            clean_label = re.sub(r'\s+\d+\s*$', '', clean_label).strip()
            if clean_label:
                return clean_label
    
    # 3. Fallback
    return f"Grupo {base}"

# ========== NOVA CAMADA DE IDENTIFICAÇÃO DE TIPOS ==========

DATE_PREFIXES = (
    "DATE", "ADATE", "SDATE", "EDATE",
    "JDATE", "DATETIME", "QYR", "WKYR", "MOYR"
)

TIME_PREFIXES = (
    "TIME", "DTIME", "MTIME"
)

def detect_physical_type(meta, df, var_name: str) -> str:
    """
    Detecta o tipo REAL da variável (string, numeric, date),
    usando:
    1) display_format do SPSS
    2) original_variable_types
    3) inspeção do dataframe (conteúdo REAL)
    """
    import re

    # ---------- 1) Formato SPSS ----------
    var_formats = getattr(meta, "variable_display_formats", {}) or {}
    fmt = str(var_formats.get(var_name, "")).upper()

    # STRING por formato Axx
    if fmt.startswith("A"):
        return "string"

    # ---------- 2) Se SPSS diz STRING ----------
    var_types = getattr(meta, "original_variable_types", {}) or {}
    original_type = var_types.get(var_name)
    if original_type and "STRING" in str(original_type).upper():
        return "string"

    # ---------- 3) Inspeção do dataframe ----------
    if var_name in df.columns:
        series = df[var_name]

        # dtype object geralmente indica texto
        if series.dtype == object:
            return "string"

        # Verificar se 80% dos valores NÃO são numéricos → string
        sample = series.dropna().astype(str).head(20)
        nonnum = 0
        for v in sample:
            try:
                float(v)
            except:
                nonnum += 1
        if len(sample) > 0 and nonnum / len(sample) > 0.5:
            return "string"

        # Verificar presença de palavras → string
        for v in sample:
            if any(c.isalpha() for c in v):
                return "string"

    # ---------- 4) Detectar datas ----------
    DATE_PREFIXES = (
        "DATE","ADATE","SDATE","EDATE","JDATE",
        "DATETIME","QYR","WKYR","MOYR"
    )
    if any(fmt.startswith(pfx) for pfx in DATE_PREFIXES):
        return "date"

    # ---------- 5) Caso nada acima → é numérica ----------
    return "numeric"

def detect_measure_type(meta, var_name: str, physical_type: str):
    """
    Retorna nominal / ordinal / scale
    apenas para variáveis numéricas.
    """
    if physical_type != "numeric":
        return None

    measures = getattr(meta, "variable_measure", {}) or {}
    measure = measures.get(var_name)

    if isinstance(measure, str):
        m = measure.lower().strip()
        if m in ("nominal", "ordinal", "scale"):
            return m

    return None

def detect_variables_universal(selected_vars, meta, valabs, df):
    """
    VERSÃO CORRIGIDA que preserva a ordem original do SPSS.
    
    Em vez de processar primeiro todos os grupos MR e depois todas as standalone,
    processa na ordem original do selected_vars, decidindo para cada posição
    se é um grupo MR ou uma variável standalone.
    """
    print(f"\n🔍 === DETECÇÃO DE VARIÁVEIS - ORDEM ORIGINAL PRESERVADA ===")
    print(f"📋 Variáveis selecionadas: {selected_vars[:5]}{'...' if len(selected_vars) > 5 else ''}")
    
    vars_meta = []
    processed_vars = set()  # Rastrear variáveis já processadas
    
    # PASSO 1: Detectar grupos MR (usando apenas variáveis selecionadas para análise)
    mr_groups, standalone_vars = detect_mr_groups_improved(selected_vars, meta, df)
    
    print(f"\n📊 Grupos MR detectados: {list(mr_groups.keys())}")
    print(f"📋 Variáveis standalone: {len(standalone_vars)}")
    
    # PASSO 2: Processar na ORDEM ORIGINAL intercalando MR e standalone
    print(f"\n🔧 Processando na ordem original do SPSS:")
    
    for i, var in enumerate(selected_vars):
        if var in processed_vars:
            continue  # Já foi processada como parte de um grupo MR
        
        # Verificar se esta variável faz parte de um grupo MR
        mr_group_for_this_var = None
        for group_name, group_info in mr_groups.items():
            if var in group_info["members"]:
                mr_group_for_this_var = (group_name, group_info)
                break
        
        if mr_group_for_this_var:
            # Esta variável é a primeira do seu grupo MR - adicionar o grupo aqui
            group_name, group_info = mr_group_for_this_var
            
            print(f"   {i+1:2d}. {group_name} (grupo MR - primeiro membro: {var})")
            
            vars_meta.append({
                "name": group_name,
                "title": group_info["title"],
                "type": "mr",
                "spss_type": "Resposta Múltipla",
                "sheet_code": group_name,
                "var_type": "multiple_response", 
                "measure": None,
                "mr_subtype": group_info["mr_subtype"],
                "stats": None
            })
            
            # Marcar todas as variáveis do grupo como processadas
            for member_var in group_info["members"]:
                if member_var == group_info.get("other_var"):
                    continue
                processed_vars.add(member_var)
            
            print(f"      ✅ Grupo MR adicionado ({group_info['mr_subtype']}) - {len(group_info['members'])} variáveis")
            
        elif var in standalone_vars:
            # Esta é uma variável standalone - processar normalmente
            print(f"   {i+1:2d}. {var} (standalone)")
            
            if var not in df.columns:
                print(f"      ⚠️ Pulando {var} (não existe no dataset)")
                processed_vars.add(var)
                continue
            
            # ── DETECÇÃO DE _S ASSOCIADA ──
            # Verificar se existe uma variável _S correspondente com respostas reais.
            # Se sim, ela será absorvida dentro desta seção (não vira seção própria).
            candidate_s = var + "_S"
            open_text_var = None
            if candidate_s in df.columns and candidate_s in standalone_vars:
                real_vals = [v for v in df[candidate_s].dropna() if str(v).strip()]
                if real_vals:
                    open_text_var = candidate_s
                    processed_vars.add(candidate_s)  # não processar como seção independente
                    print(f"      📎 _S absorvida: {candidate_s} ({len(real_vals)} respostas)")
            
            # Detectar tipo físico
            physical = detect_physical_type(meta, df, var)
            
            if physical == "string":
                vars_meta.append({
                    "name": var,
                    "title": get_var_label(meta, var),
                    "type": "string",
                    "spss_type": "Resposta Aberta",
                    "sheet_code": var,
                    "var_type": "string",
                    "measure": None,
                    "mr_subtype": None,
                    "stats": None
                })
                print(f"      ✅ Adicionado como string")
                
            elif physical == "date":
                vars_meta.append({
                    "name": var,
                    "title": get_var_label(meta, var),
                    "type": "single",
                    "spss_type": "Data",
                    "sheet_code": var,
                    "var_type": "date",
                    "measure": None,
                    "mr_subtype": None,
                    "stats": None
                })
                print(f"      ✅ Adicionado como data")
                
            else:
                # Numérico - detectar medida a partir do SPSS (sem inferência por value labels)
                measure = detect_measure_type(meta, var, physical)

                if measure == "scale":
                    # Numérica contínua (Escala) - stats serão calculadas depois com ponderação
                    vars_meta.append({
                        "name": var,
                        "title": get_var_label(meta, var),
                        "type": "single",
                        "spss_type": "Numérica (Escala)",
                        "sheet_code": var,
                        "var_type": "numeric",
                        "measure": "scale",
                        "mr_subtype": None,
                        "stats": None  # Será calculado depois com ponderação
                    })
                    print(f"      ✅ Adicionado como Numérica (Escala) seguindo SPSS")
                else:
                    # Categórica (Nominal ou Ordinal) seguindo APENAS o Measure do SPSS
                    human = "Categórica (Ordinal)" if measure == "ordinal" else "Categórica (Nominal)"
                    vars_meta.append({
                        "name": var,
                        "title": get_var_label(meta, var),
                        "type": "single",
                        "spss_type": human,
                        "sheet_code": var,
                        "var_type": "categorical",
                        "measure": measure or "nominal",
                        "mr_subtype": None,
                        "stats": None,
                        "open_text_var": open_text_var,
                    })
                    print(f"      ✅ Adicionado como {human} (Measure SPSS)")

                processed_vars.add(var)
        
        else:
            # Variável não foi classificada (não deveria acontecer normalmente)
            print(f"   {i+1:2d}. {var} (⚠️ não classificada - pulando)")
            processed_vars.add(var)
    
    # PASSO 3: Verificar se todas as variáveis foram processadas
    print(f"\n🔍 Verificação final:")
    missing_vars = set(selected_vars) - processed_vars
    if missing_vars:
        print(f"⚠️ Variáveis não processadas: {missing_vars}")
    else:
        print(f"✅ Todas as {len(selected_vars)} variáveis foram processadas")
    
    print(f"\n📈 RESUMO FINAL:")
    print(f"   Total de variáveis no dashboard: {len(vars_meta)}")
    print(f"   Grupos MR detectados: {len(mr_groups)}")
    print(f"   Variáveis standalone: {len(standalone_vars)}")
    
    # Numerar variáveis sequencialmente (P1, P2...)
    for i, vm in enumerate(vars_meta):
        vm["p_num"] = i + 1

    # Debug: mostrar ordem final CORRIGIDA
    print(f"\n✅ ORDEM FINAL PRESERVADA (CORRIGIDA):")
    for i, vm in enumerate(vars_meta):
        print(f"   {i+1:2d}. P{i+1} — {vm['name']} ({vm.get('var_type', vm['type'])})")
    
    return vars_meta, mr_groups


def build_records_and_meta(df, meta, selected_vars: List[str], filter_vars: List[str], 
                          file_source: str, client_name: str, weight_var: str = None):
    """
    Constrói:
      - created_at: timestamp
      - vars_meta: metadados das variáveis (incluindo grupos MR e stats)
      - filters_meta: metadados dos filtros
      - records: lista de dicionários prontos para o dashboard
      
    NOVO: Inclui automaticamente campos de data (submitdate, etc.) para cálculo de período de coleta
    """
    created_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    # === DETECTAR E INCLUIR CAMPOS DE DATA AUTOMATICAMENTE ===
    date_fields = []
    for col in df.columns:
        col_lower = col.lower()
        if (col_lower == 'submitdate' or 
            'submit' in col_lower or 
            ('date' in col_lower and col_lower not in ['updatedate', 'update_date']) or
            ('data' in col_lower and 'update' not in col_lower)):
            
            # Verificar se é realmente uma data
            try:
                sample_values = df[col].dropna().head(10)
                if len(sample_values) > 0:
                    for val in sample_values:
                        test_date = pd.to_datetime(val, errors='coerce')
                        if pd.notna(test_date) and test_date.year > 1900:
                            date_fields.append(col)
                            print(f"📅 Campo de data detectado: {col}")
                            break
            except:
                continue
    
    # Combinar variáveis selecionadas com campos de data (removendo duplicatas)
    all_vars_for_records = list(selected_vars)
    for date_field in date_fields:
        if date_field not in all_vars_for_records:
            all_vars_for_records.append(date_field)
            print(f"✅ Incluído automaticamente para período de coleta: {date_field}")
    
    # Mapa de value labels por variável
    valabs = get_value_labels_map(meta)

    # ----- PROCESSAMENTO DE VARIÁVEL PESO -----
    weight_values = None
    if weight_var and weight_var in df.columns:
        try:
            weight_series = pd.to_numeric(df[weight_var], errors='coerce')
            # Substituir missing/NaN por 1.0 (peso neutro)
            weight_values = weight_series.fillna(1.0)
            # Validar pesos (devem ser positivos)
            weight_values = weight_values.abs()
            weight_values = weight_values.replace(0, 1.0)  # Zero vira 1
            print(f"⚖️ Usando variável peso: {weight_var}")
            print(f"   📊 Estatísticas do peso: Média={weight_values.mean():.3f}, Min={weight_values.min():.3f}, Max={weight_values.max():.3f}")
        except Exception as e:
            print(f"⚠️ Erro ao processar peso {weight_var}: {e}. Prosseguindo sem ponderação.")
            weight_values = None
    elif weight_var:
        print(f"⚠️ Variável peso '{weight_var}' não encontrada. Prosseguindo sem ponderação.")
        
    # Função helper para aplicar pesos
    def apply_weight(base_value, index):
        if weight_values is not None and index < len(weight_values):
            return base_value * weight_values.iloc[index]
        return base_value

    # ----- ORDEM ORIGINAL DAS CATEGORIAS (labels já normalizados) -----
    def _normalize_label_for_js(lbl):
        txt = str(lbl).replace(":", "").strip()
        return _normalize_display_value(txt)

    value_orders = {}
    for var_name, labels_dict in valabs.items():
        if not labels_dict:
            continue
        # A ordem do dict de value_labels do SPSS já vem na ordem correta
        ordered_labels = [_normalize_label_for_js(lbl) for lbl in labels_dict.values()]
        value_orders[var_name] = ordered_labels
    
    # Criar mapeamento código -> label para exibição  
    code_to_label = {}
    for var_name, labels_dict in valabs.items():
        if labels_dict:
            # CORREÇÃO: converter chaves numéricas para strings
            string_mapping = {}
            for k, v in labels_dict.items():
                # Converter 0.0 → "0", 1.0 → "1", etc.
                if isinstance(k, (int, float)) and k == int(k):
                    key_str = str(int(k))
                else:
                    key_str = str(k)
                string_mapping[key_str] = str(_normalize_label_for_js(v))
            code_to_label[var_name] = string_mapping
    
    # Metadados das variáveis e grupos de múltipla resposta (FASE 1)
    vars_meta, mr_groups = detect_variables_universal(selected_vars, meta, valabs, df)

    # Detectar escalas (NPS, SAT, CSAT, CES) via tag no label da variável
    labels_map = meta.column_names_to_labels or {}
    for vm in vars_meta:
        if (vm.get('var_type') or vm.get('type') or '') == 'multiple_response':
            continue
        col_label = str(labels_map.get(vm['name']) or '')
        vmap_raw  = valabs.get(vm['name'], {})
        si = detect_scale_info(col_label, vmap_raw)
        if si:
            vm['scale_info'] = si
            vm['spss_type']  = f"{si['type']} · Escala {si['real_min']}-{si['real_max']}"
            print(f"📊 Escala detectada: {vm['name']} → {si['type']} "
                  f"{si['real_min']}-{si['real_max']}"
                  f"{' (invertida)' if si['inverted'] else ''}")
    
    # ---------- PROCESSAMENTO DE FILTROS ----------
    filters_meta = []
    for fv in filter_vars:
        if fv in df.columns:
            unique_vals = []
            print(f"🔍 DEBUG FILTRO {fv}:")
            
            # Debug: mostrar estrutura do valabs para esta variável
            var_valabs = valabs.get(fv, {})
            if var_valabs:
                print(f"   📋 Valabs keys: {list(var_valabs.keys())} (types: {[type(k).__name__ for k in var_valabs.keys()]})")
                print(f"   📋 Valabs values: {list(var_valabs.values())}")
            else:
                print(f"   ⚠️ Nenhum value_labels encontrado para {fv}")
            
            for val in df[fv].dropna().unique():
                # Usar lookup robusto para pegar o label correto
                label = safe_value_label_lookup(valabs, fv, val)
                processed_val = str(label).replace(":", "").strip()
                processed_val = _normalize_display_value(processed_val)
                unique_vals.append(processed_val)
                print(f"   {val} ({type(val).__name__}) → '{label}' → '{processed_val}'")
            
            if unique_vals:
                filters_meta.append({
                    "name": fv,
                    "title": get_var_label(meta, fv) or fv,
                    "values": safe_sorted_unique(unique_vals)
                })
                print(f"✅ Filtro {fv}: {len(unique_vals)} valores únicos")
                print(f"   Final values: {unique_vals}")
            print()
    
    # ---------- HELPERS ESPECÍFICOS DA FASE 3 ----------
    def format_spss_date(v):
        """Converte data SPSS (número de dias) em 'YYYY-MM-DD'."""
        if pd.isna(v):
            return None
        try:
            return pd.to_datetime(v, unit='d', origin='1582-10-14').strftime('%Y-%m-%d')
        except Exception:
            return None

    def add_scale_value(scale_store, var_name, value, weight=1.0):
        """Acumula valores de variáveis scale para cálculo posterior de stats."""
        if value is None:
            return
        try:
            f = float(value)
        except Exception:
            return
        if var_name not in scale_store:
            scale_store[var_name] = []
        # Armazenar como tupla (valor, peso) para estatísticas ponderadas
        scale_store[var_name].append((f, weight))

    def compute_stats(values):
        """Calcula média, mediana, desvio padrão, min, max, n com suporte a ponderação."""
        import math
        if not values:
            return None
        
        # DEBUG: Verificar tipo dos dados recebidos
        try:
            # Verificar se são tuplas (valor, peso) ou valores simples
            if values and isinstance(values[0], tuple):
                # Valores ponderados
                weighted_values = values
                total_weight = sum(weight for _, weight in weighted_values)
                
                if total_weight == 0:
                    return None
                    
                # Média ponderada
                weighted_sum = sum(value * weight for value, weight in weighted_values)
                mean = weighted_sum / total_weight
                
                # Para mediana, criar lista expandida pelos pesos (aproximação)
                expanded_values = []
                for value, weight in weighted_values:
                    # Adicionar valor repetido proporcionalmente ao peso
                    count = max(1, int(round(weight)))
                    expanded_values.extend([value] * count)
                
                expanded_values.sort()
                n_expanded = len(expanded_values)
                if n_expanded % 2 == 1:
                    median = expanded_values[n_expanded // 2]
                else:
                    median = (expanded_values[n_expanded // 2 - 1] + expanded_values[n_expanded // 2]) / 2
                
                # Desvio padrão ponderado
                var = sum(weight * (value - mean) ** 2 for value, weight in weighted_values) / total_weight
                stddev = math.sqrt(var)
                
                # Min/Max dos valores originais
                raw_values = [value for value, _ in weighted_values]
                
                return {
                    "n": int(round(total_weight)),  # Total ponderado
                    "mean": mean,
                    "median": median,
                    "stddev": stddev,
                    "min": min(raw_values),
                    "max": max(raw_values)
                }
            else:
                # Valores simples (comportamento original)
                vals = list(values)
                n = len(vals)
                vals_sorted = sorted(vals)
                mean = sum(vals) / n
                if n % 2 == 1:
                    median = vals_sorted[n // 2]
                else:
                    median = (vals_sorted[n // 2 - 1] + vals_sorted[n // 2]) / 2
                var = sum((x - mean) ** 2 for x in vals) / n
                stddev = math.sqrt(var)
                
                return {
                    "n": n,
                    "mean": mean,
                    "median": median,
                    "stddev": stddev,
                    "min": min(vals),
                    "max": max(vals)
                }
        except Exception as e:
            print(f"⚠️ Erro em compute_stats: {e}")
            print(f"   Tipo de values: {type(values)}")
            if values:
                print(f"   Primeiro elemento: {type(values[0])} = {values[0]}")
            return None
    
    # Mapeia quais variáveis são scale numéricas
    scale_vars = {
        vm["name"]: vm
        for vm in vars_meta
        if vm.get("var_type") == "numeric" and vm.get("measure") == "scale"
    }
    scale_values_store: Dict[str, List[Tuple[float, float]]] = {name: [] for name in scale_vars.keys()}
    
    # ---------- PROCESSAMENTO DE REGISTROS ----------
    records = []
    for index, row in df.iterrows():
        rec: Dict[str, Any] = {}
        
        # Adicionar peso do registro (1.0 se não há ponderação)
        current_weight = apply_weight(1.0, index)
        rec["__weight__"] = current_weight  # Campo especial para ponderação
        
        # ----- Filtros -----
        filter_debug = {}
        for fv in filter_vars:
            if fv in df.columns:
                val = row.get(fv)
                if pd.isna(val):
                    rec[fv] = None
                    filter_debug[fv] = f"NULL"
                else:
                    # Usar lookup robusto para pegar o label correto
                    label = safe_value_label_lookup(valabs, fv, val)
                    processed_val = _normalize_display_value(
                        str(label).replace(":", "").strip()
                    )
                    rec[fv] = processed_val
                    filter_debug[fv] = f"{val}→{label}→{processed_val}"
        
        # Debug para primeiros registros
        if index < 3:
            print(f"📋 Record {index}: {filter_debug}")
        
        # ----- Variáveis -----
        for vm in vars_meta:
            vname = vm["name"]
            vtype = vm.get("var_type")      # string / numeric / date / multiple_response
            measure = vm.get("measure")     # nominal/ordinal/scale/None
            base_col = vm["sheet_code"]     # nome original da coluna ou base MR
            
            # ========= STRING =========
            if vtype == "string":
                val = row.get(base_col)
                if pd.isna(val) or not str(val).strip():
                    rec[vname] = None
                else:
                    rec[vname] = format_text_response(str(val))
                continue
            
            # ========= DATE =========
            if vtype == "date":
                val = row.get(base_col)
                rec[vname] = format_spss_date(val)
                continue
            
            # ========= _S absorvida: guardar texto no registro pai =========
            otv = vm.get("open_text_var")
            if otv and otv in df.columns:
                txt = row.get(otv)
                key = vname + "__open_text__"
                if pd.notna(txt) and str(txt).strip():
                    rec[key] = format_text_response(str(txt))
                else:
                    rec[key] = None
            
            # ========= MULTIPLE RESPONSE =========
            if vtype == "multiple_response":
                group = mr_groups.get(vname, {})
                members = group.get("members", [])
                subtype = group.get("mr_subtype")

                chosen_options: List[str] = []
                for col in members:
                    val = row.get(col)
                    if pd.isna(val):
                        continue

                    vmap = valabs.get(col, {})

                    if subtype == "binary":
                        # MR binária: precisa estar marcada
                        if not mr_is_selected(val, vmap):
                            continue
                        option_text = get_mr1_label(meta, col)
                        # Binary: fallback ao nome da variável
                        if not option_text:
                            option_text = get_var_label(meta, col)
                        if not option_text:
                            option_text = col
                    else:
                        # MR por slots/categórica: basta estar preenchido
                        if not mr_is_filled(val, vmap):
                            continue
                        option_text = get_mr2_label(valabs, col, val)
                        # Slot: se não há label para este valor, ignorar —
                        # NUNCA usar o título da variável como categoria
                        if not option_text:
                            continue

                    option_text = str(option_text).strip()
                    if option_text and option_text not in chosen_options:
                        chosen_options.append(option_text)

                # Se existir variável de "outros" associada a este grupo,
                # ela entra como categoria "Outros" na MR principal.
                other_var = group.get("other_var")
                if other_var and other_var in df.columns:
                    other_val = row.get(other_var)

                    # Se é um texto preenchido válido → ativa "Outros"
                    if isinstance(other_val, str):
                        other_text = other_val.strip()

                        if other_text and other_text not in ("99", ".", "NA", "na", "N/A", "n/a", "-"):
                            if "Outros" not in chosen_options:
                                chosen_options.append("Outros")

                rec[vname] = safe_sorted_unique(chosen_options)
                continue
            
            # ========= NUMERIC (nominal / ordinal / scale) =========
            val = row.get(base_col)
            if pd.isna(val):
                rec[vname] = None
                continue
            
            # Tratar códigos de missing do usuário como ausente
            # -1 é convenção comum em SPSS para "não se aplica" / missing declarado
            try:
                if float(val) in USER_MISSING_CODES:
                    rec[vname] = None
                    continue
            except (ValueError, TypeError):
                pass
            
            # Categórico (nominal / ordinal): sempre usar o LABEL
            # VARS_VALUE_ORDER no JS também contém labels (não códigos),
            # então armazenar o label garante que filtros E ordenação ordinal funcionem.
            if measure in ("nominal", "ordinal"):
                label = safe_value_label_lookup(valabs, base_col, val)
                processed_val = str(label).replace(":", "").strip()
                processed_val = _normalize_display_value(processed_val)
                rec[vname] = processed_val
                continue
            
            # Escalar (contínuo)
            if measure == "scale":
                try:
                    num_val = float(val)
                    rec[vname] = num_val
                    add_scale_value(scale_values_store, vname, num_val, current_weight)
                except Exception:
                    rec[vname] = None
                continue
            
            # Fallback genérico
            rec[vname] = _normalize_display_value(str(val))
        
        # === PROCESSAR CAMPOS DE DATA ADICIONAIS ===
        for date_field in date_fields:
            if date_field not in rec:  # Só adicionar se não foi processado ainda
                val = row.get(date_field)
                if pd.isna(val):
                    rec[date_field] = None
                else:
                    # Tentar converter para timestamp JavaScript (formato ISO)
                    try:
                        date_obj = pd.to_datetime(val)
                        if pd.notna(date_obj):
                            # Converter para string ISO para JavaScript
                            rec[date_field] = date_obj.isoformat()
                        else:
                            rec[date_field] = None
                    except:
                        # Se não conseguir converter, manter string original
                        rec[date_field] = str(val) if val is not None else None
        
        records.append(rec)
    
    # ---------- CÁLCULO FINAL DE STATS PARA VARIÁVEIS SCALE ----------
    for vm in vars_meta:
        if vm.get("var_type") == "numeric" and vm.get("measure") == "scale":
            name = vm["name"]
            values = scale_values_store.get(name, [])
            vm["stats"] = compute_stats(values) if values else None

    # ---------- EXTRAÇÃO DE PALAVRAS‑CHAVE PARA VARIÁVEIS STRING ----------
    # Para cada variável de texto, coletar todas as respostas válidas e gerar palavras‑chave frequentes.
    try:
        for vm in vars_meta:
            if vm.get("var_type") == "string":
                vname = vm["name"]
                # Coletar respostas válidas (não nulas)
                texts = [rec.get(vname) for rec in records if rec.get(vname)]
                if texts:
                    keywords = extract_keywords_from_texts(texts)
                    # Correção somente nas palavras‑chave (rápido).
                    # NÃO corrige os textos importados, apenas padroniza o chip de keyword.
                    if SPELLCHECK_SCOPE != "none":
                        # Padronização APENAS das palavras-chave (rápido): usa pt_BR.dic carregado uma única vez.
                        for kw in keywords:
                            ww = kw.get("word")
                            if isinstance(ww, str) and ww:
                                kw["word"] = _correct_keyword_with_dic(ww)

                    vm["keywords"] = keywords  # lista de {'word': ..., 'count': ...}
                else:
                    vm["keywords"] = []
    except Exception as e:
        # Em caso de erro, não interromper o fluxo; apenas registrar no console.
        print(f"⚠️ Erro ao extrair palavras‑chave: {e}")
    
    # ---------- DETECÇÃO AUTOMÁTICA DE SKIP LOGIC ----------
    try:
        _detect_skip_logic(vars_meta, records)
    except Exception as e:
        print(f"⚠️ Erro na detecção de skip logic: {e}")

    return created_at, vars_meta, filters_meta, records, value_orders, code_to_label

# ========== DETECÇÃO DE SKIP LOGIC ==========

def detect_scale_info(var_label: str, vmap: dict) -> Optional[dict]:
    """
    Detecta tipo e metadados de escala a partir do label da variável e mapa de valores.
    Tags reconhecidas no label: (NPS), (SAT), (CSAT), (CES)
    Retorna dict com scale_info ou None se não for escala reconhecida.
    """
    if not var_label or not vmap:
        return None

    tag_match = re.search(r'\((NPS|SAT|CSAT|CES)\)', var_label, re.IGNORECASE)
    if not tag_match:
        return None

    scale_type = tag_match.group(1).upper()

    keys = sorted([k for k in vmap.keys() if isinstance(k, (int, float))])
    if len(keys) < 3:
        return None

    # Normalizar labels igual aos records (sem ":", sem espaços extras)
    vals = [str(vmap[k]).replace(':', '').strip() for k in keys]

    # Detectar NS no último label
    ns_words = ['não sabe', 'nao sabe', 'não sei', 'ns/a', 'nsnr',
                'não tenho', 'nao tenho', 'não lembro', 'nao lembro']
    has_ns   = any(w in vals[-1].lower() for w in ns_words)
    ns_label  = vals[-1] if has_ns else None
    real_vals = vals[:-1] if has_ns else vals
    real_keys = keys[:-1] if has_ns else keys

    if len(real_vals) < 3:
        return None

    # Extrair valor mínimo real do primeiro label (ex: "0- Não recomenda" → 0)
    num_match = re.match(r'^(\d+)', real_vals[0])
    real_min  = int(num_match.group(1)) if num_match else int(real_keys[0])
    real_max  = real_min + len(real_vals) - 1

    # Direção: 0-10 é sempre ascendente; demais → detectar pelo primeiro âncora
    first_lower = real_vals[0].lower()
    neg = ['ruim', 'insatisf', 'discorda', 'difícil', 'dificil', 'nenhum',
           'péssim', 'pessim', 'pior', 'inseguro', 'nao confia', 'não confia']
    pos = ['bom', 'satisfeito', 'concorda', 'fácil', 'facil', 'certeza',
           'excelente', 'recomendaria', 'definitiv', 'muito bem', 'sempre']

    if real_min == 0:
        inverted = False          # convenção universal: 0 = pior
    elif any(w in first_lower for w in neg):
        inverted = False
    elif any(w in first_lower for w in pos):
        inverted = True
    else:
        inverted = False          # fallback seguro

    # label_to_value: label normalizado → valor numérico semântico
    label_to_value = {}
    for i, v in enumerate(real_vals):
        raw_val = real_min + i
        sem_val = (real_max - raw_val + real_min) if inverted else raw_val
        label_to_value[v] = sem_val

    return {
        'type':           scale_type,
        'real_min':       real_min,
        'real_max':       real_max,
        'inverted':       inverted,
        'has_ns':         has_ns,
        'ns_label':       ns_label,
        'label_to_value': label_to_value,
    }


def _detect_skip_logic(vars_meta: List[dict], records: List[dict]) -> None:
    """
    Detecta skip logic por taxa de resposta por categoria de Qx.
    Para cada categoria de Qx, calcula qual % dos seus respondentes respondeu Qy.
    Se há uma divisão clara (algumas categorias ≈100%, outras ≈0%), é skip logic.
    Resultado embutido em varMeta['skip_logic']. Modifica vars_meta in-place.
    """
    HIGH = 0.85   # taxa de resposta mínima para categoria "incluída" no filtro
    LOW  = 0.15   # taxa de resposta máxima para categoria "excluída"
    MIN_INCLUDED = 10  # mínimo de respondentes para classificar como "incluída"
                       # (evita falso positivo por n pequeno com taxa=100% por acaso)
                       # Categorias com n pequeno e taxa baixa ainda podem ser excluídas.

    total_records = len(records)
    if total_records == 0:
        return

    # Pré-calcular respondentes de cada variável (índices onde valor não é None)
    respondents: Dict[str, set] = {}
    for vm in vars_meta:
        vname = vm["name"]
        vtype = vm.get("var_type") or vm.get("type") or "single"
        if vtype == "multiple_response":
            respondents[vname] = {i for i, r in enumerate(records)
                                  if r.get(vname) and len(r.get(vname, [])) > 0}
        else:
            respondents[vname] = {i for i, r in enumerate(records)
                                  if r.get(vname) is not None}

    for qy_idx, vm_y in enumerate(vars_meta):
        qy_name = vm_y["name"]
        resp_qy = respondents[qy_name]
        n_qy = len(resp_qy)

        # Sem sentido se todos ou nenhum respondeu
        if n_qy == 0 or n_qy == total_records:
            continue

        vtype_y = vm_y.get("var_type") or vm_y.get("type") or "single"
        if vtype_y in ("string", "date"):
            continue

        # Varrer Qx anteriores (apenas categóricas simples)
        for qx_idx in range(qy_idx):
            vm_x = vars_meta[qx_idx]
            qx_name = vm_x["name"]
            vtype_x = vm_x.get("var_type") or vm_x.get("type") or "single"
            if vtype_x in ("string", "date", "multiple_response"):
                continue

            # Para cada categoria de Qx: calcular taxa de resposta em Qy
            # n_qx_cat = registros com esse valor de Qx
            # n_resp    = desses, quantos responderam Qy
            cat_counts: Dict[str, int] = {}    # total por categoria Qx
            cat_resp:   Dict[str, int] = {}    # respondentes de Qy por categoria Qx

            for i, rec in enumerate(records):
                qx_val = rec.get(qx_name)
                if qx_val is None:
                    continue
                key = str(qx_val).strip()
                cat_counts[key] = cat_counts.get(key, 0) + 1
                if i in resp_qy:
                    cat_resp[key] = cat_resp.get(key, 0) + 1

            if not cat_counts:
                continue

            # Classificar categorias por taxa de resposta
            included, excluded, intermediate = [], [], []
            for cat, total in cat_counts.items():
                if total == 0:
                    continue  # categoria vazia — ignorar
                rate = cat_resp.get(cat, 0) / total
                if rate <= LOW:
                    excluded.append(cat)  # taxa baixa → excluída (confiável mesmo com n pequeno)
                elif rate >= HIGH and total >= MIN_INCLUDED:
                    included.append(cat)  # taxa alta E amostra suficiente → incluída
                else:
                    intermediate.append(cat)  # zona cinzenta ou n insuficiente

            # Skip logic detectado: há incluídas, excluídas, e nenhuma intermediária
            if not included or not excluded or intermediate:
                continue

            # Cobertura: % dos respondentes de Qy que vieram das categorias incluídas
            n_included_resp = sum(cat_resp.get(c, 0) for c in included)
            coverage = round(n_included_resp / n_qy * 100, 1) if n_qy > 0 else 0.0

            print(f"🔗 Skip logic detectado: {qy_name} ← {qx_name} "
                  f"incluídas={included} excluídas={excluded} cobertura={coverage}%")

            vm_y["skip_logic"] = {
                "source_var":   qx_name,
                "source_title": vm_x.get("title", qx_name),
                "categories":   included,
                "coverage":     coverage,
                "n_filtered":   round(n_qy),
                "n_total":      total_records,
            }
            break  # usar apenas a Qx anterior mais próxima que satisfaz

# ========== DICIONÁRIO DE VARIÁVEIS ==========

def generate_data_dictionary(vars_meta: List[dict], records: List[dict], meta, out_path: str) -> None:
    """
    Gera um Excel com duas abas:
      1. Dicionário de Variáveis — uma linha por pergunta (compacto)
      2. Categorias e Frequências — uma linha por categoria (expandido)
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    vvl = getattr(meta, 'variable_value_labels', {}) or {}
    labels_map = getattr(meta, 'column_names_to_labels', {}) or {}
    USER_MISSING = {-1.0}
    total = len(records)

    def n_valid_rec(vname, vtype):
        if vtype == 'multiple_response':
            return sum(1 for r in records if r.get(vname) and len(r.get(vname, [])) > 0)
        return sum(1 for r in records if r.get(vname) is not None)

    def cats_summary(vm):
        vname = vm['name']
        vtype = vm.get('var_type') or vm.get('type', '')
        if vtype == 'multiple_response':
            freq: Dict[str, int] = {}
            for r in records:
                for opt in (r.get(vname) or []):
                    if opt: freq[opt] = freq.get(opt, 0) + 1
            return '; '.join(f"{k} ({v})" for k, v in sorted(freq.items(), key=lambda x: -x[1])[:6])
        elif vtype in ('string', 'date'):
            return ''
        else:
            col = vm.get('sheet_code', vname)
            val_labels = vvl.get(col, {})
            return '; '.join(
                f"{int(k) if isinstance(k,float) and k==int(k) else k}={v}"
                for k, v in list(val_labels.items())[:6]
            )

    def get_skip_source_p(vm):
        sl = vm.get('skip_logic')
        if not sl: return ''
        src = sl.get('source_var', '')
        for v in vars_meta:
            if v['name'] == src and v.get('p_num'):
                return f"P{v['p_num']}"
        return sl.get('source_var', '')

    # Estilos
    BLUE_H  = PatternFill('solid', fgColor='1A3A5C')
    BLUE_L  = PatternFill('solid', fgColor='E6F1FB')
    ALT     = PatternFill('solid', fgColor='F8FAFC')
    WHITE   = PatternFill('solid', fgColor='FFFFFF')

    def thin():
        s = Side(style='thin', color='D0D5DD')
        return Border(left=s, right=s, top=s, bottom=s)

    def apply_sheet(ws, headers, widths, rows):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = BLUE_H
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin()
        ws.row_dimensions[1].height = 28

        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill = ALT if ri % 2 == 0 else WHITE
                cell.font = Font(bold=(ci == 1), color='1A1A2E', size=10)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin()
            ws.row_dimensions[ri].height = 28

        # Destacar col Nº
        for ri in range(2, len(rows) + 2):
            ws.cell(ri, 1).fill = BLUE_L
            ws.cell(ri, 1).font = Font(bold=True, color='0C447C', size=10)

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    wb = openpyxl.Workbook()

    # ── ABA 1: DICIONÁRIO ──
    ws1 = wb.active
    ws1.title = 'Dicionário de Variáveis'
    rows1 = []
    for vm in vars_meta:
        pn  = f"P{vm.get('p_num','')}"
        nv  = n_valid_rec(vm['name'], vm.get('var_type') or vm.get('type',''))
        nm  = total - nv
        rows1.append([
            pn,
            vm.get('sheet_code', vm['name']),
            vm.get('title', ''),
            vm.get('spss_type', ''),
            nv,
            nm,
            f"{nm/total*100:.1f}%" if total else '—',
            get_skip_source_p(vm),
            cats_summary(vm),
        ])
    apply_sheet(ws1,
        ['Nº','Variável SPSS','Pergunta','Tipo','N válidos','N missing','% missing','Filtrada por','Categorias / Opções'],
        [8, 14, 48, 20, 10, 10, 10, 12, 55],
        rows1)

    # ── ABA 2: CATEGORIAS EXPANDIDAS ──
    ws2 = wb.create_sheet('Categorias e Frequências')
    rows2 = []
    for vm in vars_meta:
        pn    = f"P{vm.get('p_num','')}"
        vname = vm['name']
        vtype = vm.get('var_type') or vm.get('type', '')
        col   = vm.get('sheet_code', vname)
        title = vm.get('title', '')
        spss  = vm.get('spss_type', '')
        nv    = n_valid_rec(vname, vtype)

        if vtype == 'multiple_response':
            freq: Dict[str, int] = {}
            for r in records:
                for opt in (r.get(vname) or []):
                    if opt: freq[opt] = freq.get(opt, 0) + 1
            for opt, cnt in sorted(freq.items(), key=lambda x: -x[1]):
                pct = f"{cnt/nv*100:.1f}%" if nv else '—'
                rows2.append([pn, col, title, spss, '—', opt, cnt, pct])
        elif vtype in ('string', 'date'):
            rows2.append([pn, col, title, spss, '—', '(texto livre)', nv, '—'])
        else:
            val_labels = vvl.get(col, {})
            freq2: Dict[str, int] = {}
            for r in records:
                v = r.get(vname)
                if v is not None: freq2[str(v)] = freq2.get(str(v), 0) + 1
            for code, lbl in val_labels.items():
                code_str = str(int(code)) if isinstance(code, float) and code == int(code) else str(code)
                cnt = freq2.get(lbl, 0)
                pct = f"{cnt/nv*100:.1f}%" if nv and cnt else '0,0%'
                rows2.append([pn, col, title, spss, code_str, lbl, cnt, pct])

    apply_sheet(ws2,
        ['Nº','Variável SPSS','Pergunta','Tipo','Código SPSS','Categoria / Opção','Frequência','%'],
        [8, 14, 42, 20, 12, 42, 12, 10],
        rows2)

    wb.save(out_path)


# ========== GERAÇÃO DE HTML ==========

def render_html_with_working_filters(file_source: str, created_at: str, client_name: str,
                                    vars_meta: List[dict], filters_meta: List[dict], 
                                    records: List[dict], value_orders: dict, code_to_label: dict,
                                    output_html_name: str = None,
                                    user_description: str = '',
                                    user_period: str = '') -> str:

    # JSON strings seguros para JavaScript
    vars_meta_json = json.dumps(vars_meta, ensure_ascii=False)
    filters_meta_json = json.dumps(filters_meta, ensure_ascii=False)
    records_json = json.dumps(records, ensure_ascii=False)
    value_orders_js = json.dumps(value_orders, ensure_ascii=False)
    code_to_label_js = json.dumps(code_to_label, ensure_ascii=False)

    # Ler client_logo e description do dashboard_overlay_config.json (mesma pasta)
    client_logo = ''
    description = user_description  # Prioridade máxima: digitado pelo usuário na GUI
    period = user_period             # Período digitado pelo usuário na GUI
    import os
    config_candidates = [
        'dashboard_overlay_config.json',
        os.path.join(os.getcwd(), 'dashboard_overlay_config.json'),
    ]
    for cfg_path in config_candidates:
        if os.path.exists(cfg_path):
            try:
                with open(cfg_path, 'r', encoding='utf-8') as _f:
                    _cfg = json.load(_f)
                client_logo = _cfg.get('client_logo', '')
                # Buscar description no item de menu
                import re as _re
                def _kw(name):
                    stops = {'de','da','do','dos','das','e','a','o','os','as',
                             'base','ponderada','html','sav','dashboard','relatorio','resultados'}
                    return {w for w in _re.findall(r'[a-zA-Z\u00C0-\u00FF]{3,}', name.lower()) if w not in stops}
                def _find_desc_exact(items, target):
                    for item in items:
                        if item.get('file', '') == target:
                            return item.get('description', '')
                        if 'children' in item:
                            r = _find_desc_exact(item['children'], target)
                            if r is not None: return r
                    return None
                def _find_desc_fuzzy(items, sav_base):
                    sw = _kw(sav_base)
                    best_s, best_d = 0, None
                    def _scan(items):
                        nonlocal best_s, best_d
                        for item in items:
                            fw = _kw(item.get('file', ''))
                            if fw and sw:
                                s = len(fw & sw) / max(len(fw), len(sw))
                                if s > best_s:
                                    best_s, best_d = s, item.get('description', '')
                            if 'children' in item: _scan(item['children'])
                    _scan(items)
                    return best_d if best_s >= 0.4 else None
                # Prioridade 1: match exato pelo HTML de saída
                if output_html_name:
                    _found = _find_desc_exact(_cfg.get('items', []), os.path.basename(output_html_name))
                    if _found: description = _found
                # Prioridade 2: match exato pelo nome derivado do SAV
                if not description:
                    _html_from_sav = os.path.basename(file_source).replace('.sav','').replace('.SAV','') + '.html'
                    _found = _find_desc_exact(_cfg.get('items', []), _html_from_sav)
                    if _found: description = _found
                # Prioridade 3: fuzzy por palavras-chave do nome do SAV
                if not description:
                    _sav_base = os.path.basename(file_source).replace('.sav','').replace('.SAV','')
                    _found = _find_desc_fuzzy(_cfg.get('items', []), _sav_base)
                    if _found: description = _found
                # Fallback: top-level description
                if not description:
                    description = _cfg.get('description', '')
            except:
                pass
            break

    # Formatação de data para nome do arquivo
    from datetime import datetime
    try:
        # Tentar extrair data de created_at ou usar data atual
        data_formatada = datetime.now().strftime("%d-%m-%Y")
        if created_at:
            # Se created_at tem formato específico, tentar parseá-lo
            data_formatada = datetime.now().strftime("%d-%m-%Y")
    except:
        data_formatada = datetime.now().strftime("%d-%m-%Y")
    
    # Nome base do arquivo sem extensão
    nome_arquivo = file_source.replace('.sav', '').replace('.SAV', '')
    titulo_pdf = f"Relatorio de resultados_{nome_arquivo}_{data_formatada}"
    
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{titulo_pdf}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js"></script>

    <script>
    // Ordem original das categorias vinda do SPSS
    const VARS_VALUE_ORDER = {value_orders_js};
    // Mapeamento código -> label para exibição
    const CODE_TO_LABEL = {code_to_label_js};
    
    // Função para formatação brasileira (vírgula decimal)
    function formatBR(number, decimals = 2) {{
        if (number === null || number === undefined || isNaN(number)) return 'N/A';
        return number.toFixed(decimals).replace('.', ',');
    }}
    </script>

    <style>
        :root {{
            --primary: #7BAFC0;
            --primary-dark: #4d8a9e;
            --teal-light: #EDF5F7;
            --green: #6FAB8A;
            --green-dark: #3d8a62;
            --red: #B83B5C;
            --success: #6FAB8A;
            --warning: #FF9800;
            --info: #9C27B0;
            --background: #f8f9fa;
            --text: #333;
            --border: #e5e5e5;
            --radius: 8px;
            --shadow: 0 1px 3px rgba(0,0,0,0.07);
            --shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}

        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: var(--background);
            color: var(--text);
            line-height: 1.6;
            padding: 15px;
            padding-top: 68px;
        }}

        /* ===== BARRA DE FILTROS COMPACTA ===== */
        .filters-container {{
            background: white;
            border-bottom: 0.5px solid var(--border);
            box-shadow: 0 1px 4px rgba(0,0,0,0.07);
            position: fixed;
            top: 0; left: 0; right: 0;
            z-index: 9999;
            width: 100%;
        }}

        .filters-bar {{
            display: flex;
            align-items: stretch;
            height: 64px;
            padding: 0 20px;
        }}

        .filter-bar-label {{
            display: flex;
            align-items: center;
            padding-right: 18px;
            border-right: 0.5px solid var(--border);
            margin-right: 18px;
            flex-shrink: 0;
            font-size: 13px;
            font-weight: 600;
            color: #6c757d;
            letter-spacing: 0.01em;
        }}

        .filters-inline {{
            display: flex;
            align-items: center;
            gap: 12px;
            flex: 1;
            overflow: visible;
        }}

        .filter-col {{
            display: flex;
            flex-direction: column;
            justify-content: center;
            gap: 4px;
            flex: 1;
            min-width: 80px;
        }}

        .filter-col-label {{
            font-size: 12px;
            color: #6c757d;
            font-weight: 500;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .filter-col-select {{
            font-size: 13px;
            padding: 4px 26px 4px 9px;
            border: 0.5px solid #ddd;
            border-radius: 5px;
            background: #f8f9fa;
            color: var(--text);
            cursor: pointer;
            height: 30px;
            appearance: none;
            -webkit-appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg width='10' height='6' viewBox='0 0 10 6' fill='none' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M1 1L5 5L9 1' stroke='%23999' stroke-width='1.5' stroke-linecap='round' stroke-linejoin='round'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 8px center;
            transition: border-color 0.15s;
            width: 100%;
        }}

        .filter-col-select:focus {{
            outline: none;
            border-color: var(--primary);
        }}

        .icon-btn.present {{ color: var(--primary-dark); }}
        .icon-btn.present:hover {{ background: var(--teal-light); color: var(--primary-dark); }}
        .icon-btn.apply {{ color: #6FAB8A; }}
        .icon-btn.apply:hover {{ background: rgba(111,171,138,0.1); color: #3d8a62; }}
        .icon-btn.clear {{ color: #B83B5C; }}
        .icon-btn.clear:hover {{ background: rgba(184,59,92,0.08); color: #8f2a44; }}
        .icon-btn.excel {{ color: #3d8a62; }}
        .icon-btn.excel:hover {{ background: rgba(111,171,138,0.1); color: #2a6245; }}
        .icon-btn.pdf {{ color: #B83B5C; }}
        .icon-btn.pdf:hover {{ background: rgba(184,59,92,0.08); color: #8f2a44; }}

        .filter-bar-actions {{
            display: flex;
            align-items: center;
            gap: 2px;
            padding-left: 14px;
            border-left: 0.5px solid var(--border);
            margin-left: 4px;
            flex-shrink: 0;
        }}

        .icon-btn {{
            width: 34px;
            height: 34px;
            border-radius: 7px;
            border: none;
            background: transparent;
            color: #bbb;
            display: flex;
            align-items: center;
            justify-content: center;
            cursor: pointer;
            transition: background 0.15s, color 0.15s;
            position: relative;
        }}

        .icon-btn:hover {{
            background: #f1f3f5;
            color: #444;
        }}

        .icon-btn svg {{
            width: 15px;
            height: 15px;
        }}

        .icon-btn .icon-tooltip {{
            display: none;
            position: absolute;
            top: calc(100% + 6px);
            left: 50%;
            transform: translateX(-50%);
            background: #222;
            color: #fff;
            font-size: 11px;
            font-weight: 400;
            padding: 4px 9px;
            border-radius: 5px;
            white-space: nowrap;
            pointer-events: none;
            z-index: 10000;
        }}

        .icon-btn:hover .icon-tooltip {{
            display: block;
        }}

        .icon-btn.loading {{
            opacity: 0.45;
            pointer-events: none;
        }}

        .content {{
            margin-top: 0;
        }}

        .custom-dropdown {{
            position: relative;
        }}

        .dropdown-button {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 10px 12px;
            background: white;
            border: 1px solid var(--border);
            border-radius: var(--radius);
            cursor: pointer;
            user-select: none;
            transition: all 0.2s ease;
            font-size: 13px;
        }}

        .dropdown-button:hover {{
            border-color: var(--primary);
            box-shadow: 0 0 0 1px rgba(123, 175, 192, 0.1);
        }}

        .dropdown-button.open {{
            border-color: var(--primary);
            box-shadow: 0 0 0 2px rgba(123, 175, 192, 0.1);
        }}

        .dropdown-content {{
            position: absolute;
            top: 100%;
            left: 0;
            right: 0;
            background: white;
            border: 1px solid var(--border);
            border-radius: var(--radius);
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            max-height: 200px;
            overflow-y: auto;
            z-index: 99999;
            display: none;
            margin-top: 2px;
        }}

        .dropdown-content.show {{
            display: block;
        }}

        .dropdown-option {{
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 8px 12px;
            cursor: pointer;
            transition: background 0.2s ease;
            font-size: 13px;
        }}

        .dropdown-option:hover {{
            background: #f8f9fa;
        }}

        .dropdown-option.select-all {{
            background: #f1f3f4;
            font-weight: 600;
            border-bottom: 1px solid var(--border);
        }}

        .dropdown-option input[type="checkbox"] {{
            margin: 0;
        }}

        .dropdown-option label {{
            cursor: pointer;
            flex: 1;
        }}

        .arrow {{
            transition: transform 0.2s ease;
            color: #666;
            font-size: 12px;
        }}

        .dropdown-button.open .arrow {{
            transform: rotate(180deg);
        }}

        .content {{
            display: flex;
            flex-direction: column;
            gap: 20px;
            padding-top: 20px;
        }}

        .section {{
            background: white;
            border-radius: var(--radius);
            box-shadow: var(--shadow);
            overflow: hidden;
            border: 1px solid var(--border);
        }}

        .section-header {{
            background: var(--teal-light);
            padding: 16px 20px;
            border-bottom: 1px solid rgba(123,175,192,0.2);
        }}

        .type-dot {{
            display: inline-block;
            width: 9px;
            height: 9px;
            border-radius: 50%;
            margin-right: 6px;
            vertical-align: middle;
            flex-shrink: 0;
        }}
        .dot-cat   {{ background: var(--primary); }}
        .dot-mr    {{ background: var(--green); }}
        .dot-open  {{ background: #aaa; }}
        .dot-scale {{ background: var(--teal-dark); }}
        .dot-date  {{ background: #aaa; }}

        .section-title {{
            font-size: 15px;
            font-weight: 500;
            margin-bottom: 5px;
            display: block;
        }}

        .section-title .p-num {{
            font-weight: 600;
            white-space: nowrap;
            margin-right: 4px;
        }}

        .section-subtitle {{
            font-size: 13px;
            color: #6c757d;
            display: flex;
            align-items: center;
            gap: 6px;
            flex-wrap: wrap;
        }}

        .type-pill {{
            display: inline-flex;
            align-items: center;
            padding: 2px 9px;
            background: rgba(123,175,192,0.12);
            border: 0.5px solid rgba(123,175,192,0.4);
            border-radius: 20px;
            font-size: 11px;
            font-weight: 500;
            color: var(--primary-dark);
            white-space: nowrap;
        }}

        .type-pill.mr {{
            background: rgba(111,171,138,0.1);
            border-color: rgba(111,171,138,0.4);
            color: var(--green-dark);
        }}

        .skip-tag {{
            position: relative;
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 2px 9px;
            background: rgba(184,59,92,0.08);
            border: 0.5px solid rgba(184,59,92,0.35);
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
            color: var(--red);
            cursor: default;
            white-space: nowrap;
            vertical-align: middle;
            margin-left: 4px;
        }}

        .skip-tag .skip-tooltip {{
            display: none;
            position: absolute;
            top: calc(100% + 6px);
            left: 0;
            background: white;
            border: 1px solid var(--border);
            border-radius: 8px;
            padding: 10px 13px;
            font-size: 12px;
            font-weight: 400;
            color: #6c757d;
            white-space: nowrap;
            box-shadow: 0 4px 12px rgba(0,0,0,0.10);
            z-index: 100;
            line-height: 1.7;
            min-width: 220px;
        }}

        .skip-tag .skip-tooltip strong {{
            color: var(--text);
            font-weight: 600;
        }}

        .skip-tag .skip-tooltip .skip-pill {{
            display: inline-block;
            padding: 1px 7px;
            background: rgba(184,59,92,0.08);
            border: 0.5px solid rgba(184,59,92,0.35);
            border-radius: 20px;
            font-size: 11px;
            color: var(--red);
            font-weight: 600;
            margin: 2px 2px 0 0;
        }}

        .skip-tag:hover .skip-tooltip {{
            display: block;
        }}

        .section-header.mr-header {{
            background: rgba(111,171,138,0.08);
            border-bottom: 1px solid rgba(111,171,138,0.2);
        }}

        .section-header.nps-header {{
            background: rgba(184,59,92,0.05);
            border-bottom-color: rgba(184,59,92,0.15);
        }}

        .type-pill.nps {{
            background: rgba(184,59,92,0.08);
            border-color: rgba(184,59,92,0.35);
            color: var(--red);
        }}

        .type-pill.scale {{
            background: rgba(123,175,192,0.12);
            border-color: rgba(123,175,192,0.4);
            color: var(--primary-dark);
        }}


        .type-pill.open {{
            background: #f1f3f5;
            border-color: #ddd;
            color: #6c757d;
        }}

        .section-header.open-header {{
            background: var(--teal-light);
            border-bottom-color: rgba(123,175,192,0.2);
        }}
        /* ===== SCALE METRIC ROWS (NPS / SAT / CSAT / CES) ===== */
        .scale-metrics {{
            border-top: 1.5px solid var(--border);
        }}
        .scale-metric-row {{
            display: flex;
            align-items: center;
            border-bottom: 0.5px solid var(--border);
        }}
        .scale-metric-row:last-child {{ border-bottom: none; }}
        .scale-metric-label {{
            padding: 7px 10px;
            font-size: 11px;
            font-weight: 600;
            color: var(--primary-dark);
            width: 45%;
            border-right: 0.5px solid var(--border);
        }}
        .scale-metric-value {{
            padding: 7px 10px;
            font-size: 13px;
            font-weight: 600;
            color: var(--primary-dark);
            flex: 1;
        }}
        .scale-metric-row.highlight {{ background: var(--teal-light); }}
        .scale-metric-row.nps-score {{
            border-top: 1px solid var(--border);
        }}
        .scale-metric-row.nps-score .scale-metric-label {{
            font-size: 12px;
        }}
        .scale-metric-row.nps-score .scale-metric-value {{
            font-size: 16px;
        }}
        /* Score negativo → vinho */
        .scale-metric-row.nps-score-neg {{ background: rgba(184,59,92,0.05); border-top-color: rgba(184,59,92,0.2); }}
        .scale-metric-row.nps-score-neg .scale-metric-label,
        .scale-metric-row.nps-score-neg .scale-metric-value {{ color: var(--red); }}
        /* Score zero → cinza */
        .scale-metric-row.nps-score-zero {{ background: #f8f9fa; }}
        .scale-metric-row.nps-score-zero .scale-metric-label,
        .scale-metric-row.nps-score-zero .scale-metric-value {{ color: #888; }}
        /* Score positivo → verde */
        .scale-metric-row.nps-score-pos {{ background: rgba(111,171,138,0.05); border-top-color: rgba(111,171,138,0.25); }}
        .scale-metric-row.nps-score-pos .scale-metric-label,
        .scale-metric-row.nps-score-pos .scale-metric-value {{ color: var(--green-dark); }}
        .scale-metric-row.nps-det .scale-metric-label,
        .scale-metric-row.nps-det .scale-metric-value {{ color: var(--red); }}
        .scale-metric-row.nps-neu .scale-metric-label,
        .scale-metric-row.nps-neu .scale-metric-value {{ color: #888; }}
        .scale-metric-row.nps-pro .scale-metric-label,
        .scale-metric-row.nps-pro .scale-metric-value {{ color: var(--green-dark); }}
        /* Cores das barras e percentuais por tipo de variável */
        .ivv-bar-bg {{ background: var(--bar-color, var(--primary)); }}
        .ivv-pct-col {{ color: var(--bar-text, var(--primary-dark)); }}


        }}

        .chart-container {{
            position: relative;
            height: 350px;
            margin-bottom: 15px;
        }}

        .table-container {{
            overflow-x: auto;
            margin-top: 15px;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }}

        th, td {{
            text-align: left;
            padding: 10px 8px;
            border-bottom: 1px solid var(--border);
        }}

        th {{
            background: #f8f9fa;
            font-weight: 600;
            font-size: 13px;
        }}

        td {{
            font-size: 13px;
        }}

        /* ===== IVV-STYLE INLINE BAR TABLE ===== */
        .ivv-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            table-layout: fixed;
            margin: 0;
        }}

        .ivv-table th {{
            background: #f8f9fa;
            font-weight: 600;
            font-size: 12px;
            color: #6c757d;
            padding: 8px 10px;
            border-bottom: 2px solid var(--border);
            text-align: left;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .ivv-table td {{
            padding: 6px 10px;
            border-bottom: 1px solid var(--border);
            font-size: 13px;
            vertical-align: middle;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .ivv-table tr.ivv-row {{
            cursor: pointer;
            transition: background 0.15s ease;
        }}

        .ivv-table tr.ivv-row:hover {{
            background: #f0f4ff;
        }}

        .ivv-table tr.ivv-row.drilldown-active {{
            background: #dbeafe;
        }}

        .ivv-table tr.ivv-total-row td {{
            font-weight: 600;
            background: #f8f9fa;
            border-top: 2px solid var(--border);
            border-bottom: none;
        }}

        .ivv-bar-cell {{
            position: relative;
            min-width: 100px;
        }}

        .ivv-bar-bg {{
            position: absolute;
            top: 4px;
            bottom: 4px;
            left: 0;
            border-radius: 3px;
            background: var(--bar-color, var(--primary));
            opacity: 0.18;
            transition: width 0.5s ease, opacity 0.2s ease;
            max-width: calc(100% - 16px);
        }}

        .ivv-bar-val {{
            position: relative;
            font-size: 13px;
            font-weight: 600;
            color: var(--bar-text, var(--primary-dark));
            padding-left: 6px;
        }}

        .ivv-pct-col {{
            text-align: left;
            font-weight: 600;
            font-size: 13px;
            color: var(--bar-text, var(--primary-dark));
            white-space: nowrap;
        }}

        .ivv-base-col {{
            text-align: left;
            font-size: 12px;
            color: #6c757d;
            white-space: nowrap;
        }}

        .ivv-label-col {{
            width: 45%;
        }}

        .section-content {{
            padding: 16px 20px 16px;
        }}

        /* Toggle NS */
        .ivv-ns-row td {{ color: #bbb; font-style: italic; }}
        .ivv-ns-row .pct, .ivv-ns-row .ivv-bar-val {{ color: #ccc !important; }}
        .ivv-ns-row .ivv-bar-bg {{ opacity: .07; }}

        .ivv-drilldown-hint {{
            font-size: 11px;
            color: #bbb;
            padding: 8px 16px 10px;
            font-style: italic;
        }}

        .ns-toggle-wrap {{
            display: inline-flex; align-items: center; gap: 6px;
            cursor: pointer; user-select: none; margin-left: 4px;
        }}
        .ns-track {{
            width: 28px; height: 16px; border-radius: 8px;
            position: relative; transition: background .2s; flex-shrink: 0;
        }}
        .ns-on     {{ background: var(--primary); }}
        .ns-on-nps {{ background: var(--red); }}
        .ns-off    {{ background: #ccc; }}
        .ns-thumb {{
            width: 12px; height: 12px; border-radius: 50%;
            background: white; position: absolute; top: 2px; transition: left .2s;
        }}
        .ns-on  .ns-thumb, .ns-on-nps .ns-thumb {{ left: 14px; }}
        .ns-off .ns-thumb {{ left: 2px; }}
        .ns-toggle-label {{ font-size: 11px; color: #6c757d; }}
            font-size: 11px;
            color: #6c757d;
            margin-top: 8px;
        }}

        .outros-sub {{
            margin: 12px 0 4px;
            border: 0.5px solid rgba(123,175,192,0.3);
            border-radius: var(--radius);
            overflow: hidden;
        }}

        .outros-sub-header {{
            background: var(--teal-light);
            padding: 8px 14px;
            border-bottom: 0.5px solid rgba(123,175,192,0.2);
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }}

        .outros-sub-label {{
            font-size: 11px;
            font-weight: 600;
            color: var(--primary-dark);
        }}

        .outros-sub-n {{
            font-size: 11px;
            color: var(--primary-dark);
            background: rgba(123,175,192,0.15);
            padding: 1px 7px;
            border-radius: 20px;
        }}

        .outros-kw {{
            padding: 2px 8px;
            background: white;
            border: 0.5px solid var(--border);
            border-radius: 20px;
            font-size: 11px;
            color: #6c757d;
            cursor: pointer;
            transition: background 0.15s, border-color 0.15s;
        }}

        .outros-kw:hover {{
            background: var(--teal-light);
            border-color: var(--primary);
            color: var(--primary-dark);
        }}

        .outros-kw.active {{
            background: var(--teal-light);
            border-color: var(--primary);
            color: var(--primary-dark);
            font-weight: 600;
        }}

        .outros-list {{
            padding: 4px 0;
            max-height: 200px;
            overflow-y: auto;
        }}

        .outros-item {{
            font-size: 13px;
            color: #444;
            padding: 8px 14px;
            border-bottom: 0.5px solid var(--border);
        }}

        .outros-item:last-child {{ border-bottom: none; }}

        /* Responsivo */
        @media (max-width: 768px) {{
            body {{
                padding: 10px;
                padding-top: 160px; /* Aumentado para mobile */
            }}
            
            .filters-grid {{
                grid-template-columns: 1fr;
                padding: 15px;
                gap: 12px;
            }}
            
            .filter-actions {{
                flex-direction: column;
                gap: 6px;
            }}
            
            .filter-btn {{
                font-size: 12px;
                padding: 6px 12px;
            }}
        }}

        @media (max-width: 480px) {{
            .filters-header {{
                flex-direction: column;
                gap: 12px;
                align-items: stretch;
            }}
            
            .filter-actions {{
                flex-direction: row;
                justify-content: center;
            }}
        }}
    </style>
</head>
<body>
    <div class="filters-container">
        <div class="filters-bar" id="filtersBar">
            <span class="filter-bar-label" id="filtrarLabel">Filtrar</span>
            <div class="filters-inline" id="filtersGrid">
                <!-- Filtros gerados dinamicamente -->
            </div>
            <div class="filter-bar-actions">
                <button id="btn-apply" class="icon-btn apply" onclick="applyFilters()" onmouseenter="this.querySelector('.icon-tooltip').style.display='block'" onmouseleave="this.querySelector('.icon-tooltip').style.display='none'">
                    <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><polyline points="2.5,8.5 6,12 13.5,4"/></svg>
                    <span class="icon-tooltip">Aplicar filtros</span>
                </button>
                <button id="btn-clear" class="icon-btn clear" onclick="clearFilters()" onmouseenter="this.querySelector('.icon-tooltip').style.display='block'" onmouseleave="this.querySelector('.icon-tooltip').style.display='none'">
                    <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="3.5" y1="3.5" x2="12.5" y2="12.5"/><line x1="12.5" y1="3.5" x2="3.5" y2="12.5"/></svg>
                    <span class="icon-tooltip">Limpar filtros</span>
                </button>
                <div id="btn-sep" class="icon-btn-sep"></div>
                <button class="icon-btn present" onclick="startPresentation()" onmouseenter="this.querySelector('.icon-tooltip').style.display='block'" onmouseleave="this.querySelector('.icon-tooltip').style.display='none'">
                    <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="1.5" y="2.5" width="13" height="9" rx="1.5"/><line x1="8" y1="11.5" x2="8" y2="14"/><line x1="5.5" y1="14" x2="10.5" y2="14"/></svg>
                    <span class="icon-tooltip">Apresentação</span>
                </button>
                <button class="icon-btn excel" onclick="exportAllTables()" onmouseenter="this.querySelector('.icon-tooltip').style.display='block'" onmouseleave="this.querySelector('.icon-tooltip').style.display='none'">
                    <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="2" y="2" width="12" height="12" rx="2"/><line x1="6" y1="2" x2="6" y2="14"/><line x1="2" y1="6" x2="14" y2="6"/><line x1="2" y1="10" x2="14" y2="10"/></svg>
                    <span class="icon-tooltip">Exportar Excel</span>
                </button>
                <button class="icon-btn pdf" onclick="exportToPDF(event)" onmouseenter="this.querySelector('.icon-tooltip').style.display='block'" onmouseleave="this.querySelector('.icon-tooltip').style.display='none'">
                    <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M4 2h6l4 4v8a1 1 0 01-1 1H4a1 1 0 01-1-1V3a1 1 0 011-1z"/><polyline points="10,2 10,6 14,6"/><line x1="5" y1="10" x2="11" y2="10"/><line x1="5" y1="12.5" x2="9" y2="12.5"/></svg>
                    <span class="icon-tooltip">Exportar PDF</span>
                </button>
            </div>
        </div>
    </div>

    <div class="content" id="content">
        <!-- Conteúdo gerado dinamicamente -->
    </div>

    <script>
        // DADOS GLOBAIS - JSONs seguros
        const VARS_META = {vars_meta_json};
        const FILTERS_META = {filters_meta_json};
        const RECORDS = {records_json};
        const FILTERS = FILTERS_META;
        const CHART_LABEL_MAX = {CHART_LABEL_MAX};
        const DESCRIPTION  = '{description}';
        const FILE_SOURCE  = '{file_source}';
        const PERIOD       = '{period}';
    // Função para quebrar rótulos longos em múltiplas linhas
    function wrapLabel(label, maxLen) {{
        if (label === null || label === undefined) return [''];
        const raw = String(label).trim();
        if (raw.length <= maxLen) return [raw];

        const words = raw.split(' ');
        const lines = [];
        let current = '';

        words.forEach(word => {{
            const tentative = (current + ' ' + word).trim();
            if (tentative.length > maxLen) {{
                if (current.trim().length > 0) {{
                    lines.push(current.trim());
                }}
                current = word;
            }} else {{
                current = tentative;
            }}
        }});

        if (current.trim().length > 0) {{
            lines.push(current.trim());
        }}
        return lines;
    }}


        // Estados globais
        let charts = {{}};

        
        // INICIALIZAÇÃO
        document.addEventListener('DOMContentLoaded', function() {{
            console.log('🌍 Dashboard SPSS Universal carregado');
            console.log('📊 ' + VARS_META.length + ' variáveis, ' + FILTERS.length + ' filtros, ' + RECORDS.length + ' registros');
            buildFilters();
            renderAll();
        }});

        // FILTROS - USANDO f em vez de filter para evitar conflitos
        function buildFilters() {{
            const container = document.getElementById('filtersGrid');
            if (!container) return;
            container.innerHTML = '';

            if (FILTERS.length === 0) {{
                // Sem filtros: ocultar label, grid e botões de aplicar/limpar
                const bar = document.getElementById('filtersBar');
                const lbl = document.getElementById('filtrarLabel');
                if (bar)  bar.style.height  = '48px';
                if (lbl)  lbl.style.display = 'none';
                const inline = document.getElementById('filtersGrid');
                if (inline) inline.style.display = 'none';
                const btnApply = document.getElementById('btn-apply');
                const btnClear = document.getElementById('btn-clear');
                const btnSep   = document.getElementById('btn-sep');
                if (btnApply) btnApply.style.display = 'none';
                if (btnClear) btnClear.style.display = 'none';
                if (btnSep)   btnSep.style.display   = 'none';
                const actions = document.querySelector('.filter-bar-actions');
                if (actions) actions.style.marginLeft = 'auto';
                return;
            }}

            FILTERS.forEach(f => {{
                const filterGroup = document.createElement('div');
                filterGroup.className = 'filter-col';

                const label = document.createElement('span');
                label.className = 'filter-col-label';
                label.textContent = f.title;
                filterGroup.appendChild(label);

                const dropdownContainer = document.createElement('div');
                dropdownContainer.className = 'custom-dropdown';

                const dropdownButton = document.createElement('div');
                dropdownButton.className = 'dropdown-button filter-col-select';
                dropdownButton.style.cssText = 'display:flex;justify-content:space-between;align-items:center;cursor:pointer;user-select:none;';
                dropdownButton.onclick = () => toggleDropdown(f.name);
                dropdownButton.innerHTML = '<span id="' + f.name + 'Text">Todos</span><span style="font-size:9px;color:#999">▼</span>';

                const dropdownContent = document.createElement('div');
                dropdownContent.className = 'dropdown-content';
                dropdownContent.id = f.name + 'Content';

                const selectAllOption = document.createElement('div');
                selectAllOption.className = 'dropdown-option select-all';
                selectAllOption.innerHTML = '<input type="checkbox" onchange="selectAllOptions(\\'' + f.name + '\\')"><label>Selecionar Todos</label>';
                dropdownContent.appendChild(selectAllOption);

                f.values.forEach(value => {{
                    const option = document.createElement('div');
                    option.className = 'dropdown-option';
                    option.dataset.filterName = f.name;
                    option.dataset.filterValue = value;
                    option.innerHTML = '<input type="checkbox" value="' + value + '" onchange="updateDropdownText(\\'' + f.name + '\\'); updateFilterCounts()"><label>' + value + ' <span class="filter-count" style="color:#aaa;font-size:11px"></span></label>';
                    dropdownContent.appendChild(option);
                }});

                dropdownContainer.appendChild(dropdownButton);
                dropdownContainer.appendChild(dropdownContent);
                filterGroup.appendChild(dropdownContainer);
                container.appendChild(filterGroup);
            }});

            // Contagens iniciais
            updateFilterCounts();
        }}

        function updateFilterCounts() {{
            // Para cada filtro, calcular contagem excluindo o próprio filtro (cross-filter)
            FILTERS.forEach(f => {{
                const content = document.getElementById(f.name + 'Content');
                if (!content) return;

                // Records filtrados excluindo este filtro
                const crossRecords = getFilteredRecords(f.name);

                // Contar por categoria (ponderado)
                const counts = {{}};
                crossRecords.forEach(r => {{
                    const v = r[f.name];
                    if (v !== null && v !== undefined && String(v).trim() !== '') {{
                        const k = String(v).trim();
                        const w = r.__weight__ || 1;
                        counts[k] = (counts[k] || 0) + w;
                    }}
                }});

                // Atualizar spans de contagem
                content.querySelectorAll('.dropdown-option:not(.select-all)').forEach(opt => {{
                    const val = opt.dataset.filterValue;
                    const span = opt.querySelector('.filter-count');
                    if (span) {{
                        const n = counts[val] || 0;
                        span.textContent = n > 0 ? '(' + Math.round(n) + ')' : '(0)';
                    }}
                }});
            }});
        }}

        function toggleDropdown(filterId) {{
            const button = event.currentTarget;
            const content = document.getElementById(filterId + 'Content');
            
            document.querySelectorAll('.dropdown-content').forEach(dropdown => {{
                if (dropdown !== content) dropdown.classList.remove('show');
            }});
            document.querySelectorAll('.dropdown-button').forEach(btn => {{
                if (btn !== button) btn.classList.remove('open');
            }});
            
            content.classList.toggle('show');
            button.classList.toggle('open');
        }}

        function selectAllOptions(filterId) {{
            const content = document.getElementById(filterId + 'Content');
            const selectAllCheckbox = content.querySelector('.select-all input');
            const checkboxes = content.querySelectorAll('.dropdown-option:not(.select-all) input');
            
            checkboxes.forEach(cb => cb.checked = selectAllCheckbox.checked);
            updateDropdownText(filterId);
        }}

        function updateDropdownText(filterId) {{
            const content = document.getElementById(filterId + 'Content');
            const textElement = document.getElementById(filterId + 'Text');
            const checkboxes = content.querySelectorAll('.dropdown-option:not(.select-all) input');
            const checkedBoxes = content.querySelectorAll('.dropdown-option:not(.select-all) input:checked');
            
            if (checkedBoxes.length === 0) {{
                textElement.textContent = 'Todos';
            }} else if (checkedBoxes.length === 1) {{
                textElement.textContent = checkedBoxes[0].nextElementSibling.textContent;
            }} else if (checkedBoxes.length === checkboxes.length) {{
                textElement.textContent = 'Todos';
            }} else {{
                textElement.textContent = checkedBoxes.length + ' selecionados';
            }}
        }}

        function getSelectedFilters() {{
            const selectedFilters = {{}};
            FILTERS.forEach(f => {{
                const content = document.getElementById(f.name + 'Content');
                if (content) {{
                    const checkedBoxes = content.querySelectorAll('.dropdown-option:not(.select-all) input:checked');
                    selectedFilters[f.name] = Array.from(checkedBoxes).map(cb => cb.value);
                }}
            }});
            return selectedFilters;
        }}

        function applyFilters() {{
            const _scrollY = window.scrollY;
            document.querySelectorAll('.dropdown-content').forEach(d => d.classList.remove('show'));
            document.querySelectorAll('.dropdown-button').forEach(b => b.classList.remove('open'));
            const cleanup = renderAll();
            updateFilterCounts();
            // Duplo rAF: 1º garante repaint, 2º aplica scroll após layout estabilizado
            requestAnimationFrame(() => requestAnimationFrame(() => {{
                window.scrollTo({{ top: _scrollY, behavior: 'instant' }});
                cleanup();
            }}));
        }}

        // ===== DRILLDOWN (clique em barras) =====
        // Regra: AND entre variáveis; OR dentro da variável (multi-seleção)
        const DRILLDOWN = new Map(); // varName -> Set(rawValue)

        function _norm(v) {{
            return String(v).trim();
        }}

        function getDrilldownSet(varName) {{
            if (!DRILLDOWN.has(varName)) DRILLDOWN.set(varName, new Set());
            return DRILLDOWN.get(varName);
        }}

        function toggleDrilldown(varName, rawValue) {{
            const s = getDrilldownSet(varName);
            const k = _norm(rawValue);
            if (s.has(k)) s.delete(k);
            else s.add(k);
            if (s.size === 0) DRILLDOWN.delete(varName);

            const anchor = document.getElementById('section-' + varName);
            const topBefore = anchor ? anchor.getBoundingClientRect().top : 0;

            const cleanup = renderAll();

            requestAnimationFrame(() => {{
                const anchorAfter = document.getElementById('section-' + varName);
                if (anchorAfter) {{
                    const topAfter = anchorAfter.getBoundingClientRect().top;
                    window.scrollBy({{ top: topAfter - topBefore, behavior: 'instant' }});
                }}
                cleanup();
            }});
        }}

        function clearAllDrilldowns() {{
            DRILLDOWN.clear();
        }}

        function clearFilters() {{
            document.querySelectorAll('.dropdown-content input[type="checkbox"]').forEach(cb => cb.checked = false);
            FILTERS.forEach(f => {{
                const textElement = document.getElementById(f.name + 'Text');
                if (textElement) textElement.textContent = 'Todos';
            }});
            // Limpa também o drilldown por clique nos gráficos
            clearAllDrilldowns();
            document.querySelectorAll('.dropdown-content').forEach(d => d.classList.remove('show'));
            document.querySelectorAll('.dropdown-button').forEach(b => b.classList.remove('open'));
            renderAll();
            updateFilterCounts();
        }}

        function getFilteredRecords(excludeVarName = null) {{
            const selectedFilters = getSelectedFilters();
            return RECORDS.filter(record => {{
                // 1) filtros do topo
                const passTop = Object.keys(selectedFilters).every(filterName => {{
                    const filterValues = selectedFilters[filterName];
                    if (filterValues.length === 0) return true;
                    const recordValue = record[filterName];
                    if (recordValue === null || recordValue === undefined) return false;

                    const normalizedRecordValue = _norm(recordValue);
                    const normalizedFilterValues = filterValues.map(v => _norm(v));
                    return normalizedFilterValues.includes(normalizedRecordValue);
                }});

                if (!passTop) return false;

                // 2) drilldowns por clique em gráficos (AND entre variáveis; OR dentro da variável)
                for (const [varName, set] of DRILLDOWN.entries()) {{
                    if (excludeVarName && varName === excludeVarName) continue;
                    if (!set || set.size === 0) continue;

                    const v = record[varName];
                    if (v === null || v === undefined) return false;

                    if (Array.isArray(v)) {{
                        let ok = false;
                        for (const item of v) {{
                            if (item === null || item === undefined) continue;
                            if (set.has(_norm(item))) {{ ok = true; break; }}
                        }}
                        if (!ok) return false;
                    }} else {{
                        if (!set.has(_norm(v))) return false;
                    }}
                }}

                return true;
            }});
        }}

        // RENDERIZAÇÃO
        function renderAll() {{
            const content = document.getElementById('content');
            // Desabilitar scroll anchoring + travar altura
            const prevAnchor = document.body.style.overflowAnchor;
            document.body.style.overflowAnchor = 'none';
            content.style.overflowAnchor = 'none';
            content.style.minHeight = content.offsetHeight + 'px';
            content.innerHTML = '';

            let pCounter = 0;
            VARS_META.forEach((varMeta) => {{
                const recordsForVar = getFilteredRecords(varMeta.name);
                const section = createSection(varMeta, recordsForVar);
                if (section) {{
                    pCounter++;
                    varMeta._rendered_p_num = pCounter;
                    const titleEl = section.querySelector('.section-title');
                    if (titleEl) {{
                        const titleColor = titleEl.style.color || 'var(--primary-dark)';
                        const pSpan = document.createElement('span');
                        pSpan.className = 'p-num';
                        pSpan.style.color = titleColor;
                        pSpan.textContent = 'P' + pCounter + '. ';
                        titleEl.prepend(pSpan);
                    }}
                    content.appendChild(section);
                }}
            }});
            // Retorna cleanup — o chamador deve invocar após restaurar o scroll
            const _cleanup = () => {{ content.style.minHeight = ''; content.style.overflowAnchor = ''; document.body.style.overflowAnchor = prevAnchor; }};
            // Se modo apresentação ativo, re-renderizar slide atual
            if (window.__presRenderCurrent) requestAnimationFrame(window.__presRenderCurrent);
            return _cleanup;
        }}


        function renderStringVariable(varMeta, records) {{
            const container = document.createElement('div');
            container.style.cssText = 'padding-top:16px;--bar-color:#aaa;';

            // Normaliza texto: tira espaços, ignora códigos de não-resposta e aplica capitalização simples
            const SKIP_CODES = new Set(['0','00','000','9','90','99','099','999','9999',
                                        'ns','nr','na','n/a','.','-','x','xx','xxx']);
            function normalizeText(text) {{
                if (text === null || text === undefined) return '';
                let t = String(text).trim().replace(/[.,;!?-]+$/, '').trim(); // remove pontuação final
                if (!t || SKIP_CODES.has(t.toLowerCase())) return '';
                return t.charAt(0).toUpperCase() + t.slice(1);
            }}

            // Coleta e normaliza as respostas
            let validResponses = records
                .map(r => normalizeText(r[varMeta.name]))
                .filter(v => v !== '');

            if (validResponses.length === 0) {{
                return null;
            }}

            // ✅ DEBUG: Verificar ordem das respostas
            console.log(`📝 ${{varMeta.name}}: Respostas de texto encontradas:`, validResponses.slice(0, 5));

            // ✅ REGRA CORRETA: Textual = Ordem alfabética
            validResponses.sort((a, b) => a.localeCompare(b, 'pt-BR'));

            // --------- BLOCO VISUAL (lista normal como antes) ----------
            const totalResponses = validResponses.length;
            const summary = document.createElement('p');
            summary.innerHTML = '<span style="font-size:13px;font-weight:600;color:#444">Total de respostas:</span> <span style="font-size:13px;color:#444">' + totalResponses + '</span>';
            summary.style.cssText = 'margin-bottom:12px;';

            const responseList = document.createElement('div');
            responseList.style.cssText =
                'max-height: 400px; overflow-y: auto; overflow-x: hidden; ' +
                'border: 0.5px solid var(--border); ' +
                'border-radius: var(--radius);';

            validResponses.forEach((response, index) => {{
                const responseItem = document.createElement('div');
                responseItem.style.cssText =
                    'padding: 10px 16px; border-bottom: 0.5px solid var(--border); ' +
                    'background: white; font-size: 13px; color: #444;';
                responseItem.innerHTML =
                    '<strong style="color:#333">' + (index + 1) + '.</strong> ' + String(response);
                responseList.appendChild(responseItem);
            }});

            // --------- TABELA OCULTA PARA EXPORTAÇÃO (USADA PELO EXCEL) ----------
            const exportTable = document.createElement('table');
            exportTable.className = 'export-text-table';
            exportTable.style.display = 'none'; // invisível para o usuário

            const thead = document.createElement('thead');
            const headRow = document.createElement('tr');
            ['Nº', 'Resposta'].forEach(h => {{
                const th = document.createElement('th');
                th.innerText = h;
                headRow.appendChild(th);
            }});
            thead.appendChild(headRow);

            const tbody = document.createElement('tbody');
            validResponses.forEach((resp, idx) => {{
                const tr = document.createElement('tr');

                const tdIndex = document.createElement('td');
                tdIndex.innerText = (idx + 1).toString();
                tr.appendChild(tdIndex);

                const tdResp = document.createElement('td');
                tdResp.innerText = resp;
                tr.appendChild(tdResp);

                tbody.appendChild(tr);
            }});

            exportTable.appendChild(thead);
            exportTable.appendChild(tbody);

            // Conteúdo principal: adiciona sumário e lista de respostas
            container.appendChild(summary);

            // --------- PALAVRAS‑CHAVE E FILTRO ---------
            const keywords = varMeta.keywords || [];
            if (keywords && keywords.length > 0) {{
                const filterContainer = document.createElement('div');
                filterContainer.style.cssText = 'margin-bottom: 8px; display: flex; flex-wrap: wrap; gap: 4px; align-items: center;';
                const filterTitle = document.createElement('span');
                filterTitle.style.fontSize = '13px';
                filterTitle.style.fontWeight = '600';
                filterTitle.textContent = 'Palavras‑chave:';
                filterContainer.appendChild(filterTitle);

                // Função auxiliar para normalizar texto para comparação: converte para minúsculas
                // e remove acentos.  Isso garante que "informacoes" corresponda a "informação".
                function normalizeForComparison(str) {{
                    if (!str) return '';
                    return String(str).toLowerCase()
                        .normalize('NFD')
                        .replace(/[\u0300-\u036f]/g, '');
                }}

                // ---------------- MULTI-SELEÇÃO (OR) ----------------
                const selectedRoots = new Set();

                // Contagens dinâmicas das keywords com base nos registros JÁ filtrados no dashboard
                const _kwCounts = new Map();
                try {{
                    const _normResponses = validResponses.map(r => normalizeForComparison(r));
                    keywords.forEach(k => {{
                        const rr = normalizeForComparison(k.root);
                        if (!rr) return;
                        let c = 0;
                        for (let i = 0; i < _normResponses.length; i++) {{
                            if (_normResponses[i].includes(rr)) c++;
                        }}
                        _kwCounts.set(k.root, c);
                    }});
                }} catch(e) {{
                    console.warn('kw count erro', e);
                }}

                // Aplica o filtro com base nas roots selecionadas (OR).
                // Se nada selecionado, mostra tudo.
                function applyKeywordMultiFilter() {{
                    const items = responseList.children;

                    if (selectedRoots.size === 0) {{
                        for (let i = 0; i < items.length; i++) items[i].style.display = '';
                        summary.innerHTML = '<span style="font-size:13px;font-weight:600;color:#444">Total de respostas:</span> <span style="font-size:13px;color:#444">' + totalResponses + '</span>';
                        return;
                    }}

                    for (let i = 0; i < items.length; i++) {{
                        const item = items[i];
                        const text = item.textContent || '';
                        const normText = normalizeForComparison(text);

                        let match = false;
                        selectedRoots.forEach(root => {{
                            if (normText.includes(root)) match = true;
                        }});

                        item.style.display = match ? '' : 'none';
                    }}

                    // Atualiza contador conforme itens visíveis
                    let visible = 0;
                    for (let i = 0; i < items.length; i++) {{
                        if (items[i].style.display !== 'none') visible++;
                    }}
                    summary.innerHTML = '<span style="font-size:13px;font-weight:600;color:#444">Total de respostas:</span> <span style="font-size:13px;color:#444">' + visible + ' (filtradas de ' + totalResponses + ')</span>';
                }}

                // Estado inicial: sem keywords selecionadas
                applyKeywordMultiFilter();

                // Toggle do chip (seleciona/deseleciona)
                function toggleKeywordRoot(root, el) {{
                    const normRoot = normalizeForComparison(root);
                    if (selectedRoots.has(normRoot)) {{
                        selectedRoots.delete(normRoot);
                        el.classList.remove('active');
                    }} else {{
                        selectedRoots.add(normRoot);
                        el.classList.add('active');
                    }}
                    applyKeywordMultiFilter();
                }}

                // Monta botões de keyword
                keywords.forEach(k => {{
                    const _c = _kwCounts.get(k.root) ?? k.count ?? 0;
                    if (_c <= 0) return;
                    const kwBtn = document.createElement('span');
                    kwBtn.className = 'outros-kw';
                    kwBtn.textContent = k.word + ' (' + _c + ')';
                    kwBtn.title = "Filtrar por " + k.word;
                    kwBtn.dataset.root = k.root;
                    kwBtn.onclick = () => toggleKeywordRoot(k.root, kwBtn);
                    filterContainer.appendChild(kwBtn);
                }});

                container.appendChild(filterContainer);
            }}

            container.appendChild(responseList);

            // adiciona a tabela escondida ao container
            container.appendChild(exportTable);


            return container;
        }}

        function renderNumericScaleVariable(varMeta, records) {{
            const container = document.createElement('div');
            container.style.cssText = 'padding-top:16px;--bar-color:#aaa;';

            // Coletar valores com seus pesos para histograma ponderado
            const weightedValues = [];
            records.forEach(r => {{
                const value = r[varMeta.name];
                const weight = r.__weight__ || 1.0;
                if (value !== null && value !== undefined && !isNaN(value)) {{
                    weightedValues.push({{value: Number(value), weight: weight}});
                }}
            }});

            if (weightedValues.length === 0) {{
                return null;
            }}

            const stats = varMeta.stats || {{}};

            // Calcular estatísticas dos records FILTRADOS (não usar varMeta.stats que é estático)
            const n      = weightedValues.length;
            const totalW = weightedValues.reduce((s, wv) => s + wv.weight, 0);
            const mean   = totalW > 0
                ? weightedValues.reduce((s, wv) => s + wv.value * wv.weight, 0) / totalW
                : 0;

            // Desvio padrão ponderado
            const variance = totalW > 1
                ? weightedValues.reduce((s, wv) => s + wv.weight * Math.pow(wv.value - mean, 2), 0) / (totalW - 1)
                : 0;
            const stddev = Math.sqrt(variance);

            // Mediana ponderada (ordenar e acumular peso até 50%)
            const sorted = [...weightedValues].sort((a, b) => a.value - b.value);
            let cumW = 0, median = sorted[0].value;
            for (const wv of sorted) {{
                cumW += wv.weight;
                if (cumW >= totalW / 2) {{ median = wv.value; break; }}
            }}
            const minValStat = sorted[0].value;
            const maxValStat = sorted[sorted.length - 1].value;

            // Cards de métricas inline
            const statsBar = document.createElement('div');
            statsBar.style.cssText = 'display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px;';

            function addStatCard(label, value, decimals=2) {{
                if (value === undefined || value === null) return;
                const card = document.createElement('div');
                card.style.cssText =
                    'background:#f8f9fa;border:0.5px solid var(--border);border-radius:6px;' +
                    'padding:6px 12px;display:flex;flex-direction:column;align-items:center;min-width:70px;';
                const lbl = document.createElement('span');
                lbl.style.cssText = 'font-size:10px;font-weight:600;color:#aaa;text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px;';
                lbl.textContent = label;
                const val = document.createElement('span');
                val.style.cssText = 'font-size:14px;font-weight:600;color:var(--primary-dark);';
                val.textContent = typeof value === 'number'
                    ? value.toLocaleString('pt-BR', {{minimumFractionDigits:0, maximumFractionDigits:decimals}})
                    : value;
                card.appendChild(lbl);
                card.appendChild(val);
                statsBar.appendChild(card);
            }}

            addStatCard('N',       n,          0);
            addStatCard('Média',   mean,       2);
            addStatCard('Mediana', median,     2);
            addStatCard('DP',      stddev,     2);
            addStatCard('Mín',     minValStat, 2);
            addStatCard('Máx',     maxValStat, 2);

            const chartContainer = document.createElement('div');
            chartContainer.className = 'chart-container';

            const canvas = document.createElement('canvas');
            chartContainer.appendChild(canvas);
            const ctx = canvas.getContext('2d');

            // Extrair apenas os valores para calcular min/max
            const values = weightedValues.map(wv => wv.value);
            const minVal = Math.min(...values);
            const maxVal = Math.max(...values);
            const binCount = 10;
            const range = maxVal - minVal || 1;
            const binSize = range / binCount;

            const bins = new Array(binCount).fill(0);
            const labels = [];

            for (let i = 0; i < binCount; i++) {{
                const start = minVal + i * binSize;
                const end = (i === binCount - 1) ? maxVal : (start + binSize);
                labels.push(`${{formatBR(start, 1)}} – ${{formatBR(end, 1)}}`);
            }}

            // Distribuir valores ponderados nos bins
            weightedValues.forEach(wv => {{
                let idx = Math.floor((wv.value - minVal) / binSize);
                if (idx < 0) idx = 0;
                if (idx >= binCount) idx = binCount - 1;
                bins[idx] += wv.weight;  // Usar peso em vez de 1
            }});

            const totalCases = weightedValues.reduce((sum, wv) => sum + wv.weight, 0);
            const percentages = bins.map(count => totalCases > 0 ? (count / totalCases * 100) : 0);
            
            // ✅ AJUSTE DINÂMICO: Eixo Y se adapta ao valor máximo
            const maxPercentage = Math.max(...percentages);
            const yAxisMax = maxPercentage > 0
                ? Math.min(100, Math.ceil(maxPercentage / 10) * 10)
                : 100;

            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: labels.map(label => wrapLabel(label, CHART_LABEL_MAX)),
                    datasets: [{{
                        data: percentages,
                        backgroundColor: 'rgba(123, 175, 192, 0.7)',
                        borderColor: 'rgba(123, 175, 192, 1)',
                        borderWidth: 1
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{ display: false }},
                        tooltip: {{
                            callbacks: {{
                                label: function(context) {{
                                    const index = context.dataIndex;
                                    const count = Math.round(bins[index]);  // Arredondar para inteiro
                                    const pct = context.parsed.y;
                                    return `${{formatBR(pct, 1)}}% (${{count}} casos)`;
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        y: {{
                            beginAtZero: true,
                            max: yAxisMax,
                            ticks: {{
                                callback: function(value) {{
                                    return value + '%';
                                }}
                            }}
                        }},
                        x: {{
                            // Para histograma: barras sem espaço entre elas
                            categoryPercentage: 1.0,
                            barPercentage: 1.0
                        }}
                    }}
                }}
            }});

            container.appendChild(statsBar);
            container.appendChild(chartContainer);

            return container;
        }}

        function renderDateVariable(varMeta, records) {{
            const container = document.createElement('div');
            container.style.cssText = 'padding-top:16px;--bar-color:#aaa;';

            const freq = {{}};
            let validCount = 0;

            records.forEach(r => {{
                const v = r[varMeta.name];
                const weight = r.__weight__ || 1.0;  // Peso do registro
                if (v !== null && v !== undefined && String(v).trim() !== '') {{
                    validCount += weight;
                    const key = String(v);
                    freq[key] = (freq[key] || 0) + weight;
                }}
            }});

            const entries = Object.entries(freq);
            if (entries.length === 0) {{
                return null;
            }}

            // ✅ DEBUG: Verificar ordem das categorias
            console.log(`📊 ${{varMeta.name}}: Categorias encontradas:`, entries.map(([label]) => label));

            // ✅ REGRA CORRETA: Datas ordenadas cronologicamente
            entries.sort((a, b) => new Date(a[0]) - new Date(b[0]));

            const labels = entries.map(([d]) => d);
            const counts = entries.map(([, c]) => c);
            const percentages = counts.map(count => validCount > 0 ? (count / validCount * 100) : 0);
            
            // ✅ AJUSTE DINÂMICO: Eixo Y se adapta ao valor máximo
            const maxPercentage = Math.max(...percentages);
            const yAxisMax = maxPercentage > 0
                ? Math.min(100, Math.ceil(maxPercentage / 10) * 10)
                : 100;

            const chartContainer = document.createElement('div');
            chartContainer.className = 'chart-container';
            const canvas = document.createElement('canvas');
            chartContainer.appendChild(canvas);
            const ctx = canvas.getContext('2d');

            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: labels.map(label => wrapLabel(label, CHART_LABEL_MAX)),
                    datasets: [{{
                        data: percentages,
                        backgroundColor: 'rgba(76, 175, 80, 0.7)',
                        borderColor: 'rgba(76, 175, 80, 1)',
                        borderWidth: 1
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{ display: false }},
                        tooltip: {{
                            callbacks: {{
                                label: function(context) {{
                                    const index = context.dataIndex;
                                    const qty = Math.round(counts[index]);  // Arredondar para inteiro
                                    const pct = context.parsed.y;
                                    return `${{formatBR(pct, 1)}}% (${{qty}} casos)`;
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        y: {{
                            beginAtZero: true,
                            max: yAxisMax,
                            ticks: {{
                                callback: function(value) {{
                                    return value + '%';
                                }}
                            }}
                        }},
                        x: {{
                            ticks: {{
                                maxRotation: 45,
                                minRotation: 0
                            }}
                        }}
                    }}
                }}
            }});

            const summary = document.createElement('p');
            summary.innerHTML = '<strong>Resumo:</strong> ' +
                entries.length + ' datas distintas';
            summary.style.marginTop = '15px';

            container.appendChild(chartContainer);
            container.appendChild(summary);

            return container;
        }}

        function renderCategoricalVariable(varMeta, records) {{
            const container = document.createElement('div');
            const freq = {{}};
            let validCount = 0;      // respondentes com pelo menos 1 resposta válida
            let totalSelections = 0; // soma de todas as opções marcadas (só MR)

            // Conta frequências
            records.forEach(r => {{
                let v = r[varMeta.name];
                const weight = r.__weight__ || 1.0;
                
                if (Array.isArray(v)) {{
                    // MR: contar o respondente UMA VEZ (base = respondentes)
                    let hasValid = false;
                    v.forEach(item => {{
                        if (item !== null && item !== undefined && String(item).trim() !== '') {{
                            const key = String(item).trim();
                            freq[key] = (freq[key] || 0) + weight;
                            totalSelections += weight;
                            hasValid = true;
                        }}
                    }});
                    if (hasValid) validCount += weight; // incrementa uma vez por respondente
                }} else {{
                    // Categórica simples: continua igual
                    if (v !== null && v !== undefined && String(v).trim() !== '') {{
                        const key = String(v).trim();
                        freq[key] = (freq[key] || 0) + weight;
                        validCount += weight;
                    }}
                }}
            }});

            // Filtrar -1 (missing) e, para escalas ou quando NS excluído, o label NS
            const nsLabelScale = varMeta.scale_info && varMeta.scale_info.ns_label;
            // NS words para detecção automática em qualquer categórica
            const nsWords = ['não sabe', 'nao sabe', 'não sei', 'nao sei', 'ns/a', 'ns/nr', 'não tenho'];
            const detectNsLabel = (label) => nsWords.some(w => String(label).toLowerCase().includes(w));
            // Excluir NS se: (a) var de escala com ns_label, ou (b) toggle NS desativado
            const nsIncluded = varMeta._nsIncluded !== false; // padrão true
            const entries = Object.entries(freq).filter(([key]) => {{
                if (key === '-1' || key === '-1.0') return false;
                // NS de escala: excluir só quando toggle OFF; senão mostrar em cinza
                if (!nsIncluded && nsLabelScale && key === nsLabelScale) return false;
                if (!nsIncluded && !nsLabelScale && detectNsLabel(key)) return false;
                return true;
            }});

            // NS do denominador: só excluir quando toggle OFF
            if (!nsIncluded) {{
                if (nsLabelScale && freq[nsLabelScale]) validCount -= freq[nsLabelScale];
                else Object.keys(freq).forEach(key => {{ if (detectNsLabel(key)) validCount -= freq[key]; }});
            }}
            if (entries.length === 0) {{
                return null;
            }}

            // ✅ DEBUG: Verificar ordem das categorias
            console.log(`📊 ${{varMeta.name}}: Categorias encontradas:`, entries.map(([label]) => label));

            // ✅ REGRAS CORRETAS DE ORDENAÇÃO baseadas no tipo da variável
            const varType = varMeta.var_type || varMeta.type || 'single';
            const measure = varMeta.measure || 'nominal';
            
            if (varType === 'multiple_response' || varMeta.type === 'mr') {{
                // 🔗 MR NOMINAL: Da maior frequência para a menor
                entries.sort((a, b) => b[1] - a[1]);
                console.log(`🔗 ${{varMeta.name}}: MR ordenado por frequência (maior→menor)`);
                
            }} else if (measure === 'ordinal') {{
                // Ordenar conforme a ordem SPSS (valueOrder contém labels como strings)
                const valueOrder = VARS_VALUE_ORDER[varMeta.name] || [];

                entries.sort((a, b) => {{
                    // Sempre comparar como string — valueOrder contém strings
                    const ia = valueOrder.indexOf(String(a[0]));
                    const ib = valueOrder.indexOf(String(b[0]));
                    // Categorias não encontradas vão para o fim
                    const orderA = ia === -1 ? 9999 : ia;
                    const orderB = ib === -1 ? 9999 : ib;
                    return orderA - orderB;
                }});
                
            }} else {{
                // 📊 SINGLE NOMINAL: Da maior frequência para a menor
                entries.sort((a, b) => b[1] - a[1]);
                console.log(`📊 ${{varMeta.name}}: Nominal ordenado por frequência (maior→menor)`);
            }}

            // Aplicar labels descritivos APÓS ordenação
            const labels = entries.map(([label]) => {{
                if (CODE_TO_LABEL[varMeta.name] && CODE_TO_LABEL[varMeta.name][label]) {{
                    const descLabel = CODE_TO_LABEL[varMeta.name][label];
                    console.log(`✅ ${{varMeta.name}}: "${{label}}" → "${{descLabel}}"`);
                    return descLabel;
                }}
                return label;
            }});
            const counts = entries.map(([,count]) => count);
            const percentages = counts.map(count => validCount > 0 ? (count / validCount * 100) : 0);
            
            // ===== IVV-STYLE INLINE BAR TABLE =====
            const isMR = (varType === 'multiple_response' || varMeta.type === 'mr');
            const sel = DRILLDOWN.get(varMeta.name) || new Set();

            // Maior percentual para escalar as barras
            const maxPct = Math.max(...percentages, 1);

            const table = document.createElement('table');
            table.className = 'ivv-table';

            // Cabeçalho
            const thead = document.createElement('thead');
            const headerRow = document.createElement('tr');
            if (isMR) {{
                headerRow.innerHTML = '<th class="ivv-label-col">Respostas</th><th>Frequência</th><th class="ivv-pct-col">% / Respondentes</th><th class="ivv-base-col">Respondentes</th>';
            }} else {{
                headerRow.innerHTML = '<th class="ivv-label-col">Respostas</th><th>Frequência</th><th class="ivv-pct-col">%</th>';
            }}            thead.appendChild(headerRow);
            table.appendChild(thead);

            const tbody = document.createElement('tbody');

            entries.forEach(([rawLabel, count], idx) => {{
                const rawKey = _norm(rawLabel);
                const isActive = sel.has(rawKey);
                const pct = validCount > 0 ? count / validCount * 100 : 0;
                const pctFmt = formatBR(pct, 1) + '%';
                const barWidth = (pct / maxPct * 100).toFixed(1);

                let displayLabel = rawLabel;
                // Ordinal vars já armazenam labels — CODE_TO_LABEL causaria duplicatas
                if (varMeta.measure !== 'ordinal' && CODE_TO_LABEL[varMeta.name] && CODE_TO_LABEL[varMeta.name][rawLabel]) {{
                    displayLabel = CODE_TO_LABEL[varMeta.name][rawLabel];
                }}

                const row = document.createElement('tr');
                const isNsRow = !!varMeta.scale_info && detectNsLabel(rawLabel);
                row.className = 'ivv-row ivv-data-row' + (isActive ? ' drilldown-active' : '') + (isNsRow ? ' ivv-ns-row' : '');
                row.dataset.label = rawLabel;
                row.dataset.label = rawLabel;
                row.title = 'Clique para filtrar por esta categoria';

                // Célula de frequência com barra embutida
                // Barra na célula de percentual, valor absoluto na frequência simples
                const barCellHTML = `
                    <td>${{Math.round(count)}}</td>
                    <td class="ivv-bar-cell">
                        <div class="ivv-bar-bg" style="width:${{barWidth}}%"></div>
                        <div class="ivv-bar-val ivv-pct-col">${{pctFmt}}</div>
                    </td>`;

                if (isMR) {{
                    row.innerHTML = `<td>${{displayLabel}}</td>${{barCellHTML}}<td class="ivv-base-col">${{Math.round(validCount)}}</td>`;
                }} else {{
                    row.innerHTML = `<td>${{displayLabel}}</td>${{barCellHTML}}`;
                }}

                // Drilldown por clique na linha
                row.addEventListener('click', () => toggleDrilldown(varMeta.name, rawKey));

                tbody.appendChild(row);
            }});

            // Linha de total
            const totalRow = document.createElement('tr');
            totalRow.className = 'ivv-total-row';
            const totalCount = Math.round(entries.reduce((sum, [, c]) => sum + c, 0));
            if (isMR) {{
                totalRow.innerHTML = `<td>Total de respostas</td><td>${{Math.round(totalSelections)}}</td><td class="ivv-bar-cell"><div class="ivv-bar-val ivv-pct-col"></div></td><td class="ivv-base-col"></td>`;
            }} else {{
                totalRow.innerHTML = `<td>Total</td><td>${{totalCount}}</td><td class="ivv-bar-cell"><div class="ivv-bar-val ivv-pct-col">100,0%</div></td>`;
            }}
            tbody.appendChild(totalRow);

            table.appendChild(tbody);

            const tableWrap = document.createElement('div');
            tableWrap.style.cssText = 'border: 0.5px solid var(--border); border-radius: var(--radius); overflow: hidden;';
            tableWrap.appendChild(table);

            // ===== SCALE METRICS (NPS / SAT / CSAT / CES) =====
            if (varMeta.scale_info && !isMR) {{
                const si  = varMeta.scale_info;
                const ltv = si.label_to_value || {{}};

                // Acumular frequências por valor semântico real
                const valueFreq = {{}};
                let scaleTotal  = 0;
                Object.entries(freq).forEach(([label, count]) => {{
                    if (si.ns_label && label === si.ns_label && !nsIncluded) return;
                    const rv = ltv[label];
                    if (rv === undefined) return;
                    valueFreq[rv] = (valueFreq[rv] || 0) + count;
                    scaleTotal   += count;
                }});

                if (scaleTotal > 0) {{
                    // Colorir linhas da tabela para NPS
                    if (si.type === 'NPS') {{
                        tbody.querySelectorAll('tr[data-label]').forEach(tr => {{
                            const rv = ltv[tr.dataset.label];
                            if (rv === undefined) return;
                            if (rv <= 6) {{
                                tr.style.setProperty('--bar-color', 'var(--red)');
                                tr.style.setProperty('--bar-text',  'var(--red)');
                            }} else if (rv <= 8) {{
                                tr.style.setProperty('--bar-color', '#bbb');
                                tr.style.setProperty('--bar-text',  '#888');
                            }} else {{
                                tr.style.setProperty('--bar-color', 'var(--green)');
                                tr.style.setProperty('--bar-text',  'var(--green-dark)');
                            }}
                        }});
                    }}

                    const metricsDiv = document.createElement('div');
                    metricsDiv.className = 'scale-metrics';

                    function addMetricRow(labelText, valueText, cssClass) {{
                        const mrow  = document.createElement('div');
                        mrow.className = 'scale-metric-row' + (cssClass ? ' ' + cssClass : '');
                        const mlbl  = document.createElement('div');
                        mlbl.className = 'scale-metric-label';
                        mlbl.textContent = labelText;
                        const mval  = document.createElement('div');
                        mval.className = 'scale-metric-value';
                        mval.textContent = valueText;
                        mrow.appendChild(mlbl);
                        mrow.appendChild(mval);
                        metricsDiv.appendChild(mrow);
                    }}

                    function fmtPct(n) {{
                        return n.toLocaleString('pt-BR', {{minimumFractionDigits:1, maximumFractionDigits:1}}) + '%';
                    }}
                    function fmtMean(n) {{
                        return n.toLocaleString('pt-BR', {{minimumFractionDigits:1, maximumFractionDigits:1}});
                    }}

                    if (si.type === 'NPS') {{
                        const det = [0,1,2,3,4,5,6].reduce((s,v) => s + (valueFreq[v]||0), 0);
                        const neu = [7,8].reduce((s,v) => s + (valueFreq[v]||0), 0);
                        const pro = [9,10].reduce((s,v) => s + (valueFreq[v]||0), 0);
                        const detPct = det / validCount * 100;
                        const neuPct = neu / validCount * 100;
                        const proPct = pro / validCount * 100;
                        const score  = Math.round(proPct - detPct);
                        addMetricRow('Detratores (0 a 6): ' + Math.round(det), fmtPct(detPct), 'nps-det');
                        addMetricRow('Neutros (7 e 8): '     + Math.round(neu), fmtPct(neuPct), 'nps-neu');
                        addMetricRow('Promotores (9 e 10): ' + Math.round(pro), fmtPct(proPct), 'nps-pro');

                        // Score NPS: verde se positivo, cinza se zero, vinho se negativo
                        const scoreClass = score > 0 ? 'nps-score nps-score-pos'
                                         : score < 0 ? 'nps-score nps-score-neg'
                                         : 'nps-score nps-score-zero';
                        addMetricRow('NPS = % Promotores − % Detratores', (score >= 0 ? '+' : '') + score, scoreClass);

                    }} else if (si.type === 'SAT') {{
                        const t3bCount = [8,9,10].reduce((s,v) => s + (valueFreq[v]||0), 0);
                        const t3bPct   = t3bCount / validCount * 100;
                        const mean     = Object.entries(valueFreq).reduce((s,[v,c]) => s + Number(v)*c, 0) / scaleTotal;
                        addMetricRow('T3B — Notas 8, 9 e 10', fmtPct(t3bPct), 'highlight');
                        addMetricRow('Média', fmtMean(mean), '');

                    }} else {{
                        // CSAT / CES: Top 2 Box = 2 maiores valores semânticos
                        const allVals = Array.from(new Set(Object.keys(valueFreq).map(Number))).sort((a,b)=>a-b);
                        const top2    = allVals.slice(-2);
                        const top2n   = top2.reduce((s,v) => s + (valueFreq[v]||0), 0);
                        const top2pct = top2n / validCount * 100;
                        const mean    = Object.entries(valueFreq).reduce((s,[v,c]) => s + Number(v)*c, 0) / scaleTotal;
                        const t2lbl   = si.type === 'CES'
                            ? 'Top 2 Box — Fácil / Muito fácil'
                            : 'Top 2 Box — Satisfeito / Muito satisfeito';
                        addMetricRow(t2lbl, fmtPct(top2pct), 'highlight');
                        addMetricRow('Média', fmtMean(mean), '');
                    }}

                    tableWrap.appendChild(metricsDiv);
                }}
            }}
            const hint = document.createElement('p');
            hint.className = 'ivv-drilldown-hint';
            hint.textContent = 'Clique em uma linha para filtrar os demais gráficos por esta categoria.';

            container.appendChild(tableWrap);
            container.appendChild(hint);

            // ── SUB-BOX DE RESPOSTAS ABERTAS (_S absorvida) ──
            if (varMeta.open_text_var) {{
                const openKey = varMeta.name + '__open_text__';
                const openTexts = records
                    .map(r => r[openKey])
                    .filter(t => t && String(t).trim());

                if (openTexts.length > 0) {{
                    // Extrair keywords simples (palavras com ≥4 chars, top 5)
                    const wordFreq = {{}};
                    openTexts.forEach(t => {{
                        String(t).toLowerCase().split(/ +/).forEach(w => {{
                            const wc = w.replace(/[^a-záéíóúâêîôûãõàèìòùç]/gi,'');
                            if (wc.length >= 4) wordFreq[wc] = (wordFreq[wc]||0)+1;
                        }});
                    }});
                    const topKw = Object.entries(wordFreq)
                        .sort((a,b)=>b[1]-a[1]).slice(0,5)
                        .filter(([,n])=>n>=2);

                    const sub = document.createElement('div');
                    sub.className = 'outros-sub';

                    const subHeader = document.createElement('div');
                    subHeader.className = 'outros-sub-header';
                    subHeader.innerHTML = `<span class="outros-sub-label">Respostas abertas para "Outro"</span><span class="outros-sub-n">${{openTexts.length}} respostas</span>`;

                    // Keyword chips
                    let activeKw = null;
                    const kwWrap = document.createElement('div');
                    kwWrap.style.cssText = 'display:flex;gap:4px;flex-wrap:wrap;margin-left:auto;';
                    topKw.forEach(([word, n]) => {{
                        const chip = document.createElement('span');
                        chip.className = 'outros-kw';
                        chip.textContent = word + ' (' + n + ')';
                        chip.onclick = () => {{
                            if (activeKw === word) {{
                                activeKw = null;
                                chip.classList.remove('active');
                            }} else {{
                                kwWrap.querySelectorAll('.outros-kw').forEach(c=>c.classList.remove('active'));
                                activeKw = word;
                                chip.classList.add('active');
                            }}
                            renderList();
                        }};
                        kwWrap.appendChild(chip);
                    }});
                    subHeader.appendChild(kwWrap);
                    sub.appendChild(subHeader);

                    const list = document.createElement('div');
                    list.className = 'outros-list';

                    function renderList() {{
                        list.innerHTML = '';
                        const filtered = activeKw
                            ? openTexts.filter(t => t.toLowerCase().includes(activeKw))
                            : openTexts;
                        filtered.forEach((t, idx) => {{
                            const item = document.createElement('div');
                            item.className = 'outros-item';
                            item.textContent = (idx+1) + '. ' + t;
                            list.appendChild(item);
                        }});
                    }}
                    renderList();

                    sub.appendChild(list);
                    container.appendChild(sub);
                }}
            }}

            return container;
        }}

        function createSection(varMeta, records) {{
            const section = document.createElement('div');
            section.className = 'section';
            
            // ===== ADICIONAR IDENTIFICADORES PARA EXPORTAÇÃO =====
            // Facilita a identificação da variável na exportação Excel/PDF
            if (varMeta.name) {{
                section.setAttribute('data-var', varMeta.name);
                section.setAttribute('data-variable', varMeta.name);
                section.id = `section-${{varMeta.name}}`;
            }}
            
            const header = document.createElement('div');
            const isNPSVar = varMeta.scale_info && varMeta.scale_info.type === 'NPS';
            const isMRSection = varMeta.type === 'mr' || varMeta.var_type === 'multiple_response';
            const isOpenSection = (varMeta.var_type || varMeta.type || '') === 'string';
            if (isNPSVar)         {{ section.style.setProperty('--bar-color','var(--red)'); }} // bar-text por linha
            else if (isMRSection) {{ section.style.setProperty('--bar-color','var(--green)'); section.style.setProperty('--bar-text','var(--green-dark)'); }}
            else if (isOpenSection){{ section.style.setProperty('--bar-color','#aaa'); section.style.setProperty('--bar-text','#6c757d'); }}
            const isOpenHeader = (varMeta.var_type || varMeta.type || '') === 'string' || (varMeta.var_type || varMeta.type || '') === 'date';
            header.className = 'section-header' + (varMeta.type === 'mr' || varMeta.var_type === 'multiple_response' ? ' mr-header' : '') + (isNPSVar ? ' nps-header' : '') + (isOpenHeader ? ' open-header' : '');
            
            const title = document.createElement('h2');
            title.className = 'section-title';
            
            const varType = varMeta.var_type || varMeta.type || "single";
            const measure = varMeta.measure || null;
            
            let icon = '';
            const isMRVar = varType === 'multiple_response' || varMeta.type === 'mr';
            if (varType === 'string') {{
                icon = '';
            }} else if (isMRVar) {{
                icon = '';
            }} else if (varType === 'date') {{
                icon = '';
            }} else if (varType === 'numeric' && measure === 'scale') {{
                icon = '';
            }} else {{
                icon = '';
            }}
            
            const pPrefix = '';  // p_num atribuído dinamicamente em renderAll
            // Cor do título segue o tipo: MR=verde, string/date=cinza, resto=teal
            const isMRType2 = varType === 'multiple_response' || varMeta.type === 'mr';
            const isOpenType = varType === 'string' || varType === 'date';
            const isNPSType  = varMeta.scale_info && varMeta.scale_info.type === 'NPS';
            const titleColor = isNPSType  ? 'var(--red)'
                             : isMRType2  ? 'var(--green-dark)'
                             : isOpenType ? '#888'
                             : 'var(--primary-dark)';
            title.style.color = titleColor;
            title.innerHTML = (icon ? icon + ' ' : '') + varMeta.title;

            const subtitle = document.createElement('div');
            subtitle.className = 'section-subtitle';

            // Pill do tipo de variável
            if (varMeta.spss_type) {{
                const typePill = document.createElement('span');
                const isMRType = varType === 'multiple_response' || varMeta.type === 'mr';
                const isNPSpill = varMeta.scale_info && varMeta.scale_info.type === 'NPS';
                const isScalePill = !!varMeta.scale_info;
                const isOpenPill = (varType === 'string' || varType === 'date') && !isMRType;
                typePill.className = 'type-pill' + (isMRType ? ' mr' : '') + (isNPSpill ? ' nps' : (isScalePill ? ' scale' : (isOpenPill ? ' open' : '')));
                typePill.textContent = varMeta.spss_type;
                subtitle.appendChild(typePill);
            }}

            // Toggle NS: para variáveis categóricas que possam ter "Não sabe"
            // Toggle NS apenas para escalas com tag (CSAT/SAT/NPS/CES)
            const isCateg = !!varMeta.scale_info;
            if (isCateg) {{
                const nsToggleWrap = document.createElement('div');
                nsToggleWrap.className = 'ns-toggle-wrap';
                nsToggleWrap.title = 'Incluir ou excluir "Não sabe/Não sei" dos cálculos';
                // Inicializar só se ainda não definido (preserva estado entre re-renders)
                if (varMeta._nsIncluded === undefined) varMeta._nsIncluded = true;
                const nsOn = varMeta._nsIncluded;
                const nsIsNPS = varMeta.scale_info && varMeta.scale_info.type === 'NPS';
                const nsOnClass = nsOn ? (nsIsNPS ? 'ns-on-nps' : 'ns-on') : 'ns-off';
                nsToggleWrap.innerHTML = `
                    <div class="ns-track ${{nsOnClass}}" id="ns-track-${{varMeta.name}}"><div class="ns-thumb"></div></div>
                    <span class="ns-toggle-label" id="ns-lbl-${{varMeta.name}}">${{nsOn ? 'Incluir "Não sei"' : 'Excluir "Não sei"'}}</span>`;
                nsToggleWrap.addEventListener('click', () => {{
                    varMeta._nsIncluded = !varMeta._nsIncluded;
                    const track = document.getElementById('ns-track-' + varMeta.name);
                    const lbl   = document.getElementById('ns-lbl-' + varMeta.name);
                    const isNps = varMeta.scale_info && varMeta.scale_info.type === 'NPS';
                    if (track) track.className = 'ns-track ' + (varMeta._nsIncluded ? (isNps ? 'ns-on-nps' : 'ns-on') : 'ns-off');
                    if (lbl)   lbl.textContent = varMeta._nsIncluded ? 'Incluir "Não sei"' : 'Excluir "Não sei"';
                    const nsAnchor = document.getElementById('section-' + varMeta.name);
                    const nsTopBefore = nsAnchor ? nsAnchor.getBoundingClientRect().top : 0;
                    const nsCleanup = renderAll();
                    requestAnimationFrame(() => {{
                        const nsAfter = document.getElementById('section-' + varMeta.name);
                        if (nsAfter) window.scrollBy({{ top: nsAfter.getBoundingClientRect().top - nsTopBefore, behavior: 'instant' }});
                        nsCleanup();
                    }});
                }});
                subtitle.appendChild(nsToggleWrap);
            }}

            // Skip logic: pill âmbar com tooltip (ao lado do type-pill)
            if (varMeta.skip_logic) {{
                const sl = varMeta.skip_logic;
                // Usar P(n) do source se disponível, senão extrair código do título
                const sourceMeta = VARS_META.find(v => v.name === sl.source_var);
                const sourceShort = sourceMeta && sourceMeta._rendered_p_num
                    ? `P${{sourceMeta._rendered_p_num}}`
                    : (sl.source_title && sl.source_title.match(/^(Q\\d+[\\w]*)[.\\s\\-]/i)
                        ? sl.source_title.match(/^(Q\\d+[\\w]*)[.\\s\\-]/i)[1]
                        : (sl.source_var || sl.source_title));
                const codeMap = CODE_TO_LABEL[sl.source_var] || {{}};
                const pillsHTML = sl.categories
                    .map(c => {{
                        // Só usar parseFloat se a string inteira for numérica (ex: "2.0" → "2")
                        // Evitar que "2 a 5 anos" vire parseFloat=2 → label errada
                        const isNumeric = !isNaN(c) && String(c).trim() !== '';
                        const label = codeMap[c]
                            || (isNumeric ? codeMap[String(parseFloat(c))] : null)
                            || c;
                        return `<span class="skip-pill">${{label}}</span>`;
                    }})
                    .join(' ');
                const tag = document.createElement('span');
                tag.className = 'skip-tag';
                tag.innerHTML = `Filtrada por ${{sourceShort}}<span class="skip-tooltip">${{sl.coverage}}% dos respondentes vieram de <strong>${{sourceShort}}</strong>:<br>${{pillsHTML}}</span>`;
                subtitle.appendChild(tag);
            }}
            
            header.appendChild(title);
            header.appendChild(subtitle);
            
            const content = document.createElement('div');
            content.className = 'section-content';
            
            // Escolha do renderizador
            let rendered = null;
            if (varType === 'string') {{
                rendered = renderStringVariable(varMeta, records);
            }} else if (varType === 'multiple_response' || varMeta.type === 'mr') {{
                rendered = renderCategoricalVariable(varMeta, records);
            }} else if (varType === 'date') {{
                rendered = renderDateVariable(varMeta, records);
            }} else if (varType === 'numeric' && measure === 'scale') {{
                rendered = renderNumericScaleVariable(varMeta, records);
            }} else {{
                rendered = renderCategoricalVariable(varMeta, records);
            }}

            // Sem dados: não exibir a seção
            if (!rendered) return null;

            content.appendChild(rendered);
            section.appendChild(header);
            section.appendChild(content);            
            return section;
        }}


        // Eventos globais
        document.addEventListener('click', function(event) {{
            if (!event.target.closest('.custom-dropdown')) {{
                document.querySelectorAll('.dropdown-content').forEach(d => d.classList.remove('show'));
                document.querySelectorAll('.dropdown-button').forEach(b => b.classList.remove('open'));
            }}
        }});

        document.addEventListener('keydown', function(event) {{
            if (event.key === 'Escape') {{
                document.querySelectorAll('.dropdown-content').forEach(d => d.classList.remove('show'));
                document.querySelectorAll('.dropdown-button').forEach(b => b.classList.remove('open'));
            }}
        }});

        function exportAllTables() {{
            const sections = document.querySelectorAll('.section');
            if (!sections.length) {{
                alert("Nenhuma tabela encontrada.");
                return;
            }}

            const wb = XLSX.utils.book_new();

            // Registros filtrados atuais
            const filteredRecs = getFilteredRecords(null);
            const totalFilt = filteredRecs.length;

            // ── ABA 1: RESUMO ──
            const now = new Date();
            const dateStr = now.toLocaleString('pt-BR');
            const activeFilters = getActiveFiltersDescription();
            const totalFiltW = filteredRecs.reduce((s,r) => s + (r.__weight__ || 1), 0);
            const totalAllW  = RECORDS.reduce((s,r) => s + (r.__weight__ || 1), 0);
            const summaryRows = [
                ['RESUMO DA EXPORTAÇÃO'],
                [],
                ['Arquivo', '{file_source}'],
                ['Gerado em', dateStr],
                ['Total de respondentes (base)', RECORDS.length],
                ['Total de respondentes ponderado (base)', Math.round(totalAllW * 100) / 100],
                ['Respondentes após filtros (n bruto)', totalFilt],
                ['Respondentes após filtros (ponderado)', Math.round(totalFiltW * 100) / 100],
                ['Variáveis analisadas', VARS_META.length],
                ['Filtros aplicados', activeFilters.join(' | ')],
            ];
            const wsSum = XLSX.utils.aoa_to_sheet(summaryRows);
            wsSum['!cols'] = [{{wch:38}}, {{wch:55}}];
            wsSum['!merges'] = [{{s:{{r:0,c:0}}, e:{{r:0,c:1}}}}];
            XLSX.utils.book_append_sheet(wb, wsSum, 'Resumo');

            // ── ABA 2: DICIONÁRIO com N válido ponderado e bruto ──
            function countValid(vm) {{
                const vname = vm.name;
                const vtype = vm.var_type || vm.type || 'single';
                if (vtype === 'multiple_response' || vtype === 'mr') {{
                    return filteredRecs.filter(r => r[vname] && r[vname].length > 0);
                }}
                return filteredRecs.filter(r => r[vname] !== null && r[vname] !== undefined);
            }}
            function weightedN(recs) {{
                return Math.round(recs.reduce((s,r) => s + (r.__weight__ || 1), 0) * 100) / 100;
            }}

            const dictRows = [['Nº','Variável SPSS','Pergunta','Tipo','N válido ponderado','N válido bruto','N missing (bruto)','Filtrada por']];
            VARS_META.forEach(vm => {{
                // P número: se foi renderizado usa o número sequencial, senão indica que não apareceu
                const pNum = vm._rendered_p_num ? 'P' + vm._rendered_p_num : '–';
                const validRecs = countValid(vm);
                const nvW = weightedN(validRecs);
                const nvBruto = validRecs.length;
                // N missing = registros filtrados que NÃO têm resposta válida
                const nm = totalFilt - nvBruto;
                const sl = vm.skip_logic;
                let filtradaPor = '';
                if (sl) {{
                    const src = VARS_META.find(v => v.name === sl.source_var);
                    filtradaPor = src && src._rendered_p_num ? 'P' + src._rendered_p_num : (sl.source_var || '');
                }}
                dictRows.push([pNum, vm.sheet_code || vm.name, vm.title || '', vm.spss_type || '', nvW, nvBruto, nm, filtradaPor]);
            }});
            const wsDic = XLSX.utils.aoa_to_sheet(dictRows);
            wsDic['!cols'] = [{{wch:6}},{{wch:14}},{{wch:55}},{{wch:20}},{{wch:20}},{{wch:14}},{{wch:10}},{{wch:12}}];
            XLSX.utils.book_append_sheet(wb, wsDic, 'Dicionário');

            const usedNames = new Set(['Resumo','Dicionário']);
            let sectionIndex = 1;

            sections.forEach(section => {{
                const titleEl = section.querySelector('.section-title');
                const table = section.querySelector('table');

                if (!table) return;

                const title = titleEl ? titleEl.innerText.trim() : "Variável";
                
                // ===== ESTRATÉGIA PARA EXTRAIR NOME DA VARIÁVEL =====
                let varName = null;
                
                // Método 1: Tentar extrair do atributo data-var (se existir)
                varName = section.getAttribute('data-var') || section.getAttribute('data-variable');
                
                // Método 2: Procurar no ID da seção (formato comum: section-P1, section-P4_1)
                if (!varName && section.id) {{
                    const idMatch = section.id.match(/section-(.+)/);
                    if (idMatch) varName = idMatch[1];
                }}
                
                // Método 3: Procurar em classes CSS (formato comum: var-P1, variable-P4_1)
                if (!varName && section.className) {{
                    const classMatch = section.className.match(/(?:var|variable)-([A-Za-z0-9_]+)/);
                    if (classMatch) varName = classMatch[1];
                }}
                
                // Método 4: Tentar extrair do conteúdo do título (formato comum: "P1. Título" ou "P4_1 - Título")
                if (!varName) {{
                    const titleMatch = title.match(/^([A-Za-z0-9_]+)[.\x20\t\x2D:|]/);
                    if (titleMatch) varName = titleMatch[1];
                }}
                
                // Método 5: Usar índice da variável de VARS_META (se disponível)
                if (!varName && typeof VARS_META !== 'undefined' && VARS_META[sectionIndex - 1]) {{
                    varName = VARS_META[sectionIndex - 1].name;
                }}
                
                // ===== CRIAR NOME DA ABA =====
                let sheetName;
                
                if (varName) {{
                    // Formato preferido: "P1 - Título" ou "P4_1 - Título"
                    let combinedTitle = `${{varName}} - ${{title}}`;
                    
                    // Se o título já contém a variável no início, evitar duplicação
                    if (title.toLowerCase().startsWith(varName.toLowerCase())) {{
                        combinedTitle = title;
                    }}
                    
                    // Limpar caracteres proibidos pelo Excel
                    combinedTitle = combinedTitle.replace(/[:\\/\\*\\[\\]\\?]/g, "");
                    
                    // Limitar a 31 caracteres (limite do Excel)
                    if (combinedTitle.length > 31) {{
                        // Tentar formato mais compacto: "P1-Título"
                        const compactTitle = `${{varName}}-${{title.replace(/[:\\/\\*\\[\\]\\?\\s]/g, "")}}`;
                        if (compactTitle.length <= 31) {{
                            sheetName = compactTitle;
                        }} else {{
                            // Cortar título mas manter variável
                            const maxTitleLength = 31 - varName.length - 1; // -1 para o hífen
                            const truncatedTitle = title.replace(/[:\\/\\*\\[\\]\\?]/g, "").substring(0, maxTitleLength);
                            sheetName = `${{varName}}-${{truncatedTitle}}`;
                        }}
                    }} else {{
                        sheetName = combinedTitle;
                    }}
                }} else {{
                    // Fallback: usar índice numérico
                    let safeName = title.replace(/[:\\/\\*\\[\\]\\?]/g, "");
                    safeName = safeName.replace(/\\s+/g, ' ').trim();
                    
                    // Adicionar número da seção
                    if (safeName.length > 27) {{ // Reservar espaço para " (X)"
                        safeName = safeName.substring(0, 27);
                    }}
                    sheetName = `${{safeName}} (${{sectionIndex}})`;
                }}
                
                // ===== GARANTIR UNICIDADE =====
                let finalName = sheetName;
                let counter = 1;
                
                // Se o nome já existe, adicionar contador
                while (usedNames.has(finalName)) {{
                    const suffix = ` (${{counter}})`;
                    const maxLength = 31 - suffix.length;
                    
                    if (sheetName.length > maxLength) {{
                        finalName = sheetName.substring(0, maxLength) + suffix;
                    }} else {{
                        finalName = sheetName + suffix;
                    }}
                    counter++;
                }}
                
                // Registrar nome usado
                usedNames.add(finalName);
                
                // ===== EXTRAIR DADOS DA TABELA =====
                const rows = [];
                table.querySelectorAll('tr').forEach(tr => {{
                    const row = [];
                    tr.querySelectorAll('th, td').forEach(cell => {{
                        row.push(cell.innerText.trim());
                    }});
                    rows.push(row);
                }});

                // ===== CRIAR PLANILHA =====
                const activeFilters = getActiveFiltersDescription();
                
                const ws = XLSX.utils.aoa_to_sheet([
                    [title], // Título completo na primeira linha
                    [activeFilters.length ? ('Filtros aplicados: ' + activeFilters.join(' | ')) : 'Filtros aplicados: Nenhum'],
                    [], // Linha vazia
                    ...rows
                ]);

                // Adicionar aba ao workbook
                XLSX.utils.book_append_sheet(wb, ws, finalName);
                
                console.log(`✅ Aba criada: "${{finalName}}" (variável: ${{varName || 'não identificada'}})`);
                sectionIndex++;
            }});

            // ===== SALVAR ARQUIVO =====
            const fileName = "tabelas_exportadas.xlsx";
            XLSX.writeFile(wb, fileName);
            
            console.log(`📊 Excel exportado com ${{usedNames.size}} abas: ${{fileName}}`);
// Excel exportado — sem popup
        }}

        // ===== FUNÇÕES DE EXPORTAÇÃO PDF =====
        
        function formatNumberBR(num) {{
            // Formatação brasileira: 1.234 (ponto para milhares)
            return num.toLocaleString('pt-BR');
        }}

        // ===== MODO APRESENTAÇÃO =====
        function startPresentation() {{
            const CLIENT_LOGO  = '{client_logo}';
            const OPINIAO_LOGO = 'https://raw.githubusercontent.com/aag1974/dashboard-ivv/main/logo.png';
            // DESCRIPTION, FILE_SOURCE e PERIOD são globais

            let sections = Array.from(document.querySelectorAll('.section'));
            const total = sections.length + 1;
            let current = 0;

            const overlay = document.createElement('div');
            overlay.id = 'pres-overlay';
            overlay.style.cssText = 'position:fixed;top:0;left:0;right:0;bottom:0;background:#EFF2F5;z-index:99999;display:flex;flex-direction:column;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;';

            const bar = document.createElement('div');
            bar.style.cssText = 'background:white;border-bottom:2px solid #7BAFC0;display:flex;align-items:center;padding:0 24px;height:46px;justify-content:space-between;flex-shrink:0;box-shadow:0 1px 4px rgba(0,0,0,0.06);position:relative;z-index:2;';
            bar.innerHTML =
                '<div style="display:flex;align-items:center;gap:10px;">' +
                  '<div style="width:8px;height:8px;border-radius:50%;background:#7BAFC0;flex-shrink:0;"></div>' +
                  '<span style="font-size:12px;font-weight:600;color:#4d8a9e;letter-spacing:.01em;">Painel de Resultados</span>' +
                  (DESCRIPTION ? '<span style="color:#ccc;font-size:11px;">|</span><span style="font-size:12px;color:#888;max-width:500px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">' + DESCRIPTION + '</span>' : '') +
                '</div>' +
                '<div style="display:flex;align-items:center;gap:6px;">' +
                  (FILTERS.length > 0 ? '<button id="pf-btn" style="display:flex;align-items:center;gap:5px;padding:0 10px;height:30px;border-radius:6px;background:#f4f6f8;border:0.5px solid #dde5ea;color:#888;cursor:pointer;font-size:12px;font-weight:500;white-space:nowrap;"><svg width="13" height="13" viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><path d="M2 4h12M4 8h8M6 12h4"/></svg> Filtrar <span id="pf-badge" style="display:none;background:#7BAFC0;color:white;font-size:10px;font-weight:700;border-radius:10px;padding:1px 6px;line-height:1.4;"></span></button><div style="width:1px;height:20px;background:#eee;margin:0 2px;"></div>' : '') +
                  '<button id="pp" style="width:30px;height:30px;border-radius:6px;background:#EDF5F7;border:0.5px solid #b8d9e3;color:#4d8a9e;cursor:pointer;font-size:15px;line-height:1;">&#8592;</button>' +
                  '<span id="pc" style="font-size:12px;color:#999;min-width:42px;text-align:center;">1 / ' + total + '</span>' +
                  '<button id="pn" style="width:30px;height:30px;border-radius:6px;background:#7BAFC0;border:none;color:white;cursor:pointer;font-size:15px;line-height:1;">&#8594;</button>' +
                  '<button id="px" style="width:30px;height:30px;border-radius:6px;background:#fdeef2;border:0.5px solid #f5bdc9;color:#B83B5C;cursor:pointer;font-size:13px;margin-left:4px;">&#10005;</button>' +
                '</div>';
            overlay.appendChild(bar);

            // ── PAINEL DE FILTROS DROPDOWN ──
            const filterPanel = document.createElement('div');
            filterPanel.id = 'pf-panel';
            filterPanel.style.cssText = 'display:none;background:white;border-bottom:2px solid #7BAFC0;padding:12px 24px;flex-shrink:0;box-shadow:0 4px 12px rgba(0,0,0,0.08);z-index:1;';

            if (FILTERS.length > 0) {{
                const filterRow = document.createElement('div');
                filterRow.style.cssText = 'display:flex;gap:14px;align-items:flex-end;flex-wrap:wrap;';

                // CSS inline para os dropdowns do painel (injetado uma vez)
                const pfStyle = document.createElement('style');
                pfStyle.textContent = '.pf-dd-content{{display:none;position:absolute;top:100%;left:0;min-width:200px;background:white;border:0.5px solid #dde5ea;border-radius:6px;box-shadow:0 4px 12px rgba(0,0,0,0.1);max-height:220px;overflow-y:auto;z-index:9999;margin-top:2px;}}.pf-dd-content.open{{display:block;}}.pf-dd-opt{{display:flex;align-items:center;gap:8px;padding:7px 12px;cursor:pointer;font-size:13px;color:#333;}}.pf-dd-opt:hover{{background:#f8f9fa;}}.pf-dd-opt.select-all{{background:#f4f6f8;font-weight:600;border-bottom:0.5px solid #eee;}}';
                document.head.appendChild(pfStyle);

                // Fechar todos os dropdowns ao clicar fora
                document.addEventListener('click', function pfOutside(e) {{
                    if (!e.target.closest('.pf-dd-wrap')) {{
                        document.querySelectorAll('.pf-dd-content.open').forEach(d => d.classList.remove('open'));
                    }}
                }});

                FILTERS.forEach(f => {{
                    const col = document.createElement('div');
                    col.style.cssText = 'display:flex;flex-direction:column;gap:4px;flex:1;min-width:140px;max-width:240px;';

                    const lbl = document.createElement('div');
                    lbl.style.cssText = 'font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.06em;';
                    lbl.textContent = f.title;

                    // Wrapper com posicionamento relativo para o dropdown
                    const ddWrap = document.createElement('div');
                    ddWrap.className = 'pf-dd-wrap';
                    ddWrap.style.cssText = 'position:relative;';

                    // Botão que mostra a seleção atual
                    const ddBtn = document.createElement('div');
                    ddBtn.id = 'pf-dd-btn-' + f.name;
                    ddBtn.style.cssText = 'height:32px;border:0.5px solid #dde5ea;border-radius:6px;font-size:13px;padding:0 28px 0 10px;background:white;color:#999;cursor:pointer;display:flex;align-items:center;user-select:none;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;position:relative;';
                    ddBtn.innerHTML = '<span id="pf-dd-txt-' + f.name + '" style="flex:1;overflow:hidden;text-overflow:ellipsis;">Todos</span><span style="position:absolute;right:8px;font-size:9px;color:#bbb;">▼</span>';

                    // Conteúdo do dropdown (checkboxes)
                    const ddContent = document.createElement('div');
                    ddContent.className = 'pf-dd-content';
                    ddContent.id = 'pf-dd-' + f.name;

                    // Opção "Selecionar Todos"
                    const allOpt = document.createElement('div');
                    allOpt.className = 'pf-dd-opt select-all';
                    const allCb = document.createElement('input');
                    allCb.type = 'checkbox';
                    allCb.id = 'pf-all-' + f.name;
                    const allLbl = document.createElement('label');
                    allLbl.htmlFor = 'pf-all-' + f.name;
                    allLbl.textContent = 'Selecionar Todos';
                    allLbl.style.cursor = 'pointer';
                    allOpt.appendChild(allCb);
                    allOpt.appendChild(allLbl);
                    ddContent.appendChild(allOpt);

                    // Opções individuais
                    const curSel = getSelectedFilters()[f.name] || [];
                    f.values.forEach(v => {{
                        const opt = document.createElement('div');
                        opt.className = 'pf-dd-opt';
                        opt.dataset.filterName = f.name;
                        opt.dataset.filterValue = v;
                        const cb = document.createElement('input');
                        cb.type = 'checkbox';
                        cb.value = v;
                        cb.checked = curSel.includes(v);
                        const lb = document.createElement('label');
                        lb.textContent = v;
                        lb.style.cursor = 'pointer';
                        opt.appendChild(cb);
                        opt.appendChild(lb);
                        ddContent.appendChild(opt);
                    }});

                    // Atualizar texto do botão
                    function pfUpdateText() {{
                        const checks = Array.from(ddContent.querySelectorAll('.pf-dd-opt:not(.select-all) input:checked'));
                        const total  = ddContent.querySelectorAll('.pf-dd-opt:not(.select-all) input').length;
                        const txt = document.getElementById('pf-dd-txt-' + f.name);
                        if (!txt) return;
                        const allChk = document.getElementById('pf-all-' + f.name);
                        if (checks.length === 0) {{
                            txt.textContent = 'Todos';
                            ddBtn.style.color = '#999';
                            ddBtn.style.borderColor = '#dde5ea';
                            if (allChk) allChk.checked = false;
                        }} else if (checks.length === total) {{
                            txt.textContent = 'Todos (' + total + ')';
                            ddBtn.style.color = '#4d8a9e';
                            ddBtn.style.borderColor = '#7BAFC0';
                            if (allChk) allChk.checked = true;
                        }} else {{
                            txt.textContent = checks.length + ' selecionado' + (checks.length > 1 ? 's' : '');
                            ddBtn.style.color = '#4d8a9e';
                            ddBtn.style.borderColor = '#7BAFC0';
                            if (allChk) allChk.checked = false;
                        }}
                    }}

                    // Checkbox "Todos"
                    allCb.onchange = () => {{
                        ddContent.querySelectorAll('.pf-dd-opt:not(.select-all) input').forEach(c => c.checked = allCb.checked);
                        pfUpdateText();
                    }};

                    // Checkboxes individuais
                    ddContent.querySelectorAll('.pf-dd-opt:not(.select-all) input').forEach(c => {{
                        c.onchange = pfUpdateText;
                    }});

                    // Toggle dropdown
                    ddBtn.onclick = e => {{
                        e.stopPropagation();
                        const isOpen = ddContent.classList.contains('open');
                        document.querySelectorAll('.pf-dd-content.open').forEach(d => d.classList.remove('open'));
                        if (!isOpen) ddContent.classList.add('open');
                    }};

                    pfUpdateText();
                    ddWrap.appendChild(ddBtn);
                    ddWrap.appendChild(ddContent);
                    col.appendChild(lbl);
                    col.appendChild(ddWrap);
                    filterRow.appendChild(col);
                }});

                // Função auxiliar: ler seleção do painel
                function pfGetSelected() {{
                    const result = {{}};
                    FILTERS.forEach(f => {{
                        const checked = Array.from(
                            document.querySelectorAll('#pf-dd-' + f.name + ' .pf-dd-opt:not(.select-all) input:checked')
                        ).map(c => c.value);
                        result[f.name] = checked;
                    }});
                    return result;
                }}

                // Botões Aplicar / Limpar
                const btnCol = document.createElement('div');
                btnCol.style.cssText = 'display:flex;gap:8px;align-items:flex-end;padding-bottom:0;';

                const applyBtn = document.createElement('button');
                applyBtn.textContent = 'Aplicar';
                applyBtn.style.cssText = 'height:32px;padding:0 18px;background:#7BAFC0;color:white;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;white-space:nowrap;';
                applyBtn.onclick = () => {{
                    const pfSel = pfGetSelected();
                    // Sincronizar com checkboxes do dashboard
                    FILTERS.forEach(f => {{
                        const vals = pfSel[f.name] || [];
                        document.querySelectorAll('.dropdown-content [data-filter-name="' + f.name + '"] input').forEach(i => {{
                            i.checked = vals.includes(i.value);
                        }});
                        updateDropdownText(f.name);
                    }});
                    applyFilters();
                    sections = Array.from(document.querySelectorAll('.section'));
                    updateFilterBtn();
                    filterPanel.style.display = 'none';
                    pfOpen = false;
                    renderSlide(current);
                }};

                const clearBtn = document.createElement('button');
                clearBtn.textContent = 'Limpar';
                clearBtn.style.cssText = 'height:32px;padding:0 12px;background:white;color:#B83B5C;border:0.5px solid #f5bdc9;border-radius:6px;font-size:12px;cursor:pointer;white-space:nowrap;';
                clearBtn.onclick = () => {{
                    // Desmarcar todos os checkboxes do painel
                    FILTERS.forEach(f => {{
                        document.querySelectorAll('#pf-dd-' + f.name + ' input').forEach(c => c.checked = false);
                        const txt = document.getElementById('pf-dd-txt-' + f.name);
                        const btn = document.getElementById('pf-dd-btn-' + f.name);
                        if (txt) txt.textContent = 'Todos';
                        if (btn) {{ btn.style.color = '#999'; btn.style.borderColor = '#dde5ea'; }}
                    }});
                    clearFilters();
                    sections = Array.from(document.querySelectorAll('.section'));
                    updateFilterBtn();
                    filterPanel.style.display = 'none';
                    pfOpen = false;
                    renderSlide(current);
                }};

                btnCol.appendChild(applyBtn);
                btnCol.appendChild(clearBtn);
                filterRow.appendChild(btnCol);
                filterPanel.appendChild(filterRow);
            }}
            overlay.appendChild(filterPanel);

            // Lógica do botão de filtro
            let pfOpen = false;
            function updateFilterBtn() {{
                const btn = document.getElementById('pf-btn');
                const badge = document.getElementById('pf-badge');
                if (!btn) return;
                const active = getSelectedFilters();
                const nActive = Object.values(active).filter(v => v.length > 0).length;
                if (nActive > 0) {{
                    btn.style.background = '#EDF5F7';
                    btn.style.borderColor = '#7BAFC0';
                    btn.style.color = '#4d8a9e';
                    if (badge) {{ badge.textContent = nActive; badge.style.display = 'inline'; }}
                }} else {{
                    btn.style.background = '#f4f6f8';
                    btn.style.borderColor = '#dde5ea';
                    btn.style.color = '#888';
                    if (badge) badge.style.display = 'none';
                }}
            }}

            const slideArea = document.createElement('div');
            slideArea.style.cssText = 'flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px 40px 16px;overflow:auto;';
            overlay.appendChild(slideArea);

            document.body.appendChild(overlay);
            if (overlay.requestFullscreen) overlay.requestFullscreen().catch(() => {{}});

            function renderSlide(idx) {{
                current = Math.max(0, Math.min(idx, total - 1));
                document.getElementById('pc').textContent = (current + 1) + ' / ' + total;
                slideArea.innerHTML = '';

                if (current === 0) {{
                    const periodo = getPeriodoColeta();
                    const card = document.createElement('div');
                    card.style.cssText = 'background:white;border-radius:12px;width:100%;max-width:860px;overflow:hidden;border:0.5px solid #dde5ea;box-shadow:0 2px 8px rgba(0,0,0,0.06);';

                    // Topo teal
                    const top = document.createElement('div');
                    top.style.cssText = 'background:#EDF5F7;padding:40px 48px 32px;';
                    const _nBruto = '<div><div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;">Respondentes</div><div style="font-size:22px;color:#444;font-weight:500;">' + RECORDS.length.toLocaleString('pt-BR') + '</div></div>';
                    const _nPondRaw = Math.round(RECORDS.reduce((s,r) => s + (r.__weight__||1), 0));
                    const _nPondHtml = Math.abs(_nPondRaw - RECORDS.length) > 0.5
                        ? '<div><div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;">N ponderado</div><div style="font-size:22px;color:#7BAFC0;font-weight:500;">' + _nPondRaw.toLocaleString('pt-BR') + '</div></div>'
                        : '';
                    // N filtrado — só aparece quando há filtro ativo
                    const _filtRecs = getFilteredRecords(null);
                    const _nFiltRaw = Math.round(_filtRecs.reduce((s,r) => s + (r.__weight__||1), 0));
                    const _hasFilter = Object.values(getSelectedFilters()).some(v => v.length > 0) || DRILLDOWN.size > 0;
                    const _nFiltHtml = _hasFilter
                        ? '<div style="border-left:2px solid #7BAFC0;padding-left:18px;"><div style="font-size:10px;color:#4d8a9e;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;font-weight:600;">N filtrado</div><div style="font-size:22px;color:#4d8a9e;font-weight:600;">' + _nFiltRaw.toLocaleString('pt-BR') + '</div></div>'
                        : '';
                    top.innerHTML =
                        '<div style="font-size:11px;font-weight:600;color:#7BAFC0;text-transform:uppercase;letter-spacing:.07em;margin-bottom:12px;">Painel de Resultados</div>' +
                        '<div style="font-size:18px;color:#444;line-height:1.6;margin-bottom:32px;max-width:560px;font-weight:500;">' + (DESCRIPTION || FILE_SOURCE) + '</div>' +
                        '<div style="display:flex;gap:36px;flex-wrap:wrap;align-items:flex-start;">' +
                          _nBruto + _nPondHtml + _nFiltHtml +
                          '<div><div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;">Perguntas</div><div style="font-size:22px;color:#444;font-weight:500;">' + VARS_META.length + '</div></div>' +
                          '<div><div style="font-size:10px;color:#aaa;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;">Período de coleta</div><div style="font-size:22px;color:#444;font-weight:500;">' + periodo + '</div></div>' +
                        '</div>';
                    card.appendChild(top);

                    // Rodapé com logos — criados via DOM para evitar problemas de aspas
                    const foot = document.createElement('div');
                    foot.style.cssText = 'display:flex;align-items:center;justify-content:space-between;padding:14px 40px;border-top:0.5px solid rgba(123,175,192,0.2);';

                    const imgClient = document.createElement('img');
                    imgClient.src = CLIENT_LOGO;
                    imgClient.style.cssText = 'height:34px;object-fit:contain;';
                    imgClient.onerror = function() {{ this.style.display = 'none'; }};
                    foot.appendChild(imgClient);

                    const opDiv = document.createElement('div');
                    opDiv.style.cssText = 'display:flex;flex-direction:column;align-items:flex-end;gap:4px;';
                    const opLbl = document.createElement('div');
                    opLbl.style.cssText = 'font-size:10px;color:#bbb;';
                    opLbl.textContent = 'Desenvolvido por';
                    const imgOp = document.createElement('img');
                    imgOp.src = OPINIAO_LOGO;
                    imgOp.style.cssText = 'height:22px;opacity:.65;';
                    imgOp.onerror = function() {{ this.style.display = 'none'; }};
                    opDiv.appendChild(opLbl);
                    opDiv.appendChild(imgOp);
                    foot.appendChild(opDiv);

                    card.appendChild(foot);
                    slideArea.appendChild(card);

                    // Dots de posição
                    const dotsBarC = document.createElement('div');
                    dotsBarC.style.cssText = 'display:flex;gap:5px;margin-top:12px;align-items:center;';
                    const maxDotsC = Math.min(total, 20);
                    for (let i = 0; i < maxDotsC; i++) {{
                        const dot = document.createElement('div');
                        dot.style.cssText = i === 0
                            ? 'width:16px;height:6px;border-radius:3px;background:#7BAFC0;cursor:pointer;'
                            : 'width:6px;height:6px;border-radius:50%;background:#c8d8df;cursor:pointer;';
                        dot.onclick = () => renderSlide(i);
                        dotsBarC.appendChild(dot);
                    }}
                    slideArea.appendChild(dotsBarC);
                }} else {{
                    const src = sections[current - 1];
                    if (!src) return;

                    const wrap = document.createElement('div');
                    wrap.style.cssText = 'background:white;border-radius:10px;width:100%;max-width:1100px;max-height:calc(100vh - 110px);overflow-y:auto;overflow-x:hidden;border:0.5px solid #dde5ea;box-shadow:0 2px 8px rgba(0,0,0,0.06);';

                    // Clonar a seção inteira — preserva CSS vars, barras, cores NPS/MR/CSAT
                    const clone = src.cloneNode(true);
                    clone.style.cssText = 'border-radius:10px;border:none;box-shadow:none;margin:0;';
                    clone.querySelectorAll('.ivv-drilldown-hint').forEach(el => el.remove());
                    clone.querySelectorAll('.ivv-row').forEach(tr => {{ tr.style.cursor = 'default'; }});

                    // ── RE-ATIVAR KEYWORD FILTER (respostas abertas) ──
                    clone.querySelectorAll('.outros-kw').forEach(chip => {{
                        chip.onclick = function() {{
                            const wasActive = chip.classList.contains('active');

                            // Toggle este chip (multi-select OR)
                            if (wasActive) {{
                                chip.classList.remove('active');
                            }} else {{
                                chip.classList.add('active');
                            }}

                            // Chips ativos no mesmo grupo
                            const siblingChips = Array.from(chip.parentElement.querySelectorAll('.outros-kw'));
                            const activeChips  = siblingChips.filter(c => c.classList.contains('active'));

                            // Normalizar raízes ativas
                            function normStr(s) {{
                                return (s || '').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'');
                            }}
                            const activeRoots = activeChips.map(c => normStr(c.dataset.root || c.textContent.replace(/[ ]*([(][0-9]+[)])[ ]*$/, '')));

                            // Determinar os itens a filtrar
                            let items;
                            const otrosSub = chip.closest('.outros-sub');
                            if (otrosSub) {{
                                // caso: sub-box "Outro" (.outros-item)
                                items = Array.from(otrosSub.querySelectorAll('.outros-item'));
                            }} else {{
                                // caso: pergunta aberta principal (irmão do filterContainer)
                                const responseList = chip.parentElement.nextElementSibling;
                                items = responseList ? Array.from(responseList.children) : [];
                            }}

                            items.forEach(item => {{
                                if (activeRoots.length === 0) {{
                                    item.style.display = '';
                                }} else {{
                                    const t = normStr(item.textContent);
                                    item.style.display = activeRoots.some(r => r && t.includes(r)) ? '' : 'none';
                                }}
                            }});
                        }};
                    }});

                    // ── RE-ATIVAR NS TOGGLE (escalas NPS/CSAT/SAT/CES) ──
                    clone.querySelectorAll('.ns-toggle-wrap').forEach(nsWrap => {{
                        nsWrap.style.pointerEvents = '';
                        nsWrap.onclick = function() {{
                            const sec = nsWrap.closest('.section');
                            const varName = sec ? sec.getAttribute('data-var') : null;
                            if (!varName) return;
                            const varMeta = VARS_META.find(v => v.name === varName);
                            if (!varMeta) return;
                            varMeta._nsIncluded = !varMeta._nsIncluded;
                            const cleanup = renderAll(); cleanup();
                            sections = Array.from(document.querySelectorAll('.section'));
                            renderSlide(current);
                        }};
                    }});

                    const foot = document.createElement('div');
                    foot.style.cssText = 'display:flex;justify-content:space-between;align-items:center;padding:8px 20px 12px;border-top:0.5px solid #eef1f4;';
                    foot.innerHTML = '<span style="font-size:10px;color:#bbb;">' + FILE_SOURCE + '</span>';
                    const footImg = document.createElement('img');
                    footImg.src = OPINIAO_LOGO;
                    footImg.style.cssText = 'height:16px;opacity:.2;';
                    footImg.onerror = function() {{ this.style.display = 'none'; }};
                    foot.appendChild(footImg);

                    wrap.appendChild(clone);
                    wrap.appendChild(foot);
                    slideArea.appendChild(wrap);

                    // Dots de posição
                    const dotsBar = document.createElement('div');
                    dotsBar.style.cssText = 'display:flex;gap:5px;margin-top:12px;align-items:center;';
                    const maxDots = Math.min(total, 20);
                    const startDot = total <= maxDots ? 0 : Math.max(0, Math.min(current - Math.floor(maxDots/2), total - maxDots));
                    for (let i = startDot; i < startDot + maxDots && i < total; i++) {{
                        const dot = document.createElement('div');
                        dot.style.cssText = i === current
                            ? 'width:16px;height:6px;border-radius:3px;background:#7BAFC0;cursor:pointer;transition:all .2s;'
                            : 'width:6px;height:6px;border-radius:50%;background:#c8d8df;cursor:pointer;transition:all .2s;';
                        dot.onclick = () => renderSlide(i);
                        dotsBar.appendChild(dot);
                    }}
                    slideArea.appendChild(dotsBar);
                }}
            }}

            document.getElementById('pp').onclick = () => renderSlide(current - 1);
            document.getElementById('pn').onclick = () => renderSlide(current + 1);
            document.getElementById('px').onclick = closePresentation;

            // Botão filtro toggle
            const pfBtn = document.getElementById('pf-btn');
            if (pfBtn) {{
                pfBtn.onclick = () => {{
                    pfOpen = !pfOpen;
                    filterPanel.style.display = pfOpen ? 'block' : 'none';
                    pfBtn.style.background = pfOpen ? '#7BAFC0' : (document.getElementById('pf-badge') && document.getElementById('pf-badge').style.display !== 'none' ? '#EDF5F7' : '#f4f6f8');
                    pfBtn.style.color = pfOpen ? 'white' : (document.getElementById('pf-badge') && document.getElementById('pf-badge').style.display !== 'none' ? '#4d8a9e' : '#888');
                    pfBtn.style.borderColor = pfOpen ? '#7BAFC0' : (document.getElementById('pf-badge') && document.getElementById('pf-badge').style.display !== 'none' ? '#7BAFC0' : '#dde5ea');
                }};
                updateFilterBtn();
            }}

            window.__presKeyHandler = e => {{
                if (e.key === 'ArrowRight' || e.key === 'ArrowDown') renderSlide(current + 1);
                else if (e.key === 'ArrowLeft' || e.key === 'ArrowUp') renderSlide(current - 1);
                else if (e.key === 'Escape') closePresentation();
            }};
            document.addEventListener('keydown', window.__presKeyHandler);
            // Hook para renderAll() atualizar o slide atual automaticamente
            window.__presRenderCurrent = () => renderSlide(current);

            renderSlide(0);
        }}

        function closePresentation() {{
            if (document.fullscreenElement) document.exitFullscreen().catch(() => {{}});
            const ov = document.getElementById('pres-overlay');
            if (ov) ov.remove();
            document.removeEventListener('keydown', window.__presKeyHandler);
            window.__presRenderCurrent = null;
        }}

        function getPeriodoColeta() {{
            // Prioridade 1: período digitado pelo usuário na geração
            if (typeof PERIOD !== 'undefined' && PERIOD && PERIOD.trim()) return PERIOD.trim();
            try {{
                const dates = [];
                // Primeiro tenta campos tipados como data no vars_meta
                const dateVarNames = new Set(
                    VARS_META.filter(v => (v.var_type || v.type) === 'date').map(v => v.name)
                );
                RECORDS.forEach(r => {{
                    Object.keys(r).forEach(k => {{
                        const isDateField = dateVarNames.has(k) ||
                            k.toLowerCase().includes('submit') ||
                            k.toLowerCase().includes('date') ||
                            k.toLowerCase().includes('data');
                        if (!isDateField) return;
                        const v = r[k];
                        if (!v) return;
                        const d = new Date(v);
                        // Rejeitar datas inválidas e conversões erradas de epoch SPSS (< 2010)
                        if (!isNaN(d.getTime()) && d.getFullYear() >= 2010 && d.getFullYear() <= 2100) {{
                            dates.push(d);
                        }}
                    }});
                }});
                if (dates.length === 0) return 'Não disponível';
                const min = new Date(Math.min(...dates));
                const max = new Date(Math.max(...dates));
                const fmt = d => d.toLocaleDateString('pt-BR');
                return min.getTime() === max.getTime() ? fmt(min) : fmt(min) + ' até ' + fmt(max);
            }} catch(e) {{ return 'Não disponível'; }}
        }}

        function getActiveFiltersDescription() {{
            const selectedFilters = getSelectedFilters();
            const activeFilters = [];

            // 1) filtros do topo
            Object.keys(selectedFilters).forEach(filterName => {{
                const filterValues = selectedFilters[filterName];
                if (filterValues.length > 0) {{
                    const filterMeta = FILTERS.find(f => f.name === filterName);
                    const filterTitle = filterMeta ? filterMeta.title : filterName;

                    if (filterValues.length === 1) {{
                        activeFilters.push(`${{filterTitle}}: ${{filterValues[0]}}`);
                    }} else {{
                        activeFilters.push(`${{filterTitle}}: ${{filterValues.length}} selecionados`);
                    }}
                }}
            }});

            // 2) drilldowns por clique (categorias selecionadas nas barras)
            for (const [varName, set] of DRILLDOWN.entries()) {{
                if (!set || set.size === 0) continue;
                const varMeta = VARS_META.find(v => v.name === varName);
                const varTitle = varMeta ? varMeta.title : varName;

                const rawVals = Array.from(set);
                const prettyVals = rawVals.map(rv => {{
                    const v = String(rv).trim();
                    if (CODE_TO_LABEL[varName] && CODE_TO_LABEL[varName][v]) return CODE_TO_LABEL[varName][v];
                    return v;
                }});

                if (prettyVals.length === 1) {{
                    activeFilters.push(`${{varTitle}} (drilldown): ${{prettyVals[0]}}`);
                }} else {{
                    activeFilters.push(`${{varTitle}} (drilldown): ${{prettyVals.length}} selecionados`);
                }}
            }}

            return activeFilters.length > 0 ? activeFilters : ['Nenhum filtro aplicado'];
        }}

        function createPDFHeader() {{
            const now = new Date();
            const dateStr = now.toLocaleString('pt-BR', {{
                day: '2-digit',
                month: '2-digit', 
                year: 'numeric',
                hour: '2-digit',
                minute: '2-digit'
            }});
            
            // Extrair informações dos dados globais
            const totalRecords = RECORDS.length;
            const totalVars = VARS_META.length;
            const activeFilters = getActiveFiltersDescription();
            
            const header = document.createElement('div');
            header.style.cssText = `
                background: white;
                padding: 0;
                margin: 0;
                font-family: Arial, sans-serif;
                page-break-after: always;
                height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
            `;
            
            header.innerHTML = `
                <div style="text-align: center; width: 100%;">
                    <h1 style="color: #7BAFC0; font-size: 32px; margin-bottom: 30px;">📋 RELATÓRIO DE RESULTADOS</h1>
                    
                    <div style="font-size: 18px; line-height: 1.8; max-width: 600px; margin: 0 auto;">
                        <p style="margin: 15px 0;"><strong>📂 Arquivo:</strong> {file_source}</p>
                        <p style="margin: 15px 0;"><strong>📅 Gerado em:</strong> ${{dateStr}}</p>
                        <p style="margin: 15px 0;"><strong>📅 Período de coleta:</strong> Calculado automaticamente</p>
                        <p style="margin: 15px 0;"><strong>👥 Respondentes:</strong> ${{formatNumberBR(totalRecords)}}</p>
                        <p style="margin: 15px 0;"><strong>📊 Variáveis analisadas:</strong> ${{formatNumberBR(totalVars)}}</p>
                        <p style="margin: 15px 0;"><strong>🔍 Filtros aplicados:</strong> ${{activeFilters.join(', ') || 'Nenhum filtro aplicado'}}</p>
                    </div>
                </div>
            `;
            
            return header;
        }}

        // ── MODAL DE ESCOLHA DO MODO PDF ──
        function showPDFModeDialog() {{
            return new Promise(resolve => {{
                const overlay = document.createElement('div');
                overlay.style.cssText = 'position:fixed;inset:0;background:rgba(0,0,0,0.35);z-index:999999;display:flex;align-items:center;justify-content:center;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;';

                const card = document.createElement('div');
                card.style.cssText = 'background:white;border-radius:14px;padding:32px 36px;max-width:520px;width:90%;box-shadow:0 8px 40px rgba(0,0,0,0.18);';

                card.innerHTML =
                    '<div style="font-size:11px;font-weight:600;color:#7BAFC0;text-transform:uppercase;letter-spacing:.07em;margin-bottom:8px;">Exportar PDF</div>' +
                    '<div style="font-size:17px;font-weight:500;color:#333;margin-bottom:6px;">Como exportar as tabelas?</div>' +
                    '<div style="font-size:13px;color:#888;margin-bottom:24px;">Tabelas longas (respostas abertas, muitos itens) podem ficar cortadas no modo dashboard.</div>' +
                    '<div style="display:flex;gap:12px;margin-bottom:20px;">' +
                      '<div id="pdf-opt-dashboard" style="flex:1;border:1.5px solid #7BAFC0;border-radius:10px;padding:16px;cursor:pointer;background:#EDF5F7;">' +
                        '<div style="font-size:13px;font-weight:600;color:#4d8a9e;margin-bottom:4px;">Modo dashboard</div>' +
                        '<div style="font-size:12px;color:#6c757d;line-height:1.5;">Mantém o layout atual. Tabelas longas ficam cortadas como no navegador.</div>' +
                      '</div>' +
                      '<div id="pdf-opt-expand" style="flex:1;border:1.5px solid #dde5ea;border-radius:10px;padding:16px;cursor:pointer;">' +
                        '<div style="font-size:13px;font-weight:600;color:#333;margin-bottom:4px;">Tabelas expandidas</div>' +
                        '<div style="font-size:12px;color:#6c757d;line-height:1.5;">Exibe todos os itens integralmente. Ideal para relatórios completos.</div>' +
                      '</div>' +
                    '</div>' +
                    '<div style="display:flex;justify-content:space-between;align-items:center;">' +
                      '<button id="pdf-cancel-btn" style="background:none;border:none;color:#aaa;font-size:13px;cursor:pointer;padding:0;">Cancelar</button>' +
                      '<button id="pdf-confirm-btn" style="background:#7BAFC0;color:white;border:none;border-radius:7px;padding:10px 24px;font-size:13px;font-weight:600;cursor:pointer;">Gerar PDF</button>' +
                    '</div>';

                overlay.appendChild(card);
                document.body.appendChild(overlay);

                let selectedMode = 'dashboard';

                const optDash = document.getElementById('pdf-opt-dashboard');
                const optExp  = document.getElementById('pdf-opt-expand');

                function selectMode(mode) {{
                    selectedMode = mode;
                    if (mode === 'dashboard') {{
                        optDash.style.border = '1.5px solid #7BAFC0';
                        optDash.style.background = '#EDF5F7';
                        optExp.style.border  = '1.5px solid #dde5ea';
                        optExp.style.background = 'white';
                    }} else {{
                        optExp.style.border  = '1.5px solid #7BAFC0';
                        optExp.style.background = '#EDF5F7';
                        optDash.style.border = '1.5px solid #dde5ea';
                        optDash.style.background = 'white';
                    }}
                }}

                optDash.onclick = () => selectMode('dashboard');
                optExp.onclick  = () => selectMode('expand');

                document.getElementById('pdf-confirm-btn').onclick = () => {{
                    overlay.remove();
                    resolve(selectedMode);
                }};
                document.getElementById('pdf-cancel-btn').onclick = () => {{
                    overlay.remove();
                    resolve(null);
                }};
            }});
        }}

        async function exportToPDF(event) {{
            const originalBtn = event ? event.target.closest('.icon-btn') : document.querySelector('.icon-btn.pdf');
            try {{
                // Mostrar modal de escolha do modo
                const pdfMode = await showPDFModeDialog();
                if (pdfMode === null) return; // cancelado

                originalBtn.classList.add('loading');
                
                console.log('=== EXPORTAÇÃO PDF INICIADA — modo:', pdfMode, '===');
                
                const contentEl = document.getElementById('content');
                const sections = contentEl.querySelectorAll('.section');
                
                if (sections.length === 0) {{
                    originalBtn.classList.remove('loading');
                    return;
                }}

                // ── EXPANSÃO DE TABELAS ──
                // Salvar e remover restrições de altura em todos os elementos com scroll
                const savedInlineStyles = [];
                if (pdfMode === 'expand') {{
                    document.querySelectorAll('#content *').forEach(el => {{
                        const s = el.style;
                        if (s.maxHeight || s.overflowY || s.overflow) {{
                            savedInlineStyles.push({{ el, maxHeight: s.maxHeight, overflow: s.overflow, overflowY: s.overflowY }});
                            s.maxHeight = 'none';
                            s.overflow  = 'visible';
                            s.overflowY = 'visible';
                        }}
                    }});
                    // Também via CSS para classes com overflow definido (.outros-list etc.)
                    const expandStyle = document.createElement('style');
                    expandStyle.id = 'pdf-expand-style';
                    expandStyle.textContent = '.outros-list {{ max-height: none !important; overflow: visible !important; }}';
                    document.head.appendChild(expandStyle);
                }}
                
                // Criar cabeçalho simples
                const headerDiv = document.createElement('div');
                headerDiv.className = 'pdf-header-temp';
                headerDiv.style.cssText = `
                    background: white;
                    padding: 0;
                    margin: 0;
                    font-family: Arial, sans-serif;
                    display: none;
                    print-color-adjust: exact;
                    -webkit-print-color-adjust: exact;
                    height: calc(100vh - 30mm);
                    align-items: center;
                    justify-content: center;
                `;
                
                // Informações do cabeçalho
                const now = new Date();
                const dateStr = now.toLocaleString('pt-BR');
                const totalRecords = RECORDS.length;
                const totalVars = VARS_META.length;
                const activeFilters = getActiveFiltersDescription();
                
                // === CALCULAR PERÍODO DE COLETA ===
                const periodoColeta = getPeriodoColeta();

                // N filtrado (quando há filtro ativo)
                const _pdfFiltRecs = getFilteredRecords(null);
                const _pdfNFilt = Math.round(_pdfFiltRecs.reduce((s,r) => s + (r.__weight__||1), 0));
                const _pdfHasFilter = Object.values(getSelectedFilters()).some(v => v.length > 0) || DRILLDOWN.size > 0;
                const _pdfNFiltLine = _pdfHasFilter
                    ? `<div style="margin: 12px 0;"><strong>👥 N filtrado:</strong> ${{formatNumberBR(_pdfNFilt)}}</div>`
                    : '';
                
                headerDiv.innerHTML = `
                    <div style="text-align: center; width: 100%;">
                        <h1 style="color: #7BAFC0; font-size: 28px; margin-bottom: 30px;">📋 RELATÓRIO DE RESULTADOS</h1>
                        
                        <div style="font-size: 16px; line-height: 1.8; max-width: 500px; margin: 0 auto;">
                            <div style="margin: 12px 0;"><strong>📂 Arquivo:</strong> {file_source}</div>
                            <div style="margin: 12px 0;"><strong>📅 Gerado em:</strong> ${{dateStr}}</div>
                            <div style="margin: 12px 0;"><strong>📅 Período de coleta:</strong> ${{periodoColeta}}</div>
                            <div style="margin: 12px 0;"><strong>👥 Respondentes:</strong> ${{formatNumberBR(totalRecords)}}</div>
                            ${{_pdfNFiltLine}}
                            <div style="margin: 12px 0;"><strong>📊 Variáveis analisadas:</strong> ${{formatNumberBR(totalVars)}}</div>
                            <div style="margin: 12px 0;"><strong>🔍 Filtros aplicados:</strong> ${{activeFilters.join('; ') || 'Nenhum filtro aplicado'}}</div>
                        </div>
                    </div>
                `;
                
                // Inserir cabeçalho no INÍCIO do content (não no body)
                const contentElement = document.getElementById('content');
                if (contentElement && contentElement.firstChild) {{
                    contentElement.insertBefore(headerDiv, contentElement.firstChild);
                }} else {{
                    contentElement.appendChild(headerDiv);
                }}

                // Criar estilo de impressão
                const printStyle = document.createElement('style');
                printStyle.id = 'pdf-print-style';
                printStyle.textContent = `
                    @media print {{
                        body {{ margin: 0; padding: 12mm 15mm; background: white; font-family: Arial, sans-serif; }}
                        .filters-container {{ display: none !important; }}
                        
                        /* CABEÇALHO */
                        .pdf-header-temp {{ 
                            display: flex !important; 
                            visibility: visible !important;
                            position: static !important;
                            margin: 0 !important;
                            padding: 0 !important;
                            page-break-after: always !important;
                            height: calc(100vh - 30mm) !important;
                            align-items: center !important;
                            justify-content: center !important;
                        }}
                        
                        /* SEÇÕES */
                        .section {{ margin-bottom: 16px; page-break-inside: avoid; border: 0.5px solid #ddd !important; border-radius: 6px; overflow: hidden; }}

                        /* HEADER DA SEÇÃO: remover fundo colorido, mostrar só borda sutil */
                        .section-header,
                        .section-header.mr-header,
                        .section-header.nps-header,
                        .section-header.open-header {{
                            background: white !important;
                            border-bottom: 0.5px solid #ddd !important;
                            padding: 10px 16px !important;
                        }}

                        /* TÍTULO: manter cor da variável mas sem underline azul */
                        .section-title {{
                            font-size: 13px !important;
                            font-weight: 600 !important;
                            margin: 0 0 4px 0 !important;
                            border-bottom: none !important;
                            padding-bottom: 0 !important;
                        }}
                        .section-subtitle {{ font-size: 11px !important; }}

                        /* TABELAS */
                        .ivv-table {{ font-size: 11px !important; }}
                        .ivv-table th {{ 
                            background: #f1f3f5 !important;
                            color: #444 !important;
                            font-weight: 600 !important;
                            padding: 6px 8px !important;
                            border-bottom: 1.5px solid #ccc !important;
                            -webkit-print-color-adjust: exact;
                        }}
                        .ivv-table td {{ 
                            padding: 5px 8px !important;
                            border-bottom: 0.5px solid #eee !important;
                            font-size: 11px !important;
                        }}
                        .ivv-table tr.ivv-total-row td {{
                            background: #f8f9fa !important;
                            border-top: 1.5px solid #ccc !important;
                            font-weight: 600 !important;
                            -webkit-print-color-adjust: exact;
                        }}
                        /* Barra de percentual: manter cor mas reduzir opacidade */
                        .ivv-bar-bg {{ opacity: 0.25 !important; -webkit-print-color-adjust: exact; }}

                        /* MÉTRICAS NPS/CSAT */
                        .scale-metrics {{ border-top: 1px solid #ddd !important; }}
                        .scale-metric-row {{ border-bottom: 0.5px solid #eee !important; }}

                        /* OCULTAR */
                        .chart-container, canvas, .ivv-drilldown-hint {{ display: none !important; }}
                        .ns-toggle-wrap {{ display: none !important; }}
                    }}
                `;
                
                document.head.appendChild(printStyle);
                
                // Aguardar um momento
                await new Promise(resolve => setTimeout(resolve, 500));
                
                // Sugerir nome do arquivo PDF baseado no título
                const originalTitle = document.title;
                const dataAtual = new Date().toLocaleDateString('pt-BR').split('/').join('-');
                const nomeArquivoLimpo = '{file_source}'.replace('.sav', '').replace('.SAV', '');
                const tituloPDF = `Relatorio de resultados_${{nomeArquivoLimpo}}_${{dataAtual}}`;
                document.title = tituloPDF;
                
                console.log('🖨️ Abrindo diálogo de impressão...');
                console.log('📄 Nome sugerido do arquivo:', tituloPDF);
                window.print();
                
                // Restaurar título original após impressão
                setTimeout(() => {{
                    document.title = originalTitle;
                }}, 2000);
                
                // Limpeza após 3 segundos
                setTimeout(() => {{
                    // Remover cabeçalho
                    document.querySelectorAll('.pdf-header-temp').forEach(el => {{
                        if (el.parentNode) el.parentNode.removeChild(el);
                    }});
                    
                    // Remover estilos de impressão
                    const style = document.getElementById('pdf-print-style');
                    if (style && style.parentNode) style.parentNode.removeChild(style);

                    // Restaurar inline styles expandidos
                    savedInlineStyles.forEach(({{ el, maxHeight, overflow, overflowY }}) => {{
                        el.style.maxHeight = maxHeight;
                        el.style.overflow  = overflow;
                        el.style.overflowY = overflowY;
                    }});
                    const expandStyle = document.getElementById('pdf-expand-style');
                    if (expandStyle) expandStyle.remove();
                    
                    console.log('🧹 Limpeza concluída');
                }}, 3000);
                
                originalBtn.classList.remove('loading');
                console.log('✅ Processo de PDF concluído com sucesso');
                
            }} catch (error) {{
                console.error('❌ Erro na exportação PDF:', error);
                const btn = event.target.closest('.icon-btn');
                if (btn) btn.classList.remove('loading');
                alert(`⚠️ Erro na exportação automática.\\n\\n` +
                      `💡 ALTERNATIVA MANUAL:\\n` +
                      `1. Pressione Ctrl+P (Cmd+P no Mac)\\n` +
                      `2. Escolha "Salvar como PDF"\\n` +
                      `3. Salve o arquivo\\n\\n` +
                      `Erro detalhado: ${{error.message}}`);
            }}
        }}
        
    </script>
</body>
</html>"""

# ========== INTERFACE GRÁFICA CORRIGIDA ==========

def run_gui() -> int:
    """Interface gráfica CORRIGIDA - exportselection=False é a chave"""
    # Spellcheck: deliberadamente desabilitado para textos importados.
    # Keywords são padronizadas via pt_BR.dic (lazy) em _correct_keyword_with_dic().
    if not GUI_AVAILABLE:
        print("❌ ERRO: tkinter não disponível!")
        print("🖥️ tkinter é necessário para a interface gráfica")
        return 1
    try:
        # 1. SELEÇÃO DO ARQUIVO
        root = tk.Tk()
        root.withdraw()
        
        in_path = filedialog.askopenfilename(
            title="Selecione o arquivo .sav (SPSS)",
            filetypes=[("SPSS files", "*.sav"), ("All files", "*.*")]
        )
        
        if not in_path:
            print("❌ Nenhum arquivo selecionado.")
            return 1
        
        print(f"📂 Carregando: {os.path.basename(in_path)}")
        
        try:
            df, meta = read_sav_auto(in_path)
            fix_labels_in_meta(meta)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar arquivo:\\n{str(e)}")
            return 2
        
        print(f"✅ Arquivo carregado: {len(df)} registros, {len(df.columns)} variáveis")
        
        # Obter labels das variáveis
        labels = {}
        for col in df.columns:
            label = get_var_label(meta, col)
            labels[col] = label if label else ""
        
        # Configurar estilo moderno para componentes ttk
        style = ttk.Style()
        style.theme_use('clam')
        
        # Estilo para combobox
        style.configure("Modern.TCombobox", 
                       fieldbackground="white",
                       background="#f8f9fa",
                       foreground="#2c3e50",
                       borderwidth=1,
                       relief="solid")
        
        # Configurar cores padrão para melhor visibilidade
        style.configure("TLabel", background="#f8f9fa", foreground="#2c3e50")
        style.configure("TFrame", background="#f8f9fa")
        
        # 2. JANELA DE SELEÇÃO - Layout moderno melhorado
        root.deiconify()
        root.title("📊 Dashboard SPSS Universal - Seleção de Variáveis")
        root.geometry("1400x800")
        root.minsize(1200, 700)
        root.configure(bg="#f8f9fa")
        
        # Configurar cores padrão para evitar problemas de sistema
        root.option_add('*TkDefaultFont', 'Segoe UI 10')
        root.option_add('*Background', '#f8f9fa')
        root.option_add('*Foreground', '#2c3e50')
        
        # Configurar grid weights para responsividade
        root.grid_columnconfigure(0, weight=1)
        root.grid_rowconfigure(0, weight=1)
        
        # Frame principal com melhor padding
        main_frame = tk.Frame(root, bg="#f8f9fa")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)
        
        # Header melhorado
        header_frame = tk.Frame(main_frame, bg="#f8f9fa")
        header_frame.pack(fill=tk.X, pady=(0, 25))
        
        # Título principal
        title_label = tk.Label(header_frame, 
                              text="📊 Dashboard SPSS Universal", 
                              font=("Segoe UI", 24, "bold"), 
                              fg="#2c3e50", bg="#f8f9fa")
        title_label.pack()
        
        # Subtítulo com informações do arquivo
        subtitle_label = tk.Label(header_frame,
                                 text=f"📁 {os.path.basename(in_path)} • {len(df):,} registros • {len(df.columns)} variáveis",
                                 font=("Segoe UI", 12), 
                                 fg="#7f8c8d", bg="#f8f9fa")
        subtitle_label.pack(pady=(5, 0))
        
        # Container principal para as listas
        content_frame = tk.Frame(main_frame, bg="#f8f9fa")
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame para listboxes lado a lado com melhor espaçamento
        lists_frame = tk.Frame(content_frame, bg="#f8f9fa")
        lists_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # VARIÁVEIS PRINCIPAIS (lado esquerdo) - Layout melhorado
        vars_frame = tk.LabelFrame(lists_frame, text="📊 VARIÁVEIS PARA O RELATÓRIO", 
                                  font=("Segoe UI", 14, "bold"), fg="#2980b9", bg="#f8f9fa",
                                  relief="solid", borderwidth=1, padx=15, pady=15)
        vars_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 15))
        
        # Informações melhoradas das variáveis principais
        vars_info = tk.Label(vars_frame, 
                            text="Selecione as variáveis para análise:\n• Ctrl/Cmd + clique: múltiplas seleções\n• Shift + clique: intervalos",
                            font=("Segoe UI", 11), fg="#5d6d7e", bg="#f8f9fa", justify=tk.LEFT)
        vars_info.pack(fill=tk.X, pady=(0, 15))
        
        # Container para listbox e scrollbar
        vars_list_frame = tk.Frame(vars_frame, bg="#f8f9fa")
        vars_list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Listbox de variáveis melhorada
        vars_listbox = tk.Listbox(vars_list_frame, selectmode=tk.EXTENDED, 
                                 font=("Consolas", 11), exportselection=False, 
                                 bg='white', fg="#2c3e50",
                                 selectbackground='#3498db', selectforeground='white',
                                 relief="solid", borderwidth=1, highlightthickness=0)
        vars_scrollbar = tk.Scrollbar(vars_list_frame, orient=tk.VERTICAL, command=vars_listbox.yview)
        vars_listbox.config(yscrollcommand=vars_scrollbar.set)
        
        vars_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        vars_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Botões de controle melhorados para variáveis
        vars_buttons_frame = tk.Frame(vars_frame, bg="#f8f9fa")
        vars_buttons_frame.pack(fill=tk.X)
        
        # Botão Selecionar Todas
        select_all_btn = tk.Button(vars_buttons_frame, text="✅ Selecionar Todas", 
                                  command=lambda: vars_listbox.select_set(0, tk.END),
                                  font=("Segoe UI", 10, "bold"), bg="#1e8449", fg="#ffffff",
                                  relief="flat", padx=15, pady=8, cursor="hand2")
        select_all_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Efeitos hover para botão Selecionar Todas
        def on_select_all_enter(e): select_all_btn.configure(bg="#239b56")
        def on_select_all_leave(e): select_all_btn.configure(bg="#1e8449")
        select_all_btn.bind("<Enter>", on_select_all_enter)
        select_all_btn.bind("<Leave>", on_select_all_leave)
        
        # Botão Limpar
        clear_btn = tk.Button(vars_buttons_frame, text="❌ Limpar", 
                             command=lambda: vars_listbox.selection_clear(0, tk.END),
                             font=("Segoe UI", 10, "bold"), bg="#c0392b", fg="#ffffff",
                             relief="flat", padx=15, pady=8, cursor="hand2")
        clear_btn.pack(side=tk.LEFT)
        
        # Efeitos hover para botão Limpar
        def on_clear_enter(e): clear_btn.configure(bg="#a93226")
        def on_clear_leave(e): clear_btn.configure(bg="#c0392b")
        clear_btn.bind("<Enter>", on_clear_enter)
        clear_btn.bind("<Leave>", on_clear_leave)
        
        # FILTROS (lado direito) - Layout melhorado
        filters_frame = tk.LabelFrame(lists_frame, text="🔍 VARIÁVEIS-FILTRO (Opcional)", 
                                     font=("Segoe UI", 14, "bold"), fg="#8e44ad", bg="#f8f9fa",
                                     relief="solid", borderwidth=1, padx=15, pady=15)
        filters_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Informações melhoradas dos filtros
        filters_info = tk.Label(filters_frame, 
                               text="Filtros para segmentação:\n• Opcional (pode deixar vazio)\n• Útil para análises específicas",
                               font=("Segoe UI", 11), fg="#5d6d7e", bg="#f8f9fa", justify=tk.LEFT)
        filters_info.pack(fill=tk.X, pady=(0, 15))
        
        # Container para listbox de filtros
        filters_list_frame = tk.Frame(filters_frame, bg="#f8f9fa")
        filters_list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Listbox de filtros melhorada
        filters_listbox = tk.Listbox(filters_list_frame, selectmode=tk.EXTENDED, 
                                    font=("Consolas", 11), exportselection=False,
                                    bg='white', fg="#2c3e50",
                                    selectbackground='#9b59b6', selectforeground='white',
                                    relief="solid", borderwidth=1, highlightthickness=0)
        filters_scrollbar = tk.Scrollbar(filters_list_frame, orient=tk.VERTICAL, command=filters_listbox.yview)
        filters_listbox.config(yscrollcommand=filters_scrollbar.set)
        
        filters_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        filters_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Botões de controle melhorados para filtros
        filters_buttons_frame = tk.Frame(filters_frame, bg="#f8f9fa")
        filters_buttons_frame.pack(fill=tk.X)
        
        # Botão Selecionar Todas (filtros)
        filters_select_all_btn = tk.Button(filters_buttons_frame, text="✅ Selecionar Todas", 
                                          command=lambda: filters_listbox.select_set(0, tk.END),
                                          font=("Segoe UI", 10, "bold"), bg="#1e8449", fg="#ffffff",
                                          relief="flat", padx=15, pady=8, cursor="hand2")
        filters_select_all_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Efeitos hover para botão Selecionar Todas (filtros)
        def on_filters_select_enter(e): filters_select_all_btn.configure(bg="#239b56")
        def on_filters_select_leave(e): filters_select_all_btn.configure(bg="#1e8449")
        filters_select_all_btn.bind("<Enter>", on_filters_select_enter)
        filters_select_all_btn.bind("<Leave>", on_filters_select_leave)
        
        # Botão Limpar (filtros)
        filters_clear_btn = tk.Button(filters_buttons_frame, text="❌ Limpar", 
                                     command=lambda: filters_listbox.selection_clear(0, tk.END),
                                     font=("Segoe UI", 10, "bold"), bg="#c0392b", fg="#ffffff",
                                     relief="flat", padx=15, pady=8, cursor="hand2")
        filters_clear_btn.pack(side=tk.LEFT)
        
        # Efeitos hover para botão Limpar (filtros)
        def on_filters_clear_enter(e): filters_clear_btn.configure(bg="#a93226")
        def on_filters_clear_leave(e): filters_clear_btn.configure(bg="#c0392b")
        filters_clear_btn.bind("<Enter>", on_filters_clear_enter)
        filters_clear_btn.bind("<Leave>", on_filters_clear_leave)
        
        # PESO/PONDERAÇÃO (nova seção melhorada abaixo dos filtros)
        weight_frame = tk.LabelFrame(filters_frame, text="⚖️ VARIÁVEL PESO (Opcional)", 
                                   font=("Segoe UI", 13, "bold"), fg="#e67e22", bg="#f8f9fa",
                                   relief="solid", borderwidth=1, padx=15, pady=10)
        weight_frame.pack(fill=tk.X, pady=(20, 0))
        
        weight_info = tk.Label(weight_frame, 
                             text="Para pesquisas por amostragem:",
                             font=("Segoe UI", 10), fg="#5d6d7e", bg="#f8f9fa")
        weight_info.pack(fill=tk.X, pady=(0, 8))
        
        # Combobox melhorado para seleção de peso
        weight_var = tk.StringVar()
        weight_combo = ttk.Combobox(weight_frame, textvariable=weight_var, 
                                  font=("Segoe UI", 11), state="readonly", width=30,
                                  style="Modern.TCombobox")
        weight_combo.pack(fill=tk.X)
        
        # POPULAR AS LISTAS COM VARIÁVEIS (preservando ordem original do SPSS)
        print(f"🔧 Preservando ordem original das {len(df.columns)} variáveis do SPSS")
        for col in df.columns:  # REMOVIDO sorted() para preservar ordem SPSS
            label_text = labels.get(col, "")
            if label_text:
                display_text = f"{col:<15} | {label_text}"
            else:
                display_text = f"{col:<15} | (sem rótulo)"
            
            vars_listbox.insert(tk.END, display_text)
            filters_listbox.insert(tk.END, display_text)
        
        # Popular combobox de peso apenas com variáveis numéricas candidatas
        weight_candidates = ["(Nenhuma - sem ponderação)"]
        for col in df.columns:
            # Detectar se é variável numérica (candidata a peso)
            if col.lower() in ['peso', 'weight', 'pond', 'ponderacao', 'factor', 'wgt']:
                weight_candidates.append(f"{col} | {labels.get(col, '(peso)')}")
            elif df[col].dtype in ['int64', 'float64'] or pd.api.types.is_numeric_dtype(df[col]):
                # Verificar se parece com peso (valores entre 0.1 e 10, média próxima de 1)
                numeric_vals = pd.to_numeric(df[col], errors='coerce').dropna()
                if len(numeric_vals) > 0:
                    mean_val = numeric_vals.mean()
                    min_val = numeric_vals.min()
                    max_val = numeric_vals.max()
                    if 0.1 <= min_val and max_val <= 20 and 0.5 <= mean_val <= 3.0:
                        weight_candidates.append(f"{col} | {labels.get(col, '(numérica)')}")
        
        weight_combo['values'] = weight_candidates
        weight_combo.current(0)  # Seleciona "Nenhuma" por padrão
        
        # Variáveis para armazenar seleções
        selected_vars = []
        selected_filters = []
        selected_weight = None
        selected_description = ""
        selected_period = ""
        success = False
        
        def on_generate():
            nonlocal selected_vars, selected_filters, selected_weight, selected_description, selected_period, success
            
            # Obter seleções
            var_indices = vars_listbox.curselection()
            filter_indices = filters_listbox.curselection()
            
            if not var_indices:
                messagebox.showwarning("Atenção", "Selecione pelo menos uma variável para o relatório!")
                return
            
            # Preservar ordem original do SPSS (REMOVIDO sorted())
            columns_list = list(df.columns)  # Ordem original preservada
            selected_vars = [columns_list[i] for i in var_indices]
            selected_filters = [columns_list[i] for i in filter_indices]
            
            # Obter variável peso selecionada
            weight_selection = weight_var.get()
            if weight_selection and not weight_selection.startswith("(Nenhuma"):
                # Extrair nome da variável do formato "PESO | descrição"
                selected_weight = weight_selection.split(" | ")[0]
                if selected_weight not in df.columns:
                    selected_weight = None
            else:
                selected_weight = None
            
            # Capturar título/descrição da pesquisa
            selected_description = desc_entry_var.get().strip()
            selected_period = period_entry_var.get().strip()
            
            # Capturar período de coleta
            selected_period = period_entry_var.get().strip()
            
            success = True
            root.quit()
        
        def on_cancel():
            nonlocal success
            success = False
            root.quit()
        
        # CAMPOS TÍTULO + PERÍODO — lado a lado
        title_section = tk.Frame(main_frame, bg="#f8f9fa")
        title_section.pack(fill=tk.X, pady=(20, 0))

        # Coluna esquerda: Título (flex maior)
        title_field_frame = tk.LabelFrame(title_section,
            text="📝 Título da Pesquisa",
            font=("Segoe UI", 11, "bold"),
            fg="#4d8a9e", bg="#f8f9fa", bd=1, relief="groove")
        title_field_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, ipady=6, padx=(0, 8))

        tk.Label(title_field_frame,
            text="Aparece na capa do modo apresentação:",
            font=("Segoe UI", 9), fg="#888", bg="#f8f9fa").pack(anchor=tk.W, padx=10, pady=(4, 2))

        desc_entry_var = tk.StringVar()
        # Pré-preencher se encontrar no config
        try:
            import json as _json2, re as _re2
            def _kw2(n):
                stops = {'de','da','do','e','a','o','base','ponderada','html','sav'}
                return {w for w in _re2.findall(r'[a-zA-Z\u00C0-\u00FF]{3,}', n.lower()) if w not in stops}
            for _cp2 in ['dashboard_overlay_config.json',
                         os.path.join(os.path.dirname(os.path.abspath(in_path)), 'dashboard_overlay_config.json'),
                         os.path.join(os.getcwd(), 'dashboard_overlay_config.json')]:
                if os.path.exists(_cp2):
                    with open(_cp2, encoding='utf-8') as _f2:
                        _cfg2 = _json2.load(_f2)
                    _sw2 = _kw2(os.path.basename(in_path))
                    def _fscan2(items):
                        for item in items:
                            fw = _kw2(item.get('file', ''))
                            if fw and _sw2 and len(fw & _sw2)/max(len(fw),len(_sw2)) >= 0.4:
                                d = item.get('description','')
                                if d: return d
                            if 'children' in item:
                                r = _fscan2(item['children'])
                                if r: return r
                        return ''
                    _pre2 = _fscan2(_cfg2.get('items', []))
                    if _pre2:
                        desc_entry_var.set(_pre2)
                    break
        except:
            pass

        desc_entry = tk.Entry(title_field_frame, textvariable=desc_entry_var,
            font=("Segoe UI", 11), bg="white", fg="#333",
            relief="flat", bd=0, highlightthickness=1,
            highlightbackground="#ccc", highlightcolor="#4d8a9e")
        desc_entry.pack(fill=tk.X, padx=10, pady=(0, 6), ipady=6)

        # Coluna direita: Período de Coleta (largura fixa)
        period_field_frame = tk.LabelFrame(title_section,
            text="📅 Período de Coleta",
            font=("Segoe UI", 11, "bold"),
            fg="#4d8a9e", bg="#f8f9fa", bd=1, relief="groove", width=220)
        period_field_frame.pack(side=tk.LEFT, fill=tk.Y, ipady=6)
        period_field_frame.pack_propagate(False)

        tk.Label(period_field_frame,
            text="Ex: mar/2026",
            font=("Segoe UI", 9), fg="#888", bg="#f8f9fa").pack(anchor=tk.W, padx=10, pady=(4, 2))

        period_entry_var = tk.StringVar()
        period_entry = tk.Entry(period_field_frame, textvariable=period_entry_var,
            font=("Segoe UI", 11), bg="white", fg="#333",
            relief="flat", bd=0, highlightthickness=1,
            highlightbackground="#ccc", highlightcolor="#4d8a9e")
        period_entry.pack(fill=tk.X, padx=10, pady=(0, 6), ipady=6)

        # BOTÕES FINAIS - Layout moderno
        buttons_section = tk.Frame(main_frame, bg="#f8f9fa")
        buttons_section.pack(fill=tk.X, pady=(30, 0))
        
        # Frame para centralizar botões
        buttons_frame = tk.Frame(buttons_section, bg="#f8f9fa")
        buttons_frame.pack(anchor=tk.CENTER)
        
        # Botão Cancelar melhorado
        cancel_btn = tk.Button(buttons_frame, text="❌ Cancelar", command=on_cancel, 
                              font=("Segoe UI", 12, "bold"), width=15, 
                              bg="#bdc3c7", fg="#2c3e50", relief="flat",
                              padx=20, pady=12, cursor="hand2")
        cancel_btn.pack(side=tk.LEFT, padx=(0, 20))
        
        # Efeitos hover para botão Cancelar
        def on_cancel_enter(e): cancel_btn.configure(bg="#95a5a6")
        def on_cancel_leave(e): cancel_btn.configure(bg="#bdc3c7")
        cancel_btn.bind("<Enter>", on_cancel_enter)
        cancel_btn.bind("<Leave>", on_cancel_leave)
        
        # Botão Gerar melhorado
        generate_btn = tk.Button(buttons_frame, text="🚀 Gerar Dashboard", command=on_generate, 
                                font=("Segoe UI", 12, "bold"), width=20, 
                                bg="#1e8449", fg="#ffffff", relief="flat",
                                padx=25, pady=12, cursor="hand2")
        generate_btn.pack(side=tk.RIGHT)
        
        # Efeitos hover para botão Gerar
        def on_generate_enter(e): generate_btn.configure(bg="#239b56")
        def on_generate_leave(e): generate_btn.configure(bg="#1e8449")
        generate_btn.bind("<Enter>", on_generate_enter)
        generate_btn.bind("<Leave>", on_generate_leave)
        
        # Instruções melhoradas
        instructions_frame = tk.Frame(main_frame, bg="#f8f9fa")
        instructions_frame.pack(fill=tk.X, pady=(20, 10))
        
        instructions_title = tk.Label(instructions_frame, 
                                     text="💡 INSTRUÇÕES DE USO",
                                     font=("Segoe UI", 12, "bold"), 
                                     fg="#34495e", bg="#f8f9fa")
        instructions_title.pack(anchor=tk.W)
        
        instructions_text = tk.Label(instructions_frame, 
                                    text="• Clique simples: seleciona um item\n"
                                         "• Ctrl/Cmd + clique: múltiplas seleções\n"
                                         "• Shift + clique: seleciona intervalo\n"
                                         "• Use os botões para facilitar a seleção", 
                                    font=("Segoe UI", 10), fg="#7f8c8d", bg="#f8f9fa", 
                                    justify=tk.LEFT)
        instructions_text.pack(anchor=tk.W, pady=(5, 0))
        
        # Executar interface
        root.mainloop()
        
        if not success:
            root.destroy()
            print("❌ Operação cancelada.")
            return 1
        
        print(f"✅ Variáveis selecionadas: {len(selected_vars)} - {selected_vars[:3]}{'...' if len(selected_vars) > 3 else ''}")
        print(f"✅ Filtros selecionados: {len(selected_filters)} - {selected_filters[:3] if selected_filters else 'Nenhum'}")
        
        # DEBUG: Mostrar detalhes das variáveis selecionadas
        print("\n🔍 === DEBUG: VARIÁVEIS SELECIONADAS ===")
        mr_candidates = []
        single_vars = []
        
        for var in selected_vars:
            if "_" in var and re.match(r'^[A-Za-z]+\d+_\d+', var):
                mr_candidates.append(var)
            else:
                single_vars.append(var)
        
        print(f"📊 Variáveis com padrão MR: {len(mr_candidates)}")
        if mr_candidates:
            for var in mr_candidates[:10]:
                print(f"   • {var}")
            if len(mr_candidates) > 10:
                print(f"   ... e mais {len(mr_candidates) - 10}")
        
        print(f"📋 Variáveis individuais: {len(single_vars)}")
        if single_vars:
            for var in single_vars[:10]:
                print(f"   • {var}")
            if len(single_vars) > 10:
                print(f"   ... e mais {len(single_vars) - 10}")
        print()
        
        root.destroy()
        
        # 3. ARQUIVO DE SAÍDA
        root2 = tk.Tk()
        root2.withdraw()
        
        default_out = os.path.splitext(in_path)[0] + "_dashboard_8_1.html"
        out_path = filedialog.asksaveasfilename(
            title="Salvar dashboard HTML como...",
            defaultextension=".html", 
            initialfile=os.path.basename(default_out),
            filetypes=[("HTML", "*.html")]
        ) or default_out
        
        root2.destroy()

        # 4. PROCESSAMENTO
        print("⚙️ Processando dados...")
        created_at, vars_meta, filters_meta, records, value_orders, code_to_label = build_records_and_meta(
            df, meta, selected_vars, selected_filters, os.path.basename(in_path), "", selected_weight
        )

        print("🎨 Gerando HTML universal...")
        html = render_html_with_working_filters(
            os.path.basename(in_path), created_at, "",
            vars_meta, filters_meta, records, value_orders, code_to_label,
            output_html_name=out_path,
            user_description=selected_description,
            user_period=selected_period
        )
        
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)

        # 5. RESULTADO
        mr_found = [v for v in vars_meta if v["type"] == "mr"]
        string_found = [v for v in vars_meta if v["type"] == "string"]
        
        result_msg = f"""✅ Dashboard criado com sucesso!

• Registros: {len(records)}
• Variáveis analisadas: {len(vars_meta)}
• Filtros: {len(filters_meta)}
• Arquivo gerado: {os.path.basename(out_path)}
"""
        
        # Informar sobre ponderação
        if selected_weight:
            result_msg += f"⚖️ Ponderação aplicada: {selected_weight}\\n"
        
        # Adiciona informações resumidas sobre tipos especiais de variáveis
        special_vars = []
        mr_count = len([v for v in vars_meta if v["type"] == "mr"])
        string_count = len([v for v in vars_meta if v["type"] == "string"])
        
        if string_count > 0:
            special_vars.append(f"🟣 {string_count} Respostas Abertas")
        if mr_count > 0:
            special_vars.append(f"🟠 {mr_count} Respostas Múltiplas")
            
        if special_vars:
            result_msg += f"\n{' | '.join(special_vars)}"

        root3 = tk.Tk()
        root3.withdraw()
        messagebox.showinfo("Dashboard Universal - Concluído", result_msg)
        root3.destroy()
        return 0

    except Exception as e:
        try:
            messagebox.showerror("Erro", f"Erro inesperado:\\n\\n{str(e)}")
        except Exception:
            print(f"Erro: {e}", file=sys.stderr)
        finally:
            try: 
                root.destroy()
            except Exception: 
                pass
        return 4

# ========== LINHA DE COMANDO ==========

def run_cli() -> int:
    import argparse
    p = argparse.ArgumentParser(description="Dashboard SPSS Universal")
    p.add_argument("input", help="Caminho do arquivo .sav")
    p.add_argument("--vars", type=str, required=True, help="Variáveis do relatório separadas por vírgula")
    p.add_argument("--filters", type=str, default="", help="Variáveis-filtro separadas por vírgula")
    p.add_argument("--cliente", type=str, default="", help="Nome do cliente para o título")
    p.add_argument("-o", "--output", default=None, help="HTML de saída")
    args = p.parse_args()

    try:
        # Spellcheck: deliberadamente desabilitado para textos importados.
        # Keywords são padronizadas via pt_BR.dic (lazy) em _correct_keyword_with_dic().
        df, meta = read_sav_auto(args.input)
        fix_labels_in_meta(meta)
        
        selected_vars = [v.strip() for v in args.vars.split(",") if v.strip()]
        filter_vars = [v.strip() for v in args.filters.split(",") if v.strip()] if args.filters else []
        
        out_path = args.output or os.path.splitext(args.input)[0] + "_dashboard_8_1.html"
        
        created_at, vars_meta, filters_meta, records, value_orders, code_to_label = build_records_and_meta(
            df, meta, selected_vars, filter_vars, os.path.basename(args.input), args.cliente, None
        )

        html = render_html_with_working_filters(
            os.path.basename(args.input), created_at, args.cliente,
            vars_meta, filters_meta, records, value_orders, code_to_label,
            output_html_name=out_path
        )
        
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)

        print(f"✅ Dashboard universal criado: {out_path}")
        return 0
        
    except Exception as e:
        print(f"❌ Erro: {e}", file=sys.stderr)
        return 1

# ========== MAIN ==========

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        sys.exit(run_cli())
    else:
        sys.exit(run_gui())
